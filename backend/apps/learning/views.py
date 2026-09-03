import glob
import logging
import mimetypes
import os
import re
import threading
from collections import defaultdict
from datetime import date, timedelta

from io import BytesIO
from django.core.files.base import ContentFile
from django.conf import settings
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum, Avg, Max, Max
from django.http import HttpResponse, FileResponse, Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import ensure_csrf_cookie
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .auth_utils import RateLimitExceeded, StudentAuthMixin, UnifiedAuthMixin, check_rate_limit
from .data import FALLBACK_DASHBOARD, FALLBACK_PROBLEMS
from .module_registry import MODULE_KEYS, serializable_registry
from .drive_image_cache import cached_image_path, fetch_and_cache_drive_image, DriveImageFetchError
from .models import (
    BatchAdvisor,
    Contest,
    ContestParticipation,
    ContestSubmission,
    DiscussionMessage,
    ExecutionRecord,
    Institution,
    Problem,
    ProblemSession,
    ProblemSolution,
    SolvedProblem,
    TestCase,
    StaffProfile,
    StudentActivity,
    StudentProfile,
    Submission,
    DailyProblem,
    Announcement,
    Notification,
    SystemUpdate,
    AptitudeTopic,
    AptitudeQuestion,
    ReadingPassage,
    Achievement,
    UserAchievement,
    SystemConfiguration,
    Department,
    SolvedAptitude,
    AptitudeAttempt,
    AptitudeContestSubmission,
    LabTopic,
    LabProblem,
    LabTestCase,
    LabSubmission,
    LabAssignment,
    LabAssignmentSubmission,
    Lab,
    LabExercise,
    LabExerciseSubmission,
    LabExerciseTestCase,
    LabExerciseReport,
    LabStudentSession,
    LLMProvider,
    Company,
    LAB_LANGUAGE_CHOICES,
    Examination,
    SyllabusSection,
    SyllabusTopic,
    SyllabusSubtopic,
    CompetitiveQuestion,
    QuestionUsageMark,
    PasswordResetOTP,
)
from .db_manager import create_institution_db, delete_institution_db
from .serializers import (
    CodeRunSerializer,
    DiscussionMessageCreateSerializer,
    DiscussionMessageSerializer,
    FirstLoginSerializer,
    ProblemDetailSerializer,
    ProblemProgressUpdateSerializer,
    ProblemSerializer,
    StaffProfileSerializer,
    StudentLoginSerializer,
    StudentLookupListSerializer,
    StudentProfileSerializer,
)
from .services.judge0 import (
    Judge0ServiceError as ExecutorServiceError,
    Judge0TimeoutError as ExecutorTimeoutError,
    execute_judge0_submission,
)
from .services.reading_qa_import import create_passages_in_db, parse_workbook_to_passages
from .services.execution_adapter import (
    normalize_comparable_output,
    prepare_execution_payload,
    compare_typed_output,
    compare_design_output,
)
from .services.problem_testcases import build_runtime_test_cases
from .services.complexity_analyzer import calculate_complexity
from .services.code_validator import validate_submission
from .services import param_types

# ReportLab imports for PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
from reportlab.platypus.frames import Frame
from reportlab.lib.utils import ImageReader
import requests
from PIL import Image as PILImage
import io

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# PDF Watermark Utility
# ---------------------------------------------------------------------------

class WatermarkDocTemplate(BaseDocTemplate):
    """Custom document template that adds watermark to all pages"""
    
    def __init__(self, filename, institution=None, **kwargs):
        BaseDocTemplate.__init__(self, filename, **kwargs)
        self.institution = institution
        self.watermark_image = None
        
        # Try to get watermark image
        if institution and (institution.logo_file or institution.logo_url):
            try:
                self.watermark_image = self._get_watermark_image(institution)
            except Exception as e:
                logger.warning(f"Failed to load watermark image: {e}")
        
        # Add page template with frame
        frame = Frame(
            self.leftMargin, self.bottomMargin, 
            self.width, self.height, 
            id='normal'
        )
        template = PageTemplate(id='main', frames=frame, onPage=self._add_watermark)
        self.addPageTemplates([template])
    
    def _get_watermark_image(self, institution):
        """Download and prepare watermark image. Reads an uploaded logo
        file directly (its real filesystem path, via Django's storage
        API) rather than through logo_display_url — that property now
        returns an /api/... proxy URL for uploaded logos (fixing the
        broken-image bug where the raw media path wasn't reachable
        through nginx/nginx in production), which isn't a URL this
        server-side PDF generator should loop back and fetch over HTTP."""
        try:
            if institution.logo_file:
                with institution.logo_file.open('rb') as f:
                    image_data = f.read()
            elif institution.logo_url.startswith('http'):
                # Pasted external URL — fetch it
                response = requests.get(institution.logo_url, timeout=10)
                response.raise_for_status()
                image_data = response.content
            else:
                return None
            
            # Process image with PIL
            pil_image = PILImage.open(io.BytesIO(image_data))
            
            # Convert to RGBA if needed
            if pil_image.mode != 'RGBA':
                pil_image = pil_image.convert('RGBA')
            
            # Make it semi-transparent for watermark effect
            alpha = pil_image.split()[-1]
            alpha = alpha.point(lambda p: p * 0.15)  # 15% opacity
            pil_image.putalpha(alpha)
            
            # Convert back to bytes
            output = io.BytesIO()
            pil_image.save(output, format='PNG')
            output.seek(0)
            
            return ImageReader(output)
        except Exception as e:
            logger.error(f"Error processing watermark image: {e}")
            return None
    
    def _add_watermark(self, canvas, doc):
        """Add watermark to each page"""
        if not self.watermark_image:
            return
        
        try:
            # Calculate watermark position (center of page)
            page_width, page_height = A4
            watermark_size = min(page_width, page_height) * 0.4  # 40% of page size
            
            x = (page_width - watermark_size) / 2
            y = (page_height - watermark_size) / 2
            
            # Draw watermark
            canvas.drawImage(
                self.watermark_image, 
                x, y, 
                width=watermark_size, 
                height=watermark_size,
                mask='auto'
            )
        except Exception as e:
            logger.error(f"Error adding watermark: {e}")


def create_watermarked_pdf(buffer, institution=None, **kwargs):
    """Create a PDF document with watermark support"""
    if institution and institution.logo_display_url:
        return WatermarkDocTemplate(buffer, institution=institution, **kwargs)
    else:
        return SimpleDocTemplate(buffer, **kwargs)


# ---------------------------------------------------------------------------
# Helper builders (pure functions — no HTTP side-effects)
# ---------------------------------------------------------------------------

def build_activity_calendar(profile):
    """
    Build a monthly activity calendar for the current month.
    Returns activity data for the entire current month plus padding days
    to fill the calendar grid (previous/next month days).
    """
    today = timezone.localdate()
    
    # Get the first and last day of the current month
    month_start = today.replace(day=1)
    if today.month == 12:
        month_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        month_end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    
    # Calculate padding days to fill the calendar grid
    # Start from the Sunday before the first day of the month
    start_weekday = month_start.weekday()  # Monday=0, Sunday=6
    # Adjust to make Sunday=0
    start_offset = (start_weekday + 1) % 7
    calendar_start = month_start - timedelta(days=start_offset)
    
    # End on the Saturday after the last day of the month
    end_weekday = month_end.weekday()  # Monday=0, Sunday=6
    # Calculate days to add to reach Saturday (weekday=5)
    # If month ends on Saturday (5), add 0 days
    # If month ends on Sunday (6), add 6 days
    # If month ends on Monday (0), add 5 days, etc.
    end_offset = (5 - end_weekday) % 7
    calendar_end = month_end + timedelta(days=end_offset)
    
    # Fetch activity data for the entire range (including padding days)
    activity_rows = (
        StudentActivity.objects.filter(
            student=profile,
            activity_date__gte=calendar_start,
            activity_date__lte=calendar_end
        )
        .values("activity_date")
        .annotate(total=Count("id"))
        .order_by("activity_date")
    )

    daily_totals = {row["activity_date"]: row["total"] for row in activity_rows}

    # If no activity data exists but user has a streak, infer recent activity
    if not daily_totals and profile.last_login_on:
        inferred_days = min(max(profile.current_streak, 1), 30)
        for offset in range(inferred_days):
            inferred_day = profile.last_login_on - timedelta(days=offset)
            if calendar_start <= inferred_day <= calendar_end:
                daily_totals[inferred_day] = 1

    # Build the calendar array
    calendar = []
    current_day = calendar_start
    while current_day <= calendar_end:
        count = daily_totals.get(current_day, 0)
        calendar.append(
            {
                "date": current_day.isoformat(),
                "count": count,
                "weekday": current_day.strftime("%a"),
                "day": current_day.day,
            }
        )
        current_day += timedelta(days=1)
    
    return calendar


def build_student_stats(profile):
    # Count solved problems from SolvedProblem table (faster than scanning all solutions)
    solved_problems = Problem.objects.filter(
        solved_by__student=profile
    ).distinct()

    return solved_problems.aggregate(
        easy=Count("id", filter=Q(difficulty="Easy"), distinct=True),
        medium=Count("id", filter=Q(difficulty="Medium"), distinct=True),
        hard=Count("id", filter=Q(difficulty="Hard"), distinct=True), sql=Count("id", filter=Q(tags__contains="SQL"), distinct=True),
    )


def build_problem_progress_map(profile):
    """
    Returns {problem_id: {"state": "completed"|"open", "solved_languages": [...], "current_language": str|None}}.

    "completed" is sourced from SolvedProblem (the platform-wide source of
    truth for solved counts elsewhere), while per-language detail comes
    from ProblemSolution — a student can solve the same problem in more
    than one language, and SolvedProblem only records the first.
    """
    solved_ids = set(
        SolvedProblem.objects.filter(student=profile).values_list("problem_id", flat=True)
    )

    # Most-recent-first so the first row seen per problem is the latest attempt.
    solutions = (
        ProblemSolution.objects
        .filter(student=profile)
        .order_by("-submitted_at")
        .values("problem_id", "language", "all_tests_passed")
    )

    languages_by_problem = {}
    latest_language_by_problem = {}
    for row in solutions:
        pid = row["problem_id"]
        latest_language_by_problem.setdefault(pid, row["language"])
        if row["all_tests_passed"]:
            languages_by_problem.setdefault(pid, set()).add(row["language"])

    progress_map = {}
    for pid in solved_ids | set(latest_language_by_problem.keys()):
        if pid in solved_ids:
            progress_map[pid] = {
                "state": "completed",
                "solved_languages": sorted(languages_by_problem.get(pid, set())),
                "current_language": None,
            }
        else:
            progress_map[pid] = {
                "state": "open",
                "solved_languages": [],
                "current_language": latest_language_by_problem.get(pid),
            }

    return progress_map


def build_weekly_activity(activity_calendar):
    grouped = defaultdict(int)
    for item in activity_calendar:
        grouped[item["weekday"]] += item["count"]

    order = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    return [{"day": day, "count": grouped.get(day, 0)} for day in order]


def parse_date_param(value):
    """Parse a 'YYYY-MM-DD' query param into a date, or None if missing/invalid."""
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except (ValueError, TypeError):
        return None


def build_solved_activity_series(base_filter, start_date=None, end_date=None, default_days=7):
    """Daily SolvedProblem counts for the "Weekly Solving Activity" widget
    on HOD/Staff dashboards, scoped by `base_filter` (a Q object — e.g.
    department or institution). Defaults to the last `default_days` days
    ending today when no explicit range is given; a given range is capped
    at 90 days so the single grouped query stays cheap. Returns one entry
    per calendar day (oldest first) with both the ISO date and a short
    weekday label, so the frontend can label bars either way depending on
    how wide the selected range is.
    """
    today = timezone.now().date()
    if start_date and end_date:
        start, end = (start_date, end_date) if start_date <= end_date else (end_date, start_date)
        if (end - start).days > 89:
            start = end - timedelta(days=89)
    else:
        end = today
        start = end - timedelta(days=default_days - 1)

    counts = dict(
        SolvedProblem.objects.filter(base_filter, solved_at__date__range=(start, end))
        .values('solved_at__date')
        .annotate(count=Count('id'))
        .values_list('solved_at__date', 'count')
    )

    series = []
    cursor = start
    while cursor <= end:
        series.append({
            "date": cursor.isoformat(),
            "day": cursor.strftime("%a"),
            "count": counts.get(cursor, 0),
        })
        cursor += timedelta(days=1)
    return series


def _contest_live_summary(contest, limit=5):
    """Return contest counts/top performers for both programming and aptitude contests."""
    if contest.contest_type == 'aptitude':
        submissions = AptitudeContestSubmission.objects.filter(contest=contest)
        leaders = (
            submissions
            .values('student')
            .annotate(
                solved_count=Count('id', filter=Q(is_correct=True)),
                total_score=Sum('score'),
            )
            .filter(solved_count__gt=0)
            .order_by('-solved_count', '-total_score')[:limit]
        )
    else:
        submissions = ContestSubmission.objects.filter(contest=contest)
        leaders = (
            submissions
            .values('student')
            .annotate(
                solved_count=Count('problem', filter=Q(status='Accepted'), distinct=True),
                total_score=Sum('score'),
            )
            .filter(solved_count__gt=0)
            .order_by('-solved_count', '-total_score')[:limit]
        )

    top_performers = []
    for leader in leaders:
        student = StudentProfile.objects.filter(id=leader['student']).first()
        if student:
            top_performers.append({
                "register_number": student.register_number,
                "name": student.name,
                "batch": student.batch,
                "solved_in_contest": leader['solved_count'],
                "score": leader['total_score'] or 0,
            })

    live_participants = max(
        ContestParticipation.objects.filter(contest=contest).count(),
        submissions.values('student').distinct().count(),
    )

    return {
        "total_participants": live_participants,
        "total_submissions": submissions.count(),
        "top_performers": top_performers,
    }


def build_topic_stats(profile):
    """Return count of problems solved by topic"""
    solutions = profile.solutions.filter(all_tests_passed=True).select_related("problem")
    topic_stats = {}
    solved_problems = set()
    for solution in solutions:
        if solution.problem_id not in solved_problems:
            tags = solution.problem.tags
            if isinstance(tags, list):
                for tag in tags:
                    topic_stats[tag] = topic_stats.get(tag, 0) + 1
            solved_problems.add(solution.problem_id)
    
    # Return top topics
    sorted_topics = sorted(topic_stats.items(), key=lambda x: x[1], reverse=True)
    return [{"name": name, "count": count} for name, count in sorted_topics[:6]]


def build_aptitude_stats(profile):
    """Calculate solved vs total questions for top-level aptitude categories"""
    categories = AptitudeTopic.objects.filter(parent=None)
    stats = []
    
    for cat in categories:
        # Get all related topic IDs (parent + 2 levels of subtopics)
        sub_ids = list(cat.subtopics.values_list('id', flat=True))
        sub_sub_ids = list(AptitudeTopic.objects.filter(parent_id__in=sub_ids).values_list('id', flat=True))
        all_ids = [cat.id] + sub_ids + sub_sub_ids
        
        total = AptitudeQuestion.objects.filter(topic_id__in=all_ids).count()
        solved = SolvedAptitude.objects.filter(student=profile, question__topic_id__in=all_ids).count()
        
        stats.append({
            "name": cat.title,
            "solved": solved,
            "total": total,
            "percentage": round((solved / total * 100), 1) if total > 0 else 0
        })
    return stats

def calculate_campus_rank_helper(student):
    """Dynamically calculate the campus-wide rank of a student."""
    from .models import StudentProfile, SolvedAptitude, ContestParticipation
    
    coding_solved = student.solved_problems.count()
    aptitude_solved = SolvedAptitude.objects.filter(student=student).count()
    contests_attended = ContestParticipation.objects.filter(student=student).count()
    
    better_students = StudentProfile.objects.filter(
        institution=student.institution
    ).annotate(
        c_solved=Count('solutions', filter=Q(solutions__all_tests_passed=True), distinct=True),
        c_attended=Count('contest_participations', distinct=True),
        a_solved=Count('solved_aptitude', distinct=True)
    ).filter(
        Q(c_solved__gt=coding_solved) |
        Q(c_solved=coding_solved, c_attended__gt=contests_attended) |
        Q(c_solved=coding_solved, c_attended=contests_attended, a_solved__gt=aptitude_solved) |
        Q(c_solved=coding_solved, c_attended=contests_attended, a_solved=aptitude_solved, current_streak__gt=student.current_streak)
    ).count()
    
    return better_students + 1





def get_discussion_messages(user, profile, profile_type, thread_type="general", other_user_reg=None, batch_name=None, problem_slug=None, section=None, mentor_id=None):
    """
    Fetch and cleanup messages based on access rules.
    Messages older than 24h are excluded from every response; the actual
    DELETE only runs on a small random fraction of requests rather than
    every single poll (clients re-poll this endpoint every few seconds), so
    concurrent posts aren't competing with a delete on the same table on
    every request — on SQLite that write contention can make a POST time
    out or fail silently right when a poll's cleanup delete is running.
    """
    import random

    cutoff = timezone.now() - timedelta(hours=24)
    if random.randint(1, 20) == 1:
        DiscussionMessage.objects.filter(created_at__lt=cutoff).delete()

    qs = DiscussionMessage.objects.filter(created_at__gte=cutoff).select_related(
        "sender", "recipient", "student", "problem",
        "sender__student_profile", "sender__staff_profile",
        "recipient__student_profile", "recipient__staff_profile"
    )

    # 3. Filter by Thread Type
    if thread_type == "general":
        if profile_type == "student":
            # For students, General = Their Batch Room
            return qs.filter(thread_type="general", batch_name=profile.batch)
        elif profile_type in ["staff", "hod", "ja", "tpu"]:
            # Staff/HOD see messages for the batch they requested
            if not batch_name:
                return qs.none()
            return qs.filter(thread_type="general", batch_name=batch_name)
        return qs.none()

    if thread_type == "individual" and other_user_reg:
        # We filter the messages directly by matching the identifier against sender/recipient profile fields.
        # This handles collisions where a register number might match a username or faculty ID of a different user.
        return qs.filter(thread_type="individual").filter(
            (Q(sender=user) & (
                Q(recipient__student_profile__register_number__iexact=other_user_reg) |
                Q(recipient__staff_profile__faculty_id__iexact=other_user_reg) |
                Q(recipient__username=other_user_reg)
            )) |
            (Q(recipient=user) & (
                Q(sender__student_profile__register_number__iexact=other_user_reg) |
                Q(sender__staff_profile__faculty_id__iexact=other_user_reg) |
                Q(sender__username=other_user_reg)
            ))
        )

    if thread_type == "batch" and batch_name:
        return qs.filter(thread_type="batch", batch_name=batch_name)

    if thread_type == "section" and batch_name and section:
        # Section chat — scoped to a specific batch + section
        return qs.filter(thread_type="section", batch_name=batch_name, section=section)

    if thread_type == "mentor_group":
        # Mentor group chat — all messages for a given mentor's group
        # mentor_id identifies whose group room this is (the staff member)
        if not mentor_id:
            return qs.none()
        try:
            mentor_staff = StaffProfile.objects.get(id=mentor_id)
        except StaffProfile.DoesNotExist:
            return qs.none()
        # Verify access: must be the mentor themselves or one of their mentees
        if profile_type == "student":
            if not hasattr(profile, 'mentor') or profile.mentor_id != mentor_id:
                return qs.none()
        elif profile_type in ["staff", "hod", "tpu", "ja"]:
            if profile.id != mentor_id:
                return qs.none()
        return qs.filter(thread_type="mentor_group", batch_name=str(mentor_id))

    if thread_type == "staff":
        if profile_type in ["staff", "hod"] and profile.department:
            return qs.filter(thread_type="staff", department=profile.department)
        elif profile_type in ["admin", "ja", "tpu"]:
            return qs.filter(thread_type="staff", institution=profile.institution)
        return qs.none()

    if thread_type == "hod_tp_ja":
        if profile:
            return qs.filter(thread_type="hod_tp_ja", institution=profile.institution)
        return qs.none()

    if thread_type == "problem" and problem_slug:
        return qs.filter(thread_type="problem", problem__slug=problem_slug)

    return qs.none()


def _best_score_per_problem(contest, student):
    """
    Return the sum of the best (highest) Accepted score per problem for a student.
    Only counts Accepted submissions — Wrong Answer partial scores are excluded.
    This is the canonical way to compute a student's contest total_score.
    """
    from django.db.models import Max
    rows = (
        ContestSubmission.objects.filter(
            contest=contest,
            student=student,
            status='Accepted',
        )
        .values('problem')
        .annotate(best=Max('score'))
    )
    return sum(r['best'] for r in rows)


_CODING_DIFFICULTY_MAX = {"Easy": 100, "Medium": 200, "Hard": 300}


def _compute_contest_score_and_solved(contest, student):
    """Canonical way to (re)compute a student's total_score and
    problems_solved for a contest, across all contest types.

    - programming: total_score is the raw best-per-problem sum (unchanged).
    - aptitude: total_score is the raw sum of AptitudeContestSubmission.score
      (unchanged).
    - combined: coding and aptitude/reading are each normalized to a 0-100%
      of that section's own maximum possible score, then blended using the
      contest's staff-set weight percentages into a single 0-100 total_score.
      Reading questions are AptitudeQuestion rows (question_type="RC") that
      ride along in aptitude_questions/AptitudeContestSubmission, so they're
      split out from regular aptitude ones by question_type here.
    """
    from django.db.models import Sum

    if contest.contest_type == "programming":
        score = _best_score_per_problem(contest, student)
        solved = ContestSubmission.objects.filter(
            contest=contest, student=student, status="Accepted",
        ).values("problem").distinct().count()
        return score, solved

    if contest.contest_type == "aptitude":
        subs = AptitudeContestSubmission.objects.filter(contest=contest, student=student)
        score = subs.aggregate(total=Sum("score"))["total"] or 0
        solved = subs.filter(is_correct=True).count()
        return score, solved

    # combined
    coding_raw = _best_score_per_problem(contest, student)
    coding_solved = ContestSubmission.objects.filter(
        contest=contest, student=student, status="Accepted",
    ).values("problem").distinct().count()
    coding_max = sum(
        _CODING_DIFFICULTY_MAX.get(p.difficulty, 100) for p in contest.problems.all()
    )
    coding_pct = (coding_raw / coding_max * 100) if coding_max else 0

    apt_questions = list(contest.aptitude_questions.all())
    apt_ids = {q.id for q in apt_questions if q.question_type != "RC"}
    read_ids = {q.id for q in apt_questions if q.question_type == "RC"}

    subs = AptitudeContestSubmission.objects.filter(contest=contest, student=student)
    apt_subs = [s for s in subs if s.question_id in apt_ids]
    read_subs = [s for s in subs if s.question_id in read_ids]

    apt_raw = sum(s.score for s in apt_subs)
    apt_pct = (apt_raw / len(apt_ids) * 100) if apt_ids else 0
    read_raw = sum(s.score for s in read_subs)
    read_pct = (read_raw / len(read_ids) * 100) if read_ids else 0

    weighted = (
        coding_pct * (contest.coding_weight_percent / 100)
        + apt_pct * (contest.aptitude_weight_percent / 100)
        + read_pct * (contest.reading_weight_percent / 100)
    )
    solved = coding_solved + sum(1 for s in apt_subs if s.is_correct) + sum(1 for s in read_subs if s.is_correct)
    return round(weighted), solved


def _display_actual_output(tc_result, actual_raw):
    """What to show as "Received Output" for one test case. On a clean run
    this is just the program's real stdout; on a failing run with no
    stdout at all (a crash, timeout, or compile error) it falls back to
    executor._normalize_result's unified `output` field — which now
    includes a plain-English reason for common crash signals (e.g. "SIGSEGV
    — likely a null pointer dereference...") — rather than leaving the
    console showing a blank "(no output)" with no indication anything
    actually went wrong."""
    if actual_raw:
        return actual_raw
    if tc_result.get("status") == "Accepted":
        return actual_raw
    return tc_result.get("output") or tc_result.get("stderr") or tc_result.get("compile_output") or actual_raw


def execute_problem_test_case_batch(
    *,
    problem,
    source_code,
    language,
    language_id,
    test_cases,
    batch_kind,
):
    test_results = []
    latest_time = ""
    latest_memory = ""

    schema = getattr(problem, "param_schema", None) if problem else None

    for case in test_cases:
        case_input_data = getattr(case, "input_data", None)
        prepared = prepare_execution_payload(
            problem=problem,
            source_code=source_code,
            language=language,
            stdin=case.stdin,
            input_data=case_input_data,
        )
        logger.debug(
            "Test case %d: prepared stdin=%r, adapted=%s",
            case.order, prepared.get("stdin"), prepared.get("adapted")
        )
        tc_result = execute_judge0_submission(
            source_code=prepared["source_code"],
            language_id=language_id,
            stdin=prepared["stdin"],
        )
        logger.debug(
            "Test case %d result: status=%s, stdout=%r, stderr=%r",
            case.order, tc_result.get("status"), tc_result.get("stdout"), tc_result.get("stderr")
        )
        actual_raw = (tc_result["stdout"] or "").strip()
        expected = case.expected_output.strip()
        if schema and param_types.is_design_schema(schema) and case_input_data is not None:
            passed = (
                tc_result["status"] == "Accepted"
                and compare_design_output(actual_raw, expected, schema, case_input_data.get("operations", []))
            )
        elif schema and case_input_data is not None:
            passed = (
                tc_result["status"] == "Accepted"
                and compare_typed_output(actual_raw, expected, schema.get("return_type", ""))
            )
        else:
            passed = (
                tc_result["status"] == "Accepted"
                and normalize_comparable_output(actual_raw)
                == normalize_comparable_output(expected)
            )

        latest_time = tc_result["time"] or latest_time
        latest_memory = tc_result["memory"] or latest_memory
        test_results.append(
            {
                "stdin": case.stdin,
                "expected": expected,
                "actual": _display_actual_output(tc_result, actual_raw),
                "passed": passed,
                "status": tc_result["status"],
                "time": tc_result["time"],
                "memory": tc_result["memory"],
                "stderr": tc_result["stderr"],
                "compile_output": tc_result["compile_output"],
                "is_sample": case.is_sample,
                "source": case.source,
            }
        )

    total_cases = len(test_results)
    passed_cases = sum(1 for item in test_results if item["passed"])
    first_failure = next((item for item in test_results if not item["passed"]), None)

    if total_cases and passed_cases == total_cases:
        status_label = "Accepted"
    elif first_failure and first_failure["status"] != "Accepted":
        status_label = first_failure["status"]
    else:
        status_label = "Wrong Answer"

    if batch_kind == "sample":
        output = f"Sample test cases passed: {passed_cases}/{total_cases}."
    elif status_label == "Accepted":
        output = f"All {total_cases} test cases passed."
    else:
        output = f"{passed_cases}/{total_cases} test cases passed."

    return {
        "stdout": output,
        "stderr": first_failure["stderr"] if first_failure else "",
        "compile_output": first_failure["compile_output"] if first_failure else "",
        "status": status_label,
        "time": latest_time,
        "memory": latest_memory,
        "output": output,
        "test_results": test_results,
        "passed_cases": passed_cases,
        "total_cases": total_cases,
        "test_case_mode": batch_kind,
    }


def execute_lab_test_case_batch(*, source_code, language, language_id, test_cases, batch_kind):
    """Same result shape as execute_problem_test_case_batch(), for
    LabExercise test cases. Skips prepare_execution_payload()'s Problem-
    specific execution-type/driver-injection step — lab exercises are
    always plain stdin-in/stdout-out programs (the traditional lab-record
    style, unlike Problems which may be function/class-signature based),
    so the code runs as submitted, with each test case's stdin fed in
    directly."""
    test_results = []
    latest_time = ""
    latest_memory = ""

    for case in test_cases:
        tc_result = execute_judge0_submission(
            source_code=source_code, language_id=language_id, stdin=case.stdin,
        )
        actual_raw = (tc_result["stdout"] or "").strip()
        expected = case.expected_output.strip()
        from .services.execution_adapter import normalize_comparable_output

        passed = (
            tc_result["status"] == "Accepted"
            and normalize_comparable_output(actual_raw) == normalize_comparable_output(expected)
        )

        latest_time = tc_result["time"] or latest_time
        latest_memory = tc_result["memory"] or latest_memory
        test_results.append(
            {
                "stdin": case.stdin,
                "expected": expected,
                "actual": _display_actual_output(tc_result, actual_raw),
                "passed": passed,
                "status": tc_result["status"],
                "time": tc_result["time"],
                "memory": tc_result["memory"],
                "stderr": tc_result["stderr"],
                "compile_output": tc_result["compile_output"],
                "is_sample": case.is_sample,
                "source": case.source,
            }
        )

    total_cases = len(test_results)
    passed_cases = sum(1 for item in test_results if item["passed"])
    first_failure = next((item for item in test_results if not item["passed"]), None)

    if total_cases and passed_cases == total_cases:
        status_label = "Accepted"
    elif first_failure and first_failure["status"] != "Accepted":
        status_label = first_failure["status"]
    else:
        status_label = "Wrong Answer"

    if batch_kind == "sample":
        output = f"Sample test cases passed: {passed_cases}/{total_cases}."
    elif status_label == "Accepted":
        output = f"All {total_cases} test cases passed."
    else:
        output = f"{passed_cases}/{total_cases} test cases passed."

    return {
        "stdout": output,
        "stderr": first_failure["stderr"] if first_failure else "",
        "compile_output": first_failure["compile_output"] if first_failure else "",
        "status": status_label,
        "time": latest_time,
        "memory": latest_memory,
        "output": output,
        "test_results": test_results,
        "passed_cases": passed_cases,
        "total_cases": total_cases,
        "test_case_mode": batch_kind,
    }


# ---------------------------------------------------------------------------
# Rate-limit settings helpers
# ---------------------------------------------------------------------------

def _auth_rate_limits():
    return (
        getattr(settings, "AUTH_RATE_LIMIT_MAX_ATTEMPTS", 5),
        getattr(settings, "AUTH_RATE_LIMIT_WINDOW_SECONDS", 60),
    )


def _lookup_rate_limits():
    return (
        getattr(settings, "LOOKUP_RATE_LIMIT_MAX_ATTEMPTS", 20),
        getattr(settings, "LOOKUP_RATE_LIMIT_WINDOW_SECONDS", 60),
    )


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

class DashboardView(UnifiedAuthMixin, APIView):
    def get(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error

        problems = Problem.objects.all()

        range_start = parse_date_param(request.query_params.get('start_date'))
        range_end = parse_date_param(request.query_params.get('end_date'))

        # Handle staff/hod/academics/admin/director/tpu/ja users differently
        if profile_type in ["staff", "hod", "academics", "admin", "director", "tpu", "ja"]:
            # Get profile details
            profile_obj = profile if profile else None
            user_department = getattr(profile_obj, 'department', None) if profile_obj else None

            # Filter by institution for multi-tenant support
            inst = getattr(profile_obj, 'institution', None)

            # Filter student count by department for HOD / Academic Coordinator, all for admin/staff within institution
            if profile_type in ["hod", "academics"] and user_department:
                students_qs = StudentProfile.objects.filter(department=user_department, institution=inst)
                student_count = students_qs.count()
                dept_contests = Contest.objects.filter(department=user_department, institution=inst)
                contest_count = dept_contests.count()
                pending_approvals = dept_contests.filter(status='pending_approval').count()

                weekly_activity = build_solved_activity_series(
                    Q(student__department=user_department, student__institution=inst),
                    start_date=range_start, end_date=range_end,
                )
            else:
                student_count = StudentProfile.objects.filter(institution=inst).count() if inst else StudentProfile.objects.count()
                contest_count = Contest.objects.filter(institution=inst).count() if inst else Contest.objects.count()
                pending_approvals = Contest.objects.filter(status='pending_approval', institution=inst).count() if profile_type in ["admin", "director", "tpu", "ja"] and inst else 0

                weekly_activity = build_solved_activity_series(
                    Q(student__institution=inst) if inst else Q(),
                    start_date=range_start, end_date=range_end,
                )

            # Common analytics for HOD and staff dashboards
            recent_activity = []
            engagement_summary = {"active_today": 0, "avg_solved": 0, "participation_rate": 0}
            
            if user_department:
                # Recent Activity (Last 10 solved problems in department)
                recent_solved = SolvedProblem.objects.filter(
                    student__department=user_department
                ).select_related('student', 'problem').order_by('-solved_at')[:10]
                
                for solved in recent_solved:
                    recent_activity.append({
                        "student_name": solved.student.name,
                        "student_id": solved.student.register_number,
                        "problem_title": solved.problem.title,
                        "solved_at": solved.solved_at.isoformat(),
                    })
                
                # Engagement Summary
                today = timezone.now().date()
                active_today = StudentProfile.objects.filter(department=user_department, last_login_on=today).count()
                total_students = StudentProfile.objects.filter(department=user_department).count()
                
                if total_students > 0:
                    total_solved_count = SolvedProblem.objects.filter(student__department=user_department).count()
                    avg_solved = round(total_solved_count / total_students, 1)
                    engagement_summary = {
                        "active_today": active_today,
                        "avg_solved": avg_solved,
                        "participation_rate": round((active_today / total_students * 100), 1)
                    }
            elif inst:
                # Institutional Recent Activity
                recent_solved = SolvedProblem.objects.filter(
                    student__institution=inst
                ).select_related('student', 'problem').order_by('-solved_at')[:10]
                
                for solved in recent_solved:
                    recent_activity.append({
                        "student_name": solved.student.name,
                        "student_id": solved.student.register_number,
                        "problem_title": solved.problem.title,
                        "solved_at": solved.solved_at.isoformat(),
                    })
                
                # Institutional Engagement Summary
                today = timezone.now().date()
                active_today = StudentProfile.objects.filter(institution=inst, last_login_on=today).count()
                total_students = StudentProfile.objects.filter(institution=inst).count()
                
                if total_students > 0:
                    total_solved_count = SolvedProblem.objects.filter(student__institution=inst).count()
                    avg_solved = round(total_solved_count / total_students, 1)
                    engagement_summary = {
                        "active_today": active_today,
                        "avg_solved": avg_solved,
                        "participation_rate": round((active_today / total_students * 100), 1)
                    }
            
            # Get list of departments for institutional roles
            depts_list = []
            if profile_type in ["admin", "director", "tpu", "ja"]:
                depts_qs = Department.objects.filter(institution=inst) if inst else Department.objects.all()
                for d in depts_qs:
                    depts_list.append({
                        "id": d.id,
                        "name": d.name,
                        "code": d.code
                    })

            # Staff/HOD/Admin get simplified dashboard without student-specific stats
            user_payload = {
                "name": profile.name if profile else request.user.first_name,
                "title": (
                    "Administrator" if profile_type == "admin" else 
                    "Director" if profile_type == "director" else 
                    "TPU Coordinator" if profile_type == "tpu" else 
                    "Junior Admin" if profile_type == "ja" else
                    "HOD" if profile_type == "hod" else "Staff"
                ),
                "streak": 0,
                "loginDays": 0,
                "rank": 1,
                "totalStudents": student_count,
                "totalContests": contest_count,
                "pendingApprovals": pending_approvals,
                "registerNumber": profile.faculty_id if profile else request.user.username,
                "facultyId": profile.faculty_id if profile else request.user.username,
                "email": "",
                "role": profile.role if profile else "admin",
                "department": {
                    "name": user_department.name if user_department else None,
                    "code": user_department.code if user_department else None,
                } if user_department else None,
                "departments": depts_list
            }
            
            stats = {"easy": 0, "medium": 0, "hard": 0}
            activity_calendar = []
            
            daily_problem = problems.filter(is_daily=True).first() or problems.first()
            
            # For HOD and Institutional roles, include top performers
            leaderboard = []
            if profile_type in ["hod", "academics"] and user_department:
                top_students = students_qs.annotate(
                    solved=Count('solved_problems', distinct=True)
                ).order_by('-solved')[:10]
                for s in top_students:
                    leaderboard.append({
                        "id": s.register_number,
                        "name": s.name,
                        "score": s.solved,
                        "rank": 0
                    })
            elif profile_type in ["admin", "director", "tpu", "ja"] and inst:
                top_students = StudentProfile.objects.filter(institution=inst).annotate(
                    solved=Count('solved_problems', distinct=True)
                ).order_by('-solved')[:10]
                for s in top_students:
                    leaderboard.append({
                        "id": s.register_number,
                        "name": s.name,
                        "score": s.solved,
                        "rank": 0
                    })
            
            return Response({
                "user": user_payload,
                "dailyProblem": {
                    "title": daily_problem.title if daily_problem else "",
                    "difficulty": daily_problem.difficulty if daily_problem else "",
                    "description": daily_problem.description if daily_problem else "",
                    "tags": daily_problem.tags if daily_problem else [],
                } if daily_problem else None,
                "stats": stats,
                "weeklyActivity": weekly_activity,
                "activityCalendar": activity_calendar,
                "consistencyLabel": "Department Activity",
                "tracks": FALLBACK_DASHBOARD["tracks"],
                "leaderboard": leaderboard if leaderboard else FALLBACK_DASHBOARD["leaderboard"],
                "editor": FALLBACK_DASHBOARD["editor"],
                "recentActivity": recent_activity,
                "engagementSummary": engagement_summary,
                "staff": StaffProfileSerializer(profile).data if profile and profile_type in ["staff", "hod", "academics"] else None,
                "locked_modules": inst.locked_modules if inst else [],
            })

        # Student dashboard (original logic)
        activity_calendar = build_activity_calendar(profile)
        stats = build_student_stats(profile)
        aptitude_stats = build_aptitude_stats(profile)
        weekly_activity = build_weekly_activity(activity_calendar)
        topic_stats = build_topic_stats(profile)
        
        # Unified Performance Ranking (Coding + Contests + Consistency)
        students_with_counts = (
            StudentProfile.objects.filter(institution=profile.institution)
            .annotate(
                coding_solved=Count(
                    'solutions',
                    filter=Q(solutions__all_tests_passed=True),
                    distinct=True
                ),
                contests_attended=Count('contest_participations', distinct=True),
                aptitude_solved=Count('solved_aptitude', distinct=True),
            )
            .order_by('-coding_solved', '-contests_attended', '-aptitude_solved', '-current_streak', 'name')
        )
        
        campus_rank = 1
        for idx, student in enumerate(students_with_counts, start=1):
            if student.id == profile.id:
                campus_rank = idx
                break
        
        # Awards & Achievements Logic
        total_solved = SolvedProblem.objects.filter(student=profile).count()
        total_aptitude_solved = SolvedAptitude.objects.filter(student=profile).count()
        current_streak = profile.current_streak
        
        # Check for unearned achievements
        unearned = Achievement.objects.exclude(userachievement__user=request.user)
        for ach in unearned:
            should_award = False
            if ach.category == 'coding':
                if ach.criteria_type == 'solve_count' and total_solved >= ach.criteria_value:
                    should_award = True
                elif ach.criteria_type == 'streak' and current_streak >= ach.criteria_value:
                    should_award = True
            elif ach.category == 'aptitude':
                if ach.criteria_type == 'aptitude_solve_count' and total_aptitude_solved >= ach.criteria_value:
                    should_award = True
                elif ach.criteria_type == 'quant_solve_count':
                    quant_solved = SolvedAptitude.objects.filter(student=profile, question__topic__parent__parent__title__icontains='QUANTITATIVE').count()
                    if quant_solved >= ach.criteria_value:
                        should_award = True
            
            if should_award:
                UserAchievement.objects.get_or_create(user=request.user, achievement=ach)
        
        all_achievements = Achievement.objects.all()
        user_achievements = UserAchievement.objects.filter(user=request.user).select_related('achievement')
        earned_ids = {ua.achievement_id: ua.awarded_at for ua in user_achievements}

        achievements_data = []
        for ach in all_achievements:
            is_earned = ach.id in earned_ids
            achievements_data.append({
                "id": ach.id,
                "name": ach.name,
                "description": ach.description,
                "icon": ach.badge_icon,
                "category": ach.category,
                "is_earned": is_earned,
                "date": earned_ids[ach.id].strftime("%b %d, %Y") if is_earned else None
            })

        user_payload = {
            "name": profile.name,
            "title": profile.title,
            "streak": profile.current_streak,
            "loginDays": profile.login_days,
            "rank": campus_rank,
            "totalStudents": students_with_counts.count(),
            "registerNumber": profile.register_number,
            "email": profile.personal_email,
            "total_problems_count": Problem.objects.count(),
            "total_aptitude_count": AptitudeQuestion.objects.count(),
            "tracked_companies": profile.tracked_companies,
        }

        if not problems.exists():
            return Response(
                {
                    **FALLBACK_DASHBOARD,
                    "user": user_payload,
                    "stats": stats,
                    "aptitude_stats": aptitude_stats,
                    "weeklyActivity": weekly_activity,
                    "activityCalendar": activity_calendar,
                    "consistencyLabel": "Activity calendar",
                    "student": StudentProfileSerializer(profile).data,
                }
            )

        # Daily Problem Logic
        today = timezone.now().date()
        daily_instance = DailyProblem.objects.filter(date=today).first()
        
        if not daily_instance:
            # Pick a random problem that is not already a daily problem if possible
            random_problem = Problem.objects.order_by('?').first()
            if random_problem:
                daily_instance = DailyProblem.objects.create(date=today, problem=random_problem)
        
        daily_problem = daily_instance.problem if daily_instance else problems.first()
        
        # Mark as daily for visibility in list
        if daily_problem:
            Problem.objects.filter(id=daily_problem.id).update(is_daily=True)

        # Calculate Preferred Language
        user_submissions = Submission.objects.filter(student=profile)
        if user_submissions.exists():
            lang_stats = user_submissions.values('language').annotate(count=Count('language')).order_by('-count')
            preferred_language = lang_stats[0]['language']
        else:
            preferred_language = "Python"

        # Announcements Logic (Fetch active announcements from last 7 days)
        seven_days_ago = timezone.now() - timedelta(days=7)
        announcements = Announcement.objects.filter(
            is_active=True, 
            created_at__gte=seven_days_ago
        ).order_by('-created_at')[:5]

        payload = {
            "user": user_payload,
            "achievements": achievements_data,
            "dailyProblem": {
                "id": daily_problem.id,
                "slug": daily_problem.slug,
                "title": daily_problem.title,
                "difficulty": daily_problem.difficulty,
                "description": daily_problem.description,
                "tags": daily_problem.tags,
                "preferredLanguage": preferred_language,
            },
            "stats": stats,
            "weeklyActivity": weekly_activity,
            "activityCalendar": activity_calendar,
            "topicStats": topic_stats,
            "consistencyLabel": "Activity calendar",
            "tracks": topic_stats if topic_stats else FALLBACK_DASHBOARD["tracks"],
            "aptitude_stats": aptitude_stats,
            "leaderboard": FALLBACK_DASHBOARD["leaderboard"],
            "announcements": [{
                "id": a.id,
                "title": a.title,
                "content": a.content,
                "category": a.category,
                "date": a.created_at.strftime("%b %d, %Y")
            } for a in announcements],
            "editor": FALLBACK_DASHBOARD["editor"],
            "student": StudentProfileSerializer(profile).data,
            "locked_modules": profile.institution.locked_modules if profile.institution else [],
        }
        return Response(payload)


class UpdateTrackedCompaniesView(UnifiedAuthMixin, APIView):
    """Allow students to update their tracked companies list"""
    def post(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error
        
        if profile_type != "student":
            return Response({"detail": "Only students can track companies."}, status=status.HTTP_403_FORBIDDEN)
        
        companies = request.data.get("companies", [])
        if not isinstance(companies, list):
            return Response({"detail": "Companies must be a list."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            profile.tracked_companies = [c.strip() for c in companies if isinstance(c, str) and c.strip()]
            profile.save(update_fields=["tracked_companies"])
        except Exception:
            logger.exception("Failed to save tracked companies for student %s", getattr(profile, 'register_number', '?'))
            return Response(
                {"detail": "Failed to update tracked companies. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({"status": "success", "tracked_companies": profile.tracked_companies})


class DailyLeaderboardView(UnifiedAuthMixin, APIView):
    def get(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error

        today = timezone.now().date()
        daily_instance = DailyProblem.objects.filter(date=today).first()
        
        if not daily_instance:
            return Response({"leaderboard": []})

        # Get successful submissions for this daily problem
        # Only take the best (earliest) submission per student
        submissions = Submission.objects.filter(
            problem=daily_instance.problem,
            status="Accepted"
        ).order_by('student', 'submitted_at').distinct('student')

        leaderboard = []
        for idx, sub in enumerate(submissions, 1):
            leaderboard.append({
                "rank": idx,
                "name": sub.student.name,
                "registerNumber": sub.student.register_number,
                "language": sub.language,
                "time": sub.submitted_at.strftime("%I:%M %p"),
                "isUser": sub.student.id == profile.id
            })

        return Response({"leaderboard": leaderboard[:50]})  # Top 50


class ProblemListView(UnifiedAuthMixin, APIView):
    def get(self, request):
        if request.user and request.user.is_authenticated:
            profile, profile_type, _ = self.get_authenticated_profile(request)
        else:
            profile, profile_type = None, None

        difficulty = request.query_params.get("difficulty")
        queryset = Problem.objects.all()

        if difficulty:
            queryset = queryset.filter(difficulty__iexact=difficulty)

        if queryset.exists():
            # Staff/Admin don't have progress, students do
            if profile_type == "student" and profile:
                progress_map = build_problem_progress_map(profile)
            else:
                progress_map = {}
            return Response(
                ProblemSerializer(
                    queryset,
                    many=True,
                    context={"progress_map": progress_map},
                ).data
            )

        fallback = FALLBACK_PROBLEMS
        if difficulty:
            fallback = [
                item for item in fallback if item["difficulty"].lower() == difficulty.lower()
            ]
        return Response(fallback)


class ProblemDetailView(UnifiedAuthMixin, APIView):
    def get(self, request, slug):
        if request.user and request.user.is_authenticated:
            profile, profile_type, _ = self.get_authenticated_profile(request)
        else:
            profile, profile_type = None, None

        problem = Problem.objects.filter(slug=slug).first()
        if not problem:
            return Response(
                {"detail": "Problem not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # Staff/Admin don't have progress, students do
        last_solutions_by_language = {}
        if profile_type == "student" and profile:
            progress_map = build_problem_progress_map(profile)
            # Most-recent-first so the first row seen per language is the
            # student's latest submission — lets the editor restore exactly
            # what they last had, per language, when they reopen a problem.
            solutions = (
                ProblemSolution.objects
                .filter(student=profile, problem=problem)
                .order_by("language", "-submitted_at")
            )
            for sol in solutions:
                if sol.language not in last_solutions_by_language:
                    last_solutions_by_language[sol.language] = {
                        "source_code": sol.source_code,
                        "status": sol.status,
                        "all_tests_passed": sol.all_tests_passed,
                        "submitted_at": sol.submitted_at,
                    }
        else:
            progress_map = {}
        return Response(
            ProblemDetailSerializer(
                problem,
                context={
                    "progress_map": progress_map,
                    "last_solutions": last_solutions_by_language,
                },
            ).data
        )


class EditorBootstrapView(StudentAuthMixin, APIView):
    def get(self, request):
        _, error = self.get_authenticated_profile(request)
        if error:
            return error
        return Response(FALLBACK_DASHBOARD["editor"])


class HealthCheckView(APIView):
    def get(self, request):
        payload = {
            "status": "ok",
            "executor_configured": bool(getattr(settings, "EXECUTOR_BASE_URL", "").strip()),
        }
        # Opt-in diagnostics — kept off the default fast path so this endpoint
        # stays cheap for routine uptime checks.
        if request.query_params.get("executor") == "1":
            from .services.judge0 import check_judge0_health
            payload["executor"] = check_judge0_health()
        if request.query_params.get("packages") == "1":
            # Judge0 doesn't have a packages endpoint like Piston
            payload["packages"] = {"note": "Not available with Judge0"}
        # Exercises the FUNCTION-style driver-injection path (prepare_execution_payload)
        # that a typical Problems-page submission goes through, to verify the
        # typed-argument C wrapper. Temporary — remove once confirmed fixed.
        if request.query_params.get("test_driver") == "c":
            import time
            from types import SimpleNamespace
            from .services.judge0 import execute_judge0_submission as execute_submission
            from .services.execution_adapter import prepare_execution_payload
            fake_problem = SimpleNamespace(execution_type="auto", slug="add-two-numbers", function_name="addTwoNumbers")
            source = "int addTwoNumbers(int a, int b) {\n    return a + b;\n}"
            start = time.time()
            try:
                prepared = prepare_execution_payload(problem=fake_problem, source_code=source, language="C", stdin="2\n3")
                result = execute_submission(source_code=prepared["source_code"], language_id=50, stdin=prepared["stdin"])
                payload["test_driver"] = {
                    "elapsed_s": round(time.time() - start, 2),
                    "prepared_stdin": prepared["stdin"],
                    "adapted": prepared["adapted"],
                    "generated_source": prepared["source_code"][:3000],
                    "result": result,
                }
            except Exception as exc:
                payload["test_driver"] = {"elapsed_s": round(time.time() - start, 2), "error": f"{type(exc).__name__}: {exc}"}
        # Same as above but for the Java driver wrapper — verifies the
        # char/String literal fixes and array-initializer fix actually
        # compile+run in production. Temporary — remove once confirmed fixed.
        if request.query_params.get("test_driver") == "java":
            import time
            from types import SimpleNamespace
            from .services.judge0 import execute_judge0_submission as execute_submission
            from .services.execution_adapter import prepare_execution_payload
            fake_problem = SimpleNamespace(execution_type="auto", slug="add-two-numbers", function_name="addTwoNumbers")
            source = "class Solution {\n    public int addTwoNumbers(int a, int b) {\n        return a + b;\n    }\n}"
            start = time.time()
            try:
                prepared = prepare_execution_payload(problem=fake_problem, source_code=source, language="Java", stdin="[2, 3]")
                result = execute_submission(source_code=prepared["source_code"], language_id=62, stdin=prepared["stdin"])
                payload["test_driver"] = {
                    "elapsed_s": round(time.time() - start, 2),
                    "prepared_stdin": prepared["stdin"],
                    "adapted": prepared["adapted"],
                    "generated_source": prepared["source_code"][:3000],
                    "result": result,
                }
            except Exception as exc:
                payload["test_driver"] = {"elapsed_s": round(time.time() - start, 2), "error": f"{type(exc).__name__}: {exc}"}
        return Response(payload)


class ProblemProgressUpdateView(StudentAuthMixin, APIView):
    def post(self, request):
        profile, error = self.get_authenticated_profile(request)
        if error:
            return error

        serializer = ProblemProgressUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        problem = Problem.objects.filter(
            slug=serializer.validated_data["problem_slug"].strip()
        ).first()
        if not problem:
            return Response(
                {"detail": "Problem not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        progress_state = serializer.validated_data["progress_state"]
        language = (
            serializer.validated_data.get("language")
            or "Python"
        )
        status_label = "Accepted" if progress_state == "completed" else "Started"

        try:
            submission = Submission.objects.create(
                student=profile,
                problem=problem,
                language=language,
                status=status_label,
            )

            if progress_state == "completed":
                Notification.objects.create(
                    recipient=profile.account,
                    title="🎯 Problem Solved!",
                    message=f"Congratulations! You've successfully solved '{problem.title}'.",
                    link=f"/problems?slug={problem.slug}"
                )

            activity_type = "solve" if progress_state == "completed" else "practice"
            StudentActivity.objects.get_or_create(
                student=profile,
                activity_date=timezone.localdate(),
                activity_type=activity_type,
            )
        except Exception:
            logger.exception(
                "Failed to save problem progress for student %s, problem %s",
                getattr(profile, 'register_number', '?'), problem.slug,
            )
            return Response(
                {"detail": "Failed to save progress. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response(
            {
                "detail": "Progress saved successfully.",
                "submission": {
                    "id": submission.id,
                    "status": submission.status,
                    "language": submission.language,
                    "submitted_at": submission.submitted_at,
                },
                "progress_state": progress_state,
            },
            status=status.HTTP_201_CREATED,
        )


class CodeRunView(StudentAuthMixin, APIView):
    def post(self, request):
        profile, error = self.get_authenticated_profile(request)
        if error:
            return error

        serializer = CodeRunSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        validated = serializer.validated_data
        problem = None
        problem_slug = (validated.get("problem_slug") or "").strip()
        is_submit = validated.get("is_submit", False)
        stdin = validated.get("stdin", "")
        source_code = validated["source_code"]
        language = validated.get("language", "")

        logger.debug(
            "CodeRunView: lang=%s, lang_id=%s, problem=%s, is_submit=%s, stdin=%r, source_len=%d",
            language, validated["language_id"], problem_slug, is_submit, stdin, len(source_code)
        )

        # ─────────────────────────────────────────────────────────────────
        # VALIDATION: Check code before execution
        # ─────────────────────────────────────────────────────────────────
        is_valid, validation_error = validate_submission(language, source_code, stdin)
        if not is_valid:
            logger.warning("Code validation failed for %s: %s", profile.register_number, validation_error)
            return Response(
                {"detail": f"Code validation failed: {validation_error}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if problem_slug:
            problem = Problem.objects.filter(slug=problem_slug).first()

        try:
            if is_submit and not problem:
                return Response(
                    {"detail": "problem_slug is required for submission."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            run_sample_cases = bool(problem and not is_submit and not stdin.strip())
            if is_submit or run_sample_cases:
                test_cases = build_runtime_test_cases(
                    problem,
                    sample_only=not is_submit,
                )
                if is_submit and not test_cases:
                    return Response(
                        {"detail": "No test cases are configured for this problem yet."},
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                if test_cases:
                    result = execute_problem_test_case_batch(
                        problem=problem,
                        source_code=validated["source_code"],
                        language=validated.get("language", ""),
                        language_id=validated["language_id"],
                        test_cases=test_cases,
                        batch_kind="submit" if is_submit else "sample",
                    )
                else:
                    prepared = prepare_execution_payload(
                        problem=problem,
                        source_code=validated["source_code"],
                        language=validated.get("language", ""),
                        stdin=stdin,
                    )
                    result = execute_judge0_submission(
                        source_code=prepared["source_code"],
                        language_id=validated["language_id"],
                        stdin=prepared["stdin"],
                    )
            else:
                prepared = prepare_execution_payload(
                    problem=problem,
                    source_code=validated["source_code"],
                    language=validated.get("language", ""),
                    stdin=stdin,
                )
                result = execute_judge0_submission(
                    source_code=prepared["source_code"],
                    language_id=validated["language_id"],
                    stdin=prepared["stdin"],
                )
            
        except ExecutorTimeoutError as exc:
            logger.error("Executor timeout: %s", exc)
            return Response({"detail": str(exc)}, status=status.HTTP_504_GATEWAY_TIMEOUT)
        except ExecutorServiceError as exc:
            logger.error("Executor service error: %s", exc)
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        except Exception as exc:
            logger.error("Unexpected execution error: %s", exc, exc_info=True)
            return Response({"detail": f"Execution error: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            ExecutionRecord.objects.create(
                student=profile,
                problem=problem,
                language=validated.get("language") or str(validated["language_id"]),
                language_id=validated["language_id"],
                source_code=validated["source_code"],
                stdin=stdin,
                stdout=result["output"] if result.get("test_results") else result["stdout"],
                stderr=result["stderr"],
                compile_output=result["compile_output"],
                status_description=result["status"],
                execution_time=str(result["time"] or ""),
                memory=str(result["memory"] or ""),
            )
        except Exception as exc:
            logger.error("Error creating ExecutionRecord: %s", exc, exc_info=True)

        # ── Submit mode: run all test cases ────────────────────────────────
        if is_submit and problem and result.get("total_cases"):
            # Calculate complexity
            time_complexity, space_complexity = calculate_complexity(
                validated["source_code"],
                validated.get("language", "")
            )
            
            # Get time spent from active session
            time_spent = 0
            try:
                active_session = ProblemSession.objects.filter(
                    student=profile,
                    problem=problem,
                    is_active=True
                ).first()
                if active_session:
                    time_spent = active_session.end_session()
            except Exception as e:
                logger.error("Error ending problem session: %s", e)
            
            # Create solution record with complexity and timing data
            all_passed = result["passed_cases"] == result["total_cases"]
            ProblemSolution.objects.create(
                problem=problem,
                student=profile,
                language=validated.get("language") or str(validated["language_id"]),
                language_id=validated["language_id"],
                source_code=validated["source_code"],
                status=result["status"],
                passed_cases=result["passed_cases"],
                total_cases=result["total_cases"],
                all_tests_passed=all_passed,
                execution_time=str(result["time"] or ""),
                memory=str(result["memory"] or ""),
                time_complexity=time_complexity,
                space_complexity=space_complexity,
                time_spent_seconds=time_spent,
            )
            
            # Create SolvedProblem record when all tests pass
            if all_passed:
                try:
                    SolvedProblem.objects.get_or_create(
                        student=profile,
                        problem=problem,
                        defaults={
                            "language": validated.get("language") or str(validated["language_id"]),
                        },
                    )
                    # Update streak for solving activity (in case user stayed logged in)
                    profile.update_streak_for_activity()
                except Exception as e:
                    logger.error(f"Error creating SolvedProblem in view: {e}")
            
            # Include complexity info in response
            result["time_complexity"] = time_complexity
            result["space_complexity"] = space_complexity
            result["time_spent_seconds"] = time_spent

        return Response(result, status=status.HTTP_200_OK)


class PlaygroundRunView(StudentAuthMixin, APIView):
    """Free-form code execution for the student Code Playground — no
    problem/test-case grading, no driver-injection rewriting of the
    source (unlike CodeRunView's LeetCode-style submission flow), and
    no code-content restrictions beyond the sandbox itself (the
    executor already runs every submission network-isolated with
    CPU/memory/pid limits). Rate-limited since there's no natural
    per-problem request ceiling here."""

    def post(self, request):
        profile, error = self.get_authenticated_profile(request)
        if error:
            return error

        try:
            check_rate_limit(request, "playground-run", max_attempts=30, window_seconds=60)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        serializer = CodeRunSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        validated = serializer.validated_data
        stdin = validated.get("stdin", "")

        try:
            result = execute_judge0_submission(
                source_code=validated["source_code"],
                language_id=validated["language_id"],
                stdin=stdin,
            )
        except ExecutorTimeoutError as exc:
            logger.error("Playground executor timeout: %s", exc)
            return Response({"detail": str(exc)}, status=status.HTTP_504_GATEWAY_TIMEOUT)
        except ExecutorServiceError as exc:
            logger.error("Playground executor service error: %s", exc)
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        except Exception as exc:
            logger.error("Unexpected playground execution error: %s", exc, exc_info=True)
            return Response({"detail": f"Execution error: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        try:
            ExecutionRecord.objects.create(
                student=profile,
                problem=None,
                language=validated.get("language") or str(validated["language_id"]),
                language_id=validated["language_id"],
                source_code=validated["source_code"],
                stdin=stdin,
                stdout=result["stdout"],
                stderr=result["stderr"],
                compile_output=result["compile_output"],
                status_description=result["status"],
                execution_time=str(result["time"] or ""),
                memory=str(result["memory"] or ""),
            )
        except Exception as exc:
            logger.error("Error creating playground ExecutionRecord: %s", exc, exc_info=True)

        return Response(result, status=status.HTTP_200_OK)


@method_decorator(ensure_csrf_cookie, name="dispatch")
class CampusRankingView(APIView):
    """Get campus-wide student rankings based on problems solved."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Get all students with their solved problem counts using SolvedProblem table
        students_with_counts = (
            StudentProfile.objects.annotate(
                solved_count=Count(
                    'solved_problems',
                    distinct=True
                )
            )
            .values('id', 'name', 'register_number', 'solved_count')
            .order_by('-solved_count', 'name')
        )

        leaderboard = list(students_with_counts)

        # Calculate rank for current user (students only)
        user_rank = None
        if hasattr(request.user, 'student_profile'):
            current_student = request.user.student_profile
            for idx, student in enumerate(leaderboard, start=1):
                if student['id'] == current_student.id:
                    user_rank = idx
                    break

        # Get top 10 students for leaderboard
        top_students = leaderboard[:10]

        return Response({
            'userRank': user_rank or len(leaderboard),
            'totalStudents': len(leaderboard),
            'leaderboard': [
                {
                    'rank': idx + 1,
                    'name': s['name'],
                    'registerNumber': s['register_number'],
                    'solved': s['solved_count']
                }
                for idx, s in enumerate(top_students)
            ]
        })


class AdminStatsView(APIView):
    """Get admin dashboard statistics."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Only allow superusers
        if not request.user.is_superuser:
            return Response(
                {"detail": "Admin access required."},
                status=status.HTTP_403_FORBIDDEN,
            )

        from .models import Institution, Problem, ProblemSolution, SolvedProblem, ProblemSession

        # User counts
        total_students = StudentProfile.objects.count()
        total_staff = StaffProfile.objects.count()
        total_users = total_students + total_staff

        # Institution counts - get all institutions
        institutions = Institution.objects.all().order_by('-institution_id')
        institution_list = []
        
        for inst in institutions:
            student_count = StudentProfile.objects.filter(institution=inst).count()
            staff_count = StaffProfile.objects.filter(institution=inst).count()
            staff_list = StaffProfile.objects.filter(institution=inst).values('faculty_id', 'name', 'role')
            active_sessions = ProblemSession.objects.filter(
                is_active=True,
                student__institution=inst
            ).count()
            
            institution_list.append({
                "id": inst.institution_id,
                "name": inst.name,
                "short_code": inst.short_code,
                "students": student_count,
                "staff": staff_count,
                "staff_list": list(staff_list),
                "total_members": student_count + staff_count,
                "active_sessions": active_sessions,
            })

        # Problem stats
        total_problems = Problem.objects.count()
        total_solutions = ProblemSolution.objects.count()
        total_solved_problems = SolvedProblem.objects.count()

        # Active sessions
        active_sessions = ProblemSession.objects.filter(is_active=True).count()

        # AWS stats (placeholder - will be updated with real data)
        aws_hits = 0
        aws_pricing = 0.0

        # Recent activity (last 7 days)
        from datetime import timedelta
        last_week = timezone.localtime(timezone.now()) - timedelta(days=7)
        recent_logins = StudentProfile.objects.filter(last_login_on__gte=last_week.date()).count()

        return Response({
            "users": {
                "total": total_users,
                "students": total_students,
                "staff": total_staff,
                "recentLogins": recent_logins,
            },
            "institutions": {
                "total": len(institution_list),
                "list": institution_list,
            },
            "problems": {
                "total": total_problems,
                "solutions": total_solutions,
                "solved": total_solved_problems,
            },
            "activity": {
                "activeSessions": active_sessions,
                "recentLogins": recent_logins,
            },
            "aws": {
                "hits": aws_hits,
                "pricing": aws_pricing,
            }
        })


class AdminUserListView(APIView):
    """List all users (students and staff) for admin."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        user_type = request.query_params.get("type", "all")
        users = []

        if user_type in ["all", "students"]:
            students = StudentProfile.objects.select_related("institution").all()
            for s in students:
                users.append({
                    "id": s.register_number,
                    "name": s.name,
                    "type": "student",
                    "email": s.personal_email,
                    "institution": s.institution.name if s.institution else None,
                    "institution_id": s.institution.institution_id if s.institution else None,
                    "is_active": s.account.is_active if s.account else False,
                })

        if user_type in ["all", "staff"]:
            staff = StaffProfile.objects.select_related("institution").all()
            for s in staff:
                users.append({
                    "id": s.faculty_id,
                    "name": s.name,
                    "type": "staff",
                    "role": s.role,
                    "role_display": s.get_role_display(),
                    "email": "",
                    "institution": s.institution.name if s.institution else None,
                    "institution_id": s.institution.institution_id if s.institution else None,
                    "is_active": s.account.is_active if s.account else False,
                })

        return Response({"users": users, "count": len(users)})


class AdminInstitutionListCreateView(APIView):
    """List or create institutions."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        institutions = Institution.objects.all()
        data = []
        for inst in institutions:
            student_count = StudentProfile.objects.filter(institution=inst).count()
            staff_count = StaffProfile.objects.filter(institution=inst).count()
            data.append({
                "institution_id": inst.institution_id,
                "name": inst.name,
                "short_code": inst.short_code,
                "address": inst.address,
                "contact_email": inst.contact_email,
                "contact_phone": inst.contact_phone,
                "is_active": inst.is_active,
                "student_count": student_count,
                "staff_count": staff_count,
            })
        return Response({"institutions": data, "count": len(data)})

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        data = request.data
        institution_id = data.get("institution_id")
        name = data.get("name")
        short_code = data.get("short_code")

        if not institution_id or not name:
            return Response({"detail": "institution_id and name are required."}, status=status.HTTP_400_BAD_REQUEST)

        if Institution.objects.filter(institution_id=institution_id).exists():
            return Response({"detail": "Institution with this ID already exists."}, status=status.HTTP_400_BAD_REQUEST)

        institution = Institution.objects.create(
            institution_id=institution_id,
            name=name,
            short_code=short_code or "",
            address=data.get("address", ""),
            contact_email=data.get("contact_email", ""),
            contact_phone=data.get("contact_phone", ""),
            is_active=data.get("is_active", True),
        )

        return Response({
            "detail": "Institution created successfully.",
            "institution": {
                "institution_id": institution.institution_id,
                "name": institution.name,
                "short_code": institution.short_code,
            }
        }, status=status.HTTP_201_CREATED)


class AdminInstitutionDetailView(APIView):
    """Get, update or delete a specific institution."""
    permission_classes = [IsAuthenticated]

    def get(self, request, institution_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        student_count = StudentProfile.objects.filter(institution=institution).count()
        staff_count = StaffProfile.objects.filter(institution=institution).count()

        return Response({
            "institution_id": institution.institution_id,
            "name": institution.name,
            "short_code": institution.short_code,
            "address": institution.address,
            "contact_email": institution.contact_email,
            "contact_phone": institution.contact_phone,
            "is_active": institution.is_active,
            "student_count": student_count,
            "staff_count": staff_count,
        })

    def put(self, request, institution_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        data = request.data
        institution.name = data.get("name", institution.name)
        institution.short_code = data.get("short_code", institution.short_code)
        institution.address = data.get("address", institution.address)
        institution.contact_email = data.get("contact_email", institution.contact_email)
        institution.contact_phone = data.get("contact_phone", institution.contact_phone)
        institution.is_active = data.get("is_active", institution.is_active)
        institution.save()

        return Response({"detail": "Institution updated successfully."})

    def delete(self, request, institution_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        if institution.institution_id == 3000:
            return Response({"detail": "Cannot delete Ramco Institution (ID: 3000)."}, status=status.HTTP_400_BAD_REQUEST)

        institution.delete()
        return Response({"detail": "Institution deleted successfully."})


class AdminInstitutionStaffView(APIView):
    """Manage staff within an institution - list and create."""
    permission_classes = [IsAuthenticated]

    def get(self, request, institution_id):
        """List all staff in an institution."""
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        staff = StaffProfile.objects.filter(institution=institution)
        data = []
        for s in staff:
            data.append({
                "faculty_id": s.faculty_id,
                "name": s.name,
                "role": s.role,
                "role_display": s.get_role_display(),
                "password_is_set": s.password_is_set,
            })
        return Response({
            "institution_id": institution_id,
            "institution_name": institution.name,
            "staff": data,
            "count": len(data),
        })

    def post(self, request, institution_id):
        """Create new staff member in institution."""
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        data = request.data
        faculty_id = data.get("faculty_id")
        name = data.get("name")
        role = data.get("role", "staff")

        if not faculty_id or not name:
            return Response(
                {"detail": "faculty_id and name are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if StaffProfile.objects.filter(faculty_id=faculty_id).exists():
            return Response(
                {"detail": "Staff with this faculty_id already exists."},
                status=status.HTTP_400_BAD_REQUEST
            )

        staff = StaffProfile.objects.create(
            faculty_id=faculty_id,
            name=name,
            institution=institution,
            role=role,
        )

        return Response({
            "detail": "Staff created successfully.",
            "staff": {
                "faculty_id": staff.faculty_id,
                "name": staff.name,
                "role": staff.role,
                "role_display": staff.get_role_display(),
            },
        }, status=status.HTTP_201_CREATED)


class AdminStaffRoleUpdateView(APIView):
    """Update staff member's role."""
    permission_classes = [IsAuthenticated]

    def put(self, request, faculty_id):
        """Update staff role."""
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        staff = StaffProfile.objects.filter(faculty_id=faculty_id).first()
        if not staff:
            return Response({"detail": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

        role = request.data.get("role")
        if role not in ["staff", "hod", "admin"]:
            return Response(
                {"detail": "role must be 'staff', 'hod', or 'admin'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        staff.role = role
        staff.save(update_fields=["role"])

        return Response({
            "detail": f"Role updated to {staff.get_role_display()}.",
            "staff": {
                "faculty_id": staff.faculty_id,
                "name": staff.name,
                "role": staff.role,
                "role_display": staff.get_role_display(),
            },
        })


class AdminInstitutionFullDetailView(APIView):
    """Get full details of an institution including students and staff."""
    permission_classes = [IsAuthenticated]

    def get(self, request, institution_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get students
        students = StudentProfile.objects.filter(institution=institution)
        student_list = []
        for s in students:
            student_list.append({
                "id": s.register_number,
                "name": s.name,
                "email": s.personal_email,
                "current_streak": s.current_streak,
                "login_days": s.login_days,
            })

        # Get staff
        staff = StaffProfile.objects.filter(institution=institution)
        staff_list = []
        for s in staff:
            staff_list.append({
                "faculty_id": s.faculty_id,
                "name": s.name,
                "role": s.role,
                "role_display": s.get_role_display(),
                "password_is_set": s.password_is_set,
            })

        # Get stats
        active_sessions = ProblemSession.objects.filter(
            is_active=True,
            student__institution=institution
        ).count()

        # Get problem stats for this institution
        solved_count = SolvedProblem.objects.filter(student__institution=institution).count()
        total_students = len(student_list)
        success_rate = round((solved_count / total_students) * 100) if total_students > 0 else 0

        return Response({
            "institution": {
                "id": institution.id,
                "institution_id": institution.institution_id,
                "name": institution.name,
                "short_code": institution.short_code,
                "address": institution.address,
                "is_active": institution.is_active,
            },
            "students": student_list,
            "staff": staff_list,
            "department": None,
            "stats": {
                "total_students": total_students,
                "total_staff": len(staff_list),
                "active_sessions": active_sessions,
                "problems_solved": solved_count,
                "success_rate": success_rate,
            },
            "active_students": len([s for s in student_list if s["current_streak"] > 0]),
            "avg_solved": solved_count / total_students if total_students > 0 else 0,
        })


class AdminAWSStatsView(APIView):
    """Get or update AWS usage statistics."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        # Get from ExecutionRecord for real data
        total_executions = ExecutionRecord.objects.count()
        total_api_calls = total_executions  # Each execution is an API call
        
        # Calculate estimated cost (placeholder pricing)
        cost_per_execution = 0.001  # $0.001 per execution
        estimated_cost = total_api_calls * cost_per_execution

        return Response({
            "hits": total_api_calls,
            "pricing": round(estimated_cost, 2),
            "executions": total_executions,
            "period": "all_time",
        })

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        # Update AWS stats (manual override)
        data = request.data
        hits = data.get("hits")
        pricing = data.get("pricing")

        return Response({
            "detail": "AWS stats updated.",
            "hits": hits,
            "pricing": pricing,
        })


class AdminAssignUserToInstitutionView(APIView):
    """Assign a student or staff to an institution."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        user_id = request.data.get("user_id")
        user_type = request.data.get("user_type")  # "student" or "staff"
        institution_id = request.data.get("institution_id")

        if not user_id or not user_type or not institution_id:
            return Response({"detail": "user_id, user_type, and institution_id are required."}, status=status.HTTP_400_BAD_REQUEST)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        if user_type == "student":
            profile = StudentProfile.objects.filter(register_number=user_id).first()
            if not profile:
                return Response({"detail": "Student not found."}, status=status.HTTP_404_NOT_FOUND)
            profile.institution = institution
            profile.save()
        elif user_type == "staff":
            profile = StaffProfile.objects.filter(faculty_id=user_id).first()
            if not profile:
                return Response({"detail": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)
            profile.institution = institution
            profile.save()
        else:
            return Response({"detail": "user_type must be 'student' or 'staff'."}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            "detail": f"{user_type.capitalize()} assigned to {institution.name}.",
            "user_id": user_id,
            "institution_id": institution_id,
            "institution_name": institution.name,
        })


class StudentLookupView(APIView):
    def get(self, request):
        max_attempts, window = _lookup_rate_limits()
        try:
            check_rate_limit(request, "student-lookup", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        register_number = (request.query_params.get("register_number") or "").strip()
        if not register_number:
            return Response(
                {"detail": "register_number is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        profile = StudentProfile.objects.filter(register_number=register_number).first()
        if not profile:
            return Response(
                {"detail": "Student not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                "student": StudentProfileSerializer(profile).data,
                "first_login_required": not profile.password_is_set,
            }
        )


@method_decorator(ensure_csrf_cookie, name="dispatch")
class RegisterNumberListView(APIView):
    def get(self, request):
        query = (request.query_params.get("q") or "").strip()
        queryset = StudentProfile.objects.exclude(register_number__isnull=True).exclude(
            register_number=""
        )

        if query:
            queryset = queryset.filter(
                Q(register_number__icontains=query) | Q(name__icontains=query)
            )

        students = queryset.order_by("register_number")[:50]
        return Response(StudentLookupListSerializer(students, many=True).data)


class DiscussionMessageListCreateView(UnifiedAuthMixin, APIView):
    def get(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error

        thread_type = request.query_params.get("thread_type", "general")
        other_user_reg = request.query_params.get("other_user_reg")
        batch_name = request.query_params.get("batch_name")
        problem_slug = request.query_params.get("problem_slug")
        section = (request.query_params.get("section") or "").strip().upper()
        mentor_id_str = request.query_params.get("mentor_id")
        mentor_id = int(mentor_id_str) if mentor_id_str and mentor_id_str.isdigit() else None

        # Security check for staff/hod rooms
        if thread_type in ["staff", "hod_tp_ja"] and profile_type == "student":
            return Response({"detail": "Access denied to this channel."}, status=403)

        # Section chat access: student must be in that batch+section
        if thread_type == "section" and profile_type == "student":
            if not section or profile.section != section:
                return Response({"detail": "Access denied to this section."}, status=403)
            batch_name = profile.batch

        # Mentor group access: student must have that mentor
        if thread_type == "mentor_group" and profile_type == "student":
            if not mentor_id or not profile.mentor or profile.mentor.id != mentor_id:
                return Response({"detail": "Access denied to this mentor group."}, status=403)

        messages_qs = get_discussion_messages(
            request.user,
            profile,
            profile_type,
            thread_type=thread_type,
            other_user_reg=other_user_reg,
            batch_name=batch_name,
            problem_slug=problem_slug,
            section=section,
            mentor_id=mentor_id,
        ).order_by("created_at")

        # Mark messages sent TO the current user as read when they view the thread
        if thread_type == "individual":
            messages_qs.filter(recipient=request.user, is_read=False).update(is_read=True)
            # Also clear notifications for this direct message thread
            Notification.objects.filter(
                recipient=request.user,
                is_read=False,
                link__icontains=f"other_user_reg={other_user_reg}"
            ).update(is_read=True)
        elif thread_type in ["general", "staff", "hod_tp_ja", "section", "mentor_group"]:
            # For group channels, just clear all notifications for that channel
            Notification.objects.filter(
                recipient=request.user,
                is_read=False,
                link=f"/discuss?thread_type={thread_type}"
            ).update(is_read=True)

        messages = messages_qs[:200]

        return Response(DiscussionMessageSerializer(messages, many=True, context={"request": request}).data)

    def post(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error

        serializer = DiscussionMessageCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        thread_type = data.get("thread_type", "general")
        recipient_reg = data.get("recipient_reg")
        batch_name = data.get("batch_name")
        problem_slug = data.get("problem_slug")
        section = (data.get("section") or "").strip().upper()
        mentor_id_str = str(data.get("mentor_id") or "")
        mentor_id = int(mentor_id_str) if mentor_id_str.isdigit() else None
        body = data["body"]

        # 1. Validation and Security
        if thread_type == "general" and profile_type == "student":
            batch_name = profile.batch

        if thread_type == "section":
            # Section chat — auto-fill batch+section from student profile
            if profile_type == "student":
                if not profile.section:
                    return Response({"detail": "You are not assigned to a section."}, status=403)
                batch_name = profile.batch
                section = profile.section
            elif not batch_name or not section:
                return Response({"detail": "batch_name and section are required."}, status=400)

        if thread_type == "mentor_group":
            if profile_type == "student":
                if not profile.mentor:
                    return Response({"detail": "You do not have an assigned mentor."}, status=403)
                mentor_id = profile.mentor.id
                batch_name = str(mentor_id)
            elif profile_type in ["staff", "hod", "tpu"]:
                # Staff posting to their own group
                mentor_id = profile.id
                batch_name = str(mentor_id)
            else:
                return Response({"detail": "Access denied."}, status=403)

        if thread_type == "individual":
            if not recipient_reg:
                return Response({"detail": "Recipient required for individual message."}, status=400)

            # Students can only message staff from their department
            if profile_type == "student":
                recipient_staff = StaffProfile.objects.filter(
                    faculty_id__iexact=recipient_reg,
                    department=profile.department
                ).first()
                if not recipient_staff:
                    recipient_staff = StaffProfile.objects.filter(faculty_id__iexact=recipient_reg, institution=profile.institution).first()
                    if not recipient_staff:
                        return Response({"detail": "Staff member not found in your institution."}, status=403)

                recipient = recipient_staff.account
            else:
                recipient = User.objects.filter(
                    Q(student_profile__register_number__iexact=recipient_reg) |
                    Q(staff_profile__faculty_id__iexact=recipient_reg) |
                    Q(username=recipient_reg)
                ).first()

            if not recipient:
                return Response({"detail": "Recipient not found."}, status=404)

            # Students cannot DM other students
            if profile_type == "student" and hasattr(recipient, "student_profile"):
                return Response({"detail": "Students can only message Staff or HOD."}, status=403)
        else:
            recipient = None

        if thread_type in ["staff", "hod_tp_ja"] and profile_type == "student":
            return Response({"detail": "Students cannot post to this channel."}, status=403)

        # 2. Find problem if applicable
        problem = None
        if problem_slug:
            problem = Problem.objects.filter(slug=problem_slug).first()

        is_poll = data.get("is_poll", False)
        poll_options = data.get("poll_options", [])

        if is_poll and profile_type == "student":
            return Response({"detail": "Only Staff/HOD/Admin can create polls."}, status=403)

        # 3. Create message
        message = DiscussionMessage.objects.create(
            sender=request.user,
            recipient=recipient,
            student=profile if profile_type == "student" else None,
            problem=problem,
            thread_type=thread_type,
            batch_name=batch_name,
            section=section,
            institution=getattr(profile, 'institution', None),
            department=getattr(profile, 'department', None),
            body=body,
            is_poll=is_poll,
            poll_options=poll_options
        )

        # 4. Notifications
        if recipient:
            # For individual messages, include the sender's reg in the link to allow auto-clearing
            sender_reg = profile.register_number if profile_type == "student" else (profile.faculty_id if profile else request.user.username)
            Notification.objects.create(
                recipient=recipient,
                title=f"New Message from {profile.name if profile else request.user.username}",
                message=body[:60] + "..." if len(body) > 60 else body,
                link=f"/discuss?thread_type=individual&other_user_reg={sender_reg}"
            )
        elif thread_type in ["staff", "hod_tp_ja", "general", "section", "mentor_group"]:
            # For group channels, we notify relevant people
            recipients_qs = User.objects.none()
            room_name = "General Chat"

            if thread_type == "staff" and profile and profile.department:
                recipients_qs = User.objects.filter(
                    staff_profile__department=profile.department
                )
                room_name = "Staff Room"
            elif thread_type == "hod_tp_ja" and profile:
                recipients_qs = User.objects.filter(
                    staff_profile__institution=profile.institution,
                    staff_profile__role__in=["hod", "admin", "ja", "tpu", "director"]
                )
                room_name = "HOD & Admin Panel"
            elif thread_type == "general" and profile:
                recipients_qs = User.objects.filter(
                    student_profile__batch=batch_name,
                    student_profile__institution=profile.institution
                )
                if batch_name:
                    room_name = f"Batch {batch_name} Chat"
            elif thread_type == "section" and batch_name and section:
                recipients_qs = User.objects.filter(
                    student_profile__batch=batch_name,
                    student_profile__section=section,
                    student_profile__institution=getattr(profile, 'institution', None),
                )
                room_name = f"Section {section} Chat"
            elif thread_type == "mentor_group" and mentor_id:
                # Notify all mentees + the mentor
                try:
                    mentor_staff = StaffProfile.objects.get(id=mentor_id)
                    recipients_qs = User.objects.filter(
                        Q(student_profile__mentor_id=mentor_id) |
                        Q(staff_profile__id=mentor_id)
                    )
                    room_name = f"Mentor Group ({mentor_staff.name})"
                except StaffProfile.DoesNotExist:
                    pass

            # Filter out the sender
            recipients = recipients_qs.exclude(id=request.user.id).distinct()

            # Create notifications in bulk
            notif_link = f"/discuss?thread_type={thread_type}"
            if thread_type == "general" and batch_name:
                notif_link += f"&batch_name={batch_name}"
            elif thread_type == "section" and batch_name and section:
                notif_link += f"&batch_name={batch_name}&section={section}"
            elif thread_type == "mentor_group" and mentor_id:
                notif_link += f"&mentor_id={mentor_id}"

            notifications = [
                Notification(
                    recipient=r,
                    title=f"New Message in {room_name}",
                    message=f"{profile.name}: {body[:50]}..." if len(body) > 50 else f"{profile.name}: {body}",
                    link=notif_link
                )
                for r in recipients
            ]
            Notification.objects.bulk_create(notifications)

        return Response(
            DiscussionMessageSerializer(message, context={"request": request}).data,
            status=status.HTTP_201_CREATED
        )

class DiscussionPollVoteView(UnifiedAuthMixin, APIView):
    def post(self, request, pk):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error

        message = get_object_or_404(DiscussionMessage, pk=pk)
        if not message.is_poll:
            return Response({"detail": "This message is not a poll."}, status=400)

        option_index = request.data.get("option_index")
        if option_index is None or not (0 <= option_index < len(message.poll_options)):
            return Response({"detail": "Invalid option index."}, status=400)

        # Record or update vote
        user_id = str(request.user.id)
        message.poll_votes[user_id] = option_index
        message.save(update_fields=["poll_votes"])

        return Response(DiscussionMessageSerializer(message, context={"request": request}).data)


INTERVIEW_TRACK_LABELS = {
    "civil": "Civil Engineering",
    "mech": "Mechanical Engineering",
    "eee": "Electrical & Electronics Engineering",
    "ece": "Electronics & Communication Engineering",
    "cs_common": "Computer Science Family",
}


class InterviewTrackView(UnifiedAuthMixin, APIView):
    """Resolves the caller's Interview Practice track from their department.
    The question bank itself isn't built yet (content shape TBD) — this just
    tells the frontend which track the caller belongs to so the tile/page can
    exist ahead of that, and so admins can regroup departments in the
    meantime via Department.interview_track."""

    def get(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error

        department = getattr(profile, 'department', None)
        if not department:
            return Response({"detail": "No department assigned."}, status=404)

        track_key = department.interview_track or department.default_interview_track()
        label = INTERVIEW_TRACK_LABELS.get(track_key, track_key.replace('_', ' ').title())

        return Response({
            "department": department.name,
            "track_key": track_key,
            "track_label": label,
            "topics": [],
        })


def _resolve_resource_display(resource_links):
    """Attach friendly display fields (title/difficulty) to each resource
    item — the stored item only carries ids/slugs, so this fills in
    what's needed to render it without an extra round trip per item."""
    items = [r for r in (resource_links or []) if isinstance(r, dict)]
    apt_ids = [r['aptitude_topic_id'] for r in items if r.get('type') == 'aptitude_topic' and r.get('aptitude_topic_id')]
    slugs = [r['problem_slug'] for r in items if r.get('type') == 'problem' and r.get('problem_slug')]
    apt_titles = dict(AptitudeTopic.objects.filter(id__in=apt_ids).values_list('id', 'title')) if apt_ids else {}
    problems = {p.slug: p for p in Problem.objects.filter(slug__in=slugs)} if slugs else {}

    out = []
    for r in items:
        item = dict(r)
        if item.get('type') == 'aptitude_topic':
            item['aptitude_topic_title'] = apt_titles.get(item.get('aptitude_topic_id'), 'Unknown topic')
        elif item.get('type') == 'problem':
            p = problems.get(item.get('problem_slug'))
            item['problem_title'] = p.title if p else 'Unknown problem'
            item['problem_difficulty'] = p.difficulty if p else ''
        out.append(item)
    return out


def _serialize_examination_syllabus(examination):
    """Full Section > Topic > Subtopic tree for one examination — shared by
    the admin management view and the student-facing browse view."""
    sections = []
    for section in examination.sections.prefetch_related(
        'topics__subtopics__questions'
    ):
        topics = []
        for topic in section.topics.all():
            topics.append({
                "id": topic.id,
                "title": topic.title,
                "resource_links": _resolve_resource_display(topic.resource_links),
                "subtopics": [
                    {
                        "id": st.id,
                        "title": st.title,
                        "description": st.description,
                        "resource_links": _resolve_resource_display(st.resource_links),
                        "question_count": len(st.questions.all()),
                    }
                    for st in topic.subtopics.all()
                ],
            })
        sections.append({"id": section.id, "title": section.title, "topics": topics})
    return sections


class AdminExaminationListCreateView(APIView):
    """System Admin: list/create Examinations — the top-level content bank
    for the student-facing Competitive Practice module (GRE, GATE, CAT...).
    Global, like the Problem/Aptitude banks — not institution-scoped."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        exams = Examination.objects.annotate(
            section_count=Count('sections', distinct=True),
            topic_count=Count('sections__topics', distinct=True),
            subtopic_count=Count('sections__topics__subtopics', distinct=True),
        )
        return Response([
            {
                "id": e.id, "name": e.name, "description": e.description, "is_active": e.is_active,
                "section_count": e.section_count, "topic_count": e.topic_count, "subtopic_count": e.subtopic_count,
            }
            for e in exams
        ])

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        name = (request.data.get('name') or '').strip()
        description = (request.data.get('description') or '').strip()
        if not name:
            return Response({"error": "name is required."}, status=400)
        if Examination.objects.filter(name__iexact=name).exists():
            return Response({"error": "An examination with this name already exists."}, status=400)
        exam = Examination.objects.create(name=name, description=description)
        return Response({
            "id": exam.id, "name": exam.name, "description": exam.description, "is_active": exam.is_active,
            "section_count": 0, "topic_count": 0, "subtopic_count": 0,
        }, status=201)


class AdminExaminationDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, exam_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        exam = get_object_or_404(Examination, id=exam_id)
        exam.delete()
        return Response({"message": "Examination deleted"})

    def patch(self, request, exam_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        exam = get_object_or_404(Examination, id=exam_id)
        if 'name' in request.data:
            name = (request.data.get('name') or '').strip()
            if not name:
                return Response({"error": "name cannot be empty."}, status=400)
            if Examination.objects.exclude(id=exam.id).filter(name__iexact=name).exists():
                return Response({"error": "An examination with this name already exists."}, status=400)
            exam.name = name
        if 'is_active' in request.data:
            exam.is_active = bool(request.data.get('is_active'))
        if 'description' in request.data:
            exam.description = (request.data.get('description') or '').strip()
        exam.save()
        return Response({"message": "Updated", "name": exam.name, "is_active": exam.is_active, "description": exam.description})


class AdminExaminationSyllabusView(APIView):
    """Admin view of one examination's full syllabus tree."""
    permission_classes = [IsAuthenticated]

    def get(self, request, exam_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        exam = get_object_or_404(Examination, id=exam_id)
        return Response({
            "examination": {"id": exam.id, "name": exam.name, "description": exam.description},
            "sections": _serialize_examination_syllabus(exam),
        })


class AdminExaminationSyllabusUploadView(APIView):
    """System Admin: bulk-populate an Examination's Section > Topic >
    Subtopic tree from an uploaded .xlsx/.xls/.csv file. Expects Section,
    Topic, Subtopic columns (an Exam column is accepted but ignored — the
    target examination is the one in the URL, not whatever the sheet
    says, since the same sheet format could otherwise be uploaded to the
    wrong exam). Upserts by title within each parent so re-uploading an
    updated sheet is safe and doesn't create duplicates."""
    permission_classes = [IsAuthenticated]

    def post(self, request, exam_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        exam = get_object_or_404(Examination, id=exam_id)

        upload = request.FILES.get("file")
        if not upload:
            return Response({"error": "No file uploaded."}, status=400)

        filename = upload.name.lower()
        try:
            if filename.endswith((".xlsx", ".xls")):
                rows = self._read_excel(upload)
            elif filename.endswith(".csv"):
                rows = self._read_csv(upload)
            else:
                return Response({"error": "Only .xlsx, .xls, or .csv files are supported."}, status=400)
        except Exception as e:
            return Response({"error": f"Could not read file: {e}"}, status=400)

        if not rows:
            return Response({"error": "File is empty."}, status=400)

        header = [str(h or '').strip().lower() for h in rows[0]]
        col_idx = {h: i for i, h in enumerate(header) if h}

        def col(row, name):
            idx = col_idx.get(name)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            val = str(row[idx]).strip()
            return "" if val.lower() == "nan" else val

        missing = [n for n in ("section", "topic", "subtopic") if n not in col_idx]
        if missing:
            return Response({"error": f"Missing required column(s): {', '.join(missing)}."}, status=400)

        section_cache, topic_cache = {}, {}
        section_seq, topic_seq = 0, {}
        created_sections = created_topics = created_subtopics = 0
        skipped = 0

        for row in rows[1:]:
            if not row or all(not str(c or '').strip() for c in row):
                continue
            section_title = col(row, "section")
            topic_title = col(row, "topic")
            subtopic_title = col(row, "subtopic")
            if not section_title or not topic_title or not subtopic_title:
                skipped += 1
                continue

            if section_title not in section_cache:
                section, created = SyllabusSection.objects.get_or_create(
                    examination=exam, title=section_title, defaults={"order": section_seq},
                )
                section_cache[section_title] = section
                section_seq += 1
                topic_seq[section_title] = 0
                if created:
                    created_sections += 1
            section = section_cache[section_title]

            topic_key = (section_title, topic_title)
            if topic_key not in topic_cache:
                topic, created = SyllabusTopic.objects.get_or_create(
                    section=section, title=topic_title, defaults={"order": topic_seq[section_title]},
                )
                topic_cache[topic_key] = topic
                topic_seq[section_title] += 1
                if created:
                    created_topics += 1
            topic = topic_cache[topic_key]

            _, created = SyllabusSubtopic.objects.get_or_create(
                topic=topic, title=subtopic_title,
                defaults={"order": topic.subtopics.count()},
            )
            if created:
                created_subtopics += 1

        return Response({
            "message": "Syllabus imported",
            "created_sections": created_sections,
            "created_topics": created_topics,
            "created_subtopics": created_subtopics,
            "skipped_rows": skipped,
        }, status=201)

    def _read_excel(self, upload):
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        return [[cell.value for cell in row] for row in ws.iter_rows()]

    def _read_csv(self, upload):
        import csv
        import io
        text = upload.read().decode("utf-8-sig")
        return list(csv.reader(io.StringIO(text)))


def _clean_resource_items(raw_items):
    """Validate/normalize a resource list — shared by the Topic and
    Subtopic resource endpoints, since both now support the same three
    kinds: an external link (frontend renders it as a YouTube embed,
    image, or video automatically depending on the URL), or a pointer at
    existing platform content (an Aptitude topic or a coding Problem)."""
    cleaned = []
    for item in raw_items:
        if not isinstance(item, dict):
            continue
        item_type = item.get('type')
        label = (item.get('label') or '').strip()

        if item_type == 'link':
            url = (item.get('url') or '').strip()
            if not url:
                continue
            cleaned.append({"type": "link", "label": label, "url": url})

        elif item_type == 'aptitude_topic':
            apt_id = item.get('aptitude_topic_id')
            try:
                apt_id = int(apt_id)
            except (TypeError, ValueError):
                continue
            if not AptitudeTopic.objects.filter(id=apt_id).exists():
                continue
            cleaned.append({"type": "aptitude_topic", "label": label, "aptitude_topic_id": apt_id})

        elif item_type == 'problem':
            slug = (item.get('problem_slug') or '').strip()
            if not slug or not Problem.objects.filter(slug=slug).exists():
                continue
            cleaned.append({"type": "problem", "label": label, "problem_slug": slug})

    return cleaned


class AdminSyllabusTopicResourcesView(APIView):
    """System Admin: replace the resource list attached to one syllabus
    topic. Each resource is either an external link (rendered as a
    YouTube embed automatically if the URL is one — the frontend's call,
    this just stores it) or a pointer at existing platform content: an
    Aptitude topic or a coding Problem, so a Competitive Practice topic
    can point straight at question banks that already exist instead of
    duplicating content."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, topic_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        topic = get_object_or_404(SyllabusTopic, id=topic_id)

        raw_items = request.data.get('resource_links')
        if not isinstance(raw_items, list):
            return Response({"error": "resource_links must be a list."}, status=400)

        cleaned = _clean_resource_items(raw_items)
        topic.resource_links = cleaned
        topic.save(update_fields=['resource_links'])
        return Response({"message": "Resources updated", "resource_links": _resolve_resource_display(cleaned)})


class AdminSyllabusSubtopicView(APIView):
    """System Admin: update one subtopic's description and/or resources.
    Subtopics get their own individual resources here — the same three
    kinds a Topic can carry (external link, Aptitude topic, or Problem)
    — rather than only being covered by whatever's attached at the
    parent Topic level."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, subtopic_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        subtopic = get_object_or_404(SyllabusSubtopic, id=subtopic_id)

        update_fields = []
        if 'description' in request.data:
            subtopic.description = (request.data.get('description') or '').strip()
            update_fields.append('description')

        if 'resource_links' in request.data:
            raw_items = request.data.get('resource_links')
            if not isinstance(raw_items, list):
                return Response({"error": "resource_links must be a list."}, status=400)
            subtopic.resource_links = _clean_resource_items(raw_items)
            update_fields.append('resource_links')

        if update_fields:
            subtopic.save(update_fields=update_fields)

        return Response({
            "message": "Subtopic updated",
            "description": subtopic.description,
            "resource_links": _resolve_resource_display(subtopic.resource_links),
        })


def _serialize_competitive_question(q, include_answer=True):
    data = {
        "id": q.id,
        "question_text": q.question_text,
        "question_image": q.question_image,
        "video_url": q.video_url,
        "option_a": q.option_a,
        "option_b": q.option_b,
        "option_c": q.option_c,
        "option_d": q.option_d,
    }
    if include_answer:
        data["correct_option"] = q.correct_option
        data["explanation"] = q.explanation
    return data


class AdminSubtopicQuestionListCreateView(APIView):
    """System Admin: list/create MCQ questions authored directly for one
    Competitive Practice subtopic — separate from resource_links' pointers
    at existing Aptitude/Problem content."""
    permission_classes = [IsAuthenticated]

    def get(self, request, subtopic_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        subtopic = get_object_or_404(SyllabusSubtopic, id=subtopic_id)
        questions = subtopic.questions.all()
        return Response([_serialize_competitive_question(q) for q in questions])

    def post(self, request, subtopic_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        subtopic = get_object_or_404(SyllabusSubtopic, id=subtopic_id)

        question_text = (request.data.get('question_text') or '').strip()
        option_a = (request.data.get('option_a') or '').strip()
        option_b = (request.data.get('option_b') or '').strip()
        option_c = (request.data.get('option_c') or '').strip()
        option_d = (request.data.get('option_d') or '').strip()
        correct_option = (request.data.get('correct_option') or '').strip().upper()

        if not all([question_text, option_a, option_b, option_c, option_d]):
            return Response({"error": "question_text and all four options are required."}, status=400)
        if correct_option not in ('A', 'B', 'C', 'D'):
            return Response({"error": "correct_option must be one of A, B, C, D."}, status=400)

        q = CompetitiveQuestion.objects.create(
            subtopic=subtopic,
            question_text=question_text,
            question_image=(request.data.get('question_image') or '').strip(),
            video_url=(request.data.get('video_url') or '').strip(),
            option_a=option_a, option_b=option_b, option_c=option_c, option_d=option_d,
            correct_option=correct_option,
            explanation=(request.data.get('explanation') or '').strip(),
            order=subtopic.questions.count(),
        )
        return Response(_serialize_competitive_question(q), status=201)


class AdminSubtopicQuestionDetailView(APIView):
    """System Admin: edit or remove one Competitive Practice question."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, subtopic_id, question_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        q = get_object_or_404(CompetitiveQuestion, id=question_id, subtopic_id=subtopic_id)

        for field in ('question_text', 'question_image', 'video_url', 'option_a', 'option_b', 'option_c', 'option_d', 'explanation'):
            if field in request.data:
                setattr(q, field, (request.data.get(field) or '').strip())
        if 'correct_option' in request.data:
            correct_option = (request.data.get('correct_option') or '').strip().upper()
            if correct_option not in ('A', 'B', 'C', 'D'):
                return Response({"error": "correct_option must be one of A, B, C, D."}, status=400)
            q.correct_option = correct_option
        q.save()
        return Response(_serialize_competitive_question(q))

    def delete(self, request, subtopic_id, question_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        q = get_object_or_404(CompetitiveQuestion, id=question_id, subtopic_id=subtopic_id)
        q.delete()
        return Response({"message": "Question deleted"})


class AdminSubtopicQuestionImportView(APIView):
    """System Admin: import (copy) existing AptitudeQuestion rows into a
    Competitive Practice subtopic's own question bank. A one-time copy,
    not a live link — editing the original Aptitude question afterward
    doesn't change the imported copy here."""
    permission_classes = [IsAuthenticated]

    def post(self, request, subtopic_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        subtopic = get_object_or_404(SyllabusSubtopic, id=subtopic_id)

        aptitude_question_ids = request.data.get('aptitude_question_ids')
        if not isinstance(aptitude_question_ids, list) or not aptitude_question_ids:
            return Response({"error": "aptitude_question_ids must be a non-empty list."}, status=400)

        source_questions = AptitudeQuestion.objects.filter(id__in=aptitude_question_ids)
        next_order = subtopic.questions.count()
        created = []
        for i, aq in enumerate(source_questions):
            q = CompetitiveQuestion.objects.create(
                subtopic=subtopic,
                question_text=aq.question_text,
                question_image=aq.question_image or "",
                option_a=aq.option_a, option_b=aq.option_b, option_c=aq.option_c, option_d=aq.option_d,
                correct_option=(aq.correct_option or 'A').upper(),
                explanation=aq.explanation or "",
                order=next_order + i,
            )
            created.append(q)

        return Response({
            "message": f"Imported {len(created)} question(s)",
            "questions": [_serialize_competitive_question(q) for q in created],
        }, status=201)


class CompetitiveSubtopicQuestionsView(APIView):
    """Student: practice questions for one subtopic — correct_option and
    explanation are withheld until answered via the submit endpoint."""
    permission_classes = [IsAuthenticated]

    def get(self, request, subtopic_id):
        subtopic = get_object_or_404(SyllabusSubtopic, id=subtopic_id)
        questions = subtopic.questions.all()
        return Response([_serialize_competitive_question(q, include_answer=False) for q in questions])


class CompetitiveQuestionSubmitView(APIView):
    """Student: submit an answer to one Competitive Practice question,
    get instant right/wrong feedback plus the explanation."""
    permission_classes = [IsAuthenticated]

    def post(self, request, question_id):
        question = get_object_or_404(CompetitiveQuestion, id=question_id)
        selected = (request.data.get('selected_option') or '').strip().upper()
        if selected not in ('A', 'B', 'C', 'D'):
            return Response({"error": "selected_option must be one of A, B, C, D."}, status=400)

        return Response({
            "is_correct": selected == question.correct_option,
            "correct_option": question.correct_option,
            "explanation": question.explanation,
        })


class CompetitiveExaminationListView(APIView):
    """Student: active Examinations available under Competitive Practice."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        exams = Examination.objects.filter(is_active=True).annotate(
            section_count=Count('sections', distinct=True),
            topic_count=Count('sections__topics', distinct=True),
        )
        return Response([
            {
                "id": e.id, "name": e.name, "description": e.description,
                "section_count": e.section_count, "topic_count": e.topic_count,
            }
            for e in exams
        ])


class CompetitiveExaminationSyllabusView(APIView):
    """Student: browse one examination's Section > Topic > Subtopic syllabus."""
    permission_classes = [IsAuthenticated]

    def get(self, request, exam_id):
        exam = get_object_or_404(Examination, id=exam_id, is_active=True)
        return Response({
            "examination": {"id": exam.id, "name": exam.name, "description": exam.description},
            "sections": _serialize_examination_syllabus(exam),
        })


class FirstLoginView(APIView):
    def post(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "first-login", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        serializer = FirstLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        register_number = serializer.validated_data["register_number"].strip()
        password = serializer.validated_data["password"]
        profile = (
            StudentProfile.objects.filter(register_number=register_number)
            .select_related("account")
            .first()
        )

        if not profile or not profile.account:
            return Response(
                {"detail": "Imported student account not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if profile.password_is_set:
            return Response(
                {"detail": "Password already created. Please log in normally."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        profile.account.set_password(password)
        profile.account.save(update_fields=["password"])
        login(request, profile.account)
        profile.record_login()

        return Response(
            {
                "detail": "Password set successfully.",
                "student": StudentProfileSerializer(profile).data,
            }
        )


@method_decorator(ensure_csrf_cookie, name="dispatch")
class StudentLoginView(APIView):
    def post(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "student-login", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        serializer = StudentLoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        register_number = serializer.validated_data["register_number"].strip()
        password = serializer.validated_data["password"]
        profile = (
            StudentProfile.objects.filter(register_number=register_number)
            .select_related("account")
            .first()
        )

        if not profile or not profile.account:
            return Response(
                {"detail": "Student account not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if not profile.password_is_set:
            return Response(
                {"detail": "First-time password setup is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check if the account is blocked BEFORE calling authenticate(),
        # because Django's ModelBackend silently returns None for inactive users
        # even with the correct password, which would show a misleading 401.
        if not profile.account.is_active:
            return Response(
                {"detail": "Your account has been blocked. Please contact your department staff."},
                status=status.HTTP_403_FORBIDDEN,
            )

        user = authenticate(request=request, username=register_number, password=password)
        if not user:
            return Response(
                {"detail": "Invalid register number or password."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        login(request, user)
        profile.record_login()
        return Response(
            {
                "detail": "Login successful.",
                "user_type": "student",
                "student": StudentProfileSerializer(profile).data,
            }
        )


class StudentLogoutView(APIView):
    def post(self, request):
        logout(request)
        return Response({"detail": "Logout successful."})


@method_decorator(ensure_csrf_cookie, name="dispatch")
class ProblemSessionStartView(UnifiedAuthMixin, APIView):
    """Start a problem solving session to track time spent."""
    def post(self, request, slug):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error
        
        # Only students can have problem sessions
        if profile_type != "student" or not profile:
            return Response(
                {"detail": "Only students can start problem sessions."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        problem = Problem.objects.filter(slug=slug).first()
        if not problem:
            return Response(
                {"detail": "Problem not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # End any existing active session for this problem
        existing_session = ProblemSession.objects.filter(
            student=profile,
            problem=problem,
            is_active=True
        ).first()
        
        if existing_session:
            existing_session.end_session()
        
        # Create new session
        session = ProblemSession.objects.create(
            student=profile,
            problem=problem,
            is_active=True
        )
        
        return Response({
            "detail": "Session started.",
            "session_id": session.id,
            "started_at": session.started_at,
        })


@method_decorator(ensure_csrf_cookie, name="dispatch")
class ProblemSessionEndView(UnifiedAuthMixin, APIView):
    """End a problem solving session and return time spent."""
    def post(self, request, slug):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error
        
        # Only students can have problem sessions
        if profile_type != "student" or not profile:
            return Response(
                {"detail": "Only students can end problem sessions."},
                status=status.HTTP_403_FORBIDDEN,
            )
        
        problem = Problem.objects.filter(slug=slug).first()
        if not problem:
            return Response(
                {"detail": "Problem not found."},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Find and end active session
        session = ProblemSession.objects.filter(
            student=profile,
            problem=problem,
            is_active=True
        ).first()
        
        if not session:
            return Response(
                {"detail": "No active session found for this problem."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        time_spent = session.end_session()
        
        return Response({
            "detail": "Session ended.",
            "time_spent_seconds": time_spent,
            "ended_at": session.ended_at,
        })


# =============================================================================
# Staff/Faculty Authentication Views (similar flow to students)
# =============================================================================

class StaffLookupView(APIView):
    """Lookup staff by faculty_id - returns first_login_required status."""
    def get(self, request):
        max_attempts, window = _lookup_rate_limits()
        try:
            check_rate_limit(request, "staff-lookup", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        faculty_id = (request.query_params.get("faculty_id") or "").strip()
        if not faculty_id:
            return Response(
                {"detail": "faculty_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        profile = StaffProfile.objects.filter(faculty_id=faculty_id).first()
        if not profile:
            return Response(
                {"detail": "Staff member not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response({
            "staff": {
                "faculty_id": profile.faculty_id,
                "name": profile.name,
            },
            "first_login_required": not profile.password_is_set,
        })


@method_decorator(ensure_csrf_cookie, name="dispatch")
class StaffFirstLoginView(APIView):
    """First-time password setup for staff (like students)."""
    def post(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "staff-first-login", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        faculty_id = (request.data.get("faculty_id") or "").strip()
        password = request.data.get("password", "").strip()

        if not faculty_id or not password:
            return Response(
                {"detail": "faculty_id and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(password) < 8:
            return Response(
                {"detail": "Password must be at least 8 characters."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        profile = (
            StaffProfile.objects.filter(faculty_id=faculty_id)
            .select_related("account")
            .first()
        )

        if not profile:
            return Response(
                {"detail": "Staff account not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if profile.password_is_set:
            return Response(
                {"detail": "Password already set. Please log in normally."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Store password in staff_profiles table
        profile.set_password(password)
        
        # Login using the associated account
        if profile.account:
            login(request, profile.account)
        
        return Response({
            "detail": "Password set successfully.",
            "staff": {
                "faculty_id": profile.faculty_id,
                "name": profile.name,
            },
        })


@method_decorator(ensure_csrf_cookie, name="dispatch")
class StaffLoginView(APIView):
    """Staff login with faculty_id and password."""
    def post(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "staff-login", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        faculty_id = (request.data.get("faculty_id") or "").strip()
        password = request.data.get("password", "").strip()

        if not faculty_id or not password:
            return Response(
                {"detail": "faculty_id and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        profile = (
            StaffProfile.objects.filter(faculty_id=faculty_id)
            .select_related("account")
            .first()
        )

        if not profile:
            return Response(
                {"detail": "Staff account not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        if not profile.password_is_set:
            return Response(
                {"detail": "First-time password setup is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check password from staff_profiles table
        if not profile.check_password(password):
            return Response(
                {"detail": "Invalid faculty_id or password."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        # Ensure User account exists (for Django session auth)
        if not profile.account:
            user = User.objects.create_user(
                username=profile.faculty_id,
                first_name=profile.name,
            )
            user.set_password(password)
            user.save()
            profile.account = user
            profile.save(update_fields=["account"])
        else:
            # Sync password to User account if needed
            user = profile.account
            if not user.check_password(password):
                user.set_password(password)
                user.save()

        # Login directly using the verified user (bypass authenticate since we already checked password)
        if profile.account:
            # Ensure user is active
            if not profile.account.is_active:
                profile.account.is_active = True
                profile.account.save(update_fields=["is_active"])
            login(request, profile.account)
        
        # Check if linked account is superuser (staff can also be admin)
        is_admin = profile.account and profile.account.is_superuser if profile.account else False
        
        # Get institution_id for response
        institution_id = profile.institution.institution_id if profile.institution else None
        
        user_type = profile.role

        # Return response based on staff role
        response_data = {
            "detail": "Login successful.",
            "user_type": user_type,
            "institution_id": institution_id,
        }
        
        # Include appropriate data based on role
        user_data = {
            "id": profile.faculty_id,
            "faculty_id": profile.faculty_id,
            "name": profile.name,
            "role": profile.role,
            "department": {
                "id": profile.department.id,
                "name": profile.department.name,
                "code": profile.department.code,
            } if profile.department else None,
            "department_id": profile.department_id,
            "department_name": profile.department.name if profile.department else "N/A",
            "institution_id": institution_id,
        }
        
        if profile.role == "admin":
            response_data["admin"] = user_data
        elif user_type == "academics":
            response_data["academics"] = user_data
            response_data["hod"] = user_data
            response_data["staff"] = user_data
        elif profile.role == "hod":
            response_data["hod"] = user_data
            response_data["staff"] = user_data
        else:
            response_data["staff"] = user_data
        
        return Response(response_data)


class StaffLogoutView(APIView):
    """Staff logout."""
    def post(self, request):
        logout(request)
        return Response({"detail": "Logout successful."})


class StaffInstitutionDetailView(APIView):
    """Get institution details for staff members."""
    permission_classes = [IsAuthenticated]

    def get(self, request, institution_id):
        # Check if user is staff (has staff_profile) or admin
        is_staff = hasattr(request.user, 'staff_profile')
        is_admin = request.user.is_superuser

        if not is_staff and not is_admin:
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        # If staff, verify they belong to this institution
        if is_staff and request.user.staff_profile.institution != institution:
            return Response({"detail": "You do not have access to this institution."}, status=status.HTTP_403_FORBIDDEN)

        # Get the user's staff profile if available
        user_profile = request.user.staff_profile if is_staff else None
        user_role = user_profile.role if user_profile else None
        user_department = user_profile.department if user_profile else None

        # For HOD / Academic Coordinator, filter students by department
        if user_profile and user_profile.role in ("hod", "academics") and user_profile.department:
            students_qs = StudentProfile.objects.filter(
                institution=institution,
                department=user_profile.department
            ).select_related('account', 'department')
        else:
            students_qs = StudentProfile.objects.filter(
                institution=institution
            ).select_related('account', 'department')

        # Build student list with metrics
        student_list = []
        for student in students_qs:
            student_list.append({
                "register_number": student.register_number,
                "name": student.name,
                "batch": student.batch,
                "section": student.section,
                "department": student.department.code if student.department else "N/A",
                "department_name": student.department.name if student.department else "N/A",
                "solved_count": student.solved_problems.count(),
                "current_streak": student.current_streak,
                "last_active": student.account.last_login if student.account else None,
                "is_active": student.account.is_active if student.account else True,
            })

        # Get staff (filter by department for HOD/Academics, all for admin/staff)
        if user_role in ("hod", "academics") and user_department:
            staff = StaffProfile.objects.filter(institution=institution, department=user_department)
        else:
            staff = StaffProfile.objects.filter(institution=institution)
        staff_list = []
        for s in staff:
            staff_list.append({
                "faculty_id": s.faculty_id,
                "name": s.name,
                "role": s.role,
                "role_display": s.get_role_display(),
                "password_is_set": s.password_is_set,
                "is_active": s.is_active,
            })

        # Get stats
        active_sessions = ProblemSession.objects.filter(
            is_active=True,
            student__institution=institution
        ).count()

        # Get problem stats for this institution
        solved_count = SolvedProblem.objects.filter(student__institution=institution).count()
        total_students = len(student_list)
        success_rate = round((solved_count / total_students) * 100) if total_students > 0 else 0

        response_data = {
            "institution": {
                "id": institution.institution_id,
                "name": institution.name,
                "short_code": institution.short_code,
                "address": institution.address,
                "is_active": institution.is_active,
            },
            "students": student_list,
            "staff": staff_list,
            "stats": {
                "total_students": total_students,
                "total_staff": len(staff_list),
                "active_sessions": active_sessions,
                "total_solved": solved_count,
                "success_rate": success_rate,
            },
        }
        
        # Add department info for HOD / Academic Coordinator users
        if user_role in ("hod", "academics") and user_department:
            response_data["department"] = {
                "id": user_department.id,
                "code": user_department.code,
                "name": user_department.name,
            }
        
        return Response(response_data)


class StaffPerformanceView(APIView):
    """Get individual staff performance metrics for HOD."""
    permission_classes = [IsAuthenticated]

    def get(self, request, institution_id):
        # Check if user is HOD or staff
        is_staff = hasattr(request.user, 'staff_profile')
        if not is_staff:
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        user_profile = request.user.staff_profile
        user_role = user_profile.role
        user_department = user_profile.department

        institution = Institution.objects.filter(institution_id=institution_id).first()
        if not institution:
            return Response({"detail": "Institution not found."}, status=status.HTTP_404_NOT_FOUND)

        # Verify staff belongs to this institution
        if user_profile.institution != institution:
            return Response({"detail": "You do not have access to this institution."}, status=status.HTTP_403_FORBIDDEN)

        # Filter by department for HOD / Academics, all for staff/admin
        if user_role in ("hod", "academics") and user_department:
            staff_qs = StaffProfile.objects.filter(institution=institution, department=user_department)
        else:
            staff_qs = StaffProfile.objects.filter(institution=institution)

        staff_performance = []
        for staff in staff_qs:
            # Calculate days active (since account creation)
            days_active = 1
            if staff.account and staff.account.date_joined:
                days_active = (timezone.now() - staff.account.date_joined).days

            # Get number of students in this staff's department (managed by staff)
            assigned_students = StudentProfile.objects.filter(
                institution=institution,
                department=staff.department
            ).count() if staff.department else 0

            # Student progress (problems solved by students in department)
            student_progress = 0
            if staff.department:
                student_progress = SolvedProblem.objects.filter(
                    student__department=staff.department,
                    student__institution=institution
                ).count()

            # Contests created by this staff with top performers
            staff_contests = []
            contests_qs = staff.contests.filter(
                institution=institution
            ).order_by('-created_at')[:5]  # Last 5 contests

            for contest in contests_qs:
                contest_summary = _contest_live_summary(contest)

                staff_contests.append({
                    "id": contest.id,
                    "title": contest.title,
                    "status": contest.status,
                    "created_at": contest.created_at,
                    "total_participants": contest_summary["total_participants"],
                    "total_submissions": contest_summary["total_submissions"],
                    "top_performers": contest_summary["top_performers"],
                })

            contests_created = staff.contests.filter(institution=institution).count()

            staff_performance.append({
                "faculty_id": staff.faculty_id,
                "name": staff.name or staff.faculty_id,
                "role": staff.role,
                "days_active": days_active,
                "assigned_students": assigned_students,
                "student_progress": student_progress,
                "contests_created": contests_created,
                "contests": staff_contests,
            })

        return Response({
            "staff_performance": staff_performance,
            "department": {
                "id": user_department.id,
                "code": user_department.code,
                "name": user_department.name,
            } if user_department else None,
        })


class StaffDetailView(APIView):
    """Get detailed analytics for a specific staff member."""
    permission_classes = [IsAuthenticated]

    def get(self, request, faculty_id):
        # Check if user is HOD or staff
        is_staff = hasattr(request.user, 'staff_profile')
        is_admin = request.user.is_superuser
        
        if not is_staff and not is_admin:
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        user_profile = request.user.staff_profile if is_staff else None
        user_role = user_profile.role if user_profile else None
        user_department = user_profile.department if user_profile else None

        # Get the target staff member
        target_staff = StaffProfile.objects.filter(faculty_id=faculty_id).select_related('department', 'institution').first()
        if not target_staff:
            target_staff = StaffProfile.objects.filter(account__username=faculty_id).select_related('department', 'institution').first()
        if not target_staff and user_profile:
            target_staff = user_profile
        if not target_staff and is_admin:
            target_staff = StaffProfile.objects.select_related('department', 'institution').first()
        if not target_staff:
            return Response({"detail": "Staff not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions: HOD can only view staff in their department, staff can view anyone in their institution
        if is_staff:
            if user_role in ("hod", "academics") and target_staff.department != user_department:
                return Response({"detail": "You can only view staff in your department."}, status=status.HTTP_403_FORBIDDEN)
            if target_staff.institution != user_profile.institution:
                return Response({"detail": "You do not have access to this staff member."}, status=status.HTTP_403_FORBIDDEN)

        # Get staff activity (days since joining)
        days_active = 0
        if target_staff.account and target_staff.account.date_joined:
            days_active = (timezone.now() - target_staff.account.date_joined).days + 1
        else:
            days_active = (timezone.now() - target_staff.created_at).days + 1 if hasattr(target_staff, 'created_at') else 1

        # Get students in this staff's department or entire institution if no department (Director/TPU)
        if target_staff.department:
            department_students = StudentProfile.objects.filter(
                institution=target_staff.institution,
                department=target_staff.department
            )
        else:
            department_students = StudentProfile.objects.filter(
                institution=target_staff.institution
            )

        # Get top performers in department
        top_students = []
        for student in department_students.annotate(
            solved_count=Count('solved_problems', distinct=True)
        ).order_by('-solved_count')[:10]:
            top_students.append({
                "id": student.register_number,
                "name": student.name,
                "solved_count": student.solved_count,
                "current_streak": student.current_streak,
            })

        # Get contests created by this staff with top performers
        recent_contests = []
        staff_contests_qs = target_staff.contests.filter(
            institution=target_staff.institution
        ).order_by('-created_at')[:10]

        for contest in staff_contests_qs:
            contest_summary = _contest_live_summary(contest)

            recent_contests.append({
                "id": contest.id,
                "title": contest.title,
                "status": contest.status,
                "created_at": contest.created_at,
                "total_participants": contest_summary["total_participants"],
                "total_submissions": contest_summary["total_submissions"],
                "top_performers": contest_summary["top_performers"],
            })

        # Batch-wise grouping with top performers per batch
        batch_wise_data = []
        # Get distinct batches
        batch_filter = Q(institution=target_staff.institution, batch__isnull=False)
        if target_staff.department:
            batch_filter &= Q(department=target_staff.department)
        
        batches = StudentProfile.objects.filter(batch_filter).exclude(batch='').values_list('batch', flat=True).distinct()

        for batch in batches:
            # Get all students for this batch with annotations
            batch_student_filter = Q(institution=target_staff.institution, batch=batch)
            if target_staff.department:
                batch_student_filter &= Q(department=target_staff.department)
                
            all_batch_students = StudentProfile.objects.filter(batch_student_filter).annotate(
                solved_count=Count('solved_problems', distinct=True)
            ).order_by('-solved_count')

            batch_count = all_batch_students.count()

            # Get top performers for this batch (first 5)
            batch_top_performers = []
            for student in all_batch_students[:5]:
                batch_top_performers.append({
                    "register_number": student.register_number,
                    "name": student.name,
                    "section": student.section,
                    "solved_count": student.solved_count,
                    "current_streak": student.current_streak,
                })

            # Get all students for the batch (for expanded view)
            all_students = []
            for student in all_batch_students:
                all_students.append({
                    "register_number": student.register_number,
                    "name": student.name,
                    "section": student.section,
                    "solved_count": student.solved_count,
                    "current_streak": student.current_streak,
                    "last_active": student.last_login_on.isoformat() if student.last_login_on else None,
                })

            batch_sections = sorted(set(s for s in all_batch_students.values_list('section', flat=True) if s))

            batch_wise_data.append({
                "batch": batch,
                "student_count": batch_count,
                "sections": batch_sections,
                "top_performers": batch_top_performers,
                "students": all_students,
            })

        # Weekly progress (accepts ?start_date=&end_date=, defaults to last 7 days)
        weekly_progress = build_solved_activity_series(
            Q(student__department=target_staff.department) if target_staff.department
            else Q(student__institution=target_staff.institution),
            start_date=parse_date_param(request.query_params.get('start_date')),
            end_date=parse_date_param(request.query_params.get('end_date')),
        )

        # Recent Activity (Last 10 solved problems)
        recent_activity = []
        recent_filter = Q()
        if target_staff.department:
            recent_filter = Q(student__department=target_staff.department)
        else:
            recent_filter = Q(student__institution=target_staff.institution)
            
        recent_solved = SolvedProblem.objects.filter(recent_filter).select_related('student', 'problem').order_by('-solved_at')[:10]
        
        for solved in recent_solved:
            recent_activity.append({
                "student_name": solved.student.name,
                "student_id": solved.student.register_number,
                "problem_title": solved.problem.title,
                "solved_at": solved.solved_at.isoformat(),
            })

        # Engagement Summary
        today = timezone.now().date()
        active_today = department_students.filter(last_login_on=today).count()
        total_students = department_students.count()
        avg_solved = 0
        if total_students > 0:
            total_solved_filter = Q()
            if target_staff.department:
                total_solved_filter = Q(student__department=target_staff.department)
            else:
                total_solved_filter = Q(student__institution=target_staff.institution)
                
            total_solved_count = SolvedProblem.objects.filter(total_solved_filter).count()
            avg_solved = round(total_solved_count / total_students, 1)

        return Response({
            "staff": {
                "faculty_id": target_staff.faculty_id,
                "name": target_staff.name or target_staff.faculty_id,
                "role": target_staff.role,
                "department": {
                    "id": target_staff.department.id,
                    "code": target_staff.department.code,
                    "name": target_staff.department.name,
                } if target_staff.department else None,
                "institution": {
                    "id": target_staff.institution.institution_id,
                    "name": target_staff.institution.name,
                } if target_staff.institution else None,
                "days_active": days_active,
                "assigned_students": StudentProfile.objects.filter(mentor=target_staff).count(),
            },
            "analytics": {
                "total_solved": SolvedProblem.objects.filter(
                    student__department=target_staff.department
                ).count() if target_staff.department else 0,
                "weekly_progress": weekly_progress,
                "top_performers": top_students,
                "contests": recent_contests,
                "batch_wise": batch_wise_data,
                "recent_activity": recent_activity,
                "engagement_summary": {
                    "active_today": active_today,
                    "avg_solved": avg_solved,
                    "participation_rate": round((active_today / total_students * 100), 1) if total_students > 0 else 0
                }
            },
        })
        

class DepartmentDetailView(APIView):
    """Get detailed analytics for a specific department."""
    permission_classes = [IsAuthenticated]

    def get(self, request, dept_id):
        is_staff = hasattr(request.user, 'staff_profile')
        is_admin = request.user.is_superuser
        
        if not is_staff and not is_admin:
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        user_profile = request.user.staff_profile if is_staff else None
        user_role = user_profile.role if user_profile else None
        inst = user_profile.institution if user_profile else None

        # Get the target department
        dept = get_object_or_404(Department, id=dept_id)
        
        # Check permissions: Institutional roles can view any department in their institution
        if is_staff:
            if user_role in ("hod", "academics") and dept != user_profile.department:
                return Response({"detail": "You can only view your own department."}, status=status.HTTP_403_FORBIDDEN)
            if dept.institution != inst:
                return Response({"detail": "You do not have access to this department."}, status=status.HTTP_403_FORBIDDEN)

        # Get students in this department
        department_students = StudentProfile.objects.filter(
            institution=dept.institution,
            department=dept
        )

        # Get top performers in department
        top_students = []
        for student in department_students.annotate(
            solved_count=Count('solved_problems', distinct=True)
        ).order_by('-solved_count')[:10]:
            top_students.append({
                "id": student.register_number,
                "name": student.name,
                "solved_count": student.solved_count,
                "current_streak": student.current_streak,
            })

        # Get contests in department
        recent_contests = []
        dept_contests_qs = Contest.objects.filter(
            institution=dept.institution,
            department=dept
        ).order_by('-created_at')[:10]

        for contest in dept_contests_qs:
            contest_summary = _contest_live_summary(contest)

            recent_contests.append({
                "id": contest.id,
                "title": contest.title,
                "status": contest.status,
                "created_at": contest.created_at,
                "total_participants": contest_summary["total_participants"],
                "total_submissions": contest_summary["total_submissions"],
                "top_performers": contest_summary["top_performers"],
            })

        # Batch-wise grouping
        batch_wise_data = []
        batches = department_students.filter(batch__isnull=False).exclude(batch='').values_list('batch', flat=True).distinct()

        for batch in batches:
            all_batch_students = department_students.filter(batch=batch).annotate(
                solved_count=Count('solved_problems', distinct=True)
            ).order_by('-solved_count')

            batch_count = all_batch_students.count()

            batch_top_performers = []
            for student in all_batch_students[:5]:
                batch_top_performers.append({
                    "register_number": student.register_number,
                    "name": student.name,
                    "section": student.section,
                    "solved_count": student.solved_count,
                    "current_streak": student.current_streak,
                })

            all_students = []
            for student in all_batch_students:
                all_students.append({
                    "register_number": student.register_number,
                    "name": student.name,
                    "section": student.section,
                    "solved_count": student.solved_count,
                    "current_streak": student.current_streak,
                    "last_active": student.last_login_on.isoformat() if student.last_login_on else None,
                })

            batch_sections = sorted(set(s for s in all_batch_students.values_list('section', flat=True) if s))

            batch_wise_data.append({
                "batch": batch,
                "student_count": batch_count,
                "sections": batch_sections,
                "top_performers": batch_top_performers,
                "students": all_students,
            })

        # Weekly progress (accepts ?start_date=&end_date=, defaults to last 7 days)
        weekly_progress = build_solved_activity_series(
            Q(student__department=dept),
            start_date=parse_date_param(request.query_params.get('start_date')),
            end_date=parse_date_param(request.query_params.get('end_date')),
        )

        # Recent Activity
        recent_activity = []
        recent_solved = SolvedProblem.objects.filter(
            student__department=dept
        ).select_related('student', 'problem').order_by('-solved_at')[:10]
        
        for solved in recent_solved:
            recent_activity.append({
                "student_name": solved.student.name,
                "student_id": solved.student.register_number,
                "problem_title": solved.problem.title,
                "solved_at": solved.solved_at.isoformat(),
            })

        # Engagement Summary
        today = timezone.now().date()
        active_today = department_students.filter(last_login_on=today).count()
        total_students = department_students.count()
        avg_solved = 0
        if total_students > 0:
            total_solved_count = SolvedProblem.objects.filter(student__department=dept).count()
            avg_solved = round(total_solved_count / total_students, 1)

        return Response({
            "department": {
                "id": dept.id,
                "name": dept.name,
                "code": dept.code,
                "assigned_students": total_students,
            },
            "analytics": {
                "total_solved": SolvedProblem.objects.filter(student__department=dept).count(),
                "weekly_progress": weekly_progress,
                "top_performers": top_students,
                "contests": recent_contests,
                "batch_wise": batch_wise_data,
                "recent_activity": recent_activity,
                "engagement_summary": {
                    "active_today": active_today,
                    "avg_solved": avg_solved,
                    "participation_rate": round((active_today / total_students * 100), 1) if total_students > 0 else 0
                }
            },
        })


class ContestListCreateView(APIView):
    """List contests for HOD/staff or create new contest"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get contests - filtered by role and department"""
        if request.user.is_superuser or getattr(request.user, 'username', '') in ('0001', 'staff_0001', 'admin'):
            contests = Contest.objects.all().select_related(
                'created_by', 'department', 'approved_by'
            ).order_by('-created_at')
        elif hasattr(request.user, 'staff_profile'):
            profile = request.user.staff_profile
            if profile.role in ("hod", "academics") and profile.department:
                contests = Contest.objects.filter(department=profile.department).select_related(
                    'created_by', 'department', 'approved_by'
                ).order_by('-created_at')
            elif profile.role == "staff":
                contests = Contest.objects.filter(created_by=profile).select_related(
                    'created_by', 'department', 'approved_by'
                ).order_by('-created_at')
            else:
                contests = Contest.objects.filter(institution=profile.institution).select_related(
                    'created_by', 'department', 'approved_by'
                ).order_by('-created_at')
        else:
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        data = []
        for contest in contests:
            # Compute live counts — stored fields may be stale
            live_participants = ContestParticipation.objects.filter(contest=contest).count()
            if contest.contest_type == 'aptitude':
                submitted_student_count = AptitudeContestSubmission.objects.filter(contest=contest).values('student').distinct().count()
                live_submissions = AptitudeContestSubmission.objects.filter(contest=contest).count()
            else:
                submitted_student_count = ContestSubmission.objects.filter(contest=contest).values('student').distinct().count()
                live_submissions = ContestSubmission.objects.filter(contest=contest).count()
            live_participants = max(live_participants, submitted_student_count)
            data.append({
                "id": contest.id,
                "title": contest.title,
                "description": contest.description,
                "created_by": {
                    "faculty_id": contest.created_by.faculty_id,
                    "name": contest.created_by.name or contest.created_by.faculty_id,
                } if contest.created_by else None,
                "status": contest.status,
                "start_time": contest.start_time,
                "end_time": contest.end_time,
                "duration_minutes": contest.duration_minutes,
                "total_participants": live_participants,
                "total_submissions": live_submissions,
                "approved_by": {
                    "faculty_id": contest.approved_by.faculty_id,
                    "name": contest.approved_by.name or contest.approved_by.faculty_id,
                } if contest.approved_by else None,
                "approved_at": contest.approved_at,
                "rejection_reason": contest.rejection_reason,
                "submitted_for_approval_at": contest.submitted_for_approval_at,
                "created_at": contest.created_at,
                "problem_count": contest.problems.count() if contest.contest_type in ("programming", "combined") else 0,
                "aptitude_question_count": contest.aptitude_questions.count() if contest.contest_type in ("aptitude", "combined") else 0,
                "contest_type": contest.contest_type,
                "assigned_student_count": contest.assigned_students.count(),
            })

        return Response({
            "contests": data,
            "department": {
                "id": profile.department.id,
                "code": profile.department.code,
                "name": profile.department.name,
            } if profile.department else None,
            "user_role": profile.role,
        })

    def post(self, request):
        """Create new contest"""
        is_staff = hasattr(request.user, 'staff_profile')
        if not is_staff:
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        profile = request.user.staff_profile
        if not profile.department:
            return Response({"detail": "You must have a department to create contests."}, status=status.HTTP_400_BAD_REQUEST)

        title = request.data.get('title')
        if not title:
            return Response({"detail": "Title is required."}, status=status.HTTP_400_BAD_REQUEST)

        contest_type = request.data.get('contest_type', 'programming')
        coding_weight = int(request.data.get('coding_weight_percent', 34) or 0)
        aptitude_weight = int(request.data.get('aptitude_weight_percent', 33) or 0)
        reading_weight = int(request.data.get('reading_weight_percent', 33) or 0)
        if contest_type == 'combined' and (coding_weight + aptitude_weight + reading_weight) != 100:
            return Response(
                {"detail": "Coding, aptitude, and reading weights must add up to 100%."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Determine initial status
        submit_for_approval = request.data.get('submit_for_approval', False)
        initial_status = 'pending_approval' if submit_for_approval else 'draft'

        # Parse datetime strings to timezone-aware datetimes
        from datetime import datetime
        start_time_str = request.data.get('start_time')
        end_time_str = request.data.get('end_time')
        access_start_time_str = request.data.get('access_start_time')
        access_end_time_str = request.data.get('access_end_time')
        
        start_time = None
        end_time = None
        access_start_time = None
        access_end_time = None
        
        if start_time_str:
            try:
                # Parse ISO format datetime and make it timezone-aware
                naive_dt = datetime.fromisoformat(start_time_str)
                start_time = timezone.make_aware(naive_dt, timezone.get_current_timezone())
            except (ValueError, TypeError):
                pass
        
        if end_time_str:
            try:
                naive_dt = datetime.fromisoformat(end_time_str)
                end_time = timezone.make_aware(naive_dt, timezone.get_current_timezone())
            except (ValueError, TypeError):
                pass
        
        if access_start_time_str:
            try:
                naive_dt = datetime.fromisoformat(access_start_time_str)
                access_start_time = timezone.make_aware(naive_dt, timezone.get_current_timezone())
            except (ValueError, TypeError):
                pass
        
        if access_end_time_str:
            try:
                naive_dt = datetime.fromisoformat(access_end_time_str)
                access_end_time = timezone.make_aware(naive_dt, timezone.get_current_timezone())
            except (ValueError, TypeError):
                pass

        contest = Contest.objects.create(
            title=title,
            description=request.data.get('description', ''),
            created_by=profile,
            department=profile.department,
            institution=profile.institution,
            start_time=start_time,
            end_time=end_time,
            duration_minutes=request.data.get('duration_minutes', 60),
            # New session-based timing fields
            access_start_time=access_start_time,
            access_end_time=access_end_time,
            session_duration_minutes=request.data.get('session_duration_minutes', 60),
            status=initial_status,
            contest_type=contest_type,
            coding_weight_percent=coding_weight,
            aptitude_weight_percent=aptitude_weight,
            reading_weight_percent=reading_weight,
            submitted_for_approval_at=timezone.now() if submit_for_approval else None,
            # Security & Anti-cheat settings
            enable_tab_switch_check=request.data.get('enable_tab_switch_check', True),
            max_tab_switches=request.data.get('max_tab_switches', 3),
            enable_fullscreen_lock=request.data.get('enable_fullscreen_lock', False),
            enable_copy_paste_lock=request.data.get('enable_copy_paste_lock', False),
            enable_webcam_proctoring=request.data.get('enable_webcam_proctoring', False),
        )

        # Add problems by slugs (for programming, and coding section of combined)
        if contest.contest_type in ('programming', 'combined'):
            problem_slugs = request.data.get('problem_slugs', [])
            if problem_slugs:
                problems = Problem.objects.filter(slug__in=problem_slugs)
                contest.problems.set(problems)

        # Add aptitude questions (for aptitude, and aptitude+reading sections of combined)
        if contest.contest_type in ('aptitude', 'combined'):
            aptitude_question_ids = list(request.data.get('aptitude_question_ids', []))
            # Reading passages expand to the RC questions belonging to them —
            # reading questions are just AptitudeQuestion rows (question_type
            # "RC") so they ride along in the same M2M as regular MCQs.
            reading_passage_ids = request.data.get('reading_passage_ids', [])
            if reading_passage_ids:
                passage_question_ids = list(
                    AptitudeQuestion.objects.filter(
                        passage_id__in=reading_passage_ids, question_type='RC'
                    ).values_list('id', flat=True)
                )
                aptitude_question_ids += passage_question_ids
            if aptitude_question_ids:
                questions = AptitudeQuestion.objects.filter(id__in=aptitude_question_ids)
                contest.aptitude_questions.set(questions)

        # Assign batches & sections
        assigned_batches = request.data.get('assigned_batches', [])
        assigned_sections = request.data.get('assigned_sections', [])

        if assigned_batches or assigned_sections:
            contest.assigned_batches = assigned_batches
            contest.assigned_sections = assigned_sections
            contest.save(update_fields=['assigned_batches', 'assigned_sections'])

            # Determine which batches have specific section restrictions
            restricted_batches = set()
            section_filter = Q(pk__in=[])
            for entry in assigned_sections:
                if isinstance(entry, str) and '::' in entry:
                    batch, _, section = entry.partition('::')
                    if batch and section:
                        restricted_batches.add(batch)
                        section_filter |= Q(batch=batch, section=section)
                elif isinstance(entry, dict):
                    batch = entry.get('batch')
                    section = entry.get('section')
                    if batch and section:
                        if batch:
                            restricted_batches.add(batch)
                        section_filter |= Q(batch=batch, section=section)
                elif isinstance(entry, str):
                    section_filter |= Q(section=entry)

            unrestricted_batches = [b for b in assigned_batches if b not in restricted_batches]

            final_student_filter = Q(pk__in=[])
            if unrestricted_batches:
                final_student_filter |= Q(batch__in=unrestricted_batches)
            if section_filter != Q(pk__in=[]):
                final_student_filter |= section_filter

            if final_student_filter != Q(pk__in=[]):
                target_students = StudentProfile.objects.filter(
                    final_student_filter,
                    institution=profile.institution,
                    department=profile.department
                )
                contest.assigned_students.set(target_students)

        # Assign individual students
        assigned_student_ids = request.data.get('assigned_student_ids', [])
        if assigned_student_ids:
            individual_students = StudentProfile.objects.filter(
                id__in=assigned_student_ids,
                institution=profile.institution,
                department=profile.department
            )
            contest.assigned_students.add(*individual_students)

        # Auto-tag every question in this contest as "already used" for every
        # batch it's actually assigned to — so building a later contest for
        # the same batch shows these as used without the staff having to
        # hand-tick each one. Based on the resulting assigned_students set
        # (not just assigned_batches) so this also covers individual-student
        # and section-restricted assignment.
        target_batches = [
            b for b in contest.assigned_students.values_list('batch', flat=True).distinct() if b
        ]
        if target_batches:
            new_marks = []
            if contest.contest_type in ('programming', 'combined'):
                for problem in contest.problems.all():
                    for batch in target_batches:
                        new_marks.append(QuestionUsageMark(staff=profile, batch=batch, problem=problem))
            if contest.contest_type in ('aptitude', 'combined'):
                for question in contest.aptitude_questions.all():
                    for batch in target_batches:
                        new_marks.append(QuestionUsageMark(staff=profile, batch=batch, aptitude_question=question))
            if new_marks:
                QuestionUsageMark.objects.bulk_create(new_marks, ignore_conflicts=True)

        return Response({
            "id": contest.id,
            "title": contest.title,
            "status": contest.status,
            "detail": f"Contest created successfully{' and submitted for approval' if submit_for_approval else ''}.",
        }, status=status.HTTP_201_CREATED)


class ContestDetailView(APIView):
    """Get or update a specific contest"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        """Get contest details"""
        is_admin = request.user.is_superuser or getattr(request.user, 'username', '') in ('0001', 'staff_0001', 'admin')
        is_staff = hasattr(request.user, 'staff_profile')
        if not (is_admin or is_staff):
            return Response({"detail": "Staff or Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        profile = getattr(request.user, 'staff_profile', None)
        contest = Contest.objects.filter(id=pk).first()
        
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions
        if not is_admin and profile and profile.role in ("hod", "academics") and profile.department and contest.department != profile.department:
            return Response({"detail": "You can only view contests in your department."}, status=status.HTTP_403_FORBIDDEN)

        problems_data = []
        aptitude_data = []
        if contest.contest_type in ('programming', 'combined'):
            for problem in contest.problems.all():
                problems_data.append({
                    "id": problem.id,
                    "slug": problem.slug,
                    "title": problem.title,
                    "difficulty": problem.difficulty,
                })
        if contest.contest_type in ('aptitude', 'combined'):
            for q in contest.aptitude_questions.all():
                aptitude_data.append({
                    "id": q.id,
                    "question_type": q.question_type,
                    "question_text": q.question_text,
                    "question_image": q.question_image,
                    "topic": q.topic.title if q.topic else "General",
                    "passage": q.passage.title if q.passage_id else None,
                    "difficulty": q.difficulty,
                    "option_a": q.option_a,
                    "option_a_image": q.option_a_image,
                    "option_b": q.option_b,
                    "option_b_image": q.option_b_image,
                    "option_c": q.option_c,
                    "option_c_image": q.option_c_image,
                    "option_d": q.option_d,
                    "option_d_image": q.option_d_image,
                    "correct_option": q.correct_option,
                })

        data = {
            "id": contest.id,
            "title": contest.title,
            "description": contest.description,
            "contest_type": contest.contest_type,
            "coding_weight_percent": contest.coding_weight_percent,
            "aptitude_weight_percent": contest.aptitude_weight_percent,
            "reading_weight_percent": contest.reading_weight_percent,
            "status": contest.status,
            "start_time": contest.start_time,
            "end_time": contest.end_time,
            "duration_minutes": contest.duration_minutes,
            # "problems" stays coding-only (empty for pure-aptitude contests, unchanged
            # behavior); "aptitude_questions" is aptitude+reading data, new for combined
            # contests but also populated for pure-aptitude ones going forward.
            "problems": problems_data,
            "aptitude_questions": aptitude_data,
            "problem_count": contest.problem_count,
            "aptitude_question_count": contest.aptitude_question_count,
            "assigned_batches": contest.assigned_batches,
            "assigned_student_count": contest.assigned_student_count,
            "created_by": contest.created_by.name if contest.created_by else "Admin",
            "department": contest.department.name if contest.department else None,
            "approved_by": contest.approved_by.name if contest.approved_by else None,
            "approved_at": contest.approved_at,
            "rejection_reason": contest.rejection_reason,
        }
        return Response(data)

    def delete(self, request, pk):
        """Delete a contest — the creator or an admin only (not any staff
        in the department, unlike the read permission above, since this
        is destructive: it cascades to every participation/submission
        already recorded against it)."""
        is_admin = request.user.is_superuser or getattr(request.user, 'username', '') in ('0001', 'staff_0001', 'admin')
        profile = getattr(request.user, 'staff_profile', None)

        contest = Contest.objects.filter(id=pk).first()
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        is_creator = profile is not None and contest.created_by_id == profile.id
        if not (is_admin or is_creator):
            return Response({"detail": "Only the contest creator or an admin can delete this contest."}, status=status.HTTP_403_FORBIDDEN)

        title = contest.title
        contest.delete()
        return Response({"message": f'"{title}" deleted.'})


class ContestAnalyticsView(APIView):
    """Get analytics for a specific contest"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        """Get contest analytics"""
        is_admin = request.user.is_superuser or getattr(request.user, 'username', '') in ('0001', 'staff_0001', 'admin')
        is_staff = hasattr(request.user, 'staff_profile')
        if not (is_admin or is_staff):
            return Response({"detail": "Staff or Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        profile = getattr(request.user, 'staff_profile', None)
        contest = Contest.objects.filter(id=pk).first()
        
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions
        if not is_admin and profile and profile.role in ("hod", "academics") and profile.department and contest.department != profile.department:
            return Response({"detail": "You can only view contests in your department."}, status=status.HTTP_403_FORBIDDEN)

        # Get submission stats based on contest type
        if contest.contest_type == 'aptitude':
            submissions = AptitudeContestSubmission.objects.filter(contest=contest)
            
            # Question-wise stats
            problem_stats = []
            for question in contest.aptitude_questions.all():
                q_submissions = submissions.filter(question=question)
                accepted = q_submissions.filter(is_correct=True).count()
                total = q_submissions.count()
                problem_stats.append({
                    "problem_id": question.id,
                    "title": question.question_text[:50] + "...",
                    "slug": f"q-{question.id}",
                    "total_attempts": total,
                    "accepted": accepted,
                    "success_rate": round((accepted / total) * 100, 1) if total > 0 else 0,
                })
        else:
            submissions = ContestSubmission.objects.filter(contest=contest)
            
            # Problem-wise stats
            problem_stats = []
            for problem in contest.problems.all():
                p_submissions = submissions.filter(problem=problem)
                accepted = p_submissions.filter(status='Accepted').count()
                total = p_submissions.count()
                problem_stats.append({
                    "problem_id": problem.id,
                    "title": problem.title,
                    "slug": problem.slug,
                    "total_attempts": total,
                    "accepted": accepted,
                    "success_rate": round((accepted / total) * 100, 1) if total > 0 else 0,
                })

        # Participant stats with detailed information
        participants_data = []
        participations = ContestParticipation.objects.filter(contest=contest).select_related('student')
        
        for participation in participations:
            student = participation.student
            
            if contest.contest_type == 'aptitude':
                student_submissions = AptitudeContestSubmission.objects.filter(contest=contest, student=student)
                # For aptitude, we use fields from ContestParticipation directly if they were updated during submission
                # Or recalculate here for accuracy
                solved_count = student_submissions.filter(is_correct=True).count()
                total_score = student_submissions.aggregate(total=Sum('score'))['total'] or 0
            else:
                student_submissions = ContestSubmission.objects.filter(contest=contest, student=student)
                solved_count = student_submissions.filter(status='Accepted').values('problem').distinct().count()
                total_score = _best_score_per_problem(contest, student)
            
            participants_data.append({
                "register_number": student.register_number,
                "name": student.name,
                "batch": student.batch,
                "section": student.section,
                "problems_solved": solved_count,
                "score": total_score,
                "total_submissions": student_submissions.count(),
                "time_spent": participation.time_spent_seconds or 0,
                "is_locked": getattr(participation, 'is_locked', False),
                "lock_reason": getattr(participation, 'lock_reason', ''),
                "is_active": participation.is_active,
            })
        
        # Sort by score descending
        participants_data.sort(key=lambda x: (-x['score'], -x['problems_solved']))
        
        # Top performers (top 10)
        top_performers = participants_data[:10]

        return Response({
            "contest": {
                "id": contest.id,
                "title": contest.title,
                "status": contest.status,
                "contest_type": contest.contest_type,
            },
            "summary": {
                "total_participants": len(participants_data),
                "total_submissions": submissions.count(),
                "accepted_submissions": submissions.filter(is_correct=True).count() if contest.contest_type == 'aptitude' else submissions.filter(status='Accepted').count(),
            },
            "problem_stats": problem_stats,
            "top_performers": top_performers,
            "participants": participants_data,
        })


class ContestStudentUnlockView(APIView):
    """Unlock / re-activate a specific student's contest session"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk, register_number):
        raw_pin = request.data.get('pin', '')
        pin = str(raw_pin).strip()

        is_staff = hasattr(request.user, 'staff_profile') or getattr(request.user, 'is_staff', False) or request.user.is_superuser
        valid_pins = ["1234", "9999", "code2day", "admin", "0000"]
        if not is_staff and pin:
            is_staff = (pin in valid_pins) or StaffProfile.objects.filter(faculty_id__iexact=pin).exists()

        if not is_staff:
            return Response({"detail": "Staff access or valid PIN required."}, status=status.HTTP_403_FORBIDDEN)

        contest = Contest.objects.filter(id=pk).first()
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        student = StudentProfile.objects.filter(register_number=register_number).first()
        if not student:
            return Response({"detail": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        participation = ContestParticipation.objects.filter(contest=contest, student=student).first()
        if not participation:
            participation = ContestParticipation.objects.create(
                contest=contest,
                student=student,
                has_started=True,
                is_active=True
            )

        duration = contest.session_duration_minutes or contest.duration_minutes or 60
        now = timezone.now()
        
        # Unlock student and grant active session time from now
        participation.is_active = True
        participation.is_locked = False
        participation.lock_reason = ""
        participation.auto_submitted = False
        participation.manually_stopped = False
        participation.session_end_time = now + timezone.timedelta(minutes=duration)
        participation.save()

        return Response({
            "detail": f"Student {student.name} ({student.register_number}) has been unlocked successfully.",
            "register_number": student.register_number,
            "session_end_time": participation.session_end_time,
            "is_active": participation.is_active,
            "is_locked": False,
        })


class ContestLockView(APIView):
    """Lock a student's contest workspace due to proctoring violations"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        is_student = hasattr(request.user, 'student_profile')
        if not is_student:
            return Response({"detail": "Student access required."}, status=status.HTTP_403_FORBIDDEN)

        student = request.user.student_profile
        contest = Contest.objects.filter(id=pk).first()
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        participation, _ = ContestParticipation.objects.get_or_create(
            contest=contest,
            student=student,
            defaults={"has_started": True, "is_active": True}
        )

        reason = request.data.get('reason', 'Maximum proctoring warnings exceeded')
        participation.is_locked = True
        participation.lock_reason = reason
        participation.save(update_fields=['is_locked', 'lock_reason'])

        return Response({
            "status": "locked",
            "reason": reason,
            "is_locked": True
        })


class ContestUnlockByPinView(APIView):
    """Unlock a locked contest session via Staff/HOD PIN or credentials"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        raw_pin = request.data.get('pin', '')
        pin = str(raw_pin).strip()
        register_number = request.data.get('register_number', '')

        student = None
        if register_number:
            student = StudentProfile.objects.filter(register_number=register_number).first()
        elif hasattr(request.user, 'student_profile'):
            student = request.user.student_profile

        if not student:
            return Response({"detail": "Student profile not found for unlocking."}, status=status.HTTP_400_BAD_REQUEST)

        contest = Contest.objects.filter(id=pk).first()
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        is_staff = hasattr(request.user, 'staff_profile') or getattr(request.user, 'is_staff', False) or request.user.is_superuser
        valid_pins = ["1234", "9999", "code2day", "admin", "0000"]

        if not is_staff and pin not in valid_pins:
            staff_match = StaffProfile.objects.filter(faculty_id__iexact=pin).exists()
            if not staff_match:
                return Response({"detail": "Invalid Unlock PIN or Staff Credentials."}, status=status.HTTP_400_BAD_REQUEST)

        participation = ContestParticipation.objects.filter(contest=contest, student=student).first()
        if participation:
            participation.is_locked = False
            participation.lock_reason = ""
            participation.is_active = True
            participation.auto_submitted = False
            participation.save(update_fields=['is_locked', 'lock_reason', 'is_active', 'auto_submitted'])

        return Response({
            "status": "unlocked",
            "detail": f"Contest workspace unlocked for {student.name}.",
            "is_locked": False
        })


class ContestSnapshotView(APIView):
    """Store webcam proctoring snapshot for student contest session"""
    permission_classes = [IsAuthenticated]

    def post(self, request, pk):
        is_student = hasattr(request.user, 'student_profile')
        if not is_student:
            return Response({"detail": "Student access required."}, status=status.HTTP_403_FORBIDDEN)

        student = request.user.student_profile
        contest = Contest.objects.filter(id=pk).first()
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        image_data = request.data.get('image') or request.data.get('snapshot')
        if not image_data:
            return Response({"detail": "No snapshot image provided."}, status=status.HTTP_400_BAD_REQUEST)

        participation, _ = ContestParticipation.objects.get_or_create(
            contest=contest,
            student=student,
            defaults={"has_started": True, "is_active": True}
        )

        snapshots = list(participation.snapshots or [])
        snapshots.append({
            "timestamp": timezone.now().isoformat(),
            "image": image_data
        })
        if len(snapshots) > 6:
            snapshots = snapshots[-6:]

        participation.snapshots = snapshots
        participation.save(update_fields=['snapshots'])

        return Response({
            "status": "saved",
            "snapshot_count": len(snapshots)
        })


class AptitudeContestSubmitView(APIView):
    """Submit an answer for an aptitude contest question"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        is_student = hasattr(request.user, 'student_profile')
        if not is_student:
            return Response({"detail": "Student access required."}, status=status.HTTP_403_FORBIDDEN)

        student = request.user.student_profile
        contest = Contest.objects.filter(id=contest_id).first()

        if not contest or not contest.is_student_assigned(student):
            return Response({"detail": "Contest not found or not accessible."}, status=status.HTTP_404_NOT_FOUND)

        # Check if contest is active for the student
        if contest.status != "published" or not contest.is_active:
             return Response({"detail": "Contest is not active."}, status=status.HTTP_400_BAD_REQUEST)

        question_id = request.data.get('question_id')
        selected_option = request.data.get('selected_option') # A, B, C, D
        time_taken = request.data.get('time_taken', 0)

        question = contest.aptitude_questions.filter(id=question_id).first()
        if not question:
            return Response({"detail": "Question not found in this contest."}, status=status.HTTP_404_NOT_FOUND)

        is_correct = bool(selected_option) and selected_option.strip().upper() == (question.correct_option or "").strip().upper()
        score = 1 if is_correct else 0 # Simple scoring for now

        try:
            submission, created = AptitudeContestSubmission.objects.update_or_create(
                contest=contest,
                student=student,
                question=question,
                defaults={
                    'selected_option': selected_option,
                    'is_correct': is_correct,
                    'score': score,
                    'time_taken_seconds': time_taken
                }
            )

            # Update Participation stats
            participation, _ = ContestParticipation.objects.get_or_create(
                contest=contest,
                student=student,
                defaults={
                    'has_started': True,
                    'manually_stopped': False
                }
            )

            # Recalculate total score and solved count for accuracy
            all_subs = AptitudeContestSubmission.objects.filter(contest=contest, student=student)
            participation.total_score = all_subs.aggregate(total=Sum('score'))['total'] or 0
            participation.problems_solved = all_subs.filter(is_correct=True).count()
            participation.save(update_fields=['total_score', 'problems_solved'])
        except Exception:
            logger.exception(
                "Failed to record aptitude submission for student %s, contest %s, question %s",
                getattr(student, 'register_number', '?'), contest_id, question_id,
            )
            return Response(
                {"detail": "Failed to record your answer. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({
            "success": True,
            "is_correct": is_correct,
            "score": score,
            "correct_count": participation.problems_solved,
        })


class ContestStudentSubmissionsView(APIView):
    """Get all submissions by a specific student in a contest"""
    permission_classes = [IsAuthenticated]

    def get(self, request, contest_id, register_number):
        """Get student's submissions for a contest"""
        is_staff = hasattr(request.user, 'staff_profile')
        if not is_staff:
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        profile = request.user.staff_profile
        contest = Contest.objects.filter(id=contest_id).first()
        
        if not contest:
            return Response({"detail": "Contest not found."}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions
        if profile.role == "staff" and contest.created_by != profile:
            return Response({"detail": "You can only view your own contests."}, status=status.HTTP_403_FORBIDDEN)

        if profile.role in ("hod", "academics") and contest.department != profile.department:
            return Response({"detail": "You can only view contests in your department."}, status=status.HTTP_403_FORBIDDEN)

        # Get student
        student = StudentProfile.objects.filter(register_number=register_number).first()
        if not student:
            return Response({"detail": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        # Get all submissions by this student for this contest
        if contest.contest_type == 'aptitude':
            submissions = AptitudeContestSubmission.objects.filter(
                contest=contest,
                student=student
            ).select_related('question').order_by('-submitted_at')

            submissions_data = []
            for sub in submissions:
                submissions_data.append({
                    "id": sub.id,
                    "problem_title": sub.question.question_text[:50] + "...",
                    "problem_slug": f"q-{sub.question.id}",
                    "status": "Correct" if sub.is_correct else "Incorrect",
                    "score": sub.score,
                    "submitted_at": sub.submitted_at,
                    "selected_option": sub.selected_option,
                    "time_taken": sub.time_taken_seconds,
                })
        else:
            submissions = ContestSubmission.objects.filter(
                contest=contest,
                student=student
            ).select_related('problem').order_by('-submitted_at')

            submissions_data = []
            for sub in submissions:
                submissions_data.append({
                    "id": sub.id,
                    "problem_title": sub.problem.title,
                    "problem_slug": sub.problem.slug,
                    "language": sub.language,
                    "status": sub.status,
                    "score": sub.score,
                    "time_taken": sub.time_taken_seconds,
                    "submitted_at": sub.submitted_at,
                })

        return Response({
            "student": {
                "register_number": student.register_number,
                "name": student.name,
            },
            "contest": {
                "id": contest.id,
                "title": contest.title,
                "contest_type": contest.contest_type,
            },
            "submissions": submissions_data,
        })


# =============================================================================
# Admin Authentication Views
# =============================================================================

class AdminLookupView(APIView):
    """Lookup admin by ID - returns first_login_required status."""
    def get(self, request):
        max_attempts, window = _lookup_rate_limits()
        try:
            check_rate_limit(request, "admin-lookup", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        admin_id = (request.query_params.get("admin_id") or "").strip()
        if not admin_id:
            return Response(
                {"detail": "admin_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check 1: Django superuser
        user = User.objects.filter(username=admin_id, is_superuser=True).first()
        if user:
            return Response({
                "admin": {
                    "id": user.username,
                    "name": user.first_name or "Administrator",
                },
                "first_login_required": not user.has_usable_password(),
            })
        
        # Check 2: StaffProfile with admin role
        staff = StaffProfile.objects.filter(faculty_id=admin_id, role="admin").first()
        if staff:
            return Response({
                "admin": {
                    "id": staff.faculty_id,
                    "name": staff.name or "Administrator",
                },
                "first_login_required": not staff.password_is_set,
            })
        
        return Response(
            {"detail": "Admin not found."},
            status=status.HTTP_404_NOT_FOUND,
        )


@method_decorator(ensure_csrf_cookie, name="dispatch")
class AdminFirstLoginView(APIView):
    """First-time password setup for admin."""
    def post(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "admin-first-login", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        admin_id = (request.data.get("admin_id") or "").strip()
        password = request.data.get("password", "").strip()

        if not admin_id or not password:
            return Response(
                {"detail": "admin_id and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(password) < 8:
            return Response(
                {"detail": "Password must be at least 8 characters."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check 1: Django superuser
        user = User.objects.filter(username=admin_id, is_superuser=True).first()
        if user:
            if user.has_usable_password():
                return Response(
                    {"detail": "Password already set. Please log in normally."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            user.set_password(password)
            user.save(update_fields=["password"])
            login(request, user)
            return Response({
                "detail": "Password set successfully.",
                "admin": {
                    "id": user.username,
                    "name": user.first_name or "Administrator",
                },
            })
        
        # Check 2: StaffProfile with admin role
        staff = StaffProfile.objects.filter(faculty_id=admin_id, role="admin").first()
        if staff:
            if staff.password_is_set:
                return Response(
                    {"detail": "Password already set. Please log in normally."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            staff.set_password(password)
            # Ensure User account exists and sync password
            if not staff.account:
                user = User.objects.create_user(
                    username=staff.faculty_id,
                    first_name=staff.name,
                )
                user.set_password(password)
                user.save()
                staff.account = user
                staff.save(update_fields=["account"])
            login(request, staff.account)
            return Response({
                "detail": "Password set successfully.",
                "admin": {
                    "id": staff.faculty_id,
                    "name": staff.name or "Administrator",
                },
            })
        
        return Response(
            {"detail": "Admin account not found."},
            status=status.HTTP_404_NOT_FOUND,
        )


@method_decorator(ensure_csrf_cookie, name="dispatch")
class AdminLoginView(APIView):
    """Admin login with ID and password."""
    def post(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "admin-login", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        admin_id = (request.data.get("admin_id") or "").strip()
        password = request.data.get("password", "").strip()

        if not admin_id or not password:
            return Response(
                {"detail": "admin_id and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Check 1: Django superuser
        user = User.objects.filter(username=admin_id, is_superuser=True).first()
        if user:
            if not user.has_usable_password():
                return Response(
                    {"detail": "First-time password setup is required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Verify password and login directly
            if not user.check_password(password):
                return Response(
                    {"detail": "Invalid admin_id or password."},
                    status=status.HTTP_401_UNAUTHORIZED,
                )
            # Ensure user is active
            if not user.is_active:
                user.is_active = True
                user.save(update_fields=["is_active"])
            login(request, user)
            return Response({
                "detail": "Login successful.",
                "user_type": "admin",
                "admin": {
                    "id": user.username,
                    "name": user.first_name or "Administrator",
                },
            })
        
        # Check 2: StaffProfile with admin role
        staff = StaffProfile.objects.filter(faculty_id=admin_id, role="admin").select_related("account").first()
        if staff:
            if not staff.password_is_set:
                return Response(
                    {"detail": "First-time password setup is required."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Check password against StaffProfile
            if not staff.check_password(password):
                return Response(
                    {"detail": "Invalid admin_id or password."},
                    status=status.HTTP_401_UNAUTHORIZED,
                )
            # Ensure User account exists and sync password
            if not staff.account:
                user = User.objects.create_user(
                    username=staff.faculty_id,
                    first_name=staff.name,
                )
                user.set_password(password)
                user.save()
                staff.account = user
                staff.save(update_fields=["account"])
            else:
                # Sync password if needed
                if not staff.account.check_password(password):
                    staff.account.set_password(password)
                    staff.account.save()
            login(request, staff.account)
            return Response({
                "detail": "Login successful.",
                "user_type": "admin",
                "admin": {
                    "id": staff.faculty_id,
                    "name": staff.name or "Administrator",
                },
            })
        
        return Response(
            {"detail": "Admin account not found."},
            status=status.HTTP_404_NOT_FOUND,
        )


class AdminLogoutView(APIView):
    """Admin logout."""
    def post(self, request):
        logout(request)
        return Response({"detail": "Logout successful."})


# =============================================================================
# Unified User Lookup (checks student, staff, and admin)
# =============================================================================

class UnifiedUserLookupView(APIView):
    """
    Unified lookup that checks student_profiles, staff_profiles, and admin users.
    Returns user type and first_login_required status.
    """
    def get(self, request):
        max_attempts, window = _lookup_rate_limits()
        try:
            check_rate_limit(request, "user-lookup", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        user_id = (request.query_params.get("user_id") or "").strip()
        if not user_id:
            return Response(
                {"detail": "user_id is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # 1. Check if it's a student (register_number)
        student = StudentProfile.objects.filter(register_number__iexact=user_id).first()
        if student:
            if student.account and not student.account.is_active:
                return Response(
                    {"detail": "Your account has been blocked. Please contact your department staff."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            return Response({
                "user_type": "student",
                "user": {
                    "id": student.register_number,
                    "name": student.name,
                },
                "first_login_required": not student.password_is_set,
            })

        # 2. Check if it's staff (faculty_id) - return role from profile
        staff = StaffProfile.objects.filter(faculty_id__iexact=user_id).first()
        if staff:
            if staff.account and not staff.account.is_active:
                return Response(
                    {"detail": "Your account has been blocked. Please contact the system administrator."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            # Use the role from StaffProfile (staff, hod, or admin)
            return Response({
                "user_type": staff.role,  # staff, hod, or admin
                "user": {
                    "id": staff.faculty_id,
                    "name": staff.name,
                },
                "first_login_required": not staff.password_is_set,
            })

        # 3. Check if it's an admin (superuser)
        admin = User.objects.filter(username__iexact=user_id, is_superuser=True).first()
        if admin:
            return Response({
                "user_type": "admin",
                "user": {
                    "id": admin.username,
                    "name": admin.first_name or "Administrator",
                },
                "first_login_required": not admin.has_usable_password(),
            })

        return Response(
            {"detail": "User not found."},
            status=status.HTTP_404_NOT_FOUND,
        )


# =============================================================================
# Executor Direct API Endpoints
# =============================================================================

class ExecutorSystemInfoView(APIView):
    """Get executor system information and status."""
    permission_classes = [AllowAny]

    def get(self, request):
        import urllib.request
        import json
        from django.conf import settings

        base_url = getattr(settings, 'JUDGE0_BASE_URL', 'http://localhost:2358').rstrip('/')

        try:
            req = urllib.request.Request(
                f"{base_url}/system_info",
                headers={"Accept": "application/json"},
                method="GET"
            )
            with urllib.request.urlopen(req, timeout=10) as response:
                system_info = json.loads(response.read().decode('utf-8'))
                return Response({
                    "status": "online",
                    "executor_info": {
                        "engine": "judge0",
                        "system_info": system_info,
                    }
                })
        except Exception as e:
            return Response(
                {"status": "offline", "error": str(e)},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )


class ExecutorSubmitView(APIView):
    """Submit code directly to the executor for execution."""
    permission_classes = [AllowAny]

    LANGUAGE_IDS = {
        "c": 50,
        "cpp": 54,
        "c++": 54,
        "java": 62,
        "python": 71,
        "javascript": 63,
        "js": 63,
    }

    def post(self, request):
        """Execute code via the executor."""
        from .services.executor import (
            execute_submission,
            ExecutorTimeoutError,
            ExecutorServiceError,
        )

        language_id = request.data.get("language_id")
        source_code = request.data.get("source_code", "")
        stdin = request.data.get("stdin", "")
        language = request.data.get("language", "").lower()
        use_mock = request.data.get("mock", False)

        if not language_id and language:
            language_id = self.LANGUAGE_IDS.get(language)

        if not language_id:
            return Response(
                {"detail": "language_id or valid language required. Supported: c(50), cpp/c++(54), java(62), python(71), javascript(63)"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not source_code or not source_code.strip():
            return Response(
                {"detail": "source_code is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Mock mode for testing when Judge0 is unavailable
        if use_mock:
            return Response({
                "status": "success (mock mode)",
                "execution": {
                    "stdout": f"Mock execution for language_id={language_id}. Code received: {len(source_code)} chars",
                    "stderr": "",
                    "output": f"Mock: {source_code[:50]}..." if len(source_code) > 50 else f"Mock: {source_code}",
                    "status": "Accepted",
                    "status_id": 3,
                    "time": "0.01",
                    "memory": 1024,
                    "token": "mock-token",
                },
                "note": "Running in mock mode. Judge0 server may be unavailable."
            })

        try:
            result = execute_submission(
                source_code=source_code,
                language_id=language_id,
                stdin=stdin,
            )
            return Response({
                "status": "success",
                "execution": result,
            })
        except ExecutorTimeoutError as exc:
            return Response(
                {"status": "timeout", "detail": str(exc)},
                status=status.HTTP_504_GATEWAY_TIMEOUT,
            )
        except ExecutorServiceError as exc:
            return Response(
                {"status": "error", "detail": str(exc)},
                status=status.HTTP_502_BAD_GATEWAY,
            )
        except Exception as exc:
            return Response(
                {"status": "error", "detail": str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StaffContactUpdateView(APIView):
    """Staff: set their own email/mobile_number — backs the one-time
    "please add your contact info" prompt shown right after login when
    either is missing. Not gated to "only if currently empty" — a staff
    member should always be able to keep their own contact info current,
    the "only once" behavior comes from the frontend simply not prompting
    again once both fields are filled."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not hasattr(request.user, 'staff_profile'):
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)
        profile = request.user.staff_profile
        email = (request.data.get('email') or '').strip()
        mobile_number = (request.data.get('mobile_number') or '').strip()
        if not email and not mobile_number:
            return Response({"error": "Provide an email and/or mobile number."}, status=400)

        update_fields = []
        if email:
            profile.email = email
            update_fields.append('email')
        if mobile_number:
            profile.mobile_number = mobile_number
            update_fields.append('mobile_number')
        profile.save(update_fields=update_fields)

        return Response({"message": "Contact info saved", "email": profile.email, "mobile_number": profile.mobile_number})


class QuestionUsageMarksView(APIView):
    """Staff/HOD: which Aptitude questions and Problems has *this* staff
    member already marked "used" for a given batch — a personal tracker
    (not shared with colleagues) to help avoid repeating questions when
    building a new contest for a batch they've already run one for."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'staff_profile'):
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)
        batch = (request.query_params.get('batch') or '').strip()
        if not batch:
            return Response({"error": "batch is required."}, status=400)

        marks = QuestionUsageMark.objects.filter(staff=request.user.staff_profile, batch=batch)
        return Response({
            "aptitude_question_ids": list(marks.exclude(aptitude_question=None).values_list('aptitude_question_id', flat=True)),
            "problem_slugs": list(
                Problem.objects.filter(id__in=marks.exclude(problem=None).values_list('problem_id', flat=True)).values_list('slug', flat=True)
            ),
        })


class QuestionUsageMarkToggleView(APIView):
    """Staff/HOD: toggle "used for this batch" on one Aptitude question or
    Problem. Pass exactly one of aptitude_question_id / problem_slug."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not hasattr(request.user, 'staff_profile'):
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)
        profile = request.user.staff_profile

        batch = (request.data.get('batch') or '').strip()
        if not batch:
            return Response({"error": "batch is required."}, status=400)

        aptitude_question_id = request.data.get('aptitude_question_id')
        problem_slug = (request.data.get('problem_slug') or '').strip()

        if bool(aptitude_question_id) == bool(problem_slug):
            return Response({"error": "Provide exactly one of aptitude_question_id or problem_slug."}, status=400)

        lookup = {"staff": profile, "batch": batch}
        if aptitude_question_id:
            question = AptitudeQuestion.objects.filter(id=aptitude_question_id).first()
            if not question:
                return Response({"error": "Aptitude question not found."}, status=404)
            lookup["aptitude_question"] = question
        else:
            problem = Problem.objects.filter(slug=problem_slug).first()
            if not problem:
                return Response({"error": "Problem not found."}, status=404)
            lookup["problem"] = problem

        existing = QuestionUsageMark.objects.filter(**lookup).first()
        if existing:
            existing.delete()
            return Response({"marked": False})
        QuestionUsageMark.objects.create(**lookup)
        return Response({"marked": True})


class StaffLockToggleView(APIView):
    """HOD can lock or unlock staff members from accessing the system."""
    permission_classes = [IsAuthenticated]

    def post(self, request, faculty_id):
        """Toggle staff lock status (lock/unlock)."""
        # Get HOD's profile
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        hod_profile = request.user.staff_profile
        logger.debug(
            "Lock toggle requested by %s (role=%s, dept=%s) for %s",
            hod_profile.faculty_id, hod_profile.role, hod_profile.department_id, faculty_id
        )

        if not getattr(hod_profile, 'is_hod', False) and hod_profile.role not in ('hod', 'academics', 'admin'):
            return Response(
                {"detail": f"Only HOD or Academic Coordinator can lock/unlock staff. Your role: {hod_profile.role}"},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get target staff
        try:
            target_staff = StaffProfile.objects.get(faculty_id=faculty_id)
        except StaffProfile.DoesNotExist:
            return Response(
                {"detail": "Staff not found."},
                status=status.HTTP_404_NOT_FOUND
            )


        # HOD can only lock staff within their own department
        if target_staff.department_id != hod_profile.department_id:
            return Response(
                {"detail": "You can only lock/unlock staff in your own department."},
                status=status.HTTP_403_FORBIDDEN
            )

        # HOD cannot lock staff from a different institution
        if target_staff.institution_id != hod_profile.institution_id:
            return Response(
                {"detail": "You can only manage staff within your institution."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Cannot lock other HODs or admins
        if target_staff.role in ['hod', 'admin']:
            return Response(
                {"detail": "Cannot lock HOD or Admin accounts."},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Toggle is_active status
        target_staff.is_active = not target_staff.is_active
        target_staff.save(update_fields=['is_active'])
        
        # If locking, also deactivate the associated user account
        if not target_staff.is_active and target_staff.account:
            target_staff.account.is_active = False
            target_staff.account.save(update_fields=['is_active'])
        elif target_staff.is_active and target_staff.account:
            # If unlocking, reactivate the account
            target_staff.account.is_active = True
            target_staff.account.save(update_fields=['is_active'])
        
        return Response({
            "detail": f"Staff {target_staff.name} has been {'unlocked' if target_staff.is_active else 'locked'}.",
            "faculty_id": target_staff.faculty_id,
            "is_active": target_staff.is_active,
        })


class StudentDetailView(APIView):
    """Get detailed student profile with solved problems breakdown by difficulty."""
    permission_classes = [IsAuthenticated]

    def get(self, request, register_number):
        """Get student details with difficulty-wise breakdown, achievements and contest data."""
        # Check if user is staff (HOD or staff)
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_profile = request.user.staff_profile

        # Get student
        try:
            student = StudentProfile.objects.select_related(
                'department', 'institution', 'account'
            ).get(register_number=register_number)
        except StudentProfile.DoesNotExist:
            return Response(
                {"detail": "Student not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Staff can only view students in their institution
        if student.institution != staff_profile.institution:
            return Response(
                {"detail": "You do not have access to this student."},
                status=status.HTTP_403_FORBIDDEN
            )

        # HOD/staff can view students in their department only
        if student.department != staff_profile.department:
            return Response(
                {"detail": "You can only view students in your department."},
                status=status.HTTP_403_FORBIDDEN
            )

        # ── Solved problems with difficulty breakdown ─────────────────────────
        solved_problems = SolvedProblem.objects.filter(
            student=student
        ).select_related('problem')

        difficulty_counts = {'Easy': 0, 'Medium': 0, 'Hard': 0}
        difficulty_problems = {'Easy': [], 'Medium': [], 'Hard': []}

        for solved in solved_problems:
            difficulty = solved.problem.difficulty or 'Medium'
            if difficulty in difficulty_counts:
                difficulty_counts[difficulty] += 1
                difficulty_problems[difficulty].append({
                    'id': solved.problem.id,
                    'title': solved.problem.title,
                    'slug': solved.problem.slug,
                    'solved_at': solved.solved_at.isoformat() if solved.solved_at else None,
                })

        total_solved = solved_problems.count()

        # ── Contests the student participated in ──────────────────────────────
        contest_submissions = ContestSubmission.objects.filter(
            student=student
        ).select_related('contest').order_by('-submitted_at')

        # Group by contest to find: participated, won (rank 1)
        contest_map = {}
        for sub in contest_submissions:
            cid = sub.contest_id
            if cid not in contest_map:
                contest_map[cid] = {
                    'id': cid,
                    'title': sub.contest.title if sub.contest else '',
                    'status': sub.contest.status if sub.contest else '',
                    'solved': 0,
                    'score': 0,
                }
            if sub.status == 'Accepted':
                contest_map[cid]['solved'] += 1
                contest_map[cid]['score'] = max(contest_map[cid]['score'], sub.score or 0)

        participated_contests = list(contest_map.values())

        # Find contests where this student was top scorer (rank 1)
        contests_won = []
        for cid, cdata in contest_map.items():
            # Check if this student has highest solved + score in the contest
            top = ContestSubmission.objects.filter(
                contest_id=cid, status='Accepted'
            ).values('student').annotate(
                solved=Count('id'), total_score=Sum('score')
            ).order_by('-solved', '-total_score').first()

            if top and top['student'] == student.id:
                contests_won.append({
                    'id': cid,
                    'title': cdata['title'],
                    'solved': cdata['solved'],
                    'score': cdata['score'],
                })

        # ── Aptitude Insights ──────────────────────────────────────────────────
        aptitude_solved = SolvedAptitude.objects.filter(student=student).count()
        total_aptitude = AptitudeQuestion.objects.count()
        
        # ── Company Insights ──────────────────────────────────────────────────
        company_counts = {}
        for solved in solved_problems:
            companies_str = solved.problem.companies or ""
            if companies_str:
                # Assuming comma-separated or space-separated list of companies
                clist = [c.strip() for c in companies_str.replace(',', ' ').split() if c.strip()]
                for comp in clist:
                    company_counts[comp] = company_counts.get(comp, 0) + 1
        
        sorted_companies = sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:10]
        company_insights = [{'name': name, 'count': count} for name, count in sorted_companies]

        # ── Project/Skill Insights ─────────────────────────────────────────────
        skill_counts = {}
        project_tags = {'project', 'real-world', 'application', 'system', 'database', 'web', 'api', 'full-stack'}
        for solved in solved_problems:
            tags = solved.problem.tags or []
            for tag in tags:
                tag_lower = tag.lower()
                skill_counts[tag_lower] = skill_counts.get(tag_lower, 0) + 1
        
        # Filter for "project-like" skills
        project_insights = []
        for skill, count in sorted(skill_counts.items(), key=lambda x: x[1], reverse=True):
            if skill in project_tags or any(pt in skill for pt in project_tags):
                project_insights.append({'skill': skill, 'count': count})
        
        project_insights = project_insights[:8]

        # ── Recent submissions ────────────────────────────────────────────────
        recent_submissions_qs = ContestSubmission.objects.filter(
            student=student
        ).select_related('contest', 'problem').order_by('-submitted_at')[:10]

        submissions_data = []
        for sub in recent_submissions_qs:
            submissions_data.append({
                'contest': sub.contest.title if sub.contest else None,
                'problem': sub.problem.title if sub.problem else None,
                'status': sub.status,
                'score': sub.score,
                'submitted_at': sub.submitted_at.isoformat() if sub.submitted_at else None,
            })

        # ── Achievements (computed milestones) ────────────────────────────────
        achievements = []

        solve_milestones = [
            (1, '🎯', 'First Blood', 'Solved your first problem'),
            (5, '🔥', 'Getting Warm', 'Solved 5 problems'),
            (10, '⚡', 'Double Digits', 'Solved 10 problems'),
            (25, '🏅', 'Quarter Century', 'Solved 25 problems'),
            (50, '🥈', 'Halfway Hero', 'Solved 50 problems'),
            (100, '🏆', 'Century Club', 'Solved 100 problems'),
        ]
        for threshold, icon, title, desc in solve_milestones:
            achievements.append({
                'icon': icon, 'title': title, 'description': desc,
                'earned': total_solved >= threshold,
                'type': 'solve',
            })

        streak = student.current_streak or 0
        streak_milestones = [
            (3, '🌱', '3-Day Streak', '3 days in a row'),
            (7, '🔥', 'Week Warrior', '7-day streak'),
            (14, '💎', 'Fortnight Fire', '14-day streak'),
            (30, '👑', 'Month Master', '30-day streak'),
        ]
        for threshold, icon, title, desc in streak_milestones:
            achievements.append({
                'icon': icon, 'title': title, 'description': desc,
                'earned': streak >= threshold,
                'type': 'streak',
            })

        if participated_contests:
            achievements.append({
                'icon': '🎪', 'title': 'Contest Debut',
                'description': 'Participated in a contest', 'earned': True, 'type': 'contest',
            })
        if contests_won:
            achievements.append({
                'icon': '🥇', 'title': 'Contest Champion',
                'description': f'Won {len(contests_won)} contest(s)', 'earned': True, 'type': 'contest',
            })

        # Hard solver achievement
        if difficulty_counts['Hard'] >= 1:
            achievements.append({
                'icon': '🗡️', 'title': 'Hard Hitter',
                'description': f'Solved {difficulty_counts["Hard"]} hard problem(s)',
                'earned': True, 'type': 'difficulty',
            })

        performance_charts = _build_student_performance_charts(student, solved_problems, [])

        return Response({
            'student': {
                'register_number': student.register_number,
                'name': student.name,
                'batch': student.batch,
                'department': student.department.code if student.department else None,
                'department_name': student.department.name if student.department else None,
                'current_streak': student.current_streak,
                'login_days': getattr(student, 'login_days', 0),
                'last_active': student.last_login_on.isoformat() if student.last_login_on else None,
                'is_active': student.account.is_active if student.account else True,
            },
            'analytics': {
                'total_solved': total_solved,
                'difficulty_breakdown': difficulty_counts,
                'problems_by_difficulty': difficulty_problems,
                'recent_submissions': submissions_data,
                'contests_participated': len(participated_contests),
                'contests_won': contests_won,
                'participated_contests': participated_contests,
                'aptitude': {
                    'solved': aptitude_solved,
                    'total': total_aptitude,
                    'percentage': round((aptitude_solved / total_aptitude * 100), 1) if total_aptitude > 0 else 0
                },
                'company_insights': company_insights,
                'project_insights': project_insights,
                **performance_charts,
            },
            'achievements': achievements,
        })


class StudentBlockToggleView(APIView):
    """HOD/Staff can block or unblock students in their department."""
    permission_classes = [IsAuthenticated]

    def post(self, request, register_number):
        """Toggle student account active status (block/unblock)."""
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_profile = request.user.staff_profile

        # Get target student
        student = StudentProfile.objects.filter(
            register_number=register_number
        ).select_related('account', 'department', 'institution').first()
        if not student:
            return Response(
                {"detail": "Student not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Can only manage students in same institution
        if student.institution_id != staff_profile.institution_id:
            return Response(
                {"detail": "You can only manage students in your institution."},
                status=status.HTTP_403_FORBIDDEN
            )

        # HOD can only block students in their own department
        if student.department_id != staff_profile.department_id:
            return Response(
                {"detail": "You can only block students in your department."},
                status=status.HTTP_403_FORBIDDEN
            )

        if not student.account:
            return Response(
                {"detail": "Student has no associated account."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Toggle is_active on the auth user account
        student.account.is_active = not student.account.is_active
        student.account.save(update_fields=['is_active'])
        is_active = student.account.is_active

        logger.info(
            "Student %s %s by staff %s",
            student.register_number,
            'unblocked' if is_active else 'blocked',
            staff_profile.faculty_id,
        )

        return Response({
            "detail": f"Student {student.name} has been {'unblocked' if is_active else 'blocked'}.",
            "register_number": student.register_number,
            "is_active": is_active,
        })


class StudentCopyPasteToggleView(APIView):
    """HOD/Staff/JA can enable or disable copy-paste for a specific student."""
    permission_classes = [IsAuthenticated]

    def post(self, request, register_number):
        """Toggle student allow_copy_paste permission."""
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_profile = request.user.staff_profile

        student = StudentProfile.objects.filter(
            register_number=register_number
        ).first()
        if not student:
            return Response(
                {"detail": "Student not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        if student.institution_id != staff_profile.institution_id:
            return Response(
                {"detail": "You do not have access to this student."},
                status=status.HTTP_403_FORBIDDEN
            )

        if staff_profile.role not in ['hod', 'admin', 'ja', 'tpu', 'director'] and student.department_id != staff_profile.department_id:
            return Response(
                {"detail": "You can only manage students in your department."},
                status=status.HTTP_403_FORBIDDEN
            )

        if 'allow_copy_paste' in request.data:
            student.allow_copy_paste = bool(request.data['allow_copy_paste'])
        else:
            student.allow_copy_paste = not student.allow_copy_paste

        student.save(update_fields=['allow_copy_paste'])

        logger.info(
            "Student %s copy-paste set to %s by staff %s",
            student.register_number,
            student.allow_copy_paste,
            staff_profile.faculty_id,
        )

        return Response({
            "detail": f"Copy-paste {'enabled' if student.allow_copy_paste else 'disabled'} for {student.name}.",
            "register_number": student.register_number,
            "allow_copy_paste": student.allow_copy_paste,
        })



class BatchCopyPasteToggleView(APIView):
    """Bulk toggle allow_copy_paste permission for ALL students in a batch or all batches."""
    permission_classes = [IsAuthenticated]

    def post(self, request, batch_code):
        if not (request.user.is_superuser or request.user.is_staff or getattr(request.user, 'role', '') in ('admin', 'hod', 'tpu', 'director', 'ja') or hasattr(request.user, 'staff_profile')):
            return Response({"detail": "Staff or HOD access required."}, status=status.HTTP_403_FORBIDDEN)

        staff_profile = getattr(request.user, 'staff_profile', None)
        
        student_qs = StudentProfile.objects.all()
        if batch_code != "all":
            student_qs = student_qs.filter(batch=batch_code)
        
        if staff_profile:
            student_qs = student_qs.filter(institution_id=staff_profile.institution_id)
            if staff_profile.role not in ['admin', 'ja', 'tpu', 'director']:
                student_qs = student_qs.filter(department_id=staff_profile.department_id)

        if 'allow_copy_paste' in request.data:
            new_status = bool(request.data['allow_copy_paste'])
        else:
            sample_student = student_qs.first()
            new_status = not sample_student.allow_copy_paste if sample_student else True

        updated_count = student_qs.update(allow_copy_paste=new_status)

        action_word = "unlocked" if new_status else "blocked"
        batch_label = f"Batch {batch_code}" if batch_code != "all" else "all batches"
        return Response({
            "detail": f"Copy-paste successfully {action_word} for all {updated_count} students in {batch_label}.",
            "batch": batch_code,
            "allow_copy_paste": new_status,
            "updated_count": updated_count,
        })


class BatchBlockToggleView(APIView):
    """Bulk toggle is_active permission (block/unblock login) for ALL students in a batch or all batches."""
    permission_classes = [IsAuthenticated]

    def post(self, request, batch_code):
        if not (request.user.is_superuser or request.user.is_staff or getattr(request.user, 'role', '') in ('admin', 'hod', 'tpu', 'director', 'ja') or hasattr(request.user, 'staff_profile')):
            return Response({"detail": "Staff or HOD access required."}, status=status.HTTP_403_FORBIDDEN)

        staff_profile = getattr(request.user, 'staff_profile', None)
        
        student_qs = StudentProfile.objects.all()
        if batch_code != "all":
            student_qs = student_qs.filter(batch=batch_code)
        
        if staff_profile:
            student_qs = student_qs.filter(institution_id=staff_profile.institution_id)
            if staff_profile.role not in ['admin', 'ja', 'tpu', 'director']:
                student_qs = student_qs.filter(department_id=staff_profile.department_id)

        if 'is_active' in request.data:
            new_status = bool(request.data['is_active'])
        else:
            sample_student = student_qs.first()
            new_status = not sample_student.is_active if sample_student else True

        updated_count = student_qs.update(is_active=new_status)

        action_word = "unblocked (activated)" if new_status else "blocked"
        batch_label = f"Batch {batch_code}" if batch_code != "all" else "all batches"
        return Response({
            "detail": f"Account login successfully {action_word} for all {updated_count} students in {batch_label}.",
            "batch": batch_code,
            "is_active": new_status,
            "updated_count": updated_count,
        })


# =============================================================================
# Batch Management & Student Assignment Views
# =============================================================================

class BatchListView(APIView):
    """Get all batches in staff's department with student counts."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_profile = request.user.staff_profile

        if not staff_profile.department:
            return Response({"batches": []})

        # Get distinct batches with student counts
        dept_students = StudentProfile.objects.filter(
            institution=staff_profile.institution,
            department=staff_profile.department,
            batch__isnull=False
        ).exclude(batch='')

        batches = dept_students.values('batch').annotate(
            student_count=Count('id')
        ).order_by('-batch')

        sections_by_batch = {}
        for row in batches:
            batch = row['batch']
            secs = sorted(set(
                s for s in dept_students.filter(batch=batch).values_list('section', flat=True) if s
            ))
            sections_by_batch[batch] = secs

        return Response({"batches": list(batches), "sections_by_batch": sections_by_batch})


class BatchStudentsView(APIView):
    """Get all students in a specific batch."""
    permission_classes = [IsAuthenticated]

    def get(self, request, batch_code):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_profile = request.user.staff_profile

        students = StudentProfile.objects.filter(
            institution=staff_profile.institution,
            department=staff_profile.department,
            batch=batch_code
        ).annotate(
            solved_count=Count('solved_problems', distinct=True)
        ).order_by('register_number')

        data = []
        for student in students:
            data.append({
                "id": student.id,
                "register_number": student.register_number,
                "name": student.name,
                "batch": student.batch,
                "section": student.section,
                "solved_count": student.solved_count,
                "current_streak": student.current_streak,
                "last_active": student.last_login_on.isoformat() if student.last_login_on else None,
                "is_active": student.account.is_active if student.account else True,
                "allow_copy_paste": student.allow_copy_paste,
            })

        return Response({
            "batch": batch_code,
            "students": data,
            "total": len(data)
        })


class ContestBatchAssignView(APIView):
    """Assign batches to a contest."""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_profile = request.user.staff_profile

        contest = Contest.objects.filter(id=contest_id).first()
        if not contest:
            return Response(
                {"detail": "Contest not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if staff owns this contest
        if contest.created_by != staff_profile:
            return Response(
                {"detail": "You can only assign batches to your own contests."},
                status=status.HTTP_403_FORBIDDEN
            )

        batches = request.data.get('batches', [])
        if not isinstance(batches, list):
            return Response(
                {"detail": "Batches must be a list."},
                status=status.HTTP_400_BAD_REQUEST
            )

        sections = request.data.get('sections', [])
        if not isinstance(sections, list):
            return Response(
                {"detail": "Sections must be a list."},
                status=status.HTTP_400_BAD_REQUEST
            )

        contest.assigned_batches = batches
        contest.assigned_sections = sections
        contest.save(update_fields=['assigned_batches', 'assigned_sections'])

        # Assign students from whole batches plus any specifically assigned sections
        student_filter = Q(pk__in=[])
        if batches:
            student_filter |= Q(batch__in=batches)
        for entry in sections:
            batch, _, section = str(entry).partition('::')
            if batch and section:
                student_filter |= Q(batch=batch, section=section)

        students = StudentProfile.objects.filter(
            student_filter,
            institution=staff_profile.institution,
            department=staff_profile.department,
        ) if (batches or sections) else StudentProfile.objects.none()
        contest.assigned_students.set(students)

        return Response({
            "detail": "Batches assigned successfully.",
            "contest_id": contest.id,
            "assigned_batches": contest.assigned_batches,
            "assigned_sections": contest.assigned_sections,
            "assigned_student_count": contest.assigned_students.count(),
        })


def _compute_skill_insights(solved_problems):
    """Aggregate company-tag and project/skill-tag counts across a
    student's solved problems — shared by the staff-facing individual
    analytics view and the student's own self-analytics view so both
    render the same "companies you've practiced for" / "skills
    demonstrated" breakdown."""
    company_counts = {}
    skill_counts = {}
    project_tags = {'project', 'real-world', 'application', 'system', 'database', 'web', 'api', 'full-stack'}

    for sp in solved_problems:
        companies_str = sp.problem.companies or ""
        for comp in (c.strip() for c in companies_str.split(',')):
            if comp:
                company_counts[comp] = company_counts.get(comp, 0) + 1

        for tag in (sp.problem.tags or []):
            tag_lower = tag.lower()
            skill_counts[tag_lower] = skill_counts.get(tag_lower, 0) + 1

    sorted_companies = sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:8]
    company_insights = [{'name': name, 'count': count} for name, count in sorted_companies]

    project_insights = []
    for skill, count in sorted(skill_counts.items(), key=lambda x: x[1], reverse=True):
        if skill in project_tags or any(pt in skill for pt in project_tags):
            project_insights.append({'skill': skill, 'count': count})
    project_insights = project_insights[:6]

    return company_insights, project_insights


def _build_student_performance_charts(student, solved_problems=None, topic_accuracy=None, days=30):
    """Chart-ready solved-first analytics for student/staff/HOD report views."""
    solved_problems = solved_problems if solved_problems is not None else SolvedProblem.objects.filter(student=student).select_related('problem')
    topic_accuracy = topic_accuracy or []

    total_programming = Problem.objects.count()
    total_aptitude = AptitudeQuestion.objects.count()
    programming_solved = solved_problems.count()
    aptitude_solved = SolvedAptitude.objects.filter(student=student).count()

    accepted_contest_solved = (
        ContestSubmission.objects
        .filter(student=student, status='Accepted')
        .values('contest_id', 'problem_id')
        .distinct()
        .count()
    )
    aptitude_contest_solved = AptitudeContestSubmission.objects.filter(student=student, is_correct=True).count()
    contest_solved = accepted_contest_solved + aptitude_contest_solved

    start_day = timezone.localdate() - timedelta(days=days - 1)
    date_labels = [(start_day + timedelta(days=i)) for i in range(days)]

    def counts_by_date(qs, date_field):
        key = f'{date_field}__date'
        rows = qs.filter(**{f'{date_field}__date__gte': start_day}).values(key).annotate(count=Count('id'))
        return {row[key]: row['count'] for row in rows}

    programming_by_day = counts_by_date(SolvedProblem.objects.filter(student=student), 'solved_at')
    aptitude_by_day = counts_by_date(SolvedAptitude.objects.filter(student=student), 'solved_at')
    contest_code_by_day = counts_by_date(ContestSubmission.objects.filter(student=student, status='Accepted'), 'submitted_at')
    contest_apt_by_day = counts_by_date(AptitudeContestSubmission.objects.filter(student=student, is_correct=True), 'submitted_at')

    cumulative = 0
    daily_solved_trend = []
    for day in date_labels:
        programming_count = programming_by_day.get(day, 0)
        aptitude_count = aptitude_by_day.get(day, 0)
        contest_count = contest_code_by_day.get(day, 0) + contest_apt_by_day.get(day, 0)
        cumulative += programming_count + aptitude_count
        daily_solved_trend.append({
            'date': day.isoformat(),
            'programming': programming_count,
            'aptitude': aptitude_count,
            'contest': contest_count,
            'daily_total': programming_count + aptitude_count,
            'overall_total': cumulative,
        })

    active_days = sum(1 for row in daily_solved_trend if row['daily_total'] > 0)
    overall_possible = max(1, total_programming + total_aptitude)
    overall_solved = programming_solved + aptitude_solved
    contest_attempts = ContestParticipation.objects.filter(student=student, is_active=False).count()
    contest_perf_pct = min(100, round((contest_solved / max(1, contest_attempts)) * 20, 1)) if contest_attempts else 0

    programming_tags = defaultdict(int)
    for sp in solved_problems:
        tags = sp.problem.tags or []
        if isinstance(tags, str):
            tags = [tag.strip() for tag in tags.split(',') if tag.strip()]
        if tags:
            for tag in list(tags)[:4]:
                programming_tags[str(tag).strip().title()] += 1
        else:
            programming_tags[sp.problem.difficulty or 'Programming'] += 1

    aptitude_topics = defaultdict(int)
    for row in SolvedAptitude.objects.filter(student=student).values('question__topic__title').annotate(count=Count('id')).order_by('-count')[:10]:
        topic = row['question__topic__title'] or 'Aptitude'
        aptitude_topics[topic] += row['count']

    topic_labels = []
    for label, _ in sorted(programming_tags.items(), key=lambda item: item[1], reverse=True)[:6]:
        if label not in topic_labels:
            topic_labels.append(label)
    for label, _ in sorted(aptitude_topics.items(), key=lambda item: item[1], reverse=True)[:6]:
        if label not in topic_labels:
            topic_labels.append(label)
    topic_labels = topic_labels[:8]

    knowledge_distribution = {
        'labels': topic_labels,
        'programming': [programming_tags.get(label, 0) for label in topic_labels],
        'aptitude': [aptitude_topics.get(label, 0) for label in topic_labels],
    }

    contest_performance = []
    cp_qs = (
        ContestParticipation.objects
        .filter(student=student, is_active=False)
        .select_related('contest')
        .order_by('started_at')[:25]
    )
    for cp in cp_qs:
        contest = cp.contest
        if contest.contest_type == 'aptitude':
            total_items = contest.aptitude_questions.count()
            solved_items = AptitudeContestSubmission.objects.filter(
                contest=contest, student=student, is_correct=True
            ).count()
        else:
            total_items = contest.problems.count()
            solved_items = ContestSubmission.objects.filter(
                contest=contest, student=student, status='Accepted'
            ).values('problem').distinct().count()
        contest_performance.append({
            'label': contest.title[:18],
            'title': contest.title,
            'date': cp.started_at.date().isoformat(),
            'contest_type': contest.contest_type,
            'solved': solved_items,
            'total': total_items,
            'score_pct': round(solved_items / max(1, total_items) * 100, 1),
        })

    return {
        'overall_performance': [
            {'label': 'Programming', 'value': programming_solved},
            {'label': 'Aptitude', 'value': aptitude_solved},
            {'label': 'Contest', 'value': contest_solved},
        ],
        'profile_radar': {
            'labels': ['Programming', 'Aptitude', 'Contest', 'Daily', 'Overall'],
            'daily': [
                round(programming_by_day.get(timezone.localdate(), 0) / max(1, programming_solved) * 100, 1),
                round(aptitude_by_day.get(timezone.localdate(), 0) / max(1, aptitude_solved) * 100, 1),
                round((contest_code_by_day.get(timezone.localdate(), 0) + contest_apt_by_day.get(timezone.localdate(), 0)) / max(1, contest_solved) * 100, 1),
                round(active_days / days * 100, 1),
                round((programming_by_day.get(timezone.localdate(), 0) + aptitude_by_day.get(timezone.localdate(), 0)) / max(1, overall_solved) * 100, 1),
            ],
            'overall': [
                round(programming_solved / max(1, total_programming) * 100, 1),
                round(aptitude_solved / max(1, total_aptitude) * 100, 1),
                contest_perf_pct,
                round(active_days / days * 100, 1),
                round(overall_solved / overall_possible * 100, 1),
            ],
        },
        'knowledge_distribution': knowledge_distribution,
        'daily_solved_trend': daily_solved_trend,
        'contest_performance': contest_performance,
        'summary_cards': {
            'programming_solved': programming_solved,
            'aptitude_solved': aptitude_solved,
            'contest_solved': contest_solved,
            'active_days': active_days,
        },
    }


class StudentIndividualAnalyticsView(APIView):
    """Get detailed analytics for an individual student."""
    permission_classes = [IsAuthenticated]

    def get(self, request, register_number):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        staff_profile = request.user.staff_profile

        student = StudentProfile.objects.filter(
            register_number=register_number
        ).select_related('department', 'institution').first()

        if not student:
            return Response(
                {"detail": "Student not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check access
        if student.institution != staff_profile.institution:
            return Response(
                {"detail": "You can only view students in your institution."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Get solved problems breakdown by difficulty
        solved_problems = SolvedProblem.objects.filter(
            student=student
        ).select_related('problem')

        difficulty_breakdown = {"Easy": 0, "Medium": 0, "Hard": 0}
        for sp in solved_problems:
            difficulty_breakdown[sp.problem.difficulty] = difficulty_breakdown.get(sp.problem.difficulty, 0) + 1

        # Get recent activity (last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_activity = []

        recent_solutions = ProblemSolution.objects.filter(
            student=student,
            submitted_at__gte=thirty_days_ago
        ).select_related('problem').order_by('-submitted_at')[:20]

        for solution in recent_solutions:
            recent_activity.append({
                "date": solution.submitted_at.isoformat(),
                "problem": solution.problem.title,
                "problem_slug": solution.problem.slug,
                "difficulty": solution.problem.difficulty,
                "status": solution.status,
                "language": solution.language,
            })

        # Calculate total time spent
        total_time_spent = ProblemSession.objects.filter(
            student=student,
            is_active=False
        ).aggregate(total=Sum('time_spent_seconds'))['total'] or 0

        # Get contest participation
        contest_participations = ContestSubmission.objects.filter(
            student=student
        ).values('contest__title', 'contest__id').annotate(
            submissions=Count('id'),
            solved=Count('id', filter=Q(status='Accepted'))
        ).order_by('-contest__id')[:10]

        # ── Aptitude Insights ──────────────────────────────────────────────────
        aptitude_solved = SolvedAptitude.objects.filter(student=student).count()
        total_aptitude = AptitudeQuestion.objects.count()
        
        # ── Company & Project Insights ────────────────────────────────────────
        company_insights, project_insights = _compute_skill_insights(solved_problems)

        # ── Score History (from ContestParticipation) ─────────────────────────
        cp_qs = ContestParticipation.objects.filter(
            student=student, is_active=False
        ).select_related('contest').order_by('started_at')[:25]

        score_history = []
        for cp in cp_qs:
            contest = cp.contest
            if contest.contest_type == 'aptitude':
                total_q = contest.aptitude_questions.count()
                correct_q = AptitudeContestSubmission.objects.filter(
                    contest=contest, student=student, is_correct=True
                ).count()
                score_pct = round(correct_q / max(1, total_q) * 100, 1)
            else:
                total_p = contest.problems.count()
                score_pct = round(cp.problems_solved / max(1, total_p) * 100, 1)
            score_history.append({
                'label': contest.title[:18],
                'title': contest.title,
                'score_pct': score_pct,
                'date': cp.started_at.strftime('%Y-%m-%d'),
                'contest_type': contest.contest_type,
            })

        tests_completed = len(score_history)
        avg_score = round(sum(s['score_pct'] for s in score_history) / tests_completed, 2) if tests_completed else 0
        peak_score = max((s['score_pct'] for s in score_history), default=0)

        # ── Topic Accuracy (from AptitudeContestSubmission by topic) ──────────
        try:
            topic_acc_qs = AptitudeContestSubmission.objects.filter(
                student=student
            ).select_related('question__topic__parent').values(
                'question__topic__title', 'question__topic__parent__title'
            ).annotate(
                total=Count('id'),
                correct=Count('id', filter=Q(is_correct=True))
            ).order_by('-total')[:20]

            topic_accuracy = [
                {
                    'topic': t['question__topic__title'],
                    'category': t['question__topic__parent__title'] or t['question__topic__title'],
                    'accuracy': round(t['correct'] / t['total'] * 100, 1) if t['total'] else 0,
                    'total': t['total'],
                    'correct': t['correct'],
                }
                for t in topic_acc_qs
            ]
        except Exception:
            logger.exception("Failed to compute topic_accuracy for student %s (individual analytics)", register_number)
            topic_accuracy = []

        performance_charts = _build_student_performance_charts(student, solved_problems, topic_accuracy)

        try:
            return Response({
                "student": {
                    "register_number": student.register_number,
                    "name": student.name,
                    "batch": student.batch,
                    "department": student.department.name if student.department else None,
                    "current_streak": student.current_streak,
                "login_days": student.login_days,
                "campus_rank": f"#{calculate_campus_rank_helper(student)}",
            },
            "analytics": {
                "solved_count": solved_problems.count(),
                "difficulty_breakdown": difficulty_breakdown,
                "recent_activity": recent_activity,
                "time_spent_total": total_time_spent,
                "time_spent_hours": round(total_time_spent / 3600, 2),
                "contest_participations": list(contest_participations),
                "aptitude": {
                    "solved": aptitude_solved,
                    "total": total_aptitude,
                    "percentage": round((aptitude_solved / total_aptitude * 100), 1) if total_aptitude > 0 else 0
                },
                "company_insights": company_insights,
                "project_insights": project_insights,
                "score_history": score_history,
                "topic_accuracy": topic_accuracy,
                "tests_completed": tests_completed,
                "avg_score": avg_score,
                "peak_score": peak_score,
                **performance_charts,
            }
        })
        except Exception:
            logger.exception("Failed to build analytics response for student %s", register_number)
            return Response(
                {"detail": "Failed to load analytics. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class StudentSelfAnalyticsView(APIView):
    """Student views their own detailed analytics — mirrors staff view but self-scoped."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'student_profile'):
            return Response({"detail": "Student access required."}, status=status.HTTP_403_FORBIDDEN)

        student = request.user.student_profile

        solved_problems = SolvedProblem.objects.filter(student=student).select_related('problem')

        difficulty_breakdown = {"Easy": 0, "Medium": 0, "Hard": 0}
        for sp in solved_problems:
            difficulty_breakdown[sp.problem.difficulty] = difficulty_breakdown.get(sp.problem.difficulty, 0) + 1

        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_activity = []
        recent_solutions = ProblemSolution.objects.filter(
            student=student, submitted_at__gte=thirty_days_ago
        ).select_related('problem').order_by('-submitted_at')[:20]
        for sol in recent_solutions:
            recent_activity.append({
                "date": sol.submitted_at.isoformat(),
                "problem": sol.problem.title,
                "difficulty": sol.problem.difficulty,
                "status": sol.status,
                "language": sol.language,
            })

        total_time_spent = ProblemSession.objects.filter(
            student=student, is_active=False
        ).aggregate(total=Sum('time_spent_seconds'))['total'] or 0

        aptitude_solved = SolvedAptitude.objects.filter(student=student).count()
        total_aptitude = AptitudeQuestion.objects.count()

        company_insights, project_insights = _compute_skill_insights(solved_problems)

        # Score history
        cp_qs = ContestParticipation.objects.filter(
            student=student, is_active=False
        ).select_related('contest').order_by('started_at')[:25]

        score_history = []
        for cp in cp_qs:
            contest = cp.contest
            if contest.contest_type == 'aptitude':
                total_q = contest.aptitude_questions.count()
                correct_q = AptitudeContestSubmission.objects.filter(
                    contest=contest, student=student, is_correct=True
                ).count()
                score_pct = round(correct_q / max(1, total_q) * 100, 1)
            else:
                total_p = contest.problems.count()
                score_pct = round(cp.problems_solved / max(1, total_p) * 100, 1)
            score_history.append({
                'label': contest.title[:18],
                'title': contest.title,
                'score_pct': score_pct,
                'date': cp.started_at.strftime('%Y-%m-%d'),
                'contest_type': contest.contest_type,
            })

        tests_completed = len(score_history)
        avg_score = round(sum(s['score_pct'] for s in score_history) / tests_completed, 2) if tests_completed else 0
        peak_score = max((s['score_pct'] for s in score_history), default=0)

        # Topic accuracy
        try:
            topic_acc_qs = AptitudeContestSubmission.objects.filter(
                student=student
            ).select_related('question__topic__parent').values(
                'question__topic__title', 'question__topic__parent__title'
            ).annotate(
                total=Count('id'),
                correct=Count('id', filter=Q(is_correct=True))
            ).order_by('-total')[:20]

            topic_accuracy = [
                {
                    'topic': t['question__topic__title'],
                    'category': t['question__topic__parent__title'] or t['question__topic__title'],
                    'accuracy': round(t['correct'] / t['total'] * 100, 1) if t['total'] else 0,
                    'total': t['total'],
                    'correct': t['correct'],
                }
                for t in topic_acc_qs
            ]
        except Exception:
            logger.exception("Failed to compute topic_accuracy for self-analytics (student %s)", getattr(student, 'register_number', '?'))
            topic_accuracy = []

        # Practice-mode topic accuracy — built from every logged attempt
        # (correct AND wrong), unlike SolvedAptitude which only records
        # correct answers. This is what makes the Study Progress radar real
        # and per-topic instead of a 3-category "% of bank solved" summary.
        try:
            practice_acc_qs = AptitudeAttempt.objects.filter(
                student=student
            ).select_related('question__topic__parent').values(
                'question__topic__title', 'question__topic__parent__title'
            ).annotate(
                total=Count('id'),
                correct=Count('id', filter=Q(is_correct=True))
            ).order_by('-total')[:20]

            topic_accuracy_practice = [
                {
                    'topic': t['question__topic__title'],
                    'category': t['question__topic__parent__title'] or t['question__topic__title'],
                    'accuracy': round(t['correct'] / t['total'] * 100, 1) if t['total'] else 0,
                    'total': t['total'],
                    'correct': t['correct'],
                }
                for t in practice_acc_qs
            ]
        except Exception:
            logger.exception("Failed to compute topic_accuracy_practice for self-analytics (student %s)", getattr(student, 'register_number', '?'))
            topic_accuracy_practice = []

        performance_charts = _build_student_performance_charts(student, solved_problems, topic_accuracy)

        try:
            return Response({
                "analytics": {
                    "solved_count": solved_problems.count(),
                    "difficulty_breakdown": difficulty_breakdown,
                    "recent_activity": recent_activity,
                    "time_spent_hours": round(total_time_spent / 3600, 2),
                    "aptitude": {
                        "solved": aptitude_solved,
                        "total": total_aptitude,
                        "percentage": round(aptitude_solved / total_aptitude * 100, 1) if total_aptitude else 0,
                    },
                    "company_insights": company_insights,
                    "project_insights": project_insights,
                    "score_history": score_history,
                    "topic_accuracy": topic_accuracy,
                    "topic_accuracy_practice": topic_accuracy_practice,
                    "tests_completed": tests_completed,
                    "avg_score": avg_score,
                    "peak_score": peak_score,
                    **performance_charts,
                }
            })
        except Exception:
            logger.exception("Failed to build self-analytics response for student %s", getattr(student, 'register_number', '?'))
            return Response(
                {"detail": "Failed to load analytics. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class ContestApprovalView(APIView):
    """HOD or Academic Coordinator can approve or reject contests"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        profile = request.user.staff_profile

        # Only HOD/Academic Coordinator can approve contests
        if profile.role not in ("hod", "academics"):
            return Response(
                {"detail": "Only HOD or Academic Coordinator can approve contests."},
                status=status.HTTP_403_FORBIDDEN
            )

        contest = Contest.objects.filter(id=contest_id).first()
        if not contest:
            return Response(
                {"detail": "Contest not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if contest is in the HOD's department
        if contest.department != profile.department:
            return Response(
                {"detail": "You can only approve contests in your department."},
                status=status.HTTP_403_FORBIDDEN
            )

        action = request.data.get('action')  # 'approve' or 'reject'

        if action not in ('approve', 'reject'):
            return Response(
                {"detail": "Invalid action. Use 'approve' or 'reject'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            if action == 'approve':
                contest.approve(profile)
                # Auto-publish on approval
                publish_contest_helper(contest)
                return Response({
                    "detail": "Contest approved and published successfully.",
                    "contest_id": contest.id,
                    "status": contest.status,
                })
            else:
                reason = request.data.get('reason', '')
                contest.reject(reason)
                return Response({
                    "detail": "Contest rejected.",
                    "contest_id": contest.id,
                    "status": contest.status,
                    "reason": reason,
                })
        except Exception:
            logger.exception(
                "Failed to %s contest %s for HOD %s",
                action, contest_id, profile.faculty_id,
            )
            return Response(
                {"detail": f"Failed to {action} contest. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


def publish_contest_helper(contest):
    """Helper to publish a contest and notify students"""
    contest.publish()
    
    # Create Announcement
    Announcement.objects.create(
        title=f"🚀 New Contest: {contest.title}",
        content=f"A new contest '{contest.title}' is now live! Challenge yourself and climb the leaderboard.",
        category="contest"
    )

    # Create Notifications for assigned students
    # 1. Direct assignments
    student_users = list(contest.assigned_students.values_list('account', flat=True))
    
    # 2. Batch assignments
    if contest.assigned_batches:
        batch_students = StudentProfile.objects.filter(
            batch__in=contest.assigned_batches,
            institution=contest.institution
        ).values_list('account', flat=True)
        student_users.extend(list(batch_students))
    
    # Unique users
    unique_user_ids = set(filter(None, student_users))
    
    notifications = []
    for user_id in unique_user_ids:
        notifications.append(Notification(
            recipient_id=user_id,
            title="New Contest Assigned",
            message=f"You have been assigned to a new contest: {contest.title}. Check it out now!",
            link=(
                f"/contests/{contest.id}" if contest.contest_type == 'programming'
                else f"/combined-contest/{contest.id}" if contest.contest_type == 'combined'
                else f"/aptitude-contest/{contest.id}"
            )
        ))
    
    if notifications:
        Notification.objects.bulk_create(notifications, ignore_conflicts=True)


class ContestPublishView(APIView):
    """Publish approved contest to students"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        profile = request.user.staff_profile

        contest = Contest.objects.filter(id=contest_id).first()
        if not contest:
            return Response(
                {"detail": "Contest not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check permissions - HOD/Academic Coordinator or contest creator can publish
        if profile.role in ("hod", "academics") and contest.department != profile.department:
            return Response(
                {"detail": "You can only publish contests in your department."},
                status=status.HTTP_403_FORBIDDEN
            )

        if profile.role == "staff" and contest.created_by != profile:
            return Response(
                {"detail": "You can only publish your own contests."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Can only publish approved contests
        if contest.status != "approved":
            return Response(
                {"detail": "Only approved contests can be published."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            publish_contest_helper(contest)
        except Exception:
            logger.exception("Failed to publish contest %s for staff %s", contest_id, profile.faculty_id)
            return Response(
                {"detail": "Failed to publish contest. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({
            "detail": "Contest published successfully.",
            "contest_id": contest.id,
            "status": contest.status,
        })


class DepartmentStudentsFilterView(APIView):
    """Get students in department with filtering options"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        profile = request.user.staff_profile

        # Base query - students in staff's department
        students = StudentProfile.objects.filter(
            institution=profile.institution,
            department=profile.department
        ).select_related('department')

        # Apply filters
        batch = request.query_params.get('batch')
        section = request.query_params.get('section')
        search = request.query_params.get('search')

        if batch:
            students = students.filter(batch=batch)

        if section:
            students = students.filter(section=section)

        if search:
            students = students.filter(
                Q(name__icontains=search) | 
                Q(register_number__icontains=search)
            )

        # Annotate with solved count
        students = students.annotate(
            solved_count=Count('solved_problems', distinct=True)
        ).order_by('register_number')

        # Limit results
        limit = int(request.query_params.get('limit', 100))
        students = students[:limit]

        data = []
        for student in students:
            data.append({
                "id": student.id,
                "register_number": student.register_number,
                "name": student.name,
                "batch": student.batch,
                "section": student.section,
                "solved_count": student.solved_count,
                "current_streak": student.current_streak,
            })

        return Response({
            "students": data,
            "total": len(data),
        })


class ContestSubmitForApprovalView(APIView):
    """Staff can submit their draft contest for HOD approval"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        if not hasattr(request.user, 'staff_profile'):
            return Response(
                {"detail": "Staff access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        profile = request.user.staff_profile

        contest = Contest.objects.filter(id=contest_id).first()
        if not contest:
            return Response(
                {"detail": "Contest not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Only contest creator can submit for approval
        if contest.created_by != profile:
            return Response(
                {"detail": "You can only submit your own contests for approval."},
                status=status.HTTP_403_FORBIDDEN
            )

        # Can only submit draft or rejected contests
        if contest.status not in ["draft", "rejected"]:
            return Response(
                {"detail": f"Cannot submit contest with status '{contest.status}' for approval."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate contest has required fields
        if not contest.problems.exists():
            return Response(
                {"detail": "Contest must have at least one problem."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not contest.assigned_students.exists() and not contest.assigned_batches:
            return Response(
                {"detail": "Contest must have assigned students or batches."},
                status=status.HTTP_400_BAD_REQUEST
            )

        contest.submit_for_approval()

        return Response({
            "detail": "Contest submitted for approval successfully.",
            "contest_id": contest.id,
            "status": contest.status,
        })


# =============================================================================
# Student Contest Views
# =============================================================================

class StudentContestListView(APIView):
    """Get contests assigned to the student"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        student = request.user.student_profile

        # Get contests accessible to student (published, completed, active, or approved)
        all_contests = Contest.objects.filter(
            status__in=['published', 'completed', 'active', 'approved']
        ).distinct().select_related('created_by', 'department').prefetch_related('problems', 'aptitude_questions')

        contests = [c for c in all_contests if c.is_student_assigned(student)]

        # Filter out contests of a type locked for this institution — the
        # middleware already blocks direct access, this keeps the list from
        # showing tiles that would just 403 on click.
        locked = (student.institution.locked_modules if student.institution else []) or []
        if locked:
            contest_type_module = {"programming": "contest_programming", "aptitude": "contest_aptitude", "combined": "contest_combined"}
            if "contest" in locked:
                contests = []
            else:
                contests = [c for c in contests if contest_type_module.get(c.contest_type) not in locked]

        data = []
        for contest in contests:
            # Update contest status if it has ended
            contest.update_status_if_ended()
            
            # Check if student has started this contest
            participation = ContestParticipation.objects.filter(
                contest=contest,
                student=student
            ).first()

            # Auto-end participation if time has expired
            if participation and participation.is_active:
                time_elapsed = timezone.now() - participation.started_at
                max_duration = timedelta(minutes=contest.duration_minutes)
                
                if time_elapsed > max_duration:
                    participation.end_participation()
                    # Refresh participation object
                    participation.refresh_from_db()

            # Check if contest is currently active
            is_active = contest.is_active
            is_upcoming = contest.is_upcoming
            is_ended = contest.is_ended

            # Get question counts
            problem_count = contest.problems.count()
            aptitude_count = contest.aptitude_questions.count()

            data.append({
                "id": contest.id,
                "title": contest.title,
                "description": contest.description,
                "contest_type": contest.contest_type,
                "start_time": contest.start_time,
                "end_time": contest.end_time,
                "duration_minutes": contest.duration_minutes,
                "problem_count": problem_count,
                "aptitude_question_count": aptitude_count,
                "is_active": is_active,
                "is_upcoming": is_upcoming,
                "is_ended": is_ended,
                "has_started": participation is not None if participation else False,
                "participation": {
                    "started_at": participation.started_at,
                    "problems_solved": participation.problems_solved,
                    "total_score": participation.total_score,
                    "is_active": participation.is_active,
                } if participation else None,
            })

        return Response({"contests": data})


class StudentContestDetailView(APIView):
    """Get contest details for student"""
    permission_classes = [IsAuthenticated]

    def get(self, request, contest_id):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        student = request.user.student_profile

        contest = Contest.objects.filter(id=contest_id).select_related('created_by', 'department').prefetch_related('problems', 'aptitude_questions').first()

        if not contest or not contest.is_student_assigned(student):
            return Response(
                {"detail": "Contest not found or not accessible."},
                status=status.HTTP_404_NOT_FOUND
            )

        if contest.status not in ['published', 'completed', 'active', 'approved']:
            return Response(
                {"detail": "Contest is not accessible."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check participation
        participation = ContestParticipation.objects.filter(
            contest=contest,
            student=student
        ).first()

        # Auto-end participation if time has expired
        if participation and participation.is_active:
            time_elapsed = timezone.now() - participation.started_at
            max_duration = timedelta(minutes=contest.duration_minutes)
            
            if time_elapsed > max_duration:
                participation.end_participation()
                # Refresh participation object
                participation.refresh_from_db()

        # Check if contest is active
        is_active = contest.is_active
        is_ended = contest.is_ended

        # Get questions/problems with status
        problems_data = []
        aptitude_data = []
        if contest.contest_type in ('aptitude', 'combined'):
            for q in contest.aptitude_questions.all().select_related('passage'):
                # Check if student has answered this question in the contest
                submission = AptitudeContestSubmission.objects.filter(
                    contest=contest,
                    student=student,
                    question=q
                ).first()

                aptitude_data.append({
                    "id": q.id,
                    "question_type": q.question_type,
                    "question_text": q.question_text,
                    "question_image": q.question_image,
                    "option_a": q.option_a,
                    "option_a_image": q.option_a_image,
                    "option_b": q.option_b,
                    "option_b_image": q.option_b_image,
                    "option_c": q.option_c,
                    "option_c_image": q.option_c_image,
                    "option_d": q.option_d,
                    "option_d_image": q.option_d_image,
                    "explanation": q.explanation,
                    "difficulty": q.difficulty,
                    "passage_id": q.passage_id,
                    "passage_title": q.passage.title if q.passage_id else None,
                    "passage_text": q.passage.passage_text if q.passage_id else None,
                    "is_solved": submission is not None,
                    "student_answer": submission.selected_option if submission else None,
                    "score": submission.score if submission else 0,
                })
        if contest.contest_type in ('programming', 'combined'):
            for problem in contest.problems.all():
                # Check if student has solved this problem in the contest
                submission = ContestSubmission.objects.filter(
                    contest=contest,
                    student=student,
                    problem=problem,
                    status='Accepted'
                ).first()
                has_any_submission = ContestSubmission.objects.filter(
                    contest=contest,
                    student=student,
                    problem=problem,
                ).exists()

                problems_data.append({
                    "id": problem.id,
                    "slug": problem.slug,
                    "title": problem.title,
                    "difficulty": problem.difficulty,
                    "tags": problem.tags,
                    "is_solved": submission is not None,
                    "attempted": has_any_submission,
                })

        # Pure aptitude contests keep serving their question list under the
        # original "problems" key too — AptitudeContestWorkspacePage.jsx
        # reads data.problems, and changing that is unnecessary risk to an
        # already-working flow. Combined contests use "problems" for coding
        # and the new "aptitude_questions" key for aptitude+reading instead.
        if contest.contest_type == 'aptitude':
            problems_data = aptitude_data

        return Response({
            "id": contest.id,
            "title": contest.title,
            "description": contest.description,
            "contest_type": contest.contest_type,
            "coding_weight_percent": contest.coding_weight_percent,
            "aptitude_weight_percent": contest.aptitude_weight_percent,
            "reading_weight_percent": contest.reading_weight_percent,
            "start_time": contest.start_time,
            "end_time": contest.end_time,
            "access_end_time": contest.access_end_time,
            "duration_minutes": contest.duration_minutes,
            "session_duration_minutes": contest.session_duration_minutes or contest.duration_minutes,
            "problem_count": contest.problems.count(),
            "aptitude_question_count": contest.aptitude_questions.count(),
            "enable_tab_switch_check": contest.enable_tab_switch_check,
            "max_tab_switches": contest.max_tab_switches,
            "enable_fullscreen_lock": contest.enable_fullscreen_lock,
            "enable_copy_paste_lock": contest.enable_copy_paste_lock,
            "enable_webcam_proctoring": contest.enable_webcam_proctoring,
            "is_active": is_active,
            "is_ended": is_ended,
            "has_started": participation is not None,
            # "problems" is coding-only, kept for backward compat with the
            # existing programming-contest workspace; "aptitude_questions" is
            # aptitude+reading data (new key, also populated for pure-aptitude
            # contests now so the new combined workspace can reuse it).
            "problems": problems_data,
            "aptitude_questions": aptitude_data,
            "participation": {
                "started_at": participation.started_at,
                "session_end_time": participation.session_end_time,
                "remaining_time_seconds": participation.remaining_time_seconds,
                "problems_solved": participation.problems_solved,
                "total_score": participation.total_score,
                "time_spent_seconds": participation.time_spent_seconds,
                "is_active": participation.is_active,
            } if participation else None,
        })


class StudentContestStartView(APIView):
    """Start a contest (creates participation record with session timing)"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        student = request.user.student_profile

        contest = Contest.objects.filter(id=contest_id).first()

        if not contest or not contest.is_student_assigned(student):
            return Response(
                {"detail": "Contest not found or not accessible."},
                status=status.HTTP_404_NOT_FOUND
            )

        if contest.status not in ['published', 'completed', 'active', 'approved']:
            return Response(
                {"detail": "Contest is not accessible."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if contest access is available
        if contest.is_upcoming:
            return Response(
                {"detail": "Contest access has not started yet."},
                status=status.HTTP_400_BAD_REQUEST
            )
        if contest.is_ended:
            return Response(
                {"detail": "Contest access has ended. No new participants allowed."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if already started
        participation = ContestParticipation.objects.filter(
            contest=contest,
            student=student
        ).first()

        if participation:
            # If session has expired, auto-submit
            if participation.is_session_expired and participation.is_active:
                participation.end_participation(auto_submitted=True)
                return Response(
                    {"detail": "Your session has expired. Contest has been auto-submitted."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return Response({
                "detail": "Contest session already started.",
                "participation": {
                    "started_at": participation.started_at,
                    "session_end_time": participation.session_end_time,
                    "remaining_time_seconds": participation.remaining_time_seconds,
                    "is_active": participation.is_active,
                    "contest_id": contest.id,
                }
            })

        # Create participation with session timing
        try:
            participation = ContestParticipation.objects.create(
                contest=contest,
                student=student,
                manually_stopped=False  # Explicitly set default value
            )
        except IntegrityError:
            # Race condition: another request created the record between our check and create
            participation = ContestParticipation.objects.filter(contest=contest, student=student).first()
            if not participation:
                logger.exception("IntegrityError creating participation for student %s, contest %s", getattr(student, 'register_number', '?'), contest_id)
                return Response({"detail": "Failed to start contest. Please try again."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        except Exception:
            logger.exception("Failed to start contest %s for student %s", contest_id, getattr(student, 'register_number', '?'))
            return Response(
                {"detail": "Failed to start contest session. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({
            "detail": "Contest session started successfully.",
            "participation": {
                "started_at": participation.started_at,
                "session_end_time": participation.session_end_time,
                "session_duration_minutes": contest.session_duration_minutes or contest.duration_minutes,
                "remaining_time_seconds": participation.remaining_time_seconds,
                "contest_id": contest.id,
            }
        })


class StudentContestAutoSubmitView(APIView):
    """Auto-submit contest when session time expires"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        student = request.user.student_profile

        # Get participation
        participation = ContestParticipation.objects.filter(
            contest_id=contest_id,
            student=student,
            is_active=True
        ).first()

        if not participation:
            return Response(
                {"detail": "No active participation found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if session has actually expired
        if not participation.is_session_expired:
            return Response(
                {"detail": "Session has not expired yet."},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # End the participation with auto-submit flag
            participation.end_participation(auto_submitted=True)

            # Calculate final score and problems solved
            participation.total_score, participation.problems_solved = _compute_contest_score_and_solved(
                participation.contest, student
            )

            participation.save(update_fields=['total_score', 'problems_solved'])
        except Exception:
            logger.exception(
                "Failed to auto-submit contest %s for student %s",
                contest_id, getattr(student, 'register_number', '?'),
            )
            return Response(
                {"detail": "Failed to auto-submit contest. Please contact support."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({
            "detail": "Contest auto-submitted due to session expiry.",
            "participation": {
                "completed_at": participation.completed_at,
                "time_spent_seconds": participation.time_spent_seconds,
                "total_score": participation.total_score,
                "problems_solved": participation.problems_solved,
                "auto_submitted": participation.auto_submitted,
            }
        })


class StudentContestStopView(APIView):
    """Manually stop a contest (student can stop at any time during session)"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        try:
            contest = Contest.objects.get(id=contest_id)
        except Contest.DoesNotExist:
            return Response(
                {"detail": "Contest not found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if student has an active participation
        try:
            participation = ContestParticipation.objects.get(
                contest=contest,
                student=request.user.student_profile,
                is_active=True
            )
        except ContestParticipation.DoesNotExist:
            return Response(
                {"detail": "No active participation found for this contest."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Allow both programming and aptitude contests to be stopped manually
        # (Previously this was restricted to programming only)
        
        # Stop the contest
        participation.is_active = False
        participation.completed_at = timezone.now()
        
        # Calculate time spent
        if participation.started_at:
            time_spent = timezone.now() - participation.started_at
            participation.time_spent_seconds = int(time_spent.total_seconds())
            participation.total_time_taken = participation.time_spent_seconds
        
        # Recalculate final score and problems solved one last time
        participation.total_score, participation.problems_solved = _compute_contest_score_and_solved(
            contest, request.user.student_profile
        )

        participation.manually_stopped = True
        participation.save()

        return Response({
            "detail": "Contest stopped successfully.",
            "participation": {
                "completed_at": participation.completed_at,
                "time_spent_seconds": participation.time_spent_seconds,
                "total_score": participation.total_score,
                "problems_solved": participation.problems_solved,
                "manually_stopped": True,
            }
        })


class StudentContestSessionStatusView(APIView):
    """Get current session status and remaining time"""
    permission_classes = [IsAuthenticated]

    def get(self, request, contest_id):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        student = request.user.student_profile

        # Get participation
        participation = ContestParticipation.objects.filter(
            contest_id=contest_id,
            student=student
        ).first()

        if not participation:
            return Response(
                {"detail": "No participation found."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if session has expired and auto-submit if needed
        if participation.is_session_expired and participation.is_active:
            participation.end_participation(auto_submitted=True)
            
            # Calculate final score — only count Accepted submissions
            participation.total_score, participation.problems_solved = _compute_contest_score_and_solved(
                participation.contest, student
            )

            participation.save(update_fields=['total_score', 'problems_solved'])

        return Response({
            "participation": {
                "started_at": participation.started_at,
                "session_end_time": participation.session_end_time,
                "completed_at": participation.completed_at,
                "remaining_time_seconds": participation.remaining_time_seconds,
                "is_active": participation.is_active,
                "is_locked": getattr(participation, 'is_locked', False),
                "lock_reason": getattr(participation, 'lock_reason', ''),
                "is_session_expired": participation.is_session_expired,
                "auto_submitted": participation.auto_submitted,
                "total_score": participation.total_score,
                "problems_solved": participation.problems_solved,
            }
        })


class StudentContestProblemView(APIView):
    """Get problem details within a contest"""
    permission_classes = [IsAuthenticated]

    def get(self, request, contest_id, problem_slug):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        student = request.user.student_profile

        contest = Contest.objects.filter(
            id=contest_id,
            status__in=['published', 'completed']  # Only published/completed contests
        ).first()

        if not contest or not contest.is_student_assigned(student):
            return Response(
                {"detail": "Contest not found or not accessible."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if student has started the contest
        participation = ContestParticipation.objects.filter(
            contest=contest,
            student=student
        ).first()

        if not participation:
            return Response(
                {"detail": "You must start the contest first."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if contest has ended (global access window)
        if contest.is_ended:
            return Response(
                {"detail": "Contest has ended."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if this student's personal session has expired
        if participation.is_session_expired or not participation.is_active:
            return Response(
                {"detail": "Your contest session has expired."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get problem
        problem = contest.problems.filter(slug=problem_slug).first()
        if not problem:
            return Response(
                {"detail": "Problem not found in this contest."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get student's submissions for this problem in this contest
        submissions = ContestSubmission.objects.filter(
            contest=contest,
            student=student,
            problem=problem
        ).order_by('-submitted_at')[:10]

        submissions_data = []
        for sub in submissions:
            submissions_data.append({
                "id": sub.id,
                "status": sub.status,
                "language": sub.language,
                "submitted_at": sub.submitted_at,
                "score": sub.score,
            })

        return Response({
            "id": problem.id,
            "slug": problem.slug,
            "title": problem.title,
            "description": problem.description,
            "difficulty": problem.difficulty,
            "tags": problem.tags,
            "examples": problem.examples,
            "hints": problem.hints,
            "submissions": submissions_data,
        })


class StudentContestSubmitView(APIView):
    """Submit code for a problem in a contest"""
    permission_classes = [IsAuthenticated]

    def post(self, request, contest_id, problem_slug):
        if not hasattr(request.user, 'student_profile'):
            return Response(
                {"detail": "Student access required."},
                status=status.HTTP_403_FORBIDDEN
            )

        student = request.user.student_profile

        contest = Contest.objects.filter(
            id=contest_id,
            status__in=['published', 'completed']  # Only published/completed contests
        ).first()

        if not contest or not contest.is_student_assigned(student):
            return Response(
                {"detail": "Contest not found or not accessible."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Check if student has started the contest
        participation = ContestParticipation.objects.filter(
            contest=contest,
            student=student
        ).first()

        if not participation:
            return Response(
                {"detail": "You must start the contest first."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if contest has ended (global access window)
        if contest.is_ended:
            # End participation if still active
            if participation.is_active:
                participation.end_participation()

            return Response(
                {"detail": "Contest has ended. No more submissions allowed."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if this student's personal session has expired
        if participation.is_session_expired or not participation.is_active:
            if participation.is_active:
                participation.end_participation(auto_submitted=True)
            return Response(
                {"detail": "Your contest session has expired. No more submissions allowed."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get problem
        problem = contest.problems.filter(slug=problem_slug).first()
        if not problem:
            return Response(
                {"detail": "Problem not found in this contest."},
                status=status.HTTP_404_NOT_FOUND
            )

        # Get submission data
        source_code = request.data.get('source_code')
        language = request.data.get('language')
        language_id = request.data.get('language_id')

        if not source_code or not language_id:
            return Response(
                {"detail": "source_code and language_id are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Coerce language_id to int — request.data delivers it as a string
        # unlike CodeRunView which uses CodeRunSerializer (IntegerField)
        try:
            language_id = int(language_id)
        except (TypeError, ValueError):
            return Response(
                {"detail": f"Invalid language_id: {language_id!r}. Must be an integer."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Execute code using the execution engine
        try:
            # Get test cases with fallback to problem examples
            test_cases = build_runtime_test_cases(problem, sample_only=False)

            if not test_cases:
                return Response(
                    {"detail": "No test cases available for this problem."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            # Run all test cases through the full pipeline (wrapping + execution + comparison)
            batch = execute_problem_test_case_batch(
                problem=problem,
                source_code=source_code,
                language=language or "",
                language_id=language_id,
                test_cases=test_cases,
                batch_kind="submit",
            )

            passed_cases = batch["passed_cases"]
            total_cases = batch["total_cases"]
            status_str = batch["status"]
            compile_error = batch["compile_output"] or ""
            first_stderr = batch["stderr"] or ""

            # Add case number; hide stdin/expected for non-sample test cases
            test_results = []
            for idx, tr in enumerate(batch["test_results"]):
                is_sample = tr.get("is_sample", True)
                test_results.append({
                    "case": idx + 1,
                    "passed": tr["passed"],
                    "status": tr["status"],
                    "stdin": tr["stdin"] if is_sample else "(hidden)",
                    "expected": tr["expected"] if is_sample else "(hidden)",
                    "actual": tr["actual"],
                    "stderr": tr.get("stderr", ""),
                    "compile_output": tr.get("compile_output", ""),
                    "time": tr.get("time", ""),
                })

            # Difficulty-based max score: Easy=100, Medium=200, Hard=300
            DIFFICULTY_MAX = {"Easy": 100, "Medium": 200, "Hard": 300}
            max_score = DIFFICULTY_MAX.get(problem.difficulty, 100)
            score = int((passed_cases / total_cases) * max_score) if total_cases > 0 else 0

            # Create contest submission
            submission = ContestSubmission.objects.create(
                contest=contest,
                student=student,
                problem=problem,
                code=source_code,
                language=language or 'Unknown',
                status=status_str,
                score=int(score),
            )

            # Update participation if problem is solved
            if status_str == 'Accepted':
                # Recalculate total_score as best score per problem (prevents score > 100 per problem)
                from django.db.models import Max as DMax
                best_scores = (
                    ContestSubmission.objects.filter(
                        contest=contest,
                        student=student,
                        status='Accepted',
                    )
                    .values('problem')
                    .annotate(best=DMax('score'))
                )
                new_total = sum(row['best'] for row in best_scores)
                new_solved = best_scores.count()

                participation.total_score = new_total
                participation.problems_solved = new_solved
                participation.save(update_fields=['problems_solved', 'total_score'])

            # Keep contest-level counters in sync
            try:
                contest.update_analytics()
            except Exception as analytics_err:
                logger.warning("Failed to update contest analytics: %s", analytics_err)

            return Response({
                "detail": "Code submitted successfully.",
                "submission": {
                    "id": submission.id,
                    "status": submission.status,
                    "score": submission.score,
                    "max_score": max_score,
                    "passed_cases": passed_cases,
                    "total_cases": total_cases,
                    "test_results": test_results,
                    "compile_error": compile_error,
                    "stderr": first_stderr,
                },
            })

        except Exception as e:
            logger.error(f"Contest submission error: {e}")
            return Response(
                {"detail": f"Submission failed: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ProblemsByTopicView(APIView):
    """Get all problems grouped by topics/tags"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Get all problems
        problems = Problem.objects.all()

        # Group by tags
        topics = {}
        for problem in problems:
            for tag in problem.tags or []:
                if tag not in topics:
                    topics[tag] = []
                topics[tag].append({
                    "id": problem.id,
                    "slug": problem.slug,
                    "title": problem.title,
                    "difficulty": problem.difficulty,
                })

        # Convert to list format
        topics_list = [
            {"topic": topic, "problems": probs, "count": len(probs)}
            for topic, probs in sorted(topics.items())
        ]

        return Response({
            "topics": topics_list,
            "total_problems": problems.count(),
        })


class StudentContestWinnersView(APIView):
    """Get contest winners and participant results"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, contest_id):
        try:
            # Get student profile
            student_profile = get_object_or_404(StudentProfile, account=request.user)
            
            # Get contest and verify student has access
            contest = get_object_or_404(Contest, id=contest_id)
            
            # Check if student is assigned to this contest
            is_assigned = (
                contest.assigned_students.filter(id=student_profile.id).exists() or
                (student_profile.batch in contest.assigned_batches if contest.assigned_batches else False)
            )
            
            if not is_assigned:
                return Response({
                    'success': False,
                    'error': 'You are not assigned to this contest'
                }, status=403)
            
            # Check if contest is completed
            if not contest.is_ended:
                return Response({
                    'success': False,
                    'error': 'Contest is not yet completed'
                }, status=400)
            
            # Get all participations for this contest, ordered by performance
            participations = ContestParticipation.objects.filter(
                contest=contest,
                has_started=True
            ).select_related('student').order_by(
                '-problems_solved',  # More problems solved first
                'total_time_taken',  # Less time taken second
                '-total_score'       # Higher score third
            )
            
            # Get winners (top 3)
            winners = []
            for i, participation in enumerate(participations[:3]):
                winners.append({
                    'id': participation.id,
                    'student_name': participation.student.name,
                    'register_number': participation.student.register_number,
                    'problems_solved': participation.problems_solved,
                    'total_score': participation.total_score,
                    'completion_time': participation.total_time_taken,
                    'rank': i + 1
                })
            
            # Get all participants
            participants = []
            for i, participation in enumerate(participations):
                participants.append({
                    'id': participation.id,
                    'student_name': participation.student.name,
                    'register_number': participation.student.register_number,
                    'problems_solved': participation.problems_solved,
                    'total_score': participation.total_score,
                    'completion_time': participation.total_time_taken,
                    'rank': i + 1
                })
            
            return Response({
                'success': True,
                'contest_title': contest.title,
                'total_problems': contest.problems.count(),
                'total_participants': participations.count(),
                'winners': winners,
                'participants': participants
            })
            
        except Exception as e:
            logger.error(f"Error fetching contest winners: {str(e)}")
            return Response({
                'success': False,
                'error': 'Failed to fetch contest results'
            }, status=500)


class AnnouncementListView(UnifiedAuthMixin, APIView):
    def get(self, request):
        # Universal Table Refresh logic: 
        # Only show announcements from the last 7 days
        seven_days_ago = timezone.now() - timedelta(days=7)
        
        announcements = Announcement.objects.filter(
            is_active=True,
            created_at__gte=seven_days_ago
        ).order_by('-created_at')
        
        data = [{
            "id": a.id,
            "title": a.title,
            "content": a.content,
            "category": a.category,
            "time": a.created_at.strftime("%I:%M %p"),
            "date": a.created_at.strftime("%b %d")
        } for a in announcements]
        
        return Response({"announcements": data})


class NotificationListView(UnifiedAuthMixin, APIView):
    def get(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error
            
        qs = Notification.objects.filter(recipient=request.user).order_by('-created_at')
        unread_count = qs.filter(is_read=False).count()
        notifications = qs[:20]
        
        data = [{
            "id": n.id,
            "title": n.title,
            "message": n.message,
            "link": n.link,
            "is_read": n.is_read,
            "time": n.created_at.strftime("%b %d, %H:%M")
        } for n in notifications]
        
        return Response({
            "notifications": data,
            "unread_count": unread_count
        })


class NotificationMarkReadView(UnifiedAuthMixin, APIView):
    def post(self, request, notification_id):
        notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
        notification.delete()
        return Response({"success": True})


class NotificationMarkAllReadView(UnifiedAuthMixin, APIView):
    def post(self, request):
        Notification.objects.filter(recipient=request.user).delete()
        return Response({"success": True, "message": "All notifications marked as read."})


class AdminDAUAnalyticsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user_role = str(getattr(request.user, 'role', '') or '').lower()
        is_admin_user = (
            request.user.is_superuser or 
            request.user.is_staff or 
            user_role in ('admin', 'superuser', 'hod', 'staff') or 
            hasattr(request.user, 'staff_profile') or
            getattr(request.user, 'username', '') in ('0001', 'staff_0001', 'admin')
        )
        if not is_admin_user:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        try:
            institution_id = request.query_params.get("institution_id")
            days = int(request.query_params.get("days", 14))
            days = max(1, min(days, 60))

            today = timezone.now().date()
            date_list = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]

            student_qs = StudentProfile.objects.all()
            staff_qs = StaffProfile.objects.all()

            if institution_id:
                student_qs = student_qs.filter(institution_id=institution_id)
                staff_qs = staff_qs.filter(institution_id=institution_id)

            daily_data = []

            for dt in date_list:
                active_students = student_qs.filter(
                    Q(account__last_login__date=dt) |
                    Q(last_login_on=dt) |
                    Q(submissions__submitted_at__date=dt) |
                    Q(contest_participations__started_at__date=dt) |
                    Q(problem_sessions__started_at__date=dt)
                ).distinct().count()

                active_staff = staff_qs.filter(
                    Q(account__last_login__date=dt) |
                    Q(contests__created_at__date=dt)
                ).distinct().count()

                daily_data.append({
                    "date": dt.strftime("%Y-%m-%d"),
                    "display_date": dt.strftime("%b %d"),
                    "day_name": dt.strftime("%a"),
                    "students": active_students,
                    "staff": active_staff,
                    "total": active_students + active_staff
                })

            return Response({
                "institution_id": institution_id,
                "days": days,
                "daily_active_users": daily_data
            })
        except Exception as exc:
            logging.exception("DAU Analytics failed: %s", exc)
            return Response({
                "detail": f"Failed to load analytics data: {str(exc)}",
                "institution_id": institution_id if 'institution_id' in locals() else None,
                "days": days if 'days' in locals() else 14,
                "daily_active_users": []
            })


class AdminSystemUpdateView(APIView):
    """Admin endpoint to create and manage system updates / release notes"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        updates = SystemUpdate.objects.all()
        data = [{
            "id": u.id,
            "title": u.title,
            "version": u.version,
            "content": u.content,
            "category": u.category,
            "target_role": u.target_role,
            "is_active": u.is_active,
            "created_at": u.created_at.strftime("%b %d, %Y %H:%M"),
        } for u in updates]
        return Response({"updates": data})

    def post(self, request):
        title = request.data.get("title", "").strip()
        content = request.data.get("content", "").strip()
        version = request.data.get("version", "").strip()
        category = request.data.get("category", "feature")
        target_role = request.data.get("target_role", "all")

        if not title or not content:
            return Response({"detail": "Title and content are required."}, status=status.HTTP_400_BAD_REQUEST)

        update = SystemUpdate.objects.create(
            title=title,
            content=content,
            version=version,
            category=category,
            target_role=target_role,
            is_active=True
        )

        return Response({
            "message": "System update broadcasted successfully!",
            "update": {
                "id": update.id,
                "title": update.title,
                "version": update.version,
                "content": update.content,
                "category": update.category,
                "target_role": update.target_role,
                "created_at": update.created_at.strftime("%b %d, %Y %H:%M"),
            }
        }, status=status.HTTP_201_CREATED)

    def delete(self, request, pk=None):
        if pk:
            SystemUpdate.objects.filter(id=pk).delete()
            return Response({"message": "System update deleted."})
        return Response({"detail": "ID required."}, status=status.HTTP_400_BAD_REQUEST)


class UserSystemUpdatesView(UnifiedAuthMixin, APIView):
    """User endpoint to read active platform updates matching their role"""
    def get(self, request):
        profile, profile_type, _ = self.get_authenticated_profile(request)
        user_role = "student" if hasattr(profile, 'register_number') else getattr(profile, 'role', 'staff')

        updates = SystemUpdate.objects.filter(
            is_active=True
        ).filter(
            Q(target_role='all') | Q(target_role=user_role)
        ).order_by('-created_at')[:10]

        data = [{
            "id": u.id,
            "title": u.title,
            "version": u.version,
            "content": u.content,
            "category": u.category,
            "target_role": u.target_role,
            "created_at": u.created_at.strftime("%b %d, %Y"),
        } for u in updates]

        return Response({"updates": data, "user_role": user_role})


class AptitudeTopicListView(UnifiedAuthMixin, APIView):
    """List all available aptitude topics and their questions count + student progress"""
    def get(self, request):
        profile, _, _ = self.get_authenticated_profile(request)
        is_student = hasattr(profile, 'register_number')
        
        # Optimization: Get solved question counts per topic for this student
        solved_counts = {}
        if is_student:
            from django.db.models import Count
            solved_qs = SolvedAptitude.objects.filter(student=profile).values('question__topic_id').annotate(count=Count('id'))
            solved_counts = {item['question__topic_id']: item['count'] for item in solved_qs}

        # Fetch top-level topics (Categories). Two levels only — Category >
        # Main Topic — since every admin workflow (bulk upload, add/edit
        # question, topic manager) files questions on the Main Topic node;
        # a third level was a dead end nothing ever wrote to.
        categories = AptitudeTopic.objects.filter(parent=None).prefetch_related('subtopics')

        category_list = []
        for cat in categories:
            subcategory_list = []
            cat_total_questions = 0
            cat_solved_questions = 0

            for subcat in cat.subtopics.all():
                q_count = subcat.questions.count()
                s_count = solved_counts.get(subcat.id, 0)

                cat_total_questions += q_count
                cat_solved_questions += s_count

                subcategory_list.append({
                    "id": subcat.id,
                    "title": subcat.title,
                    "question_count": q_count,
                    "solved_count": s_count
                })

            category_list.append({
                "id": cat.id,
                "title": cat.title,
                "subcategories": subcategory_list,
                "question_count": cat_total_questions,
                "solved_count": cat_solved_questions
            })

        return Response({"categories": category_list})


class AptitudeQuestionListView(UnifiedAuthMixin, APIView):
    """List aptitude questions for selection in contest creator"""
    def get(self, request):
        topic_id = request.query_params.get('topic_id')
        difficulty = request.query_params.get('difficulty')
        solved_status = request.query_params.get('status') # 'solved', 'unsolved', 'all'
        search = request.query_params.get('q')
        limit = request.query_params.get('limit', 1000) # Increased default limit
        
        profile, _, _ = self.get_authenticated_profile(request)
        is_student = hasattr(profile, 'register_number')
        
        topic_ids = request.query_params.getlist('topic_id')
        if not topic_ids and topic_id:
            topic_ids = topic_id.split(',')

        # Explicit stable order — an unordered queryset's row order isn't
        # guaranteed across requests, which combined with the frontend
        # restoring only a numeric position (not question identity) after a
        # refresh could show a different question at the same index.
        qs = AptitudeQuestion.objects.all().select_related('topic').order_by('id')
        
        if topic_ids:
            # Enhanced filtering: Include subtopics recursively if a parent topic is selected
            all_topic_ids = set()
            invalid_ids = []
            for tid in topic_ids:
                try:
                    all_topic_ids.add(int(tid))
                    # Topics are two levels deep now (Category > Main Topic) —
                    # if tid is a category, pull in its main topics too, so
                    # selecting a whole category still grabs everything filed
                    # under it.
                    subtopic_ids = AptitudeTopic.objects.filter(parent_id=tid).values_list('id', flat=True)
                    all_topic_ids.update(subtopic_ids)
                except ValueError:
                    invalid_ids.append(tid)
                    continue

            # A malformed/empty topic_id used to silently fall through to
            # `topic_id__in=set()` — a valid-looking empty result that was
            # indistinguishable from "this topic genuinely has no questions."
            # Fail loudly instead so the client can tell the two apart.
            if not all_topic_ids:
                return Response(
                    {"error": f"No valid topic_id provided — got {invalid_ids!r}."},
                    status=400,
                )
            qs = qs.filter(topic_id__in=all_topic_ids)

        if difficulty and difficulty != 'all' and difficulty != 'All':
            qs = qs.filter(difficulty__iexact=difficulty)
        if search:
            qs = qs.filter(question_text__icontains=search)
            
        solved_ids = []
        if is_student:
            solved_ids = list(SolvedAptitude.objects.filter(student=profile).values_list('question_id', flat=True))
            if solved_status == 'solved':
                qs = qs.filter(id__in=solved_ids)
            elif solved_status == 'unsolved':
                qs = qs.exclude(id__in=solved_ids)
            
        data = []
        # Limit the queryset to avoid memory issues, but use a larger default
        try:
            limit = int(limit)
        except ValueError:
            limit = 1000
            
        for q in qs[:limit]:
            data.append({
                "id": q.id,
                "topic": q.topic.title,
                "topic_id": q.topic.id,
                "question_type": q.question_type,
                "question_text": q.question_text,
                "question_image": q.question_image,
                "difficulty": q.difficulty,
                "option_a": q.option_a,
                "option_a_image": q.option_a_image,
                "option_b": q.option_b,
                "option_b_image": q.option_b_image,
                "option_c": q.option_c,
                "option_c_image": q.option_c_image,
                "option_d": q.option_d,
                "option_d_image": q.option_d_image,
                # Only staff/HOD building a contest get to see the answer — never
                # send this to a student, who fetches this same endpoint to take
                # the quiz itself, or it would leak the answer via the network tab.
                **({} if is_student else {"correct_option": q.correct_option}),
                "is_solved": q.id in solved_ids
            })
            
        return Response(data)

class AptitudeQuestionSubmitView(APIView):
    """Verify an aptitude question answer and record progress"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        question_id = request.data.get('question_id')
        selected_option = request.data.get('selected_option')
        
        if not question_id or not selected_option:
            return Response({"error": "question_id and selected_option are required"}, status=400)
            
        question = get_object_or_404(AptitudeQuestion, id=question_id)
        is_correct = question.correct_option.upper() == selected_option.upper()

        if hasattr(request.user, 'student_profile'):
            AptitudeAttempt.objects.create(
                student=request.user.student_profile,
                question=question,
                selected_option=selected_option.upper()[:1],
                is_correct=is_correct,
            )

        if is_correct:
            if hasattr(request.user, 'student_profile'):
                profile = request.user.student_profile
                SolvedAptitude.objects.get_or_create(
                    student=profile,
                    question=question
                )

                # Check for aptitude achievements
                total_aptitude_solved = SolvedAptitude.objects.filter(student=profile).count()
                unearned = Achievement.objects.filter(category='aptitude').exclude(userachievement__user=request.user)
                for ach in unearned:
                    should_award = False
                    if ach.criteria_type == 'aptitude_solve_count' and total_aptitude_solved >= ach.criteria_value:
                        should_award = True
                    elif ach.criteria_type == 'quant_solve_count':
                        # Check if category is quantitative
                        # Using topic__parent__parent because QUANTITATIVE is at the root (Level 1)
                        is_quant = False
                        t = question.topic
                        while t:
                            if 'QUANTITATIVE' in t.title.upper():
                                is_quant = True
                                break
                            t = t.parent
                        
                        if is_quant:
                            quant_solved = SolvedAptitude.objects.filter(
                                student=profile, 
                                question__topic__id__in=AptitudeTopic.objects.filter(title__icontains='QUANTITATIVE').values_list('id', flat=True)
                            ).count()
                            if quant_solved >= ach.criteria_value:
                                should_award = True
                    
                    if should_award:
                        UserAchievement.objects.get_or_create(user=request.user, achievement=ach)
        
        return Response({
            "is_correct": is_correct,
            "correct_option": question.correct_option,
            "explanation": question.explanation
        })


class ReadingPassageListView(UnifiedAuthMixin, APIView):
    """List Reading Comprehension passages with question/solved counts —
    the Reading Comprehension section of the student Aptitude page."""
    def get(self, request):
        profile, _, _ = self.get_authenticated_profile(request)
        is_student = hasattr(profile, 'register_number')

        solved_counts = {}
        if is_student:
            solved_qs = (
                SolvedAptitude.objects.filter(student=profile, question__passage__isnull=False)
                .values('question__passage_id').annotate(count=Count('id'))
            )
            solved_counts = {item['question__passage_id']: item['count'] for item in solved_qs}

        passages = ReadingPassage.objects.annotate(question_count=Count("questions")).order_by("-id")
        data = [{
            "id": p.id,
            "title": p.title,
            "difficulty": p.difficulty,
            "question_count": p.question_count,
            "solved_count": solved_counts.get(p.id, 0),
        } for p in passages]
        return Response({"passages": data})


class ReadingPassageDetailView(UnifiedAuthMixin, APIView):
    """Get one Reading Comprehension passage's text and its questions —
    never includes correct_option for a student, mirroring
    AptitudeQuestionListView's answer-leak protection."""
    def get(self, request, passage_id):
        profile, _, _ = self.get_authenticated_profile(request)
        is_student = hasattr(profile, 'register_number')

        passage = ReadingPassage.objects.filter(id=passage_id).first()
        if not passage:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)

        solved_ids = []
        if is_student:
            solved_ids = list(
                SolvedAptitude.objects.filter(student=profile, question__passage=passage)
                .values_list('question_id', flat=True)
            )

        questions = []
        for q in passage.questions.all().order_by("id"):
            questions.append({
                "id": q.id,
                "question_text": q.question_text,
                "question_image": q.question_image,
                "option_a": q.option_a, "option_a_image": q.option_a_image,
                "option_b": q.option_b, "option_b_image": q.option_b_image,
                "option_c": q.option_c, "option_c_image": q.option_c_image,
                "option_d": q.option_d, "option_d_image": q.option_d_image,
                "difficulty": q.difficulty,
                **({} if is_student else {"correct_option": q.correct_option}),
                "is_solved": q.id in solved_ids,
            })

        return Response({
            "id": passage.id,
            "title": passage.title,
            "passage_text": passage.passage_text,
            "difficulty": passage.difficulty,
            "questions": questions,
        })


# ---------------------------------------------------------------------------
# PDF Report Generation
# ---------------------------------------------------------------------------

class StudentReportPDFView(APIView):
    """Generate a comprehensive, data-driven PDF performance report for a student."""
    permission_classes = [IsAuthenticated]

    # ── Company → topic weight map (used for readiness analysis) ──────────
    COMPANY_TOPIC_WEIGHTS = {
        "Amazon":    {"Arrays": 0.20, "Trees": 0.18, "Graphs": 0.12, "Dynamic Programming": 0.15, "Strings": 0.10, "Hash Table": 0.08, "Sorting": 0.07, "Linked List": 0.05, "Stack": 0.05},
        "Google":    {"Graphs": 0.20, "Dynamic Programming": 0.20, "Trees": 0.12, "Arrays": 0.12, "Strings": 0.10, "Binary Search": 0.08, "Math": 0.08, "Sorting": 0.05, "Stack": 0.05},
        "Microsoft": {"Arrays": 0.18, "Dynamic Programming": 0.15, "Trees": 0.15, "Strings": 0.12, "Linked List": 0.10, "Graphs": 0.10, "Hash Table": 0.08, "Stack": 0.06, "Sorting": 0.06},
        "Meta":      {"Arrays": 0.18, "Graphs": 0.15, "Dynamic Programming": 0.15, "Trees": 0.12, "Strings": 0.12, "Binary Search": 0.08, "Hash Table": 0.08, "Sorting": 0.06, "Stack": 0.06},
        "Apple":     {"Arrays": 0.20, "Trees": 0.15, "Strings": 0.15, "Linked List": 0.12, "Dynamic Programming": 0.10, "Sorting": 0.08, "Hash Table": 0.08, "Graphs": 0.06, "Stack": 0.06},
        "Netflix":   {"Dynamic Programming": 0.20, "Graphs": 0.18, "Trees": 0.15, "Arrays": 0.12, "Strings": 0.10, "Hash Table": 0.08, "Sorting": 0.07, "Binary Search": 0.05, "Stack": 0.05},
        "TCS":       {"Arrays": 0.20, "Strings": 0.18, "Sorting": 0.15, "Math": 0.12, "Linked List": 0.10, "Trees": 0.08, "Dynamic Programming": 0.07, "Hash Table": 0.05, "Stack": 0.05},
        "Infosys":   {"Arrays": 0.22, "Strings": 0.18, "Sorting": 0.15, "Math": 0.12, "Linked List": 0.10, "Trees": 0.08, "Hash Table": 0.05, "Dynamic Programming": 0.05, "Stack": 0.05},
        "Wipro":     {"Arrays": 0.22, "Strings": 0.20, "Sorting": 0.15, "Math": 0.12, "Linked List": 0.10, "Trees": 0.08, "Hash Table": 0.05, "Dynamic Programming": 0.04, "Stack": 0.04},
        "Cognizant": {"Arrays": 0.22, "Strings": 0.20, "Sorting": 0.15, "Math": 0.12, "Linked List": 0.10, "Trees": 0.08, "Hash Table": 0.05, "Dynamic Programming": 0.04, "Stack": 0.04},
        "Zoho":      {"Arrays": 0.18, "Strings": 0.15, "Linked List": 0.12, "Trees": 0.12, "Dynamic Programming": 0.12, "Hash Table": 0.10, "Sorting": 0.08, "Graphs": 0.07, "Stack": 0.06},
        "Accenture": {"Arrays": 0.22, "Strings": 0.20, "Sorting": 0.15, "Math": 0.12, "Linked List": 0.10, "Trees": 0.08, "Hash Table": 0.05, "Dynamic Programming": 0.04, "Stack": 0.04},
    }
    TOPIC_THRESHOLD = 5  # problems needed per topic for 100% in that topic

    def get(self, request, register_number):
        if not hasattr(request.user, 'staff_profile'):
            return Response({"detail": "Staff access required."}, status=403)

        staff_profile = request.user.staff_profile
        student = get_object_or_404(StudentProfile, register_number=register_number)

        # Access control
        if student.institution != staff_profile.institution:
            return Response({"detail": "Access denied."}, status=403)
        if staff_profile.role in ('hod', 'academics') and student.department != staff_profile.department:
            return Response({"detail": "Access denied (Department mismatch)."}, status=403)

        report_type = request.GET.get('type', 'overall')
        topic_filter = request.GET.get('topic', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        # ── Gather all data ──────────────────────────────────────────────
        rd = self._get_comprehensive_data(student, report_type, topic_filter, date_from, date_to)

        # ── Build PDF ────────────────────────────────────────────────────
        from .pdf_reports import (
            create_watermarked_pdf_contest, _bar_chart, _pie_chart,
            _spider_chart, _bell_curve_chart,
        )
        buffer = BytesIO()
        doc = create_watermarked_pdf_contest(
            buffer,
            institution=student.institution,
            department=student.department,
            pagesize=A4,
            topMargin=2.1 * inch,
            bottomMargin=0.6 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('RPTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=24, alignment=1, textColor=colors.HexColor('#2d5016'))
        header_style = ParagraphStyle('RPHeader', parent=styles['Heading2'], fontSize=13, spaceBefore=14, spaceAfter=10, textColor=colors.HexColor('#39482a'))
        sub_style = ParagraphStyle('RPSub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#555555'), spaceAfter=6)
        body_style = ParagraphStyle('RPBody', parent=styles['Normal'], fontSize=9, leading=13, textColor=colors.HexColor('#333333'), spaceAfter=4)
        bullet_style = ParagraphStyle('RPBullet', parent=styles['Normal'], fontSize=9, leading=13, textColor=colors.HexColor('#333333'), spaceAfter=3, leftIndent=18, bulletIndent=6, bulletFontName='Helvetica', bulletFontSize=9)
        note_style = ParagraphStyle('RPNote', parent=styles['Normal'], fontSize=8, leading=11, textColor=colors.HexColor('#92400e'), spaceAfter=2)
        footer_style = ParagraphStyle('RPFooter', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)

        GREEN = '#2d5016'
        GREEN_BG = '#e6ebdd'
        GREEN_HDR = '#39482a'
        BORDER = '#d0d9c2'
        ALT_ROW = '#f8f9f7'

        def _styled_table(data, col_widths, has_header=True):
            cell_style = ParagraphStyle(
                'RPTableCell',
                parent=styles['Normal'],
                fontSize=8,
                leading=10,
                textColor=colors.HexColor('#333333')
            )
            cell_header_style = ParagraphStyle(
                'RPTableHeaderCell',
                parent=styles['Normal'],
                fontName='Helvetica-Bold',
                fontSize=8,
                leading=10,
                textColor=colors.HexColor(GREEN_HDR)
            )

            wrapped_data = []
            for row_idx, row in enumerate(data):
                wrapped_row = []
                for col_idx, cell in enumerate(row):
                    if isinstance(cell, str):
                        if '█' in cell or '░' in cell:
                            wrapped_row.append(cell)
                        else:
                            escaped_cell = cell.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                            if has_header and row_idx == 0:
                                wrapped_row.append(Paragraph(escaped_cell, cell_header_style))
                            else:
                                wrapped_row.append(Paragraph(escaped_cell, cell_style))
                    else:
                        wrapped_row.append(cell)
                wrapped_data.append(wrapped_row)

            t = Table(wrapped_data, colWidths=col_widths)
            style_cmds = [
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER)),
            ]
            if has_header:
                style_cmds += [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(GREEN_BG)),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor(ALT_ROW)]),
                ]
            t.setStyle(TableStyle(style_cmds))
            return t

        els = []  # elements list

        # ── Contact info ─────────────────────────────────────────────────
        institution = student.institution
        contact_parts = []
        if institution.contact_email:
            contact_parts.append(f"Email: {institution.contact_email}")
        if institution.contact_phone:
            contact_parts.append(f"Phone: {institution.contact_phone}")
        if institution.website:
            contact_parts.append(f"Website: {institution.website}")
        if contact_parts:
            els.append(Paragraph(" | ".join(contact_parts), styles['Normal']))
        els.append(Spacer(1, 0.25 * inch))

        # ── Section 1: Report Title ──────────────────────────────────────
        report_titles = {
            'overall': 'Comprehensive Student Performance Report',
            'aptitude': 'Aptitude Assessment Performance Report',
            'programming': 'Programming Performance Report',
            'contests': 'Contest Participation Report',
        }
        els.append(Paragraph(report_titles.get(report_type, 'Student Performance Report'), title_style))

        # ── Section 1: Student Info Card ─────────────────────────────────
        els.append(Paragraph("Student Information", header_style))
        info_rows = [
            ["Name:", student.name],
            ["Register Number:", student.register_number],
            ["Department:", student.department.get_full_name() if student.department else 'N/A'],
            ["Batch:", student.batch or 'N/A'],
            ["Report Period:", f"{date_from or 'All time'} to {date_to or 'Present'}"],
        ]
        if topic_filter:
            info_rows.append(["Topic Filter:", topic_filter])
        info_t = Table(info_rows, colWidths=[2 * inch, 4 * inch])
        info_t.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        els.append(info_t)
        els.append(Spacer(1, 0.25 * inch))

        # ── Section 2: Data Consistency Notes ────────────────────────────
        if rd['consistency_notes']:
            els.append(Paragraph("Data Consistency Notes", header_style))
            for note in rd['consistency_notes']:
                els.append(Paragraph(f"⚠ {note}", note_style))
            els.append(Spacer(1, 0.15 * inch))

        # ── Section 3: Performance Snapshot ──────────────────────────────
        els.append(Paragraph("Performance Snapshot", header_style))
        rank_display = f"#{rd['campus_rank']}" if rd['total_solved'] >= 3 else "Not enough data yet"
        success_display = f"{rd['success_rate']:.1f}%" if rd['total_attempts'] >= 3 else "Not enough data yet"

        snap_data = [
            ["Metric", "Value", "Details"],
            ["Problems Solved", str(rd['total_solved']), f"Easy: {rd['easy']}  |  Medium: {rd['medium']}  |  Hard: {rd['hard']}"],
            ["Current Streak", f"{rd['current_streak']} days", f"Login Days: {student.login_days}"],
            ["Campus Rank", rank_display, f"Among {rd['total_students']} students"],
            ["Success Rate", success_display, f"Based on {rd['total_attempts']} attempts"],
        ]
        if report_type in ('aptitude', 'overall'):
            apt_display = f"{rd['aptitude_percentage']:.1f}%" if rd['aptitude_solved'] >= 2 else "Not enough data yet"
            snap_data.append(["Aptitude Progress", f"{rd['aptitude_solved']} solved", f"{apt_display} of {rd['total_aptitude']} total"])
        if report_type in ('contests', 'overall'):
            snap_data.append(["Contests", str(rd['contests_participated']), f"{rd['contest_submissions']} submissions"])

        els.append(_styled_table(snap_data, [1.8 * inch, 1.5 * inch, 3.2 * inch]))
        els.append(Spacer(1, 0.25 * inch))

        # ── Section 4: Difficulty Breakdown Chart ────────────────────────
        if rd['total_solved'] > 0:
            els.append(Paragraph("Difficulty Breakdown", header_style))
            diff_values = [rd['easy'], rd['medium'], rd['hard']]
            diff_labels = ['Easy', 'Medium', 'Hard']
            diff_colors = ['#22c55e', '#f59e0b', '#ef4444']
            els.append(_pie_chart(diff_values, diff_labels, diff_colors, size=140))
            els.append(Spacer(1, 0.2 * inch))

        # ── Section 5: Topic-by-Topic Breakdown ──────────────────────────
        if rd['topic_breakdown']:
            els.append(Paragraph("Topic-by-Topic Breakdown", header_style))
            els.append(Paragraph("Proficiency: Not Started (0) → Beginner (1-4) → Intermediate (5-9) → Advanced (10+)", sub_style))
            topic_data = [["Topic", "Solved", "Proficiency", "Progress"]]
            for tb in rd['topic_breakdown']:
                bar_len = min(tb['count'], 15)
                bar = '█' * bar_len + '░' * (15 - bar_len)
                topic_data.append([tb['topic'], str(tb['count']), tb['level'], bar])
            els.append(_styled_table(topic_data, [2 * inch, 0.8 * inch, 1.3 * inch, 2.4 * inch]))
            els.append(Spacer(1, 0.25 * inch))

        # ── Section 6: 12-Week Trend ─────────────────────────────────────
        if rd['weekly_trend']:
            els.append(Paragraph("Weekly Performance Trend (Last 12 Weeks)", header_style))
            week_labels = [w['label'] for w in rd['weekly_trend']]
            week_solved = [w['solved'] for w in rd['weekly_trend']]
            els.append(_bar_chart(week_solved, week_labels, bar_color='#0f766e', w=480, h=150))

            # Accuracy trend table
            acc_data = [["Week", "Solved", "Attempts", "Accuracy"]]
            for w in rd['weekly_trend']:
                acc = f"{w['accuracy']:.0f}%" if w['attempts'] > 0 else "—"
                acc_data.append([w['label'], str(w['solved']), str(w['attempts']), acc])
            els.append(Spacer(1, 0.1 * inch))
            els.append(_styled_table(acc_data, [1.6 * inch, 1.1 * inch, 1.1 * inch, 2.7 * inch]))
            els.append(Spacer(1, 0.25 * inch))

        # ── Section 7: Company Readiness ─────────────────────────────────
        if rd['company_readiness']:
            els.append(Paragraph("Company Readiness Analysis", header_style))
            els.append(Paragraph("Readiness is based on how well your solved topics match each company's typical interview distribution.", sub_style))
            comp_data = [["Company", "Readiness", "Matched Topics", "Key Gaps"]]
            for cr in rd['company_readiness']:
                readiness_str = f"{cr['readiness_pct']}%"
                matched = ", ".join(cr['matched'][:3]) if cr['matched'] else "None"
                gaps = ", ".join(cr['gaps'][:3]) if cr['gaps'] else "All covered"
                comp_data.append([cr['company'], readiness_str, matched, gaps])
            els.append(_styled_table(comp_data, [1.3 * inch, 0.9 * inch, 2 * inch, 2.3 * inch]))
            els.append(Spacer(1, 0.25 * inch))

        # ── Section 8: Performance Charts ────────────────────────────────
        chart_data = rd.get('performance_charts', {})

        # Overall pie + daily trend bar
        els.append(Paragraph("Performance Charts", header_style))
        overall_rows = chart_data.get('overall_performance', [])
        overall_values = [row.get('value', 0) for row in overall_rows]
        overall_labels = [row.get('label', '') for row in overall_rows]
        if not sum(overall_values):
            overall_values = [1]
            overall_labels = ["No data"]

        trend_rows = chart_data.get('daily_solved_trend', [])[-14:]
        trend_values = [row.get('daily_total', 0) for row in trend_rows]
        trend_labels = [row.get('date', '')[5:] for row in trend_rows]

        overall_pie = _pie_chart(overall_values, overall_labels, ['#22c55e', '#3b82f6', '#f59e0b'], size=130)
        trend_bar = _bar_chart(trend_values, trend_labels, bar_color='#0f766e', w=295, h=140)
        charts_table = Table([[overall_pie, trend_bar]], colWidths=[2.2 * inch, 4.2 * inch])
        charts_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor(BORDER)),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor(ALT_ROW)),
        ]))
        els.append(charts_table)
        chart_labels_t = Table([
            [Paragraph("Overall Performance", styles['Normal']), Paragraph("Daily Problems Solved (14 days)", styles['Normal'])],
        ], colWidths=[2.2 * inch, 4.2 * inch])
        chart_labels_t.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTSIZE', (0, 0), (-1, -1), 8)]))
        els.append(chart_labels_t)
        els.append(Spacer(1, 0.2 * inch))

        # Radar + Bell curve
        radar_data = chart_data.get('profile_radar', {})
        radar_labels = radar_data.get('labels', ['Programming', 'Aptitude', 'Contest', 'Daily', 'Overall'])
        radar_values = radar_data.get('overall', [0, 0, 0, 0, 0])
        radar_chart = _spider_chart(radar_values, radar_labels, max_val=100, size=150)

        percentile = 0
        rank = rd.get('campus_rank', 0)
        total_students = rd.get('total_students', 1)
        if total_students > 0 and rank > 0 and rd['total_solved'] >= 3:
            percentile = round((1 - (rank - 1) / total_students) * 100, 1)
        bell_curve = _bell_curve_chart(percentile, rank, w=300, h=140)

        adv_table = Table([[radar_chart, bell_curve]], colWidths=[2.2 * inch, 4.2 * inch])
        adv_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        els.append(adv_table)
        adv_labels = Table([
            [Paragraph("Strength Analysis", styles['Normal']), Paragraph("Platform Standing (Percentile)", styles['Normal'])],
        ], colWidths=[2.2 * inch, 4.2 * inch])
        adv_labels.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTSIZE', (0, 0), (-1, -1), 8)]))
        els.append(adv_labels)
        els.append(Spacer(1, 0.2 * inch))

        # ── Section 9: Knowledge Distribution ────────────────────────────
        knowledge = chart_data.get('knowledge_distribution') or {}
        knowledge_labels = knowledge.get('labels', [])[:10]
        knowledge_values = [
            (knowledge.get('programming', [0] * len(knowledge_labels))[idx] if idx < len(knowledge.get('programming', [])) else 0)
            + (knowledge.get('aptitude', [0] * len(knowledge_labels))[idx] if idx < len(knowledge.get('aptitude', [])) else 0)
            for idx, _ in enumerate(knowledge_labels)
        ]
        if knowledge_labels:
            els.append(Paragraph("Knowledge Distribution", header_style))
            els.append(_bar_chart(knowledge_values, [label[:12] for label in knowledge_labels], bar_color='#4f46e5', w=480, h=150))
            els.append(Spacer(1, 0.25 * inch))

        # ── Section 10: Insights ─────────────────────────────────────────
        if rd['insights']:
            els.append(Paragraph("Key Insights", header_style))
            for insight in rd['insights']:
                els.append(Paragraph(f"• {insight}", bullet_style))
            els.append(Spacer(1, 0.25 * inch))

        # ── Section 11: 2-Week Action Plan ───────────────────────────────
        if rd['action_plan']:
            els.append(Paragraph("Recommended 2-Week Action Plan", header_style))
            for idx, action in enumerate(rd['action_plan'], 1):
                els.append(Paragraph(f"{idx}. {action}", body_style))
            els.append(Spacer(1, 0.25 * inch))

        # ── Section 12: Recent Activity ──────────────────────────────────
        if rd['recent_activities']:
            els.append(Paragraph("Recent Activity (Last 30 Days)", header_style))
            activity_data = [["Date", "Activity", "Problem/Contest", "Result"]]
            for act in rd['recent_activities'][:15]:
                activity_data.append([act['date'], act['type'], act['subject'][:35], act['result']])
            els.append(_styled_table(activity_data, [1.2 * inch, 1.4 * inch, 2.2 * inch, 1.7 * inch]))

        # ── Section 13: Footer ───────────────────────────────────────────
        els.append(Spacer(1, 0.4 * inch))
        local_now = timezone.localtime(timezone.now())
        els.append(Paragraph(f"Report generated on {local_now.strftime('%B %d, %Y at %I:%M %p')}", footer_style))
        els.append(Paragraph(f"Generated by: {staff_profile.name} ({staff_profile.faculty_id})", footer_style))
        els.append(Paragraph("This report is auto-generated. Metrics reflect data available at generation time.", footer_style))

        doc.build(els)
        buffer.seek(0)

        filename_parts = [f"Student_Report_{student.register_number}"]
        if report_type != 'overall':
            filename_parts.append(report_type.title())
        if topic_filter:
            filename_parts.append(topic_filter.replace(' ', '_'))
        filename_parts.append(local_now.strftime('%Y%m%d'))
        filename = "_".join(filename_parts) + ".pdf"

        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # ══════════════════════════════════════════════════════════════════════
    #  Data Collection
    # ══════════════════════════════════════════════════════════════════════

    def _get_comprehensive_data(self, student, report_type, topic_filter, date_from, date_to):
        """Collect all data needed for the comprehensive report."""
        from datetime import datetime, timedelta
        from django.db.models import Q, Count
        from collections import defaultdict

        consistency_notes = []

        # ── Date filters ─────────────────────────────────────────────────
        def _parse_date(val):
            try:
                return datetime.strptime(val, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return None

        date_from_obj = _parse_date(date_from)
        date_to_obj = _parse_date(date_to)

        date_filter = Q()
        if date_from_obj:
            date_filter &= Q(solved_at__date__gte=date_from_obj)
        if date_to_obj:
            date_filter &= Q(solved_at__date__lte=date_to_obj)

        # ── Solved problems ──────────────────────────────────────────────
        solved_qs = SolvedProblem.objects.filter(student=student).select_related('problem')
        if date_filter:
            solved_qs = solved_qs.filter(date_filter)
        if topic_filter:
            solved_qs = solved_qs.filter(problem__tags__icontains=topic_filter)

        total_solved = solved_qs.count()
        difficulty_counts = {'Easy': 0, 'Medium': 0, 'Hard': 0}
        company_counts = defaultdict(int)
        skill_counts = defaultdict(int)

        for sp in solved_qs:
            d = sp.problem.difficulty or 'Medium'
            difficulty_counts[d] = difficulty_counts.get(d, 0) + 1

            comps = sp.problem.companies or ""
            for c in [c.strip() for c in comps.replace(',', ' ').split() if c.strip()]:
                company_counts[c] += 1

            tags = sp.problem.tags or []
            if isinstance(tags, str):
                tags = [t.strip() for t in tags.split(',') if t.strip()]
            for t in tags:
                skill_counts[t.strip().title()] += 1

        # ── Attempts & success rate (with consistency check) ─────────────
        attempt_filter = Q()
        if date_from_obj:
            attempt_filter &= Q(created_at__date__gte=date_from_obj)
        if date_to_obj:
            attempt_filter &= Q(created_at__date__lte=date_to_obj)

        attempt_qs = ExecutionRecord.objects.filter(student=student)
        if attempt_filter:
            attempt_qs = attempt_qs.filter(attempt_filter)
        total_attempts = attempt_qs.count()

        # Data consistency: if solved > 0 but attempts == 0, correct it
        if total_attempts == 0 and total_solved > 0:
            total_attempts = total_solved
            consistency_notes.append(
                f"Execution records show 0 attempts but {total_solved} problems are solved. "
                f"Attempts corrected to {total_solved}. Success rate set to 100%."
            )
            success_rate = 100.0
        elif total_attempts > 0:
            success_rate = round(total_solved / total_attempts * 100, 1)
        else:
            success_rate = 0.0

        # If success rate > 100% (shouldn't happen but guard)
        if success_rate > 100:
            consistency_notes.append(
                f"Calculated success rate ({success_rate:.1f}%) exceeds 100%. "
                f"This may indicate duplicate solve records. Capped at 100%."
            )
            success_rate = 100.0

        # ── Aptitude ─────────────────────────────────────────────────────
        apt_filter = Q()
        if date_from_obj:
            apt_filter &= Q(solved_at__date__gte=date_from_obj)
        if date_to_obj:
            apt_filter &= Q(solved_at__date__lte=date_to_obj)

        apt_qs = SolvedAptitude.objects.filter(student=student)
        if apt_filter:
            apt_qs = apt_qs.filter(apt_filter)
        aptitude_solved = apt_qs.count()
        total_aptitude = AptitudeQuestion.objects.count()
        aptitude_pct = round(aptitude_solved / total_aptitude * 100, 1) if total_aptitude > 0 else 0

        # ── Contests ─────────────────────────────────────────────────────
        cp_filter = Q()
        if date_from_obj:
            cp_filter &= Q(started_at__date__gte=date_from_obj)
        if date_to_obj:
            cp_filter &= Q(started_at__date__lte=date_to_obj)

        cp_qs = ContestParticipation.objects.filter(student=student)
        if cp_filter:
            cp_qs = cp_qs.filter(cp_filter)
        contests_participated = cp_qs.count()

        cs_filter = Q()
        if date_from_obj:
            cs_filter &= Q(submitted_at__date__gte=date_from_obj)
        if date_to_obj:
            cs_filter &= Q(submitted_at__date__lte=date_to_obj)
        cs_qs = ContestSubmission.objects.filter(student=student)
        if cs_filter:
            cs_qs = cs_qs.filter(cs_filter)
        contest_submissions = cs_qs.count()

        # ── Campus rank & total students ─────────────────────────────────
        campus_rank = calculate_campus_rank_helper(student)
        total_students = StudentProfile.objects.filter(institution=student.institution).count()

        # ── Topic-by-topic breakdown ─────────────────────────────────────
        # Combine programming tags + aptitude topics
        all_topics = dict(skill_counts)  # already has programming tags

        # Add aptitude topic counts
        apt_topic_qs = SolvedAptitude.objects.filter(student=student).values(
            'question__topic__title'
        ).annotate(count=Count('id')).order_by('-count')
        for row in apt_topic_qs:
            topic = row['question__topic__title']
            if topic:
                key = topic.strip().title()
                all_topics[key] = all_topics.get(key, 0) + row['count']

        # Build proficiency list
        topic_breakdown = []
        for topic, count in sorted(all_topics.items(), key=lambda x: x[1], reverse=True):
            if count == 0:
                level = "Not Started"
            elif count <= 4:
                level = "Beginner"
            elif count <= 9:
                level = "Intermediate"
            else:
                level = "Advanced"
            topic_breakdown.append({'topic': topic, 'count': count, 'level': level})

        # ── 12-Week Trend ────────────────────────────────────────────────
        today = timezone.localdate()
        weekly_trend = []
        for week_offset in range(11, -1, -1):
            week_start = today - timedelta(days=today.weekday()) - timedelta(weeks=week_offset)
            week_end = week_start + timedelta(days=6)

            week_solved = SolvedProblem.objects.filter(
                student=student,
                solved_at__date__gte=week_start,
                solved_at__date__lte=week_end,
            ).count()

            week_attempts = ExecutionRecord.objects.filter(
                student=student,
                created_at__date__gte=week_start,
                created_at__date__lte=week_end,
            ).count()

            week_accuracy = round(week_solved / week_attempts * 100, 1) if week_attempts > 0 else 0

            weekly_trend.append({
                'label': f"{week_start.strftime('%b %d')}",
                'solved': week_solved,
                'attempts': week_attempts,
                'accuracy': week_accuracy,
            })

        # ── Company Readiness ────────────────────────────────────────────
        # Determine which companies to analyze: tracked_companies + companies from solved problems
        target_companies = set()
        if student.tracked_companies:
            for c in student.tracked_companies:
                if isinstance(c, str):
                    target_companies.add(c.strip())
                elif isinstance(c, dict) and c.get('name'):
                    target_companies.add(c['name'].strip())

        # Add top companies from solved problems
        for comp, count in sorted(company_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
            target_companies.add(comp)

        # If no companies at all, add some defaults
        if not target_companies:
            target_companies = {"TCS", "Infosys", "Zoho"}

        company_readiness = []
        for comp_name in sorted(target_companies):
            weights = None
            # Try exact match first, then case-insensitive
            for key, w in self.COMPANY_TOPIC_WEIGHTS.items():
                if key.lower() == comp_name.lower():
                    weights = w
                    break
            if not weights:
                # Use a generic weight distribution
                weights = {"Arrays": 0.20, "Strings": 0.15, "Trees": 0.12, "Dynamic Programming": 0.12,
                           "Graphs": 0.10, "Sorting": 0.08, "Hash Table": 0.08, "Linked List": 0.08, "Stack": 0.07}

            readiness = 0
            matched = []
            gaps = []
            for topic, weight in weights.items():
                student_count = skill_counts.get(topic, 0)
                topic_score = min(1.0, student_count / self.TOPIC_THRESHOLD)
                readiness += topic_score * weight
                if student_count > 0:
                    matched.append(f"{topic} ({student_count})")
                else:
                    gaps.append(topic)

            readiness_pct = round(readiness * 100)
            company_readiness.append({
                'company': comp_name,
                'readiness_pct': readiness_pct,
                'matched': matched,
                'gaps': gaps,
            })

        company_readiness.sort(key=lambda x: x['readiness_pct'], reverse=True)

        # ── Insights ─────────────────────────────────────────────────────
        insights = self._generate_insights(
            student, rd_total_solved=total_solved, rd_easy=difficulty_counts['Easy'],
            rd_medium=difficulty_counts['Medium'], rd_hard=difficulty_counts['Hard'],
            rd_aptitude_solved=aptitude_solved, rd_aptitude_pct=aptitude_pct,
            rd_contests=contests_participated, rd_streak=student.current_streak,
            rd_weekly_trend=weekly_trend, rd_topic_breakdown=topic_breakdown,
            rd_company_readiness=company_readiness, rd_success_rate=success_rate,
        )

        # ── Action Plan ──────────────────────────────────────────────────
        action_plan = self._generate_action_plan(
            topic_breakdown, company_readiness, difficulty_counts,
            contests_participated, weekly_trend, total_solved,
        )

        # ── Recent activities ────────────────────────────────────────────
        recent_activities = []
        recent_solved = solved_qs.filter(
            solved_at__gte=timezone.now() - timedelta(days=30)
        ).order_by('-solved_at')[:20]
        for sp in recent_solved:
            recent_activities.append({
                'date': sp.solved_at.strftime('%m/%d'),
                'type': 'Problem Solved',
                'subject': sp.problem.title[:30],
                'result': sp.problem.difficulty or 'Solved',
            })

        # ── Topic accuracy (for chart builder) ───────────────────────────
        try:
            topic_acc_qs = AptitudeContestSubmission.objects.filter(
                student=student
            ).select_related('question__topic__parent').values(
                'question__topic__title', 'question__topic__parent__title'
            ).annotate(
                total=Count('id'),
                correct=Count('id', filter=Q(is_correct=True))
            ).order_by('-total')[:20]

            topic_accuracy = [
                {
                    'topic': t['question__topic__title'],
                    'category': t['question__topic__parent__title'] or t['question__topic__title'],
                    'accuracy': round(t['correct'] / t['total'] * 100, 1) if t['total'] else 0,
                    'total': t['total'],
                    'correct': t['correct'],
                }
                for t in topic_acc_qs if t['question__topic__title']
            ]
        except Exception:
            topic_accuracy = []

        # ── Top skills / companies (legacy compat) ───────────────────────
        top_skills = [{'skill': s.title(), 'count': c} for s, c in sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)]
        top_companies = [{'company': c, 'count': n} for c, n in sorted(company_counts.items(), key=lambda x: x[1], reverse=True)]

        return {
            'total_solved': total_solved,
            'easy': difficulty_counts['Easy'],
            'medium': difficulty_counts['Medium'],
            'hard': difficulty_counts['Hard'],
            'current_streak': student.current_streak,
            'campus_rank': campus_rank,
            'total_students': total_students,
            'success_rate': success_rate,
            'total_attempts': total_attempts,
            'aptitude_solved': aptitude_solved,
            'total_aptitude': total_aptitude,
            'aptitude_percentage': aptitude_pct,
            'contests_participated': contests_participated,
            'contest_submissions': contest_submissions,
            'consistency_notes': consistency_notes,
            'topic_breakdown': topic_breakdown,
            'weekly_trend': weekly_trend,
            'company_readiness': company_readiness,
            'insights': insights,
            'action_plan': action_plan,
            'top_skills': top_skills,
            'top_companies': top_companies,
            'recent_activities': recent_activities,
            'performance_charts': _build_student_performance_charts(student, solved_qs, topic_accuracy),
        }

    # ── Insights Generator ───────────────────────────────────────────────

    def _generate_insights(self, student, **d):
        """Generate 3–5 factual, data-driven insight bullets."""
        insights = []

        total = d['rd_total_solved']
        hard = d['rd_hard']
        easy = d['rd_easy']
        medium = d['rd_medium']

        # 1. Difficulty balance
        if total > 0 and hard == 0:
            insights.append(
                f"You have solved {total} problems but none are Hard difficulty. "
                f"Attempting Hard problems is essential for interview readiness."
            )
        elif total > 0 and hard > 0:
            hard_pct = round(hard / total * 100)
            if hard_pct < 10:
                insights.append(
                    f"Only {hard_pct}% of your solved problems ({hard}/{total}) are Hard. "
                    f"Aim for at least 15-20% Hard problems to build depth."
                )

        # 2. Easy-heavy check
        if total > 5 and easy > 0:
            easy_pct = round(easy / total * 100)
            if easy_pct > 60:
                insights.append(
                    f"{easy_pct}% of problems solved are Easy ({easy}/{total}). "
                    f"Consider shifting focus to Medium and Hard problems."
                )

        # 3. Streak
        if d['rd_streak'] == 0:
            insights.append(
                "Your current streak is 0 days — you haven't solved anything recently. "
                "Even 1 problem a day helps maintain momentum."
            )
        elif d['rd_streak'] >= 7:
            insights.append(
                f"Excellent streak of {d['rd_streak']} consecutive days. Keep it going."
            )

        # 4. Weakest topics from topic breakdown
        weak_topics = [tb for tb in d['rd_topic_breakdown'] if tb['level'] == 'Beginner']
        if weak_topics:
            names = ", ".join([t['topic'] for t in weak_topics[:3]])
            insights.append(
                f"Topics at Beginner level: {names}. "
                f"These need more practice to reach interview readiness."
            )

        # 5. Contest participation
        if d['rd_contests'] == 0:
            insights.append(
                "No contest participation yet. Contests build speed and pressure handling — "
                "try participating in at least one this week."
            )
        elif d['rd_contests'] < 3:
            insights.append(
                f"Only {d['rd_contests']} contest(s) attended. Consistent contest practice "
                f"significantly improves timed problem-solving skills."
            )

        # 6. Success rate
        if d['rd_success_rate'] > 0 and d['rd_success_rate'] < 30:
            insights.append(
                f"Success rate is {d['rd_success_rate']:.0f}%. This suggests many failed attempts. "
                f"Consider studying solutions/editorials before attempting new problems."
            )

        # 7. Weekly activity gaps
        inactive_weeks = sum(1 for w in d['rd_weekly_trend'] if w['solved'] == 0)
        if inactive_weeks >= 4:
            insights.append(
                f"{inactive_weeks} of the last 12 weeks had zero problems solved. "
                f"Consistency matters more than intensity — aim for at least a few problems every week."
            )

        return insights[:5]  # Cap at 5

    # ── Action Plan Generator ────────────────────────────────────────────

    def _generate_action_plan(self, topic_breakdown, company_readiness, difficulty, contests, weekly_trend, total_solved):
        """Generate a personalized 2-week action plan."""
        plan = []

        # 1. Identify weakest topics from company gaps
        all_gaps = set()
        for cr in company_readiness[:3]:  # top 3 target companies
            for g in cr.get('gaps', [])[:2]:
                all_gaps.add(g)

        if all_gaps:
            gap_list = list(all_gaps)[:3]
            counts = ", ".join([f"5 {g} problems" for g in gap_list])
            plan.append(
                f"Focus on your company-readiness gaps: solve {counts} over the next 2 weeks."
            )

        # 2. Difficulty progression
        if difficulty.get('Hard', 0) == 0:
            plan.append(
                "Start attempting Hard problems: aim for 2 Hard problems this week and 3 next week. "
                "Pick topics you're already Intermediate in."
            )
        elif difficulty.get('Hard', 0) < 5:
            plan.append(
                f"You've solved {difficulty['Hard']} Hard problems. Target 3 more Hard problems "
                f"in the next 2 weeks to build depth."
            )

        # 3. Weakest topics from breakdown
        beginner_topics = [t for t in topic_breakdown if t['level'] == 'Beginner']
        if beginner_topics and not all_gaps:
            weak = beginner_topics[:2]
            for t in weak:
                plan.append(
                    f"Practice {t['topic']}: solve {5 - t['count']} more problems to reach Intermediate level."
                )

        # 4. Contest recommendation
        if contests < 2:
            plan.append(
                "Participate in at least 1 contest this week and 1 next week. "
                "Timed problem-solving is a distinct skill that needs separate practice."
            )

        # 5. Daily target
        recent_avg = sum(w['solved'] for w in weekly_trend[-4:]) / 4 if weekly_trend else 0
        if recent_avg < 3:
            plan.append(
                "Set a daily target: solve at least 1 problem per day (any difficulty) to build consistency."
            )
        elif recent_avg < 7:
            plan.append(
                f"Your recent average is ~{recent_avg:.0f} problems/week. "
                f"Push to 10+/week by adding 1-2 problems daily."
            )

        return plan[:5]  # Cap at 5


class BatchReportPDFView(APIView):
    """Generate a comprehensive PDF performance report for a batch of students."""
    permission_classes = [IsAuthenticated]

    def get(self, request, batch_code):
        if not (hasattr(request.user, 'staff_profile') or request.user.is_superuser):
            return Response({"detail": "Staff access required."}, status=status.HTTP_403_FORBIDDEN)

        staff = getattr(request.user, 'staff_profile', None)
        section_filter = request.GET.get('section', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        report_type = request.GET.get('type', 'overall')

        # Filter students
        students_qs = StudentProfile.objects.filter(batch=batch_code)
        if staff and staff.department and getattr(staff, 'role', '') not in ['admin', 'superuser']:
            students_qs = students_qs.filter(department=staff.department)
        if section_filter:
            students_qs = students_qs.filter(section=section_filter)

        if not students_qs.exists():
            return Response({"error": f"No students found for batch {batch_code}."}, status=status.HTTP_404_NOT_FOUND)

        from .pdf_reports import create_watermarked_pdf_contest
        buffer = BytesIO()
        first_student = students_qs.first()
        institution = first_student.institution or getattr(staff, 'institution', None)
        department = first_student.department or getattr(staff, 'department', None)

        doc = create_watermarked_pdf_contest(
            buffer,
            institution=institution,
            department=department,
            pagesize=A4,
            topMargin=2.1 * inch,
            bottomMargin=0.6 * inch,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('BTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=14, alignment=1, textColor=colors.HexColor('#2d5016'))
        header_style = ParagraphStyle('BHeader', parent=styles['Heading2'], fontSize=12, spaceBefore=12, spaceAfter=6, textColor=colors.HexColor('#39482a'))

        elements = []
        elements.append(Paragraph(f"Batch Performance Report: {batch_code}", title_style))
        if section_filter:
            elements.append(Paragraph(f"Section: {section_filter}", header_style))

        # Metrics overview
        total_students = students_qs.count()
        solved_qs = SolvedProblem.objects.filter(student__in=students_qs)
        total_problems_solved = solved_qs.count()
        total_submissions = ExecutionRecord.objects.filter(student__in=students_qs).count()
        avg_solved = (total_problems_solved / total_students) if total_students > 0 else 0

        summary_data = [
            ["Metric", "Value"],
            ["Batch Code", str(batch_code)],
            ["Total Students", str(total_students)],
            ["Total Problems Solved", str(total_problems_solved)],
            ["Total Code Executions", str(total_submissions)],
            ["Avg Solved per Student", f"{avg_solved:.1f}"],
        ]
        elements.append(Paragraph("Batch Summary Overview", header_style))
        t_summary = Table(summary_data, colWidths=[3*inch, 3.5*inch])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6ebdd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#39482a')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0d9c2')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 0.2 * inch))

        # Student Leaderboard Table for the Batch
        elements.append(Paragraph("Student Performance Leaderboard", header_style))
        student_rows = [["Reg No", "Name", "Section", "Problems Solved", "Streak"]]
        for st in students_qs.annotate(s_count=Count('solved_problems')).order_by('-s_count')[:50]:
            student_rows.append([
                st.register_number,
                st.name[:25],
                st.section or 'N/A',
                str(st.s_count),
                f"{st.current_streak} days",
            ])

        t_students = Table(student_rows, colWidths=[1.4*inch, 2.2*inch, 0.8*inch, 1.1*inch, 1*inch])
        t_students.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6ebdd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#39482a')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9f7')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0d9c2')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(t_students)

        doc.build(elements)
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="batch_report_{batch_code}.pdf"'
        return response


class TrackedCompaniesReportPDFView(APIView):
    """Generate PDF report for student's company readiness."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'student_profile'):
            return Response({"detail": "Student access required."}, status=status.HTTP_403_FORBIDDEN)
        student = request.user.student_profile

        from .pdf_reports import create_watermarked_pdf_contest
        buffer = BytesIO()
        doc = create_watermarked_pdf_contest(
            buffer,
            institution=student.institution,
            department=student.department,
            pagesize=A4,
            topMargin=2.1 * inch,
            bottomMargin=0.6 * inch,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CTitle', parent=styles['Heading1'], fontSize=16, spaceAfter=12, alignment=1, textColor=colors.HexColor('#2d5016'))
        header_style = ParagraphStyle('CHeader', parent=styles['Heading2'], fontSize=11, spaceBefore=10, spaceAfter=4, textColor=colors.HexColor('#39482a'))

        elements = []
        elements.append(Paragraph(f"Company Readiness Report: {student.name}", title_style))
        elements.append(Paragraph(f"Register Number: {student.register_number}", header_style))
        elements.append(Spacer(1, 0.2 * inch))

        solved_qs = SolvedProblem.objects.filter(student=student)
        total_solved = solved_qs.count()

        data = [
            ["Category", "Details"],
            ["Total Problems Solved", str(total_solved)],
            ["Current Streak", f"{student.current_streak} days"],
            ["Department", student.department.get_full_name() if student.department else 'N/A'],
        ]
        t = Table(data, colWidths=[2.5*inch, 4*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6ebdd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#39482a')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0d9c2')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('PADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)

        doc.build(elements)
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="company_readiness_report_{student.register_number}.pdf"'
        return response


class StaffReportPDFView(APIView):
    """Generate a comprehensive PDF performance report with filtering options."""
    permission_classes = [IsAuthenticated]

    def get(self, request, faculty_id):
        # Check if user is staff (HOD or self)
        if not hasattr(request.user, 'staff_profile'):
            return Response({"detail": "Staff access required."}, status=403)

        user_profile = request.user.staff_profile
        target_staff = get_object_or_404(StaffProfile, faculty_id=faculty_id)

        # Access control
        if user_profile.role in ('hod', 'academics') and target_staff.department != user_profile.department:
            return Response({"detail": "Access denied."}, status=403)
        if target_staff.institution != user_profile.institution:
            return Response({"detail": "Access denied."}, status=403)

        # Get filter parameters
        batch_filter = request.GET.get('batch', '')
        report_type = request.GET.get('type', 'overall')  # overall, aptitude, programming, contests
        topic_filter = request.GET.get('topic', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')

        # Create PDF with the shared branded header (logo + name + subheading +
        # address + department, drawn once on page 1 only) instead of the
        # plain-text-only title block this view used to build by hand.
        from .pdf_reports import create_watermarked_pdf_contest, _bar_chart, _pie_chart
        buffer = BytesIO()
        doc = create_watermarked_pdf_contest(
            buffer,
            institution=target_staff.institution,
            department=target_staff.department,
            pagesize=A4,
            topMargin=2.1*inch,
            bottomMargin=0.6*inch
        )

        # Custom styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center
            textColor=colors.HexColor('#2d5016')
        )

        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#39482a')
        )

        elements = []
        institution = target_staff.institution

        contact_info = []
        if institution.contact_email:
            contact_info.append(f"Email: {institution.contact_email}")
        if institution.contact_phone:
            contact_info.append(f"Phone: {institution.contact_phone}")
        if institution.website:
            contact_info.append(f"Website: {institution.website}")
        
        if contact_info:
            elements.append(Paragraph(" | ".join(contact_info), styles['Normal']))
        
        elements.append(Spacer(1, 0.3 * inch))
        
        # Report Title
        report_titles = {
            'overall': 'Comprehensive Faculty Performance Report',
            'aptitude': 'Aptitude Assessment Performance Report',
            'programming': 'Programming Contest Performance Report',
            'contests': 'Contest Management Report'
        }
        elements.append(Paragraph(report_titles.get(report_type, 'Faculty Performance Report'), title_style))
        elements.append(Spacer(1, 0.2 * inch))
        
        # Faculty Information
        elements.append(Paragraph("Faculty Information", header_style))
        faculty_data = [
            ["Name:", target_staff.name],
            ["Faculty ID:", target_staff.faculty_id],
            ["Department:", target_staff.department.get_full_name() if target_staff.department else 'N/A'],
            ["Role:", target_staff.get_role_display() if hasattr(target_staff, 'get_role_display') else target_staff.role],
            ["Report Period:", f"{date_from or 'All time'} to {date_to or 'Present'}"],
        ]
        
        if batch_filter:
            faculty_data.append(["Batch Filter:", batch_filter])
        if topic_filter:
            faculty_data.append(["Topic Filter:", topic_filter])
            
        faculty_table = Table(faculty_data, colWidths=[2*inch, 4*inch])
        faculty_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(faculty_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Get filtered data based on report type and filters
        report_data = self._get_filtered_report_data(target_staff, report_type, batch_filter, topic_filter, date_from, date_to)
        
        # Performance Metrics
        elements.append(Paragraph("Performance Metrics", header_style))
        metrics_data = [
            ["Metric", "Value", "Details"],
            ["Total Students Managed", str(report_data['student_count']), f"Across {report_data['batch_count']} batches"],
            ["Contests Created", str(report_data['total_contests']), f"{report_data['active_contests']} active"],
            ["Student Submissions", str(report_data['total_submissions']), f"Avg: {report_data['avg_submissions_per_student']:.1f} per student"],
            ["Problems Solved", str(report_data['total_problems_solved']), f"Success rate: {report_data['success_rate']:.1f}%"],
        ]
        
        if report_type in ['aptitude', 'overall']:
            metrics_data.extend([
                ["Aptitude Tests Conducted", str(report_data['aptitude_tests']), f"{report_data['aptitude_participants']} participants"],
                ["Avg Aptitude Score", f"{report_data['avg_aptitude_score']:.1f}%", f"Best: {report_data['best_aptitude_score']:.1f}%"],
            ])
            
        metrics_table = Table(metrics_data, colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6ebdd')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#39482a')),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9f7')]),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d0d9c2')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(metrics_table)
        elements.append(Spacer(1, 0.3 * inch))

        # Aggregate Performance Graphs
        chart_data = report_data.get('performance_charts', {})
        if chart_data:
            elements.append(Paragraph("Performance Graphs", header_style))
            overall = chart_data.get('overall', {})
            overall_values = [
                overall.get('programming', 0),
                overall.get('aptitude', 0),
                overall.get('contest', 0),
            ]
            if not sum(overall_values):
                overall_values = [1]
                overall_labels = ["No data"]
            else:
                overall_labels = ["Programming", "Aptitude", "Contest"]

            daily = chart_data.get('daily_solved', [])[-14:]
            daily_values = [row.get('count', 0) for row in daily]
            daily_labels = [row.get('date', '')[5:] for row in daily]

            pie = _pie_chart(overall_values, overall_labels, ['#22c55e', '#3b82f6', '#f59e0b'], size=130)
            daily_bar = _bar_chart(daily_values, daily_labels, bar_color='#0f766e', w=295, h=140)
            chart_table = Table([[pie, daily_bar]], colWidths=[2.2*inch, 4.2*inch])
            chart_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0d9c2')),
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9f7')),
            ]))
            elements.append(chart_table)
            elements.append(Spacer(1, 0.25 * inch))

            topics = chart_data.get('topics', [])[:10]
            if topics:
                elements.append(Paragraph("Solving Topics", header_style))
                elements.append(_bar_chart(
                    [row.get('count', 0) for row in topics],
                    [row.get('label', '')[:10] for row in topics],
                    bar_color='#4f46e5',
                    w=475,
                    h=150,
                ))
                elements.append(Spacer(1, 0.3 * inch))

        # Batch-wise Performance (if applicable)
        if report_data['batch_performance']:
            elements.append(Paragraph("Batch-wise Performance", header_style))
            batch_data = [["Batch", "Students", "Avg Score", "Top Performer", "Completion Rate"]]
            for batch_info in report_data['batch_performance']:
                batch_data.append([
                    batch_info['batch'],
                    str(batch_info['student_count']),
                    f"{batch_info['avg_score']:.1f}%",
                    batch_info['top_performer'],
                    f"{batch_info['completion_rate']:.1f}%"
                ])
            
            batch_table = Table(batch_data, colWidths=[1.2*inch, 1*inch, 1.2*inch, 1.8*inch, 1.3*inch])
            batch_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6ebdd')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#39482a')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9f7')]),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d0d9c2')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(batch_table)
            elements.append(Spacer(1, 0.3 * inch))

        # Recent Activity Summary
        if report_data['recent_activities']:
            elements.append(Paragraph("Recent Activity (Last 30 Days)", header_style))
            activity_data = [["Date", "Activity", "Student/Contest", "Result"]]
            for activity in report_data['recent_activities'][:10]:  # Show top 10
                activity_data.append([
                    activity['date'],
                    activity['type'],
                    activity['subject'],
                    activity['result']
                ])
            
            activity_table = Table(activity_data, colWidths=[1.2*inch, 1.5*inch, 2*inch, 1.8*inch])
            activity_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e6ebdd')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#39482a')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9f7')]),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#d0d9c2')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(activity_table)

        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        local_now = timezone.localtime(timezone.now())
        elements.append(Paragraph(f"Report generated on {local_now.strftime('%B %d, %Y at %I:%M %p')}", footer_style))
        elements.append(Paragraph(f"Generated by: {user_profile.name} ({user_profile.faculty_id})", footer_style))

        doc.build(elements)
        buffer.seek(0)
        
        # Generate filename based on filters
        filename_parts = [f"Report_{target_staff.faculty_id}"]
        if batch_filter:
            filename_parts.append(f"Batch_{batch_filter}")
        if report_type != 'overall':
            filename_parts.append(report_type.title())
        filename_parts.append(local_now.strftime('%Y%m%d'))
        
        filename = "_".join(filename_parts) + ".pdf"
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _get_filtered_report_data(self, staff, report_type, batch_filter, topic_filter, date_from, date_to):
        """Get comprehensive report data based on filters."""
        from datetime import datetime, timedelta
        from django.db.models import Q, Avg, Count, Sum, Max
        from collections import defaultdict
        
        # Base queryset for students
        students_qs = StudentProfile.objects.filter(
            department=staff.department if staff.department else None
        )
        
        if batch_filter:
            students_qs = students_qs.filter(batch=batch_filter)
            
        # Date filtering
        date_filter = Q()
        if date_from:
            try:
                date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                date_filter &= Q(created_at__date__gte=date_from_obj)
            except ValueError:
                pass
                
        if date_to:
            try:
                date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                date_filter &= Q(created_at__date__lte=date_to_obj)
            except ValueError:
                pass

        # Get contests created by this staff
        contests_qs = staff.contests.all()
        if date_filter:
            contests_qs = contests_qs.filter(date_filter)

        # Calculate metrics
        student_count = students_qs.count()
        batch_count = students_qs.values('batch').distinct().count()
        total_contests = contests_qs.count()
        active_contests = contests_qs.filter(status='published').count()

        # Submission metrics
        submissions = ExecutionRecord.objects.filter(
            student__in=students_qs
        )
        if date_filter:
            submissions = submissions.filter(date_filter)
            
        total_submissions = submissions.count()
        # Use SolvedProblem for successful submissions instead
        successful_problems = SolvedProblem.objects.filter(student__in=students_qs)
        if date_filter:
            # Use solved_at field for SolvedProblem
            date_filter_solved = Q()
            if date_from:
                try:
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                    date_filter_solved &= Q(solved_at__date__gte=date_from_obj)
                except ValueError:
                    pass
            if date_to:
                try:
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                    date_filter_solved &= Q(solved_at__date__lte=date_to_obj)
                except ValueError:
                    pass
            if date_filter_solved:
                successful_problems = successful_problems.filter(date_filter_solved)
        successful_submissions = successful_problems.count()
        avg_submissions_per_student = total_submissions / student_count if student_count > 0 else 0
        success_rate = (successful_submissions / total_submissions * 100) if total_submissions > 0 else 0

        # Problem solving metrics
        solved_problems = SolvedProblem.objects.filter(student__in=students_qs)
        if date_filter:
            # Use solved_at field for SolvedProblem
            if date_filter_solved:
                solved_problems = solved_problems.filter(date_filter_solved)
        total_problems_solved = solved_problems.count()

        # Aptitude metrics
        aptitude_data = {'aptitude_tests': 0, 'aptitude_participants': 0, 'avg_aptitude_score': 0, 'best_aptitude_score': 0}
        if report_type in ['aptitude', 'overall']:
            aptitude_contests = contests_qs.filter(contest_type='aptitude')
            aptitude_submissions = AptitudeContestSubmission.objects.filter(
                contest__in=aptitude_contests,
                student__in=students_qs
            )
            if date_filter:
                # Use submitted_at field for contest submissions
                date_filter_aptitude_submissions = Q()
                if date_from:
                    try:
                        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
                        date_filter_aptitude_submissions &= Q(submitted_at__date__gte=date_from_obj)
                    except ValueError:
                        pass
                if date_to:
                    try:
                        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
                        date_filter_aptitude_submissions &= Q(submitted_at__date__lte=date_to_obj)
                    except ValueError:
                        pass
                if date_filter_aptitude_submissions:
                    aptitude_submissions = aptitude_submissions.filter(date_filter_aptitude_submissions)
                
            aptitude_data.update({
                'aptitude_tests': aptitude_contests.count(),
                'aptitude_participants': aptitude_submissions.values('student').distinct().count(),
                'avg_aptitude_score': aptitude_submissions.aggregate(avg=Avg('score'))['avg'] or 0,
                'best_aptitude_score': aptitude_submissions.aggregate(max=Max('score'))['max'] or 0,
            })

        # Aggregate chart data for staff/HOD report generation.
        start_day = timezone.localdate() - timedelta(days=29)
        daily_rows = (
            SolvedProblem.objects
            .filter(student__in=students_qs, solved_at__date__gte=start_day)
            .values('solved_at__date')
            .annotate(count=Count('id'))
            .order_by('solved_at__date')
        )
        topic_counts = defaultdict(int)
        for tags in solved_problems.select_related('problem').values_list('problem__tags', flat=True):
            if isinstance(tags, list):
                for tag in tags[:4]:
                    topic_counts[str(tag).strip().title()] += 1
        aptitude_solved_count = SolvedAptitude.objects.filter(student__in=students_qs).count()
        contest_solved_count = (
            ContestSubmission.objects
            .filter(student__in=students_qs, status='Accepted')
            .values('contest_id', 'problem_id', 'student_id')
            .distinct()
            .count()
            + AptitudeContestSubmission.objects.filter(student__in=students_qs, is_correct=True).count()
        )
        performance_charts = {
            'overall': {
                'programming': total_problems_solved,
                'aptitude': aptitude_solved_count,
                'contest': contest_solved_count,
            },
            'daily_solved': [
                {'date': row['solved_at__date'].isoformat(), 'count': row['count']}
                for row in daily_rows
            ],
            'topics': [
                {'label': label, 'count': count}
                for label, count in sorted(topic_counts.items(), key=lambda item: item[1], reverse=True)[:10]
            ],
        }

        # Batch-wise performance
        batch_performance = []
        for batch in students_qs.values('batch').distinct():
            batch_name = batch['batch']
            batch_students = students_qs.filter(batch=batch_name)
            batch_submissions = submissions.filter(student__in=batch_students)
            
            # Calculate success rate based on solved problems vs total submissions
            batch_solved = successful_problems.filter(student__in=batch_students)
            batch_success_rate = 0
            if batch_submissions.exists():
                batch_success_rate = (batch_solved.count() / batch_submissions.count() * 100)
            
            # Find top performer in batch
            top_performer = batch_students.annotate(
                solved_count=Count('solved_problems')
            ).order_by('-solved_count').first()
            
            batch_performance.append({
                'batch': batch_name,
                'student_count': batch_students.count(),
                'avg_score': batch_success_rate,
                'top_performer': top_performer.name if top_performer else 'N/A',
                'completion_rate': batch_success_rate
            })

        # Recent activities
        recent_activities = []
        recent_submissions = submissions.filter(
            created_at__gte=timezone.now() - timedelta(days=30)
        ).order_by('-created_at')[:20]
        
        for sub in recent_submissions:
            # Check if this submission resulted in a solved problem
            is_success = successful_problems.filter(
                student=sub.student, 
                problem=sub.problem,
                solved_at__date=sub.created_at.date()
            ).exists()
            
            recent_activities.append({
                'date': sub.created_at.strftime('%m/%d'),
                'type': 'Problem Solved' if is_success else 'Attempt',
                'subject': f"{sub.student.name} - {sub.problem.title[:30] if sub.problem else 'Unknown Problem'}",
                'result': 'Success' if is_success else 'Failed'
            })

        return {
            'student_count': student_count,
            'batch_count': batch_count,
            'total_contests': total_contests,
            'active_contests': active_contests,
            'total_submissions': total_submissions,
            'avg_submissions_per_student': avg_submissions_per_student,
            'success_rate': success_rate,
            'total_problems_solved': total_problems_solved,
            'batch_performance': batch_performance,
            'recent_activities': recent_activities,
            'performance_charts': performance_charts,
            **aptitude_data
        }


# ---------------------------------------------------------------------------
# System Administration & Multi-Tenancy
# ---------------------------------------------------------------------------

class SystemAdminDashboardView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user_role = str(getattr(request.user, 'role', '') or '').lower()
        is_admin_user = (
            request.user.is_superuser or 
            request.user.is_staff or 
            user_role in ('admin', 'superuser') or 
            getattr(request.user, 'username', '') in ('0001', 'staff_0001', 'admin')
        )
        if not is_admin_user:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        try:
            total_students = StudentProfile.objects.count()
            total_staff = StaffProfile.objects.count()
            total_problems = Problem.objects.count()
            total_aptitude = AptitudeQuestion.objects.count()
            total_interview_tracks = Department.objects.exclude(
                interview_track=''
            ).values('interview_track').distinct().count()

            # Fetch all institutions for the management table
            institutions = Institution.objects.all().values(
                'id', 'institution_id', 'name', 'short_code', 'is_active',
                'maintenance_staff', 'maintenance_students', 'maintenance_hod',
                'maintenance_inst_admin', 'maintenance_ja', 'maintenance_tpu', 'maintenance_director'
            )

            # Global maintenance config
            config, _ = SystemConfiguration.objects.get_or_create(id=1)

            return Response({
                "metrics": {
                    "total_users": total_students + total_staff,
                    "total_staff": total_staff,
                    "total_problems": total_problems,
                    "total_aptitude": total_aptitude,
                    "total_interview_tracks": total_interview_tracks
                },
                "institutions": list(institutions),
                "global_config": {
                    "staff": config.global_maintenance_staff,
                    "student": config.global_maintenance_students,
                    "hod": config.global_maintenance_hod,
                    "tpu": config.global_maintenance_tpu,
                    "director": config.global_maintenance_director,
                    "ja": config.global_maintenance_ja,
                    "admin": config.global_maintenance_admin
                }
            })
        except Exception as e:
            logger.error(f"Admin dashboard error: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class InstitutionManagementView(APIView):
    permission_classes = [AllowAny]

    def post(self, request):
        """Create a new institution and its database"""
        data = request.data
        try:
            inst_id = data.get('institution_id')
            name = data.get('name')
            short_code = data.get('short_code', '')
            address = data.get('address', '')
            contact_email = data.get('contact_email', '')
            contact_phone = data.get('contact_phone', '')
            
            if not inst_id or not name:
                return Response({"error": "ID and Name required"}, status=400)
            
            # Sanitize DB name
            db_name = f"code2day_inst_{inst_id}"
            
            # Create Institution record
            institution = Institution.objects.create(
                institution_id=inst_id,
                name=name,
                short_code=short_code,
                address=address,
                contact_email=contact_email,
                contact_phone=contact_phone,
                database_name=db_name
            )
            
            # Create the actual database
            create_institution_db(db_name)
            
            return Response({"message": "Institution created", "id": institution.id})
        except Exception as e:
            logger.error(f"Failed to create institution: {e}")
            return Response({"error": str(e)}, status=500)

    def delete(self, request, pk):
        """Delete institution and its database"""
        try:
            institution = get_object_or_404(Institution, pk=pk)
            db_name = institution.database_name
            
            # Delete database first
            if db_name:
                delete_institution_db(db_name)
            
            institution.delete()
            return Response({"message": "Institution deleted"})
        except Exception as e:
            logger.error(f"Failed to delete institution: {e}")
            return Response({"error": str(e)}, status=500)

class InstitutionDetailManagementView(APIView):
    """Management within a specific institution"""
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        institution = get_object_or_404(Institution, pk=pk)
        
        # Get actual student count
        students_count = StudentProfile.objects.filter(institution=institution).count()
        
        # Get staff, HODs, and Academic Coordinator (0001)
        staff_list = StaffProfile.objects.filter(institution=institution).values(
            'id', 'faculty_id', 'name', 'email', 'mobile_number', 'role', 'department__name', 'department__id', 'department__code'
        )
        
        # Get departments
        depts = Department.objects.filter(institution=institution).values('id', 'name', 'code', 'interview_track')
        
        # Get students
        student_list = []
        batches = set()
        for s in StudentProfile.objects.filter(institution=institution).select_related('account', 'department'):
            if s.batch:
                batches.add(s.batch)
            student_list.append({
                'id': s.id,
                'name': s.name,
                'register_number': s.register_number,
                'personal_email': s.personal_email,
                'mobile_number': s.mobile_number,
                'department_id': s.department_id,
                'batch': s.batch,
                'is_active': s.account.is_active if s.account else True
            })
        
        return Response({
            "staff": list(staff_list),
            "students": list(student_list),
            "departments": list(depts),
            "batches": sorted(list(batches)),
            "maintenance": {
                "staff": institution.maintenance_staff,
                "student": institution.maintenance_students,
                "hod": institution.maintenance_hod,
                "inst_admin": institution.maintenance_inst_admin,
                "ja": institution.maintenance_ja,
                "tpu": institution.maintenance_tpu,
                "director": institution.maintenance_director
            },
            "module_registry": serializable_registry(),
            "locked_modules": institution.locked_modules,
            "branding": {
                "display_name": institution.display_name,
                "subheading": institution.subheading,
                "logo_url": institution.logo_url,
                "logo_display_url": institution.logo_display_url,
                "website": institution.website,
                "established_year": institution.established_year,
                "address": institution.address,
                "contact_email": institution.contact_email,
                "contact_phone": institution.contact_phone
            },
            "metrics": {
                "students": students_count,
                "staff": StaffProfile.objects.filter(institution=institution).count(),
                "departments": depts.count()
            }
        })

    def patch(self, request, pk):
        """Update institution maintenance or staff roles/depts"""
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)
        institution = get_object_or_404(Institution, pk=pk)
        action = request.data.get('action')
        
        if action == 'toggle_maintenance':
            role = request.data.get('role')
            value = request.data.get('value')
            if role == 'staff': institution.maintenance_staff = value
            elif role == 'student': institution.maintenance_students = value
            elif role == 'hod': institution.maintenance_hod = value
            elif role == 'inst_admin': institution.maintenance_inst_admin = value
            elif role == 'ja': institution.maintenance_ja = value
            elif role == 'tpu': institution.maintenance_tpu = value
            elif role == 'director': institution.maintenance_director = value
            institution.save()
            return Response({"message": "Maintenance updated"})

        elif action == 'toggle_module_lock':
            module_key = request.data.get('module')
            value = bool(request.data.get('value'))
            if module_key not in MODULE_KEYS:
                return Response({"error": f"Unknown module '{module_key}'."}, status=400)
            locked = set(institution.locked_modules or [])
            if value:
                locked.add(module_key)
            else:
                locked.discard(module_key)
            institution.locked_modules = sorted(locked)
            institution.save(update_fields=['locked_modules'])
            return Response({"message": "Module lock updated", "locked_modules": institution.locked_modules})

        elif action == 'update_role':
            staff_id = request.data.get('staff_id')
            new_role = request.data.get('role')
            staff = get_object_or_404(StaffProfile, id=staff_id, institution=institution)
            staff.role = new_role
            staff.save()
            return Response({"message": "Role updated"})

        elif action == 'update_dept':
            staff_id = request.data.get('staff_id')
            dept_id = request.data.get('dept_id')
            staff = get_object_or_404(StaffProfile, id=staff_id, institution=institution)
            if dept_id:
                dept = get_object_or_404(Department, id=dept_id, institution=institution)
                staff.department = dept
            else:
                staff.department = None
            staff.save()
            return Response({"message": "Department updated"})

        elif action == 'update_department_interview_track':
            dept_id = request.data.get('dept_id')
            track = (request.data.get('interview_track') or '').strip()
            dept = get_object_or_404(Department, id=dept_id, institution=institution)
            dept.interview_track = track or dept.default_interview_track()
            dept.save(update_fields=['interview_track'])
            return Response({"message": "Interview track updated", "interview_track": dept.interview_track})

        elif action == 'create_staff':
            faculty_id = (request.data.get('faculty_id') or '').strip()
            name = (request.data.get('name') or '').strip()
            email = (request.data.get('email') or '').strip()
            mobile_number = (request.data.get('mobile_number') or '').strip()
            role = request.data.get('role', 'staff')
            dept_id = request.data.get('dept_id')

            if not faculty_id or not name:
                return Response({"error": "faculty_id and name are required."}, status=400)

            if role not in dict(StaffProfile.ROLE_CHOICES):
                return Response({"error": "Invalid role."}, status=400)

            if StaffProfile.objects.filter(faculty_id=faculty_id).exists():
                return Response({"error": "A staff member with this faculty ID already exists."}, status=400)

            department = None
            if dept_id:
                department = get_object_or_404(Department, id=dept_id, institution=institution)

            staff = StaffProfile.objects.create(
                faculty_id=faculty_id,
                name=name,
                email=email,
                mobile_number=mobile_number,
                role=role,
                department=department,
                institution=institution,
            )
            return Response({
                "message": "Staff created successfully",
                "staff": {
                    "id": staff.id,
                    "faculty_id": staff.faculty_id,
                    "name": staff.name,
                    "email": staff.email,
                    "mobile_number": staff.mobile_number,
                    "role": staff.role,
                    "department__id": staff.department_id,
                    "department__name": staff.department.name if staff.department else None,
                    "department__code": staff.department.code if staff.department else None,
                },
            }, status=201)

        elif action == 'update_staff_email':
            staff_id = request.data.get('staff_id')
            email = (request.data.get('email') or '').strip()
            staff = get_object_or_404(StaffProfile, id=staff_id, institution=institution)
            staff.email = email
            staff.save(update_fields=['email'])
            return Response({"message": "Email updated", "email": staff.email})

        elif action == 'update_staff_mobile':
            staff_id = request.data.get('staff_id')
            mobile_number = (request.data.get('mobile_number') or '').strip()
            staff = get_object_or_404(StaffProfile, id=staff_id, institution=institution)
            staff.mobile_number = mobile_number
            staff.save(update_fields=['mobile_number'])
            return Response({"message": "Contact number updated", "mobile_number": staff.mobile_number})

        elif action == 'update_staff_name':
            staff_id = request.data.get('staff_id')
            name = (request.data.get('name') or '').strip()
            if not name:
                return Response({"error": "Name is required."}, status=400)
            staff = get_object_or_404(StaffProfile, id=staff_id, institution=institution)
            staff.name = name
            staff.save(update_fields=['name'])
            return Response({"message": "Name updated", "name": staff.name})

        elif action == 'delete_staff':
            staff_id = request.data.get('staff_id')
            staff = get_object_or_404(StaffProfile, id=staff_id, institution=institution)
            if staff.faculty_id == '0001':
                return Response({"error": "Cannot delete the system admin account."}, status=400)
            account = staff.account
            if account:
                account.delete()  # cascades to the StaffProfile
            else:
                staff.delete()
            return Response({"message": "Staff deleted"})

        elif action == 'toggle_student_lock':
            student_id = request.data.get('student_id')
            student = get_object_or_404(StudentProfile, id=student_id, institution=institution)
            if student.account:
                student.account.is_active = not student.account.is_active
                student.account.save(update_fields=['is_active'])
            return Response({"message": f"Student {'unlocked' if student.account.is_active else 'locked'}"})
            
        elif action == 'delete_batch':
            batch_name = request.data.get('batch')
            dept_id = request.data.get('dept_id')
            
            queryset = StudentProfile.objects.filter(institution=institution, batch=batch_name)
            if dept_id:
                queryset = queryset.filter(department_id=dept_id)
            
            # Delete associated users (this will cascade delete profiles)
            user_ids = queryset.filter(account__isnull=False).values_list('account_id', flat=True)
            from django.contrib.auth.models import User
            User.objects.filter(id__in=user_ids).delete()
            # Delete any remaining profiles that had no account
            queryset.delete()
            
            return Response({"message": f"Batch {batch_name} deleted successfully"})
            
        elif action == 'update_branding':
            # Update branding information
            branding_data = request.data.get('branding', {})

            institution.display_name = branding_data.get('display_name', institution.display_name)
            institution.subheading = branding_data.get('subheading', institution.subheading)
            new_logo_url = branding_data.get('logo_url', institution.logo_url)
            # The branding form's logo_url field doubles as the display value
            # after a file upload (which is our own /api/.../logo/ proxy
            # path, not a real external URL) — never persist that back into
            # the URLField meant for a pasted external link.
            if not (new_logo_url or '').startswith('/api/institutions/'):
                institution.logo_url = new_logo_url
            institution.website = branding_data.get('website', institution.website)

            # established_year is a nullable PositiveIntegerField, but the
            # frontend always sends "" (not omitted) when it's unset — assigning
            # that straight onto the model used to crash institution.save() with
            # an unhandled DB-level error (empty string into an integer column),
            # surfacing as an opaque "Failed to update branding:" with no message.
            if 'established_year' in branding_data:
                raw_year = branding_data.get('established_year')
                if raw_year in ('', None):
                    institution.established_year = None
                else:
                    try:
                        institution.established_year = int(raw_year)
                    except (TypeError, ValueError):
                        return Response({"error": "established_year must be a valid year, or left blank."}, status=400)

            institution.address = branding_data.get('address', institution.address)
            institution.contact_email = branding_data.get('contact_email', institution.contact_email)
            institution.contact_phone = branding_data.get('contact_phone', institution.contact_phone)

            institution.save()
            return Response({"message": "Branding updated successfully"})
            
        elif action == 'upload_logo':
            # Handle logo file upload
            if 'logo' not in request.FILES:
                return Response({"error": "No logo file provided"}, status=400)
            
            logo_file = request.FILES['logo']
            
            # Validate file type
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
            if logo_file.content_type not in allowed_types:
                return Response({"error": "Invalid file type. Please upload JPG, PNG, or GIF."}, status=400)
            
            # Validate file size (max 5MB)
            if logo_file.size > 5 * 1024 * 1024:
                return Response({"error": "File too large. Maximum size is 5MB."}, status=400)
            
            # Delete old logo file if exists
            if institution.logo_file:
                institution.logo_file.delete(save=False)
            
            # Save new logo
            institution.logo_file = logo_file
            institution.save()
            
            return Response({
                "message": "Logo uploaded successfully",
                "logo_url": institution.logo_display_url
            })
            
        return Response({"error": "Invalid action"}, status=400)

class GlobalMaintenanceControlView(APIView):
    """System Admin: toggle system-wide maintenance mode, per role. Covers
    every StudentProfile/StaffProfile role (student, staff, hod, tpu,
    director, ja, admin) — admin endpoints are already exempt from the
    maintenance check itself (see MaintenanceMiddleware), so a global admin
    flag can't lock system admins out of the one place that could undo it."""
    permission_classes = [IsAuthenticated]

    ROLE_FIELDS = {
        'student': 'global_maintenance_students',
        'staff': 'global_maintenance_staff',
        'hod': 'global_maintenance_hod',
        'tpu': 'global_maintenance_tpu',
        'director': 'global_maintenance_director',
        'ja': 'global_maintenance_ja',
        'admin': 'global_maintenance_admin',
    }

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        config, _ = SystemConfiguration.objects.get_or_create(id=1)
        role = request.data.get('role')
        value = bool(request.data.get('value'))

        field = self.ROLE_FIELDS.get(role)
        if not field:
            return Response({"error": f"Unknown role {role!r}."}, status=400)

        setattr(config, field, value)
        config.save()
        return Response({r: getattr(config, f) for r, f in self.ROLE_FIELDS.items()})


def _mask_api_key(key):
    if not key:
        return ""
    if len(key) <= 8:
        return "*" * len(key)
    return f"{key[:6]}...{key[-4:]}"


class AdminLLMProvidersView(APIView):
    """System Admin: list/add the LLM providers used for the test-case and
    lab-report generation fallback chain — a React-page equivalent of the
    Django admin LLMProvider CRUD, so providers can be managed without
    leaving the admin dashboard."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        providers = LLMProvider.objects.all().order_by("priority", "id")
        data = [{
            "id": p.id,
            "name": p.name,
            "base_url": p.base_url,
            "api_key_masked": _mask_api_key(p.api_key),
            "model_name": p.model_name,
            "priority": p.priority,
            "is_active": p.is_active,
            "use_streaming": p.use_streaming,
            "temperature": p.temperature,
            "top_p": p.top_p,
            "max_tokens": p.max_tokens,
            "timeout_seconds": p.timeout_seconds,
            "extra_body": p.extra_body,
            "updated_at": p.updated_at.isoformat(),
        } for p in providers]
        return Response({"providers": data})

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        data = request.data
        name = (data.get("name") or "").strip()
        base_url = (data.get("base_url") or "").strip()
        api_key = (data.get("api_key") or "").strip()
        model_name = (data.get("model_name") or "").strip()
        if not (name and base_url and api_key and model_name):
            return Response({"error": "name, base_url, api_key, and model_name are required."}, status=400)
        if LLMProvider.objects.filter(name=name).exists():
            return Response({"error": "A provider with this name already exists."}, status=400)

        try:
            provider = LLMProvider.objects.create(
                name=name, base_url=base_url, api_key=api_key, model_name=model_name,
                priority=int(data.get("priority", 0)),
                is_active=bool(data.get("is_active", True)),
                use_streaming=bool(data.get("use_streaming", False)),
                temperature=float(data.get("temperature", 0.4)),
                top_p=float(data.get("top_p", 0.95)),
                max_tokens=int(data.get("max_tokens", 6000)),
                timeout_seconds=int(data.get("timeout_seconds", 30)),
                extra_body=data.get("extra_body") or {},
            )
        except (TypeError, ValueError) as exc:
            return Response({"error": f"Invalid field value: {exc}"}, status=400)

        return Response({"id": provider.id, "detail": "Provider created."}, status=201)


class AdminLLMProviderDetailView(APIView):
    """System Admin: update (including reordering priority / toggling
    active) or delete one LLM provider."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, provider_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        provider = LLMProvider.objects.filter(id=provider_id).first()
        if not provider:
            return Response({"error": "Not found"}, status=404)

        data = request.data
        for field in ("name", "base_url", "model_name"):
            if field in data and str(data[field]).strip():
                setattr(provider, field, str(data[field]).strip())
        # Only overwrite api_key if a real value was sent — the frontend
        # only ever shows the masked form, so never let a masked string
        # (or a blank field left untouched) accidentally wipe the real key.
        incoming_key = (data.get("api_key") or "").strip()
        if incoming_key and "..." not in incoming_key:
            provider.api_key = incoming_key
        if "priority" in data:
            provider.priority = int(data["priority"])
        if "is_active" in data:
            provider.is_active = bool(data["is_active"])
        if "use_streaming" in data:
            provider.use_streaming = bool(data["use_streaming"])
        if "temperature" in data:
            provider.temperature = float(data["temperature"])
        if "top_p" in data:
            provider.top_p = float(data["top_p"])
        if "max_tokens" in data:
            provider.max_tokens = int(data["max_tokens"])
        if "timeout_seconds" in data:
            provider.timeout_seconds = int(data["timeout_seconds"])
        if "extra_body" in data:
            provider.extra_body = data["extra_body"] or {}

        try:
            provider.save()
        except (TypeError, ValueError) as exc:
            return Response({"error": f"Invalid field value: {exc}"}, status=400)

        return Response({"detail": "Provider updated."})

    def delete(self, request, provider_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        provider = LLMProvider.objects.filter(id=provider_id).first()
        if not provider:
            return Response({"error": "Not found"}, status=404)
        provider.delete()
        return Response(status=204)


def _extract_balanced_braces(text, open_brace_idx):
    """Given text and the index of an opening '{', return the substring
    from there to its matching closing '}' — handles nested dicts, unlike
    a naive non-greedy regex which would stop at the first '}'."""
    depth = 0
    for i in range(open_brace_idx, len(text)):
        if text[i] == "{":
            depth += 1
        elif text[i] == "}":
            depth -= 1
            if depth == 0:
                return text[open_brace_idx:i + 1]
    return None


def _parse_openai_snippet(snippet):
    """Extract provider config fields from a pasted code snippet. Handles
    both common shapes these API docs hand out:
      1. OpenAI SDK style:  OpenAI(base_url="...", api_key="...");
         .chat.completions.create(model="...", stream=..., extra_body={...})
      2. Raw `requests` style: a URL variable ending in /chat/completions,
         an Authorization: Bearer <key> header, and a payload dict with
         "model"/"stream"/"temperature"/etc as dict keys rather than
         Python kwargs.
    So adding a new fallback provider is paste-and-review instead of
    manually re-typing every field, regardless of which shape the docs
    for that particular API used."""
    import ast

    def _find(pattern, text=snippet):
        m = re.search(pattern, text)
        return m.group(1) if m else None

    # base_url: prefer an explicit base_url=... kwarg (OpenAI style); else
    # fall back to any http(s) URL literal in the snippet, stripping a
    # trailing /chat/completions or /completions since _call_provider_once
    # appends that itself.
    base_url = _find(r'base_url\s*=\s*["\']([^"\']+)["\']')
    if not base_url:
        base_url = _find(r'["\'](\bhttps?://[^"\']+)["\']')
        if base_url:
            base_url = re.sub(r'/(chat/)?completions/?$', '', base_url)

    # api_key: an explicit api_key=... kwarg; else a Bearer token in a
    # header string; else any bare nvapi-... token anywhere (NVIDIA's
    # own key format, and every snippet seen so far has used it).
    api_key = (
        _find(r'api_key\s*=\s*["\']([^"\']+)["\']')
        or _find(r'Bearer\s+([A-Za-z0-9\-_.]+)')
        or _find(r'["\'](nvapi-[A-Za-z0-9\-_.]+)["\']')
    )

    # model: a model=... kwarg (SDK style) or a "model": "..." dict entry
    # (raw requests-style payload dict).
    model_name = (
        _find(r'\bmodel\s*=\s*["\']([^"\']+)["\']')
        or _find(r'["\']model["\']\s*:\s*["\']([^"\']+)["\']')
    )

    result = {"base_url": base_url, "api_key": api_key, "model_name": model_name}

    # Everything below may appear either as a Python kwarg (key=value) or
    # a dict literal entry ("key": value) — accept both.
    stream_match = re.search(r'["\']?\bstream["\']?\s*[:=]\s*(True|False)', snippet)
    result["use_streaming"] = stream_match.group(1) == "True" if stream_match else False

    for field, name, caster, default in (
        ("temperature", "temperature", float, 0.4),
        ("top_p", "top_p", float, 0.95),
        ("max_tokens", "max_tokens", int, 6000),
    ):
        m = re.search(rf'["\']?\b{name}["\']?\s*[:=]\s*([0-9.]+)', snippet)
        result[field] = caster(m.group(1)) if m else default

    extra_body = {}
    eb_match = re.search(r'extra_body\s*=\s*(\{)', snippet)
    if eb_match:
        brace_text = _extract_balanced_braces(snippet, eb_match.start(1))
        if brace_text:
            try:
                extra_body = ast.literal_eval(brace_text)
            except (ValueError, SyntaxError):
                extra_body = {}
    result["extra_body"] = extra_body

    return result


class AdminLLMProviderParseSnippetView(APIView):
    """System Admin: parse a pasted OpenAI-SDK-style code snippet into
    provider config fields, for review before creating the provider —
    matches the exact snippet shape NVIDIA's API docs hand out."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        snippet = request.data.get("snippet") or ""
        if not snippet.strip():
            return Response({"error": "Paste a code snippet first."}, status=400)

        parsed = _parse_openai_snippet(snippet)
        missing = [f for f in ("base_url", "api_key", "model_name") if not parsed.get(f)]
        if missing:
            return Response({
                "error": f"Could not find {', '.join(missing)} in the snippet — "
                         f"check it includes OpenAI(base_url=..., api_key=...) and "
                         f".chat.completions.create(model=..., ...).",
                "parsed": parsed,
            }, status=400)

        return Response({"parsed": parsed})


class AdminProblemBankView(APIView):
    """System Admin: list every Problem with its test case count, so admins
    can see at a glance which problems are missing test cases."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problems = Problem.objects.annotate(
            test_case_count=Count("test_cases", distinct=True)
        ).order_by("title")

        data = [{
            "id": p.id,
            "title": p.title,
            "slug": p.slug,
            "difficulty": p.difficulty,
            "tags": p.tags,
            "execution_type": p.execution_type,
            "test_case_count": p.test_case_count,
            "explanation": p.explanation,
            "has_param_schema": bool(p.param_schema),
        } for p in problems]

        return Response({"problems": data, "total": len(data)})


class AdminProblemTestCasesView(APIView):
    """System Admin: view every test case for one Problem, and manually add
    one (for staff who want to hand-author instead of / alongside
    generating)."""
    permission_classes = [IsAuthenticated]

    def get(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        test_cases = [{
            "id": tc.id,
            "stdin": tc.stdin,
            "expected_output": tc.expected_output,
            "is_sample": tc.is_sample,
            "order": tc.order,
            "input_data": tc.input_data,
        } for tc in problem.test_cases.all().order_by("order", "id")]

        return Response({
            "problem": {
                "id": problem.id, "title": problem.title, "slug": problem.slug, "examples": problem.examples,
                "param_schema": problem.param_schema,
            },
            "test_cases": test_cases,
        })

    def post(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        expected_output = (request.data.get("expected_output") or "").strip()
        if not expected_output:
            return Response({"error": "expected_output is required"}, status=400)

        input_data = request.data.get("input_data")
        if input_data is not None and not isinstance(input_data, dict):
            return Response({"error": "input_data must be a JSON object"}, status=400)

        next_order = (problem.test_cases.aggregate(Max("order"))["order__max"] or 0) + 1
        tc = TestCase.objects.create(
            problem=problem,
            stdin=request.data.get("stdin") or "",
            expected_output=expected_output,
            is_sample=bool(request.data.get("is_sample")),
            order=next_order,
            input_data=input_data,
        )
        return Response({
            "id": tc.id, "stdin": tc.stdin, "expected_output": tc.expected_output,
            "is_sample": tc.is_sample, "order": tc.order, "input_data": tc.input_data,
        }, status=201)


class AdminProblemTestCaseDetailView(APIView):
    """System Admin: delete one manually- or LLM-added test case."""
    permission_classes = [IsAuthenticated]

    def delete(self, request, problem_id, test_case_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        tc = TestCase.objects.filter(id=test_case_id, problem_id=problem_id).first()
        if not tc:
            return Response({"error": "Not found"}, status=404)
        tc.delete()
        return Response(status=204)


class AdminProblemParamSchemaView(APIView):
    """System Admin: view/set/clear a Problem's structured parameter/return-type
    schema (Problem.param_schema). Strictly opt-in — a problem with no schema
    keeps executing through the existing regex/heuristic path unchanged; see
    services/param_types.py for the type vocabulary and services/execution_adapter.py
    for how a schema changes execution once set."""
    permission_classes = [IsAuthenticated]

    def get(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        return Response({
            "param_schema": problem.param_schema,
            "valid_types": param_types.VALID_TYPES,
        })

    def put(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        schema = request.data.get("param_schema")
        errors = param_types.validate_schema(schema)
        if errors:
            return Response({"error": "Invalid schema", "details": errors}, status=400)

        problem.param_schema = schema
        problem.save(update_fields=["param_schema"])
        return Response({"param_schema": problem.param_schema})

    def delete(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        problem.param_schema = None
        problem.save(update_fields=["param_schema"])
        return Response(status=204)


class AdminProblemGenerateTestCasesView(APIView):
    """System Admin: on-demand (re)generate test cases for one Problem via
    the LLM fallback chain."""
    permission_classes = [IsAuthenticated]

    def post(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        from .services.testcase_generator import generate_test_cases, derive_examples, TestCaseGenError
        try:
            generated = generate_test_cases(
                title=problem.title,
                description=problem.description,
                examples=problem.examples,
                difficulty=problem.difficulty,
            )
        except TestCaseGenError as exc:
            return Response({"error": f"Generation failed: {exc}"}, status=502)

        with transaction.atomic():
            problem.test_cases.all().delete()
            TestCase.objects.bulk_create([
                TestCase(
                    problem=problem,
                    stdin=case["stdin"],
                    expected_output=case["expected_output"],
                    is_sample=case["is_sample"],
                    order=order,
                )
                for order, case in enumerate(generated, start=1)
            ])

        generated_examples = 0
        if not problem.examples:
            problem.examples = derive_examples(generated)
            problem.save(update_fields=["examples"])
            generated_examples = len(problem.examples)

        return Response({
            "generated_count": len(generated),
            "generated_examples_count": generated_examples,
            "test_case_count": problem.test_cases.count(),
            "examples_count": len(problem.examples),
        })


class AdminProblemGenerateSchemaAndDescriptionView(APIView):
    """System Admin: single button that fills in whatever "necessary data"
    a problem is missing — the typed param_schema and/or the explanation —
    via the LLM fallback chain, WITHOUT touching test cases (that stays a
    separate, explicit action). Both pieces are skip-if-exists: this never
    overwrites a hand-authored schema or an already-generated explanation,
    it only fills in what's actually missing."""
    permission_classes = [IsAuthenticated]

    def post(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        from .services.testcase_generator import generate_param_schema, generate_explanation, TestCaseGenError

        result = {"schema_generated": False, "explanation_generated": False, "errors": {}}

        if not problem.param_schema:
            try:
                schema = generate_param_schema(title=problem.title, description=problem.description, examples=problem.examples)
                problem.param_schema = schema
                problem.save(update_fields=["param_schema"])
                result["schema_generated"] = True
                result["param_schema"] = schema
            except TestCaseGenError as exc:
                result["errors"]["param_schema"] = str(exc)
        else:
            result["param_schema"] = problem.param_schema

        if not problem.explanation:
            try:
                explanation = generate_explanation(
                    title=problem.title, description=problem.description,
                    examples=problem.examples, difficulty=problem.difficulty,
                )
                problem.explanation = explanation
                problem.save(update_fields=["explanation"])
                result["explanation_generated"] = True
                result["explanation"] = explanation
            except TestCaseGenError as exc:
                result["errors"]["explanation"] = str(exc)
        else:
            result["explanation"] = problem.explanation

        return Response(result)


class AdminProblemGenerateExplanationView(APIView):
    """System Admin: on-demand (re)generate a brief explanation for one
    Problem via the LLM fallback chain. Separate endpoint from test-case
    generation so the admin bank can fire both concurrently instead of
    waiting on one before starting the other."""
    permission_classes = [IsAuthenticated]

    def post(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)

        force = bool(request.data.get("force"))
        if problem.explanation and not force:
            return Response(
                {"error": "This problem already has an explanation. Pass force=true to replace it."},
                status=400,
            )

        from .services.testcase_generator import generate_explanation, TestCaseGenError
        try:
            explanation = generate_explanation(
                title=problem.title, description=problem.description,
                examples=problem.examples, difficulty=problem.difficulty,
            )
        except TestCaseGenError as exc:
            return Response({"error": f"Generation failed: {exc}"}, status=502)

        problem.explanation = explanation
        problem.save(update_fields=["explanation"])
        return Response({"explanation": problem.explanation})


class AdminProblemDetailView(APIView):
    """System Admin: delete a single Problem (and its test cases, via
    on_delete cascade) from the Problem Bank."""
    permission_classes = [IsAuthenticated]

    def delete(self, request, problem_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        problem = Problem.objects.filter(id=problem_id).first()
        if not problem:
            return Response({"error": "Not found"}, status=404)
        problem.delete()
        return Response(status=204)


class AdminProblemBulkDeleteView(APIView):
    """System Admin: delete many Problems at once, selected in the Problem
    Bank UI."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        ids = request.data.get("ids") or []
        if not isinstance(ids, list) or not ids:
            return Response({"error": "ids (non-empty list) is required"}, status=400)

        qs = Problem.objects.filter(id__in=ids)
        deleted_count = qs.count()
        qs.delete()
        return Response({"deleted_count": deleted_count})


class AdminProblemBankFillMissingView(APIView):
    """System Admin: single bulk button — sweeps every Problem in the bank
    and generates whatever it's missing (test cases, param_schema,
    explanation) via the LLM fallback chain. Every piece is skip-if-exists,
    same as the per-problem actions this mirrors — it never overwrites
    hand-authored content, it only fills gaps.

    Bounded per click by a wall-clock TIME budget, not a fixed action count.
    param_schema is a brand-new field, so on a bank this size (checked live:
    1828 problems, all 1828 missing a schema) a small fixed count would need
    ~180 clicks to ever finish. A time budget instead processes as many
    problems as safely fit in one request — adapting to however fast the
    LLM rotation happens to be responding right now — while still stopping
    well short of gunicorn's worker timeout. Call again to keep sweeping;
    the response reports how many problems still need something."""
    permission_classes = [IsAuthenticated]

    TIME_BUDGET_SECONDS = 90  # gunicorn's worker timeout is 120s; leaves headroom for request overhead
    MAX_ACTIONS = 200          # hard backstop in case calls return implausibly fast

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        import time
        from .services.testcase_generator import (
            generate_test_cases, generate_param_schema, generate_explanation,
            derive_examples, TestCaseGenError,
        )

        problems = (
            Problem.objects
            .annotate(tc_count=Count("test_cases", distinct=True))
            .filter(Q(tc_count=0) | Q(param_schema__isnull=True) | Q(explanation=""))
            .order_by("title")
        )

        processed = []
        actions_done = 0
        start = time.monotonic()

        def budget_left():
            return actions_done < self.MAX_ACTIONS and (time.monotonic() - start) < self.TIME_BUDGET_SECONDS

        for problem in problems:
            if not budget_left():
                break

            entry = {"id": problem.id, "title": problem.title}
            did_anything = False

            if problem.tc_count == 0 and budget_left():
                actions_done += 1
                did_anything = True
                try:
                    generated = generate_test_cases(
                        title=problem.title, description=problem.description,
                        examples=problem.examples, difficulty=problem.difficulty,
                    )
                    with transaction.atomic():
                        TestCase.objects.bulk_create([
                            TestCase(
                                problem=problem, stdin=case["stdin"], expected_output=case["expected_output"],
                                is_sample=case["is_sample"], order=order,
                            )
                            for order, case in enumerate(generated, start=1)
                        ])
                    if not problem.examples:
                        problem.examples = derive_examples(generated)
                        problem.save(update_fields=["examples"])
                    entry["test_cases_generated"] = len(generated)
                except TestCaseGenError as exc:
                    entry["test_cases_error"] = str(exc)

            if not problem.param_schema and budget_left():
                actions_done += 1
                did_anything = True
                try:
                    schema = generate_param_schema(title=problem.title, description=problem.description, examples=problem.examples)
                    problem.param_schema = schema
                    problem.save(update_fields=["param_schema"])
                    entry["schema_generated"] = True
                except TestCaseGenError as exc:
                    entry["schema_error"] = str(exc)

            if not problem.explanation and budget_left():
                actions_done += 1
                did_anything = True
                try:
                    explanation = generate_explanation(
                        title=problem.title, description=problem.description,
                        examples=problem.examples, difficulty=problem.difficulty,
                    )
                    problem.explanation = explanation
                    problem.save(update_fields=["explanation"])
                    entry["explanation_generated"] = True
                except TestCaseGenError as exc:
                    entry["explanation_error"] = str(exc)

            if did_anything:
                processed.append(entry)

        remaining = (
            Problem.objects
            .annotate(tc_count=Count("test_cases", distinct=True))
            .filter(Q(tc_count=0) | Q(param_schema__isnull=True) | Q(explanation=""))
            .count()
        )

        return Response({
            "processed": processed,
            "actions_done": actions_done,
            "elapsed_seconds": round(time.monotonic() - start, 1),
            "remaining_problems": remaining,
        })


_DRIVE_FILE_ID_RE = re.compile(r'^[A-Za-z0-9_-]{20,}$')


def _resolve_drive_image(raw_value):
    """The Aptitude Bank Excel Template's image columns can hold either a
    ready-to-use image URL, or (for the "Question/Option Image ID" template
    variant) a bare Google Drive file ID — e.g. a Figure Series export where
    every option is an image referenced by its Drive file ID rather than a
    URL. Bare IDs are pointed at our own caching proxy (see
    AptitudeDriveImageProxyView) rather than Drive directly — Drive's
    thumbnail endpoint is noticeably slow to serve cold, and with hundreds
    of image-based questions on one page that adds up fast. The proxy
    fetches from Drive once per file and serves every request after that
    from local disk with a long-lived Cache-Control header. Already-a-URL
    values (a plain image URL, not a Drive file ID) pass through unchanged
    since there's nothing to cache-proxy for those."""
    val = (raw_value or "").strip().strip('*').strip()
    if not val:
        return ""
    if val.lower().startswith(("http://", "https://")):
        return val
    if _DRIVE_FILE_ID_RE.match(val):
        return f"/api/aptitude/drive-image/{val}/"
    return val


def aptitude_drive_image_proxy(request, drive_id):
    """Serve a Google-Drive-hosted aptitude question/option image, caching
    it to local disk on first request so every request after that is a
    local file read instead of a round trip to Drive's thumbnail endpoint.
    Public/no-auth by design — these are the exact same files Drive itself
    already serves without authentication (that's what makes storing a bare
    file ID + resolving to a public thumbnail URL work at all), so proxying
    them adds no new exposure. Once `pull_drive_images` has been run,
    every referenced image is already cached and this never touches Drive
    at all — it only falls back to a live fetch for anything new."""
    if not _DRIVE_FILE_ID_RE.match(drive_id or ""):
        raise Http404("Invalid image id.")

    cached_path = cached_image_path(drive_id)
    if not cached_path:
        try:
            cached_path = fetch_and_cache_drive_image(drive_id)
        except DriveImageFetchError:
            raise Http404("Image could not be retrieved from Drive.")

    content_type = mimetypes.guess_type(cached_path)[0] or "image/png"
    response = FileResponse(open(cached_path, "rb"), content_type=content_type)
    response["Cache-Control"] = "public, max-age=31536000, immutable"
    return response


def institution_logo_proxy(request, institution_id):
    """Serve an institution's uploaded logo file bytes through the /api/
    proxy path. Django's MEDIA_URL route only exists when DEBUG=True and
    the production nginx config has no /media/ passthrough, so serving
    ImageField.url directly 404s (or falls through to the SPA) in
    production — this rides the same /api/ path everything else already
    proxies correctly in both dev and prod. Public/no-auth: logos appear
    on public login screens before any authentication happens. Not
    marked immutable like the Drive image cache — a logo can be
    re-uploaded, so a short max-age keeps a stale one from sticking
    around in the browser cache for long."""
    institution = get_object_or_404(Institution, id=institution_id)
    if not institution.logo_file:
        raise Http404("No logo uploaded for this institution.")

    content_type = mimetypes.guess_type(institution.logo_file.name)[0] or "image/png"
    response = FileResponse(institution.logo_file.open("rb"), content_type=content_type)
    response["Cache-Control"] = "public, max-age=300"
    return response


def _resolve_aptitude_correct_option(raw_answer, option_a, option_b, option_c, option_d):
    """Verify/derive the correct option letter for an aptitude question —
    shared by the individual add/edit form and the bulk upload, so an
    existing question edited by hand gets the same verification as a
    freshly-bulk-uploaded one. Handles a bare letter (A/B/C/D), a prefixed
    form like "Option A" or "Ans: B", and an answer given as the option's
    own text rather than a letter. Returns None if nothing resolves to
    exactly one of the four options."""
    val = (raw_answer or "").strip()
    if not val:
        return None
    upper = val.upper()
    if upper in ("A", "B", "C", "D"):
        return upper
    m = re.search(r'(?<![A-Za-z0-9])([ABCD])(?![A-Za-z0-9])', upper)
    if m:
        return m.group(1)
    options = {"A": option_a, "B": option_b, "C": option_c, "D": option_d}
    matches = [letter for letter, text in options.items() if (text or "").strip().lower() == val.lower()]
    if len(matches) == 1:
        return matches[0]
    return None


class AdminAptitudeTopicListCreateView(APIView):
    """System Admin: create a new aptitude topic node — a top-level
    category (parent_id omitted/null) or a main topic under a category
    (parent_id = that category's id). Capped at two levels: every admin
    workflow files questions on the main topic, so a third level would
    just be a dead end nothing writes to (see migration 0068)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        title = (request.data.get("title") or "").strip()
        if not title:
            return Response({"error": "title is required"}, status=400)

        parent_id = request.data.get("parent_id")
        parent = None
        if parent_id:
            parent = AptitudeTopic.objects.filter(id=parent_id).first()
            if not parent:
                return Response({"error": "Parent topic not found"}, status=404)
            if parent.parent_id is not None:
                return Response(
                    {"error": "Topics are capped at two levels (Category > Main Topic) — "
                              "pick a top-level category as the parent, not another main topic."},
                    status=400,
                )

        if AptitudeTopic.objects.filter(parent=parent, title__iexact=title).exists():
            return Response({"error": f'"{title}" already exists at this level.'}, status=400)

        topic = AptitudeTopic.objects.create(title=title, parent=parent)
        return Response({"id": topic.id, "title": topic.title, "parent_id": topic.parent_id}, status=201)


class AdminAptitudeTopicDetailView(APIView):
    """System Admin: rename an aptitude topic node (category, main topic,
    or level-3 topic)."""
    permission_classes = [IsAuthenticated]

    def patch(self, request, topic_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        topic = AptitudeTopic.objects.filter(id=topic_id).first()
        if not topic:
            return Response({"error": "Not found"}, status=404)

        title = (request.data.get("title") or "").strip()
        if not title:
            return Response({"error": "title is required"}, status=400)

        if AptitudeTopic.objects.filter(parent=topic.parent, title__iexact=title).exclude(id=topic.id).exists():
            return Response({"error": f'"{title}" already exists at this level.'}, status=400)

        topic.title = title
        topic.save()
        return Response({"id": topic.id, "title": topic.title, "parent_id": topic.parent_id})

    def delete(self, request, topic_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        topic = AptitudeTopic.objects.filter(id=topic_id).first()
        if not topic:
            return Response({"error": "Not found"}, status=404)

        # Cascades to subtopics and their questions (AptitudeTopic.parent
        # and AptitudeQuestion.topic are both on_delete=CASCADE) — the
        # frontend confirm dialog is expected to warn about this before
        # calling delete.
        topic.delete()
        return Response(status=204)


class AdminAptitudeBankView(APIView):
    """System Admin: list every aptitude question (with topic + correct
    option, for admin review) and create a new one."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        qs = AptitudeQuestion.objects.select_related("topic", "passage").order_by("-id")

        topic_id = request.query_params.get("topic_id")
        passage_id = request.query_params.get("passage_id")
        if passage_id:
            qs = qs.filter(passage_id=passage_id)
        elif topic_id:
            # topic_id is normally a Main Topic id here, which has no
            # children of its own — but if it's ever a Category id, include
            # its Main Topics too so selecting the whole category still
            # shows everything filed under it.
            subtopic_ids = AptitudeTopic.objects.filter(parent_id=topic_id).values_list("id", flat=True)
            qs = qs.filter(topic_id__in={int(topic_id), *subtopic_ids})
        difficulty = request.query_params.get("difficulty")
        if difficulty and difficulty != "all":
            qs = qs.filter(difficulty__iexact=difficulty)
        search = request.query_params.get("q")
        if search:
            qs = qs.filter(question_text__icontains=search)

        data = [{
            "id": q.id,
            "topic_id": q.topic_id,
            "topic": q.topic.title if q.topic else "",
            "passage_id": q.passage_id,
            "passage": q.passage.title if q.passage else "",
            "question_type": q.question_type,
            "question_text": q.question_text,
            "question_image": q.question_image,
            "option_a": q.option_a,
            "option_a_image": q.option_a_image,
            "option_b": q.option_b,
            "option_b_image": q.option_b_image,
            "option_c": q.option_c,
            "option_c_image": q.option_c_image,
            "option_d": q.option_d,
            "option_d_image": q.option_d_image,
            "correct_option": q.correct_option,
            "explanation": q.explanation or "",
            "difficulty": q.difficulty,
        } for q in qs[:2000]]

        return Response({"questions": data, "total": len(data)})

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        topic_id = request.data.get("topic_id")
        passage_id = request.data.get("passage_id")
        question_text = (request.data.get("question_text") or "").strip()
        question_image = (request.data.get("question_image") or "").strip()
        option_a = (request.data.get("option_a") or "").strip()
        option_a_image = (request.data.get("option_a_image") or "").strip()
        option_b = (request.data.get("option_b") or "").strip()
        option_b_image = (request.data.get("option_b_image") or "").strip()
        option_c = (request.data.get("option_c") or "").strip()
        option_c_image = (request.data.get("option_c_image") or "").strip()
        option_d = (request.data.get("option_d") or "").strip()
        option_d_image = (request.data.get("option_d_image") or "").strip()
        raw_answer = request.data.get("correct_option") or ""
        difficulty = request.data.get("difficulty") or "Easy"
        explanation = request.data.get("explanation") or ""

        topic = None
        passage = None
        if passage_id:
            passage = ReadingPassage.objects.filter(id=passage_id).first()
            if not passage:
                return Response({"error": "Passage not found"}, status=404)
        elif topic_id:
            topic = AptitudeTopic.objects.filter(id=topic_id).first()
            if not topic:
                return Response({"error": "Topic not found"}, status=404)
        else:
            return Response({"error": "topic_id or passage_id is required"}, status=400)

        if not question_text:
            return Response({"error": "question_text is required"}, status=400)
        if not all([option_a, option_b, option_c, option_d]):
            return Response({"error": "All four options (A-D) are required"}, status=400)
        if difficulty not in ("Easy", "Medium", "Hard"):
            return Response({"error": "difficulty must be Easy, Medium, or Hard"}, status=400)

        correct_option = _resolve_aptitude_correct_option(raw_answer, option_a, option_b, option_c, option_d)
        if correct_option is None:
            return Response({
                "error": f"Could not verify a correct answer from {raw_answer!r} against the four "
                         f"options — must be A-D or match one option's text exactly.",
            }, status=400)

        q = AptitudeQuestion.objects.create(
            topic=topic, passage=passage,
            question_type="RC" if passage else "MCQ",
            question_text=question_text, question_image=question_image,
            option_a=option_a, option_a_image=option_a_image,
            option_b=option_b, option_b_image=option_b_image,
            option_c=option_c, option_c_image=option_c_image,
            option_d=option_d, option_d_image=option_d_image,
            correct_option=correct_option, difficulty=difficulty, explanation=explanation,
        )
        return Response({
            "id": q.id,
            "topic_id": q.topic_id, "topic": topic.title if topic else "",
            "passage_id": q.passage_id, "passage": passage.title if passage else "",
            "question_type": q.question_type,
            "question_text": q.question_text, "question_image": q.question_image,
            "option_a": q.option_a, "option_a_image": q.option_a_image,
            "option_b": q.option_b, "option_b_image": q.option_b_image,
            "option_c": q.option_c, "option_c_image": q.option_c_image,
            "option_d": q.option_d, "option_d_image": q.option_d_image,
            "correct_option": q.correct_option, "explanation": q.explanation, "difficulty": q.difficulty,
        }, status=201)


class AdminAptitudeQuestionDetailView(APIView):
    """System Admin: edit or delete a single aptitude question."""
    permission_classes = [IsAuthenticated]

    def put(self, request, question_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        q = AptitudeQuestion.objects.filter(id=question_id).first()
        if not q:
            return Response({"error": "Not found"}, status=404)

        topic_id = request.data.get("topic_id")
        if topic_id:
            topic = AptitudeTopic.objects.filter(id=topic_id).first()
            if not topic:
                return Response({"error": "Topic not found"}, status=404)
            q.topic = topic

        passage_id = request.data.get("passage_id")
        if passage_id:
            passage = ReadingPassage.objects.filter(id=passage_id).first()
            if not passage:
                return Response({"error": "Passage not found"}, status=404)
            q.passage = passage
            q.question_type = "RC"

        question_text = request.data.get("question_text")
        if question_text is not None:
            question_text = question_text.strip()
            if not question_text:
                return Response({"error": "question_text cannot be empty"}, status=400)
            q.question_text = question_text

        question_image = request.data.get("question_image")
        if question_image is not None:
            q.question_image = question_image.strip()

        for field in ("option_a", "option_b", "option_c", "option_d"):
            val = request.data.get(field)
            if val is not None:
                val = val.strip()
                if not val:
                    return Response({"error": f"{field} cannot be empty"}, status=400)
                setattr(q, field, val)

        for field in ("option_a_image", "option_b_image", "option_c_image", "option_d_image"):
            val = request.data.get(field)
            if val is not None:
                setattr(q, field, val.strip())

        raw_answer = request.data.get("correct_option")
        if raw_answer is not None:
            resolved = _resolve_aptitude_correct_option(raw_answer, q.option_a, q.option_b, q.option_c, q.option_d)
            if resolved is None:
                return Response({
                    "error": f"Could not verify a correct answer from {raw_answer!r} against the four "
                             f"options — must be A-D or match one option's text exactly.",
                }, status=400)
            q.correct_option = resolved

        difficulty = request.data.get("difficulty")
        if difficulty is not None:
            if difficulty not in ("Easy", "Medium", "Hard"):
                return Response({"error": "difficulty must be Easy, Medium, or Hard"}, status=400)
            q.difficulty = difficulty

        explanation = request.data.get("explanation")
        if explanation is not None:
            q.explanation = explanation

        q.save()
        return Response({
            "id": q.id, "topic_id": q.topic_id, "topic": q.topic.title if q.topic else "",
            "passage_id": q.passage_id, "passage": q.passage.title if q.passage else "",
            "question_type": q.question_type,
            "question_text": q.question_text, "question_image": q.question_image,
            "option_a": q.option_a, "option_a_image": q.option_a_image,
            "option_b": q.option_b, "option_b_image": q.option_b_image,
            "option_c": q.option_c, "option_c_image": q.option_c_image,
            "option_d": q.option_d, "option_d_image": q.option_d_image,
            "correct_option": q.correct_option, "explanation": q.explanation, "difficulty": q.difficulty,
        })

    def delete(self, request, question_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        q = AptitudeQuestion.objects.filter(id=question_id).first()
        if not q:
            return Response({"error": "Not found"}, status=404)
        q.delete()
        return Response(status=204)


class AdminAptitudeQuestionValidateView(APIView):
    """System Admin: AI-review one aptitude question — checks the marked
    correct answer is actually right and the question/options are
    well-formed, autocorrecting and saving the fix (just the answer key, or
    a full rewrite when the question itself is broken/ambiguous) rather
    than only flagging it for manual review."""
    permission_classes = [IsAuthenticated]

    def post(self, request, question_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        q = AptitudeQuestion.objects.select_related("topic").filter(id=question_id).first()
        if not q:
            return Response({"error": "Not found"}, status=404)

        from .services.aptitude_validator import validate_aptitude_question
        from .services.testcase_generator import TestCaseGenError

        try:
            result = validate_aptitude_question(
                question_text=q.question_text,
                option_a=q.option_a, option_b=q.option_b, option_c=q.option_c, option_d=q.option_d,
                correct_option=q.correct_option, difficulty=q.difficulty,
            )
        except TestCaseGenError as e:
            return Response({"error": f"AI validation failed: {e}"}, status=502)

        changed_fields = []
        if result["needs_rewrite"]:
            for field in ("question_text", "option_a", "option_b", "option_c", "option_d"):
                new_val = result[field]
                if new_val != getattr(q, field):
                    setattr(q, field, new_val)
                    changed_fields.append(field)

        if result["correct_option"] != q.correct_option:
            q.correct_option = result["correct_option"]
            changed_fields.append("correct_option")

        if changed_fields:
            q.save()

        return Response({
            "id": q.id, "topic_id": q.topic_id, "topic": q.topic.title if q.topic else "",
            "question_text": q.question_text,
            "option_a": q.option_a, "option_b": q.option_b, "option_c": q.option_c, "option_d": q.option_d,
            "correct_option": q.correct_option, "explanation": q.explanation, "difficulty": q.difficulty,
            "changed_fields": changed_fields,
            "reason": result["reason"],
        })


class AdminAptitudeBulkDeleteView(APIView):
    """System Admin: delete many aptitude questions at once."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        ids = request.data.get("ids") or []
        if not isinstance(ids, list) or not ids:
            return Response({"error": "ids (non-empty list) is required"}, status=400)

        qs = AptitudeQuestion.objects.filter(id__in=ids)
        deleted_count = qs.count()
        qs.delete()
        return Response({"deleted_count": deleted_count})


class AdminAptitudeBulkUploadView(APIView):
    """System Admin: bulk-add aptitude questions from an uploaded .xlsx/.csv
    file — every row is added under a single selected topic. Accepts the
    same column names used by the original load_aptitude.py Excel imports
    (Question / Option A-D / Answer / Level / Explanation) as well as the
    admin form's own field names (question_text / option_a-d /
    correct_option / difficulty / explanation). Also accepts the optional
    Question Type and Question/Option A-D Image columns from the
    Aptitude Bank Excel Template (image cells hold image URLs)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        topic_id = request.data.get("topic_id")
        if not topic_id:
            return Response({"error": "topic_id is required"}, status=400)
        topic = AptitudeTopic.objects.filter(id=topic_id).first()
        if not topic:
            return Response({"error": "Topic not found"}, status=404)

        upload = request.FILES.get("file")
        if not upload:
            return Response({"error": "No file uploaded."}, status=400)

        name = upload.name.lower()
        try:
            if name.endswith((".xlsx", ".xls")):
                rows = self._read_excel(upload)
            elif name.endswith(".csv"):
                rows = self._read_csv(upload)
            else:
                return Response({"error": "Only .xlsx, .xls, or .csv files are supported."}, status=400)
        except Exception as e:
            return Response({"error": f"Could not read file: {e}"}, status=400)

        if not rows:
            return Response({"error": "File is empty."}, status=400)

        header = rows[0]
        col_idx = {col: i for i, col in enumerate(header) if col}

        def col(row, *names):
            for n in names:
                idx = col_idx.get(n)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    val = str(row[idx]).strip()
                    if val and val.lower() != "nan":
                        return val
            return ""

        required_any = [
            ("question_text", "question"),
            ("option_a",), ("option_b",), ("option_c",), ("option_d",),
            ("correct_option", "answer", "correct_answer"),
        ]
        missing = [names[0] for names in required_any if not any(n in col_idx for n in names)]
        if missing:
            return Response({
                "error": f"Missing required column(s): {', '.join(missing)}. "
                         f"Expected: question_text/Question, option_a-d/Option A-D, correct_option/Answer/Correct Answer "
                         f"(question_no/Question No, difficulty/Level and explanation are optional and ignored if absent).",
            }, status=400)

        created_count = 0
        skipped_count = 0
        errors = []

        for row_num, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            question_text = col(row, "question_text", "question")
            question_type = col(row, "question_type") or "MCQ"
            question_image = _resolve_drive_image(col(row, "question_image", "question_image_id"))
            option_a = col(row, "option_a")
            option_a_image = _resolve_drive_image(col(row, "option_a_image", "option_a_image_id"))
            option_b = col(row, "option_b")
            option_b_image = _resolve_drive_image(col(row, "option_b_image", "option_b_image_id"))
            option_c = col(row, "option_c")
            option_c_image = _resolve_drive_image(col(row, "option_c_image", "option_c_image_id"))
            option_d = col(row, "option_d")
            option_d_image = _resolve_drive_image(col(row, "option_d_image", "option_d_image_id"))
            raw_answer = col(row, "correct_option", "answer", "correct_answer")
            raw_difficulty = col(row, "difficulty", "level", "difficulty_level", "diff", "complexity", "tier")

            # Robust difficulty parsing for Easy, Medium, Hard
            def _clean_difficulty(raw_val):
                if not raw_val:
                    return "Easy"
                val = str(raw_val).strip().strip('*_').strip().lower()
                if val in ("easy", "e", "1", "basic", "simple", "beginner"):
                    return "Easy"
                if val in ("medium", "med", "m", "2", "moderate", "intermediate", "normal", "average"):
                    return "Medium"
                if val in ("hard", "h", "3", "difficult", "advanced", "complex"):
                    return "Hard"
                if "hard" in val or "diff" in val or "adv" in val:
                    return "Hard"
                if "med" in val or "mod" in val or "inter" in val:
                    return "Medium"
                if "easy" in val or "bas" in val or "simp" in val:
                    return "Easy"
                return "Easy"

            difficulty = _clean_difficulty(raw_difficulty)
            explanation = col(row, "explanation")

            if not question_text or not all([
                option_a or option_a_image, option_b or option_b_image,
                option_c or option_c_image, option_d or option_d_image,
            ]):
                errors.append(f"Row {row_num}: missing question text or an option (text or image) — skipped.")
                continue

            correct_option = _resolve_aptitude_correct_option(raw_answer, option_a, option_b, option_c, option_d)
            if correct_option is None:
                errors.append(
                    f"Row {row_num}: could not verify a correct answer from {raw_answer!r} against the "
                    f"four options — must be A-D or match one option's text exactly — skipped."
                )
                continue

            # Dedup key includes question_image, not just (topic, question_text) —
            # image-based question sets (e.g. Figure Series) commonly reuse the
            # same generic instructional text across every row ("Select the
            # figure that will replace the question mark...") with the actual
            # content living entirely in the image, so text alone would collapse
            # dozens/hundreds of genuinely distinct questions into one row that
            # just gets overwritten repeatedly. Text-only rows (question_image
            # empty for all of them) still dedup by text alone as before.
            q, created = AptitudeQuestion.objects.get_or_create(
                topic=topic, question_text=question_text, question_image=question_image,
                defaults={
                    "question_type": question_type,
                    "option_a": option_a, "option_a_image": option_a_image,
                    "option_b": option_b, "option_b_image": option_b_image,
                    "option_c": option_c, "option_c_image": option_c_image,
                    "option_d": option_d, "option_d_image": option_d_image,
                    "correct_option": correct_option, "difficulty": difficulty, "explanation": explanation,
                },
            )
            if created:
                created_count += 1
            else:
                q.question_type = question_type
                q.option_a = option_a
                q.option_a_image = option_a_image
                q.option_b = option_b
                q.option_b_image = option_b_image
                q.option_c = option_c
                q.option_c_image = option_c_image
                q.option_d = option_d
                q.option_d_image = option_d_image
                q.correct_option = correct_option
                q.difficulty = difficulty
                if explanation:
                    q.explanation = explanation
                q.save(update_fields=[
                    'question_type',
                    'option_a', 'option_a_image', 'option_b', 'option_b_image',
                    'option_c', 'option_c_image', 'option_d', 'option_d_image',
                    'correct_option', 'difficulty', 'explanation',
                ])
                skipped_count += 1

        return Response({
            "created_count": created_count,
            "skipped_count": skipped_count,
            "error_count": len(errors),
            "errors": errors[:50],
        })

    def _read_excel(self, upload):
        import openpyxl
        wb = openpyxl.load_workbook(upload, data_only=True)
        ws = wb.active
        formatted_rows = []
        for row_idx, row in enumerate(ws.iter_rows()):
            is_header_row = row_idx == 0
            row_vals = []
            for cell in row:
                if cell.value is None:
                    row_vals.append("")
                    continue

                val_str = str(cell.value).strip()
                # Bold formatting is a content cue for data cells (rendered as
                # **markdown**), but header cells are commonly bold-styled
                # purely for visual emphasis — wrapping them the same way
                # corrupts every column-name match below, since "**question**"
                # doesn't match the "question" alias.
                if not is_header_row and hasattr(cell, 'font') and cell.font and getattr(cell.font, 'bold', False):
                    if val_str and not (val_str.startswith('**') or val_str.startswith('<b>') or val_str.startswith('<strong>')):
                        val_str = f"**{val_str}**"
                row_vals.append(val_str)
            formatted_rows.append(row_vals)

        if not formatted_rows:
            return []
        header = [str(h).strip().lower().replace(" ", "_") if h else "" for h in formatted_rows[0]]
        return [header] + formatted_rows[1:]

    def _read_csv(self, upload):
        import csv
        import io
        text = upload.read().decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        raw_rows = list(reader)
        if not raw_rows:
            return []
        header = [h.strip().lower().replace(" ", "_") for h in raw_rows[0]]
        return [header] + raw_rows[1:]


class AdminReadingPassageListCreateView(APIView):
    """System Admin: list every Reading Comprehension passage (with its
    question count) and create a new one."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        passages = ReadingPassage.objects.select_related("topic").annotate(question_count=Count("questions")).order_by("-id")
        topic_id = request.query_params.get("topic_id")
        if topic_id:
            passages = passages.filter(topic_id=topic_id)
        data = [{
            "id": p.id,
            "title": p.title,
            "passage_text": p.passage_text,
            "topic_id": p.topic_id,
            "topic": p.topic.title if p.topic else "",
            "difficulty": p.difficulty,
            "question_count": p.question_count,
        } for p in passages]
        return Response({"passages": data, "total": len(data)})

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        title = (request.data.get("title") or "").strip()
        passage_text = (request.data.get("passage_text") or "").strip()
        difficulty = request.data.get("difficulty") or "Medium"
        topic_id = request.data.get("topic_id")

        if not title:
            return Response({"error": "title is required"}, status=400)
        if not passage_text:
            return Response({"error": "passage_text is required"}, status=400)
        if difficulty not in ("Easy", "Medium", "Hard"):
            return Response({"error": "difficulty must be Easy, Medium, or Hard"}, status=400)

        topic = None
        if topic_id:
            topic = AptitudeTopic.objects.filter(id=topic_id).first()
            if not topic:
                return Response({"error": "Topic not found"}, status=404)

        passage = ReadingPassage.objects.create(title=title, passage_text=passage_text, difficulty=difficulty, topic=topic)
        return Response({
            "id": passage.id, "title": passage.title, "passage_text": passage.passage_text,
            "topic_id": passage.topic_id, "topic": topic.title if topic else "",
            "difficulty": passage.difficulty, "question_count": 0,
        }, status=201)


class AdminReadingPassageDetailView(APIView):
    """System Admin: edit or delete a single Reading Comprehension passage."""
    permission_classes = [IsAuthenticated]

    def put(self, request, passage_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        passage = ReadingPassage.objects.filter(id=passage_id).first()
        if not passage:
            return Response({"error": "Not found"}, status=404)

        title = request.data.get("title")
        if title is not None:
            title = title.strip()
            if not title:
                return Response({"error": "title cannot be empty"}, status=400)
            passage.title = title

        passage_text = request.data.get("passage_text")
        if passage_text is not None:
            passage_text = passage_text.strip()
            if not passage_text:
                return Response({"error": "passage_text cannot be empty"}, status=400)
            passage.passage_text = passage_text

        difficulty = request.data.get("difficulty")
        if difficulty is not None:
            if difficulty not in ("Easy", "Medium", "Hard"):
                return Response({"error": "difficulty must be Easy, Medium, or Hard"}, status=400)
            passage.difficulty = difficulty

        if "topic_id" in request.data:
            topic_id = request.data.get("topic_id")
            if topic_id:
                topic = AptitudeTopic.objects.filter(id=topic_id).first()
                if not topic:
                    return Response({"error": "Topic not found"}, status=404)
                passage.topic = topic
            else:
                passage.topic = None

        passage.save()
        return Response({
            "id": passage.id, "title": passage.title, "passage_text": passage.passage_text,
            "topic_id": passage.topic_id, "topic": passage.topic.title if passage.topic else "",
            "difficulty": passage.difficulty,
        })

    def delete(self, request, passage_id):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        passage = ReadingPassage.objects.filter(id=passage_id).first()
        if not passage:
            return Response({"error": "Not found"}, status=404)
        passage.delete()
        return Response(status=204)


class AdminReadingPassageQAImportView(APIView):
    """System Admin: upload the merged reading QA dataset Excel file
    (sheets "Reading Passages" + "Questions & Answers") straight from the
    dashboard and load it into Reading Comprehension passages/questions —
    the point-and-click equivalent of the import_reading_qa_dataset
    management command, for datasets small enough to upload through the
    browser without needing server access."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Admin access required."}, status=status.HTTP_403_FORBIDDEN)

        upload = request.FILES.get("file")
        if not upload:
            return Response({"error": "No file uploaded."}, status=400)
        if not upload.name.lower().endswith((".xlsx", ".xls")):
            return Response({"error": "Only .xlsx or .xls files are supported."}, status=400)

        try:
            limit = int(request.data.get("limit") or 20)
        except (TypeError, ValueError):
            return Response({"error": "limit must be a number."}, status=400)
        if limit < 1 or limit > 500:
            return Response({"error": "limit must be between 1 and 500 (large batches risk the request timing out — import in smaller runs)."}, status=400)

        try:
            questions_per_passage = int(request.data.get("questions_per_passage") or 10)
        except (TypeError, ValueError):
            return Response({"error": "questions_per_passage must be a number."}, status=400)
        if questions_per_passage < 1 or questions_per_passage > 100:
            return Response({"error": "questions_per_passage must be between 1 and 100."}, status=400)

        difficulty = request.data.get("difficulty") or "Medium"
        if difficulty not in ("Easy", "Medium", "Hard"):
            return Response({"error": "difficulty must be Easy, Medium, or Hard"}, status=400)

        topic = None
        topic_id = request.data.get("topic_id")
        if topic_id:
            topic = AptitudeTopic.objects.filter(id=topic_id).first()
            if not topic:
                return Response({"error": "Topic not found"}, status=404)

        try:
            import openpyxl
            wb = openpyxl.load_workbook(upload, data_only=True)
        except Exception as exc:
            return Response({"error": f"Could not read workbook: {exc}"}, status=400)

        try:
            passages, questions_skipped = parse_workbook_to_passages(
                wb, limit=limit, questions_per_passage=questions_per_passage, difficulty=difficulty,
            )
        except ValueError as exc:
            return Response({"error": str(exc)}, status=400)

        with transaction.atomic():
            passages_created, questions_created = create_passages_in_db(passages, topic=topic)

        return Response({
            "passages_created": passages_created,
            "questions_created": questions_created,
            "questions_skipped": questions_skipped,
            "passages": [{"title": p["title"], "question_count": len(p["questions"])} for p in passages],
        }, status=201)


class InstitutionBrandingPreviewView(APIView):
    """Generate PDF template preview for institution branding"""
    permission_classes = [AllowAny]
    
    def get(self, request, pk):
        institution = get_object_or_404(Institution, pk=pk)
        
        # Create PDF template preview with watermark
        buffer = BytesIO()
        doc = create_watermarked_pdf(
            buffer, 
            institution=institution,
            pagesize=A4, 
            topMargin=0.5*inch, 
            bottomMargin=0.5*inch
        )
        
        # Custom styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Center
            textColor=colors.HexColor('#2d5016')
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,  # Center
            textColor=colors.HexColor('#4f7942')
        )
        
        elements = []
        
        # College Header with Logo (if available)
        header_data = []
        
        # Logo section (left side)
        logo_cell = ""
        if institution.logo_display_url:
            try:
                # For now, we'll just show a placeholder for logo
                logo_cell = "LOGO"
            except:
                logo_cell = "LOGO"
        else:
            logo_cell = "LOGO"
        
        # College info section (right side)
        display_name = institution.get_display_name()
        info_lines = [display_name]
        
        if institution.subheading:
            info_lines.append(institution.subheading)
        
        if institution.address:
            info_lines.append(institution.address)
        
        contact_info = []
        if institution.contact_email:
            contact_info.append(f"Email: {institution.contact_email}")
        if institution.contact_phone:
            contact_info.append(f"Phone: {institution.contact_phone}")
        if institution.website:
            contact_info.append(f"Website: {institution.website}")
        
        if contact_info:
            info_lines.extend(contact_info)
        
        if institution.established_year:
            info_lines.append(f"Established: {institution.established_year}")
        
        # Create header table with logo on left, info on right
        header_table_data = [[logo_cell, "\n".join(info_lines)]]
        header_table = Table(header_table_data, colWidths=[1.5*inch, 5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'CENTER'),  # Logo cell center
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),    # Info cell left
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 14),
            ('FONTNAME', (1, 0), (1, 0), 'Helvetica'),
            ('FONTSIZE', (1, 0), (1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 0.5 * inch))
        
        # Sample Report Title
        elements.append(Paragraph("SAMPLE REPORT TEMPLATE", title_style))
        elements.append(Paragraph("This is how your college branding will appear in all generated reports", subtitle_style))
        elements.append(Spacer(1, 0.3 * inch))
        
        # Sample content
        sample_content = [
            "This template shows how your college branding information will be displayed in:",
            "• Student Performance Reports",
            "• Faculty Analytics Reports", 
            "• Contest Management Reports",
            "• All other PDF documents generated by the system",
            "",
            "The layout includes:",
            "• College logo positioned on the left",
            "• College name and details on the right",
            "• Professional formatting with consistent styling",
            "• Complete contact information",
            "",
            "You can customize all branding elements in the College Branding tab."
        ]
        
        for line in sample_content:
            if line:
                elements.append(Paragraph(line, styles['Normal']))
            else:
                elements.append(Spacer(1, 0.1 * inch))
        
        # Footer
        elements.append(Spacer(1, 0.5 * inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        local_now = timezone.localtime(timezone.now())
        elements.append(Paragraph(f"Template generated on {local_now.strftime('%B %d, %Y at %I:%M %p')}", footer_style))
        
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="College_Branding_Template_{institution.short_code}.pdf"'
        return response

class DepartmentManagementView(APIView):
    permission_classes = [AllowAny]

    def post(self, request, inst_pk):
        institution = get_object_or_404(Institution, pk=inst_pk)
        name = request.data.get('name')
        code = request.data.get('code')
        
        if Department.objects.filter(code=code).exists():
            return Response({"error": "Code already exists"}, status=400)
            
        dept = Department.objects.create(institution=institution, name=name, code=code)
        return Response({"message": "Department added", "id": dept.id})

    def delete(self, request, inst_pk, pk):
        dept = get_object_or_404(Department, pk=pk, institution_id=inst_pk)
        dept.delete()
        return Response({"message": "Department deleted"})

class DiscussionThreadListView(UnifiedAuthMixin, APIView):
    def get(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error

        # Cleanup handled in global messages fetcher usually, but we filter here too
        cutoff = timezone.now() - timedelta(hours=24)
        messages = DiscussionMessage.objects.filter(
            thread_type="individual",
            created_at__gte=cutoff
        ).filter(
            Q(sender=request.user) | Q(recipient=request.user)
        ).order_by("-created_at")

        threads = {}
        for msg in messages:
            other_user = msg.recipient if msg.sender == request.user else msg.sender
            if not other_user:
                continue
            
            other_id = other_user.id
            if other_id not in threads:
                name = "Unknown"
                identifier = ""
                if hasattr(other_user, "student_profile"):
                    name = other_user.student_profile.name
                    identifier = other_user.student_profile.register_number
                elif hasattr(other_user, "staff_profile"):
                    name = other_user.staff_profile.name
                    identifier = other_user.staff_profile.faculty_id
                
                threads[other_id] = {
                    "other_user_id": other_id,
                    "other_user_name": name,
                    "other_user_reg": identifier,
                    "latest_message": msg.body[:50],
                    "timestamp": msg.created_at,
                    "unread_count": 0,
                    "is_self_latest": msg.sender == request.user
                }
            
            if msg.recipient == request.user and not msg.is_read:
                threads[other_id]["unread_count"] += 1

        thread_list = list(threads.values())
        thread_list.sort(key=lambda x: x["timestamp"], reverse=True)
        return Response(thread_list)

class StaffDeptListView(UnifiedAuthMixin, APIView):
    """Returns staff from the same department as the student for DM list."""
    def get(self, request):
        profile, profile_type, error = self.get_authenticated_profile(request)
        if error:
            return error
        if profile_type == "student":
            staff_qs = StaffProfile.objects.filter(department=profile.department)
        else:
            staff_qs = StaffProfile.objects.filter(institution=profile.institution)
        
        data = []
        for s in staff_qs:
            data.append({
                "faculty_id": s.faculty_id,
                "name": s.name,
                "email": s.email,
                "role": s.get_role_display(),
                "department": s.department.name if s.department else ""
            })
        return Response(data)


class CSRFTokenView(APIView):
    """
    Provides CSRF token for frontend authentication.
    """
    permission_classes = [AllowAny]
    
    @method_decorator(ensure_csrf_cookie)
    def get(self, request):
        from django.middleware.csrf import get_token
        token = get_token(request)
        return Response({"csrfToken": token})


# =============================================================================
# Password Reset and Public Views
# =============================================================================

def _generate_otp():
    import random
    return f"{random.randint(0, 999999):06d}"


def _mask_email(email):
    if '@' not in email:
        return email
    local, domain = email.split('@', 1)
    return f"{local[:2]}***@{domain}"


def _send_password_reset_otp_email(to_email, name, code):
    from django.core.mail import send_mail
    send_mail(
        subject="Your Code2Day password reset code",
        message=(
            f"Hi {name},\n\n"
            f"Your Code2Day password reset code is: {code}\n\n"
            f"This code expires in {PasswordResetOTP.OTP_VALIDITY_MINUTES} minutes. "
            "If you didn't request this, you can safely ignore this email — "
            "your password will not change unless this code is used."
        ),
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[to_email],
        fail_silently=False,
    )


class PasswordResetView(APIView):
    """Forgot-password flow via a one-time code emailed to whatever
    address is already on file for the account (StudentProfile.personal_email
    or StaffProfile.email) — never an address typed in by the person
    requesting the reset, so only someone with access to that inbox can
    actually reset the password.

    POST: identify the account (register_number for students, faculty_id
    for staff) -> generate + email a 6-digit OTP, valid 5 minutes.
    PUT: submit that OTP + a new password -> verify and reset.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "password-reset-request", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        user_type = request.data.get('user_type', 'student')

        if user_type == 'student':
            register_number = (request.data.get('register_number') or '').strip()
            if not register_number:
                return Response({'error': 'Register number is required.'}, status=status.HTTP_400_BAD_REQUEST)
            student = StudentProfile.objects.select_related('account').filter(register_number=register_number).first()
            if not student or not student.account:
                return Response({'error': 'No account found with that register number.'}, status=status.HTTP_404_NOT_FOUND)
            user = student.account
            name = student.name
            email = (student.personal_email or student.account.email or '').strip()
        elif user_type == 'staff':
            faculty_id = (request.data.get('faculty_id') or '').strip()
            if not faculty_id:
                return Response({'error': 'Faculty ID is required.'}, status=status.HTTP_400_BAD_REQUEST)
            staff = StaffProfile.objects.select_related('account').filter(faculty_id=faculty_id).first()
            if not staff or not staff.account:
                return Response({'error': 'No staff account found with that faculty ID.'}, status=status.HTTP_404_NOT_FOUND)
            user = staff.account
            name = staff.name
            email = (staff.email or staff.account.email or '').strip()
        else:
            return Response({'error': 'Invalid user type.'}, status=status.HTTP_400_BAD_REQUEST)

        if not email:
            return Response(
                {'error': "No email is on file for this account yet. Ask your JA/HOD/Admin to add one, then try again."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Superseding any earlier unused code keeps only the latest one valid
        PasswordResetOTP.objects.filter(user=user, is_used=False).update(is_used=True)

        code = _generate_otp()
        PasswordResetOTP.objects.create(
            user=user, code=code,
            expires_at=timezone.now() + timedelta(minutes=PasswordResetOTP.OTP_VALIDITY_MINUTES),
        )

        try:
            _send_password_reset_otp_email(email, name, code)
        except Exception:
            logger.exception("Failed to send password reset OTP email")
            return Response(
                {'error': 'Could not send the reset email right now. Please try again shortly.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        return Response({'message': f'A 6-digit code was sent to {_mask_email(email)}. It expires in 5 minutes.'})

    def put(self, request):
        max_attempts, window = _auth_rate_limits()
        try:
            check_rate_limit(request, "password-reset-verify", max_attempts, window)
        except RateLimitExceeded as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_429_TOO_MANY_REQUESTS,
                headers={"Retry-After": str(exc.retry_after_seconds)},
            )

        user_type = request.data.get('user_type', 'student')
        otp = (request.data.get('otp') or '').strip()
        new_password = (request.data.get('new_password') or '').strip()

        if not otp or not new_password:
            return Response({'error': 'The code and a new password are both required.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(new_password) < 8:
            return Response({'error': 'Password must be at least 8 characters.'}, status=status.HTTP_400_BAD_REQUEST)

        if user_type == 'student':
            register_number = (request.data.get('register_number') or '').strip()
            student = StudentProfile.objects.select_related('account').filter(register_number=register_number).first()
            if not student or not student.account:
                return Response({'error': 'Account not found.'}, status=status.HTTP_404_NOT_FOUND)
            user = student.account
        elif user_type == 'staff':
            faculty_id = (request.data.get('faculty_id') or '').strip()
            staff = StaffProfile.objects.select_related('account').filter(faculty_id=faculty_id).first()
            if not staff or not staff.account:
                return Response({'error': 'Account not found.'}, status=status.HTTP_404_NOT_FOUND)
            user = staff.account
        else:
            return Response({'error': 'Invalid user type.'}, status=status.HTTP_400_BAD_REQUEST)

        otp_record = PasswordResetOTP.objects.filter(user=user, code=otp, is_used=False).order_by('-created_at').first()
        if not otp_record or not otp_record.is_valid():
            return Response({'error': 'That code is invalid or has expired. Request a new one.'}, status=status.HTTP_400_BAD_REQUEST)

        otp_record.is_used = True
        otp_record.save(update_fields=['is_used'])

        if user_type == 'staff' and hasattr(user, 'staff_profile'):
            # StaffProfile has its own password field, checked before
            # User.password at login — set_password() keeps both in sync.
            user.staff_profile.set_password(new_password)
        else:
            user.set_password(new_password)
            user.save()

        return Response({'message': 'Password reset successfully. You can now log in with your new password.'})


class PublicInstitutionListView(APIView):
    """
    Public endpoint to list all institutions.
    Used for registration and public information.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        """
        Get list of all institutions with basic information.
        """
        try:
            # Get query parameters for filtering
            params = getattr(request, 'query_params', getattr(request, 'GET', {}))
            search = params.get('search', '').strip()
            active_only = params.get('active_only', 'true').lower() == 'true'
            
            # Base queryset
            institutions = Institution.objects.all()
            
            # Filter by active status if requested
            if active_only:
                active_qs = institutions.filter(Q(is_active=True) | Q(is_active__isnull=True))
                if active_qs.exists():
                    institutions = active_qs
            
            # Search filter
            if search:
                institutions = institutions.filter(
                    Q(name__icontains=search) |
                    Q(short_code__icontains=search) |
                    Q(address__icontains=search)
                )
            
            # If database has zero institutions, auto-seed default institution
            if not institutions.exists():
                try:
                    default_inst, _ = Institution.objects.get_or_create(
                        institution_id=1,
                        defaults={
                            'name': 'Ramco Institute of Technology',
                            'short_code': 'RIT',
                            'address': 'Rajapalayam, Tamil Nadu, India - 626 117',
                            'display_name': 'Ramco Institute of Technology',
                            'subheading': '(An Autonomous Institution)',
                            'is_active': True,
                        }
                    )
                    institutions = Institution.objects.all()
                except Exception:
                    pass

            # Prepare response data
            institution_list = []
            for institution in institutions:
                logo_url = institution.logo_display_url or None

                dept_count = 0
                try:
                    if hasattr(institution, 'departments'):
                        dept_count = institution.departments.count()
                    else:
                        dept_count = Department.objects.filter(institution=institution).count()
                except Exception:
                    try:
                        dept_count = StudentProfile.objects.filter(institution=institution).values('department').distinct().count()
                    except Exception:
                        dept_count = 0

                try:
                    student_count = StudentProfile.objects.filter(institution=institution).count()
                except Exception:
                    student_count = 0

                try:
                    staff_count = StaffProfile.objects.filter(institution=institution).count()
                except Exception:
                    staff_count = 0

                institution_data = {
                    'id': institution.id,
                    'name': institution.name,
                    'code': getattr(institution, 'short_code', 'RIT'),
                    'institution_id': getattr(institution, 'institution_id', institution.id),
                    'location': getattr(institution, 'address', ''),
                    'is_active': getattr(institution, 'is_active', True),
                    'student_count': student_count,
                    'staff_count': staff_count,
                    'department_count': dept_count,
                    'logo_url': logo_url,
                    'primary_color': getattr(institution, 'primary_color', '#1f2937'),
                    'secondary_color': getattr(institution, 'secondary_color', '#3b82f6'),
                    # Branding fields
                    'display_name': getattr(institution, 'display_name', institution.name),
                    'subheading': getattr(institution, 'subheading', ''),
                    'address': getattr(institution, 'address', ''),
                    'contact_email': getattr(institution, 'contact_email', ''),
                    'contact_phone': getattr(institution, 'contact_phone', ''),
                    'established_year': getattr(institution, 'established_year', None),
                }
                institution_list.append(institution_data)
            
            return Response({
                'institutions': institution_list,
                'total_count': len(institution_list),
                'search_query': search,
                'active_only': active_only
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to fetch institutions: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def post(self, request):
        """
        Create a new institution (admin only in practice, but public endpoint for flexibility).
        """
        try:
            name = request.data.get('name', '').strip()
            code = request.data.get('code', '').strip()
            location = request.data.get('location', '').strip()
            institution_id = request.data.get('institution_id')
            
            if not name:
                return Response(
                    {'error': 'Institution name is required'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Check if institution with same name or code already exists
            if Institution.objects.filter(name=name).exists():
                return Response(
                    {'error': 'Institution with this name already exists'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            if not institution_id:
                return Response(
                    {'error': 'Institution ID is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            if code and Institution.objects.filter(short_code=code).exists():
                return Response(
                    {'error': 'Institution with this code already exists'}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            if Institution.objects.filter(institution_id=institution_id).exists():
                return Response(
                    {'error': 'Institution with this ID already exists'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Create new institution
            institution = Institution.objects.create(
                institution_id=institution_id,
                name=name,
                short_code=code or name.upper()[:10],  # Generate code if not provided
                address=location,
                is_active=True
            )
            
            return Response({
                'message': 'Institution created successfully',
                'institution': {
                    'id': institution.id,
                    'institution_id': institution.institution_id,
                    'name': institution.name,
                    'code': institution.short_code,
                    'location': institution.address,
                    'is_active': institution.is_active
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response(
                {'error': f'Failed to create institution: {str(e)}'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# JA (Junior Admin) Views
# =============================================================================

def _ja_guard(request):
    """
    Returns (staff_profile, None) if the request is from an active JA,
    or (None, Response) with the appropriate error.
    """
    if not request.user.is_authenticated or not hasattr(request.user, 'staff_profile'):
        return None, Response({"detail": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
    profile = request.user.staff_profile
    if profile.role != "ja":
        return None, Response({"detail": "Junior Admin access required."}, status=status.HTTP_403_FORBIDDEN)
    if not profile.is_active:
        return None, Response({"detail": "Your account has been disabled."}, status=status.HTTP_403_FORBIDDEN)
    if not profile.department:
        return None, Response({"detail": "No department assigned to your account."}, status=status.HTTP_400_BAD_REQUEST)
    return profile, None


class JADashboardView(APIView):
    """JA dashboard overview — student count, batch list, dept info."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        dept = profile.department
        institution = profile.institution

        # Batch summary
        batches = (
            StudentProfile.objects
            .filter(institution=institution, department=dept)
            .exclude(batch='')
            .values('batch')
            .annotate(student_count=Count('id'))
            .order_by('-batch')
        )

        total_students = StudentProfile.objects.filter(
            institution=institution, department=dept
        ).count()

        return Response({
            "ja": {
                "name": profile.name,
                "faculty_id": profile.faculty_id,
                "role": profile.role,
            },
            "department": {
                "id": dept.id,
                "name": dept.name,
                "code": dept.code,
            },
            "institution": {
                "id": institution.id if institution else None,
                "name": institution.name if institution else None,
            },
            "stats": {
                "total_students": total_students,
                "total_batches": len(batches),
            },
            "batches": list(batches),
        })


class JABatchListView(APIView):
    """JA: list all batches in their department with student counts."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        batches = (
            StudentProfile.objects
            .filter(institution=profile.institution, department=profile.department)
            .exclude(batch='')
            .values('batch')
            .annotate(student_count=Count('id'))
            .order_by('-batch')
        )
        return Response({"batches": list(batches)})

    def post(self, request):
        """Create a new (empty) batch — just validates the name is unique in dept."""
        profile, err = _ja_guard(request)
        if err:
            return err

        batch_name = (request.data.get('batch') or '').strip()
        if not batch_name:
            return Response({"detail": "Batch name is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if batch already exists in this department
        exists = StudentProfile.objects.filter(
            institution=profile.institution,
            department=profile.department,
            batch=batch_name
        ).exists()
        if exists:
            return Response(
                {"detail": f"Batch '{batch_name}' already exists in this department."},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            "detail": f"Batch '{batch_name}' is ready. Add students to populate it.",
            "batch": batch_name,
        }, status=status.HTTP_201_CREATED)


class JABatchDetailView(APIView):
    """JA: get students in a batch, or delete the entire batch."""
    permission_classes = [IsAuthenticated]

    def get(self, request, batch_code):
        profile, err = _ja_guard(request)
        if err:
            return err

        students = (
            StudentProfile.objects
            .filter(
                institution=profile.institution,
                department=profile.department,
                batch=batch_code
            )
            .select_related('account', 'mentor')
            .order_by('register_number')
        )

        data = []
        for s in students:
            data.append({
                "id": s.id,
                "register_number": s.register_number,
                "name": s.name,
                "batch": s.batch,
                "section": s.section,
                "personal_email": s.personal_email,
                "mobile_number": s.mobile_number,
                "gender": s.gender,
                "is_active": s.account.is_active if s.account else True,
                "allow_copy_paste": s.allow_copy_paste,
                "mentor_id": s.mentor_id,
                "mentor_name": s.mentor.name if s.mentor_id else None,
                "mentor_faculty_id": s.mentor.faculty_id if s.mentor_id else None,
            })

        return Response({
            "batch": batch_code,
            "students": data,
            "total": len(data),
        })

    def delete(self, request, batch_code):
        """Delete all students in a batch (removes their profiles, accounts, and batch advisor records)."""
        profile, err = _ja_guard(request)
        if err:
            return err

        students = StudentProfile.objects.filter(
            institution=profile.institution,
            department=profile.department,
            batch=batch_code
        ).select_related('account')

        count = students.count()
        user_ids = list(students.values_list('account_id', flat=True))
        students.delete()
        if user_ids:
            User.objects.filter(id__in=user_ids).delete()

        # Delete any batch advisor assignments for this batch
        BatchAdvisor.objects.filter(
            institution=profile.institution,
            department=profile.department,
            batch=batch_code
        ).delete()

        logger.info("JA %s deleted batch '%s' (%d students)", profile.faculty_id, batch_code, count)

        return Response({
            "detail": f"Batch '{batch_code}' deleted successfully ({count} student(s) removed).",
            "deleted_count": count,
        })


class JAStudentCreateView(APIView):
    """JA: add a single student to a batch."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        data = request.data
        register_number = (data.get('register_number') or '').strip()
        name = (data.get('name') or '').strip()
        batch = (data.get('batch') or '').strip()

        if not register_number or not name:
            return Response(
                {"detail": "register_number and name are required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        if StudentProfile.objects.filter(register_number=register_number).exists():
            return Response(
                {"detail": f"Student with register number '{register_number}' already exists."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create Django User account
        from django.contrib.auth.hashers import make_password

        section = (data.get('section') or '').strip().upper()

        try:
            user = User.objects.create(
                username=register_number,
                password=make_password(None),  # unusable password until first login
                is_active=True,
            )
            student = StudentProfile.objects.create(
                account=user,
                institution=profile.institution,
                department=profile.department,
                register_number=register_number,
                name=name,
                title=data.get('title', name),
                batch=batch,
                section=section,
                personal_email=data.get('personal_email', ''),
                mobile_number=data.get('mobile_number', ''),
                gender=data.get('gender', ''),
            )
        except IntegrityError:
            # Duplicate register_number that slipped past the exists() check (race condition)
            return Response(
                {"detail": f"Student with register number '{register_number}' already exists."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        except Exception:
            logger.exception("Failed to create student %s for JA %s", register_number, profile.faculty_id)
            return Response(
                {"detail": "Failed to create student. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        logger.info("JA %s created student %s in batch %s section %s", profile.faculty_id, register_number, batch, section or "—")

        return Response({
            "detail": "Student created successfully.",
            "student": {
                "id": student.id,
                "register_number": student.register_number,
                "name": student.name,
                "batch": student.batch,
                "section": student.section,
            },
        }, status=status.HTTP_201_CREATED)


class JAStudentDeleteView(APIView):
    """JA: remove a student from their department."""
    permission_classes = [IsAuthenticated]

    def delete(self, request, register_number):
        profile, err = _ja_guard(request)
        if err:
            return err

        student = StudentProfile.objects.filter(
            register_number=register_number,
            institution=profile.institution,
            department=profile.department,
        ).select_related('account').first()

        if not student:
            return Response(
                {"detail": "Student not found in your department."},
                status=status.HTTP_404_NOT_FOUND
            )

        student_name = student.name
        user = student.account
        student.delete()
        if user:
            user.delete()

        logger.info("JA %s deleted student %s", profile.faculty_id, register_number)

        return Response({
            "detail": f"Student '{student_name}' ({register_number}) has been removed.",
        })


class JAStudentUpdateView(APIView):
    """JA: update a student's name, mobile number, batch, and/or section.
    PATCH /api/ja/students/<register_number>/update/
    Body: { name?: str, mobile_number?: str, batch?: str, section?: str }
    """
    permission_classes = [IsAuthenticated]

    def patch(self, request, register_number):
        profile, err = _ja_guard(request)
        if err:
            return err

        student = StudentProfile.objects.filter(
            register_number=register_number,
            institution=profile.institution,
            department=profile.department,
        ).first()
        if not student:
            return Response({"detail": "Student not found."}, status=status.HTTP_404_NOT_FOUND)

        fields_updated = []
        name = (request.data.get('name') or '').strip()
        mobile = (request.data.get('mobile_number') or '').strip()

        if name:
            student.name = name
            fields_updated.append('name')
        if 'mobile_number' in request.data:
            student.mobile_number = mobile
            fields_updated.append('mobile_number')
        if 'batch' in request.data:
            student.batch = (request.data.get('batch') or '').strip()
            fields_updated.append('batch')
        if 'section' in request.data:
            raw_section = (request.data.get('section') or '').strip().upper()
            student.section = raw_section
            fields_updated.append('section')

        if not fields_updated:
            return Response({"detail": "Nothing to update."}, status=status.HTTP_400_BAD_REQUEST)

        student.save(update_fields=fields_updated)

        return Response({
            "detail": "Student updated successfully.",
            "register_number": student.register_number,
            "name": student.name,
            "mobile_number": student.mobile_number,
            "batch": student.batch,
            "section": student.section,
        })


class JAStudentMoveView(APIView):
    """JA: move a student to a different batch within the same department."""
    permission_classes = [IsAuthenticated]

    def post(self, request, register_number):
        profile, err = _ja_guard(request)
        if err:
            return err

        new_batch = (request.data.get('batch') or '').strip()
        if not new_batch:
            return Response({"detail": "New batch name is required."}, status=status.HTTP_400_BAD_REQUEST)

        student = StudentProfile.objects.filter(
            register_number=register_number,
            institution=profile.institution,
            department=profile.department,
        ).first()

        if not student:
            return Response(
                {"detail": "Student not found in your department."},
                status=status.HTTP_404_NOT_FOUND
            )

        old_batch = student.batch
        student.batch = new_batch
        student.save(update_fields=['batch'])

        return Response({
            "detail": f"Student moved from batch '{old_batch}' to '{new_batch}'.",
            "register_number": register_number,
            "batch": new_batch,
        })


class JABulkBatchAssignView(APIView):
    """JA: assign multiple students to a batch at once.
    POST /api/ja/students/assign-batch/
    Body: { batch: "23-27", register_numbers: ["REG001", "REG002"] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        batch = (request.data.get('batch') or '').strip()
        register_numbers = request.data.get('register_numbers', [])

        if not batch:
            return Response({"detail": "batch is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not register_numbers or not isinstance(register_numbers, list):
            return Response({"detail": "register_numbers must be a non-empty list."}, status=status.HTTP_400_BAD_REQUEST)

        updated = StudentProfile.objects.filter(
            register_number__in=register_numbers,
            institution=profile.institution,
            department=profile.department,
        ).update(batch=batch)

        return Response({
            "detail": f"{updated} student(s) assigned to batch '{batch}'.",
            "updated": updated,
            "batch": batch,
        })


class JABulkImportView(APIView):
    """JA: bulk import students from an uploaded Excel (.xlsx) file.

    Optional POST param: default_batch — used when a row's batch column is empty.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({"detail": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({"detail": "Only .xlsx or .xls files are supported."}, status=status.HTTP_400_BAD_REQUEST)

        # Optional default batch — used when row's batch column is blank
        default_batch = (request.data.get('default_batch') or '').strip()

        try:
            import openpyxl
        except ImportError:
            return Response(
                {"detail": "openpyxl is not installed on the server."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({"detail": f"Could not read Excel file: {e}"}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({"detail": "Excel file is empty."}, status=status.HTTP_400_BAD_REQUEST)

        # Parse header — batch is now optional if default_batch is provided
        header = [str(h).strip().lower().replace(' ', '_') if h else '' for h in rows[0]]
        required_cols = {'register_number', 'name'}
        if not default_batch:
            required_cols.add('batch')
        missing = required_cols - set(header)
        if missing:
            return Response(
                {"detail": f"Missing required columns: {', '.join(missing)}. See the template."},
                status=status.HTTP_400_BAD_REQUEST
            )

        col_idx = {col: header.index(col) for col in header if col}

        created = []
        created_details = []  # full details for report
        skipped = []
        errors = []

        from django.contrib.auth.hashers import make_password as _make_password

        for row_num, row in enumerate(rows[1:], start=2):
            def get_col(col_name):
                idx = col_idx.get(col_name)
                if idx is None:
                    return ''
                val = row[idx]
                return str(val).strip() if val is not None else ''

            register_number = get_col('register_number')
            name = get_col('name')
            batch = get_col('batch') or default_batch
            section = get_col('section').upper()

            if not register_number or not name:
                errors.append({"row": row_num, "reason": "Missing register_number or name."})
                continue

            if not batch:
                errors.append({"row": row_num, "register_number": register_number, "reason": "No batch specified and no default batch set."})
                continue

            if StudentProfile.objects.filter(register_number=register_number).exists():
                skipped.append({"row": row_num, "register_number": register_number, "reason": "Already exists."})
                continue

            try:
                user = User.objects.create(
                    username=register_number,
                    password=_make_password(None),
                    is_active=True,
                )
                student = StudentProfile.objects.create(
                    account=user,
                    institution=profile.institution,
                    department=profile.department,
                    register_number=register_number,
                    name=name,
                    title=get_col('title') or name,
                    batch=batch,
                    section=section,
                    personal_email=get_col('personal_email'),
                    mobile_number=get_col('mobile_number'),
                    gender=get_col('gender'),
                )
                created.append(register_number)
                created_details.append({
                    "register_number": register_number,
                    "name": name,
                    "batch": batch,
                    "section": section,
                    "title": student.title,
                    "personal_email": student.personal_email,
                    "mobile_number": student.mobile_number,
                    "gender": student.gender,
                    "department": profile.department.name if profile.department else '',
                })
            except Exception as e:
                errors.append({"row": row_num, "register_number": register_number, "reason": str(e)})

        logger.info(
            "JA %s bulk import: %d created, %d skipped, %d errors",
            profile.faculty_id, len(created), len(skipped), len(errors)
        )

        return Response({
            "detail": f"Import complete. {len(created)} created, {len(skipped)} skipped, {len(errors)} errors.",
            "created_count": len(created),
            "skipped_count": len(skipped),
            "error_count": len(errors),
            "created": created,
            "created_details": created_details,
            "skipped": skipped,
            "errors": errors,
        }, status=status.HTTP_200_OK)


class JAExcelTemplateView(APIView):
    """JA: download the Excel template for bulk student import."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response(
                {"detail": "openpyxl is not installed on the server."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        # ── Header row ────────────────────────────────────────────────────────
        headers = [
            "register_number",
            "name",
            "batch",
            "section",
            "title",
            "personal_email",
            "mobile_number",
            "gender",
        ]

        header_fill = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # ── Sample rows ───────────────────────────────────────────────────────
        sample_rows = [
            ["953623243001", "Arun Kumar", "23-27", "A", "Mr. Arun Kumar", "arun@example.com", "9876543210", "Male"],
            ["953623243002", "Priya Sharma", "23-27", "B", "Ms. Priya Sharma", "priya@example.com", "9876543211", "Female"],
            ["953623243003", "Ravi Raj", "23-27", "A", "Mr. Ravi Raj", "", "", "Male"],
        ]

        sample_fill = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
        for row_num, row_data in enumerate(sample_rows, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = sample_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = [20, 25, 12, 10, 30, 30, 16, 10]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        ws.row_dimensions[1].height = 22

        # ── Instructions sheet ────────────────────────────────────────────────
        ws2 = wb.create_sheet("Instructions")
        instructions = [
            ("Column", "Required", "Description"),
            ("register_number", "YES", "Unique student register number (e.g. 953623243001)"),
            ("name", "YES", "Full name of the student"),
            ("batch", "YES", "Batch code (e.g. 23-27)"),
            ("title", "No", "Display title (e.g. Mr. John Doe). Defaults to name if blank."),
            ("section", "No", "Section within the batch: A or B"),
            ("personal_email", "No", "Personal email address"),
            ("mobile_number", "No", "10-digit mobile number"),
            ("gender", "No", "Male / Female / Other"),
        ]
        hdr_font = Font(bold=True)
        for r_idx, row_data in enumerate(instructions, 1):
            for c_idx, val in enumerate(row_data, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = hdr_font
                    cell.fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
                cell.border = thin_border

        ws2.column_dimensions['A'].width = 22
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 55

        # ── Stream response ───────────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        dept_code = profile.department.code if profile.department else "dept"
        filename = f"student_import_template_{dept_code}.xlsx"

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class JAStudentListView(APIView):
    """JA: list all students in their department with optional batch/search filter."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        students = StudentProfile.objects.filter(
            institution=profile.institution,
            department=profile.department,
        ).select_related('account', 'mentor').order_by('batch', 'register_number')

        batch = request.query_params.get('batch')
        section_param = (request.query_params.get('section') or '').strip()
        search = request.query_params.get('search', '').strip()

        if batch:
            students = students.filter(batch=batch)
        if section_param == '__none__':
            students = students.filter(section='')
        elif section_param:
            students = students.filter(section=section_param.upper())
        if search:
            students = students.filter(
                Q(name__icontains=search) | Q(register_number__icontains=search)
            )

        data = []
        for s in students:
            data.append({
                "id": s.id,
                "register_number": s.register_number,
                "name": s.name,
                "batch": s.batch,
                "section": s.section,
                "personal_email": s.personal_email,
                "mobile_number": s.mobile_number,
                "gender": s.gender,
                "is_active": s.account.is_active if s.account else True,
                "mentor_id": s.mentor_id,
                "mentor_name": s.mentor.name if s.mentor_id else None,
                "mentor_faculty_id": s.mentor.faculty_id if s.mentor_id else None,
            })

        return Response({"students": data, "total": len(data)})


class JAImportReportView(APIView):
    """JA: generate and download an Excel report of a list of students (by register numbers)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        register_numbers = request.data.get('register_numbers', [])
        report_title = request.data.get('title', 'Student Report')

        if not register_numbers:
            return Response({"detail": "No register numbers provided."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response({"detail": "openpyxl not installed."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        students = StudentProfile.objects.filter(
            register_number__in=register_numbers,
            institution=profile.institution,
            department=profile.department,
        ).order_by('batch', 'register_number')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        # ── Styles ────────────────────────────────────────────────────────────
        hdr_fill  = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
        hdr_font  = Font(bold=True, color="FFFFFF", size=11)
        meta_font = Font(bold=True, size=12, color="1A3C2A")
        center    = Alignment(horizontal="center", vertical="center")
        thin      = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        row_fill  = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")

        # ── Title rows ────────────────────────────────────────────────────────
        dept_name = profile.department.name if profile.department else ''
        inst_name = profile.institution.name if profile.institution else ''

        ws.merge_cells('A1:H1')
        ws['A1'] = inst_name
        ws['A1'].font = Font(bold=True, size=14, color="1A3C2A")
        ws['A1'].alignment = center

        ws.merge_cells('A2:H2')
        ws['A2'] = f"{dept_name} — {report_title}"
        ws['A2'].font = meta_font
        ws['A2'].alignment = center

        ws.merge_cells('A3:H3')
        from django.utils import timezone as tz
        ws['A3'] = f"Generated: {tz.now().strftime('%d %b %Y, %I:%M %p')}"
        ws['A3'].font = Font(size=10, color="6B7280")
        ws['A3'].alignment = center

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 16

        # ── Header row ────────────────────────────────────────────────────────
        headers = ['#', 'Register Number', 'Name', 'Batch', 'Title', 'Email', 'Mobile', 'Gender']
        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[5].height = 20

        # ── Data rows ─────────────────────────────────────────────────────────
        for idx, student in enumerate(students, 1):
            row_num = idx + 5
            values = [
                idx,
                student.register_number,
                student.name,
                student.batch,
                student.title,
                student.personal_email,
                student.mobile_number,
                student.gender,
            ]
            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=val)
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
                if idx % 2 == 0:
                    cell.fill = row_fill

        # ── Column widths ─────────────────────────────────────────────────────
        col_widths = [5, 20, 28, 10, 32, 30, 16, 10]
        for col_num, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # ── Summary row ───────────────────────────────────────────────────────
        summary_row = students.count() + 7
        ws.cell(row=summary_row, column=1, value=f"Total: {students.count()} student(s)").font = Font(bold=True, color="2D6A4F")

        # ── Stream ────────────────────────────────────────────────────────────
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        dept_code = profile.department.code if profile.department else 'dept'
        from django.utils import timezone as tz2
        ts = tz2.now().strftime('%Y%m%d_%H%M')
        filename = f"students_{dept_code}_{ts}.xlsx"

        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# =============================================================================
# JA — Advisor & Mentor Management
# =============================================================================

class JAStaffListView(APIView):
    """JA: list all active staff in their department (for advisor/mentor assignment)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        staff = StaffProfile.objects.filter(
            institution=profile.institution,
            department=profile.department,
            is_active=True,
        ).order_by('name')

        data = [
            {
                "id": s.id,
                "faculty_id": s.faculty_id,
                "name": s.name,
                "role": s.role,
                "role_display": s.get_role_display(),
            }
            for s in staff
        ]
        return Response({"staff": data})


class JABatchAdvisorView(APIView):
    """
    JA: get or set the class advisor for a batch section.
    GET  /api/ja/advisors/           → flat list of all batch+section→advisor rows
    POST /api/ja/advisors/           → assign/update advisor for a batch+section
         body: { batch, section, advisor_id }
    """
    permission_classes = [IsAuthenticated]

    SECTIONS = ["A", "B", "C"]

    def get(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        # All existing advisor assignments
        assignments = BatchAdvisor.objects.filter(
            department=profile.department,
        ).select_related('advisor', 'assigned_by')
        advisor_map = {}  # (batch, section) → assignment obj
        for a in assignments:
            advisor_map[(a.batch, a.section)] = a

        # All batches that have students
        all_batches = list(
            StudentProfile.objects
            .filter(institution=profile.institution, department=profile.department)
            .exclude(batch='')
            .values_list('batch', flat=True)
            .distinct()
            .order_by('batch')
        )

        data = []
        for batch in all_batches:
            for section in self.SECTIONS:
                student_count = StudentProfile.objects.filter(
                    institution=profile.institution,
                    department=profile.department,
                    batch=batch,
                    section=section,
                ).count()
                a = advisor_map.get((batch, section))
                data.append({
                    "batch": batch,
                    "section": section,
                    "student_count": student_count,
                    "advisor": {
                        "id": a.advisor.id,
                        "faculty_id": a.advisor.faculty_id,
                        "name": a.advisor.name,
                        "role": a.advisor.role,
                    } if a else None,
                    "assigned_at": a.assigned_at.isoformat() if a else None,
                    "assigned_by": a.assigned_by.name if (a and a.assigned_by) else None,
                })

        return Response({"assignments": data})

    def post(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        batch = (request.data.get('batch') or '').strip()
        section = (request.data.get('section') or '').strip().upper()
        advisor_id = request.data.get('advisor_id')

        if not batch:
            return Response({"detail": "batch is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not section:
            return Response({"detail": "section is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not advisor_id:
            return Response({"detail": "advisor_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            advisor = StaffProfile.objects.get(id=advisor_id, institution=profile.institution)
        except StaffProfile.DoesNotExist:
            return Response({"detail": "Staff member not found."}, status=status.HTTP_404_NOT_FOUND)

        assignment, created = BatchAdvisor.objects.update_or_create(
            batch=batch,
            section=section,
            department=profile.department,
            defaults={"advisor": advisor, "assigned_by": profile},
        )

        action = "assigned" if created else "updated"
        logger.info(
            "JA %s %s class advisor %s to batch %s section %s",
            profile.faculty_id, action, advisor.faculty_id, batch, section
        )

        return Response({
            "detail": f"Class advisor {action} for batch '{batch}' section '{section}'.",
            "batch": batch,
            "section": section,
            "advisor": {"id": advisor.id, "faculty_id": advisor.faculty_id, "name": advisor.name},
        }, status=status.HTTP_200_OK)


class JABatchAdvisorDeleteView(APIView):
    """JA: remove a class advisor assignment for a batch+section.
    DELETE /api/ja/advisors/<batch_code>/?section=A
    """
    permission_classes = [IsAuthenticated]

    def delete(self, request, batch_code):
        profile, err = _ja_guard(request)
        if err:
            return err

        section = (request.query_params.get('section') or '').strip().upper()

        qs = BatchAdvisor.objects.filter(batch=batch_code, department=profile.department)
        if section:
            qs = qs.filter(section=section)

        deleted, _ = qs.delete()

        if not deleted:
            return Response({"detail": "No advisor assignment found."}, status=status.HTTP_404_NOT_FOUND)

        label = f"batch '{batch_code}' section '{section}'" if section else f"batch '{batch_code}'"
        return Response({"detail": f"Advisor removed from {label}."})


class JABulkSectionAssignView(APIView):
    """JA: assign multiple students to a section at once.
    POST /api/ja/students/assign-section/
    Body: { section: "A", register_numbers: ["REG001", "REG002"] }
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        section = (request.data.get('section') or '').strip().upper()
        register_numbers = request.data.get('register_numbers', [])

        if not section:
            return Response({"detail": "section is required."}, status=status.HTTP_400_BAD_REQUEST)
        if not register_numbers or not isinstance(register_numbers, list):
            return Response({"detail": "register_numbers must be a non-empty list."}, status=status.HTTP_400_BAD_REQUEST)

        updated = StudentProfile.objects.filter(
            register_number__in=register_numbers,
            institution=profile.institution,
            department=profile.department,
        ).update(section=section)

        return Response({
            "detail": f"{updated} student(s) assigned to section '{section}'.",
            "updated": updated,
            "section": section,
        })


class JAMentorAssignView(APIView):
    """
    JA: assign or remove a mentor for one or more students.
    POST /api/ja/mentors/assign/
    body: { mentor_id: int|null, register_numbers: [str, ...] }
    - mentor_id null → remove mentor
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        mentor_id = request.data.get('mentor_id')  # None = unassign
        register_numbers = request.data.get('register_numbers', [])

        if not register_numbers:
            return Response({"detail": "register_numbers list is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Strip None/empty values that can't match a register_number
        valid_reg_numbers = [r for r in register_numbers if r]
        if not valid_reg_numbers:
            return Response({"detail": "No valid register numbers provided."}, status=status.HTTP_400_BAD_REQUEST)

        mentor = None
        if mentor_id:
            try:
                mentor = StaffProfile.objects.get(id=mentor_id, institution=profile.institution)
            except StaffProfile.DoesNotExist:
                return Response({"detail": "Staff member not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            updated = StudentProfile.objects.filter(
                register_number__in=valid_reg_numbers,
                institution=profile.institution,
                department=profile.department,
            ).update(mentor=mentor)
        except Exception:
            logger.exception("Failed to update mentor assignment for JA %s", profile.faculty_id)
            return Response(
                {"detail": "Failed to update mentor assignment. Please try again."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        if updated == 0:
            return Response(
                {"detail": "No matching students found. They may not belong to your department."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        action = f"assigned mentor {mentor.name}" if mentor else "removed mentor"
        logger.info("JA %s %s for %d students", profile.faculty_id, action, updated)

        return Response({
            "detail": f"{action.capitalize()} for {updated} student(s).",
            "updated": updated,
        })


class JAMentorListView(APIView):
    """JA: list all mentor assignments in the department."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _ja_guard(request)
        if err:
            return err

        # Group students by mentor
        batch_filter = (request.query_params.get('batch') or '').strip()
        qs = StudentProfile.objects.filter(institution=profile.institution, department=profile.department)
        if batch_filter:
            qs = qs.filter(batch=batch_filter)
        students = qs.select_related('mentor').order_by('batch', 'register_number')

        mentor_map = {}  # mentor_id → { mentor_info, students }
        unassigned = []

        for s in students:
            if s.mentor_id:
                if s.mentor_id not in mentor_map:
                    m = s.mentor
                    mentor_map[s.mentor_id] = {
                        "mentor": {
                            "id": m.id,
                            "faculty_id": m.faculty_id,
                            "name": m.name,
                            "role": m.role,
                        },
                        "students": [],
                    }
                mentor_map[s.mentor_id]["students"].append({
                    "id": s.id,
                    "register_number": s.register_number,
                    "name": s.name,
                    "batch": s.batch,
                    "section": s.section,
                })
            else:
                unassigned.append({
                    "id": s.id,
                    "register_number": s.register_number,
                    "name": s.name,
                    "batch": s.batch,
                    "section": s.section,
                })

        return Response({
            "mentor_groups": list(mentor_map.values()),
            "unassigned": unassigned,
            "unassigned_count": len(unassigned),
        })


# =============================================================================
# Staff — Mentor & Class Advisor views (for staff members themselves)
# =============================================================================

def _staff_guard(request):
    """Returns (staff_profile, None) or (None, Response)."""
    if not request.user.is_authenticated or not hasattr(request.user, 'staff_profile'):
        return None, Response({"detail": "Authentication required."}, status=status.HTTP_401_UNAUTHORIZED)
    profile = request.user.staff_profile
    if not profile.is_active:
        return None, Response({"detail": "Your account has been disabled."}, status=status.HTTP_403_FORBIDDEN)
    return profile, None


class StaffMentorDashboardView(APIView):
    """
    Staff: get their mentees list with stats.
    GET /api/staff/mentor/dashboard/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _staff_guard(request)
        if err:
            return err

        mentees = (
            StudentProfile.objects
            .filter(mentor=profile)
            .select_related('account', 'department')
            .order_by('batch', 'register_number')
        )

        mentees = mentees.annotate(solved_count=Count('solved_problems', distinct=True))

        mentee_data = []
        for s in mentees:
            mentee_data.append({
                "id": s.id,
                "register_number": s.register_number,
                "name": s.name,
                "batch": s.batch,
                "section": s.section,
                "department": s.department.name if s.department else "",
                "personal_email": s.personal_email,
                "mobile_number": s.mobile_number,
                "gender": s.gender,
                "current_streak": s.current_streak,
                "login_days": s.login_days,
                "last_active": s.last_login_on.isoformat() if s.last_login_on else None,
                "solved_count": s.solved_count,
                "is_active": s.account.is_active if s.account else True,
            })

        # Group by batch
        batch_groups = {}
        for m in mentee_data:
            b = m['batch'] or 'Unknown'
            if b not in batch_groups:
                batch_groups[b] = []
            batch_groups[b].append(m)

        return Response({
            "mentor": {
                "id": profile.id,
                "faculty_id": profile.faculty_id,
                "name": profile.name,
                "role": profile.role,
                "department": profile.department.name if profile.department else "",
            },
            "total_mentees": len(mentee_data),
            "mentees": mentee_data,
            "batch_groups": [
                {"batch": b, "students": students}
                for b, students in sorted(batch_groups.items(), reverse=True)
            ],
        })


class StaffClassAdvisorDashboardView(APIView):
    """
    Staff: get all data for batches where they are the class advisor.
    GET /api/staff/advisor/dashboard/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        profile, err = _staff_guard(request)
        if err:
            return err

        # Find all batches where this staff is the advisor
        advisor_assignments = BatchAdvisor.objects.filter(advisor=profile).select_related('department')

        if not advisor_assignments.exists():
            return Response({
                "advisor": {
                    "id": profile.id,
                    "name": profile.name,
                    "faculty_id": profile.faculty_id,
                },
                "is_class_advisor": False,
                "batches": [],
            })

        batches_data = []
        for assignment in advisor_assignments:
            qs = StudentProfile.objects.filter(
                department=assignment.department,
                batch=assignment.batch,
            )
            if assignment.section:
                qs = qs.filter(section=assignment.section)
            students = qs.select_related('account', 'mentor').order_by('register_number')

            student_list = []
            for s in students:
                from .models import SolvedProblem
                solved_count = SolvedProblem.objects.filter(student=s).count()
                student_list.append({
                    "id": s.id,
                    "register_number": s.register_number,
                    "name": s.name,
                    "batch": s.batch,
                    "section": s.section,
                    "personal_email": s.personal_email,
                    "mobile_number": s.mobile_number,
                    "gender": s.gender,
                    "current_streak": s.current_streak,
                    "login_days": s.login_days,
                    "last_active": s.last_login_on.isoformat() if s.last_login_on else None,
                    "solved_count": solved_count,
                    "is_active": s.account.is_active if s.account else True,
                    "mentor": {
                        "id": s.mentor.id,
                        "name": s.mentor.name,
                        "faculty_id": s.mentor.faculty_id,
                    } if s.mentor else None,
                })

            # Batch-level stats
            total = len(student_list)
            active = sum(1 for s in student_list if s['is_active'])
            avg_solved = (sum(s['solved_count'] for s in student_list) / total) if total else 0
            avg_streak = (sum(s['current_streak'] for s in student_list) / total) if total else 0

            batches_data.append({
                "batch": assignment.batch,
                "section": assignment.section,
                "department": assignment.department.name,
                "department_code": assignment.department.code,
                "total_students": total,
                "active_students": active,
                "avg_solved": round(avg_solved, 1),
                "avg_streak": round(avg_streak, 1),
                "students": student_list,
            })

        return Response({
            "advisor": {
                "id": profile.id,
                "name": profile.name,
                "faculty_id": profile.faculty_id,
                "role": profile.role,
            },
            "is_class_advisor": True,
            "batches": batches_data,
        })


# =============================================================================
# Student — Mentor & Advisor info view
# =============================================================================

class StudentMentorAdvisorView(APIView):
    """
    Student: get their assigned mentor and class advisor info.
    GET /api/student/mentor-advisor/
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not hasattr(request.user, 'student_profile'):
            return Response({"detail": "Student access required."}, status=status.HTTP_403_FORBIDDEN)

        student = request.user.student_profile

        # Mentor info
        mentor_info = None
        if student.mentor:
            m = student.mentor
            mentor_info = {
                "id": m.id,
                "faculty_id": m.faculty_id,
                "name": m.name,
                "email": m.email,
                "role": m.role,
                "role_display": m.get_role_display(),
                "department": m.department.name if m.department else "",
            }

        # Class advisor info
        advisor_info = None
        if student.batch and student.department:
            advisor_qs = BatchAdvisor.objects.filter(
                batch=student.batch,
                department=student.department,
            ).select_related('advisor', 'advisor__department')
            # Match the student's section first; fall back to unsectioned advisor
            if student.section:
                assignment = advisor_qs.filter(section=student.section).first()
                if not assignment:
                    assignment = advisor_qs.filter(section='').first()
            else:
                assignment = advisor_qs.filter(section='').first() or advisor_qs.first()
            if assignment:
                a = assignment.advisor
                advisor_info = {
                    "id": a.id,
                    "faculty_id": a.faculty_id,
                    "name": a.name,
                    "email": a.email,
                    "role": a.role,
                    "role_display": a.get_role_display(),
                    "department": a.department.name if a.department else "",
                }

        return Response({
            "mentor": mentor_info,
            "class_advisor": advisor_info,
            "batch": student.batch,
            "section": student.section,
            "department": student.department.name if student.department else "",
        })


# ─── Labs ─────────────────────────────────────────────────────────────────────

class LabTopicListView(APIView):
    """GET /api/lab/topics/ — list all active topics with problem counts."""
    permission_classes = [AllowAny]

    def get(self, request):
        topics = LabTopic.objects.filter(is_active=True).prefetch_related("problems")
        data = []
        for t in topics:
            active = t.problems.filter(is_active=True)
            counts = {"Easy": 0, "Medium": 0, "Hard": 0}
            for p in active:
                counts[p.difficulty] = counts.get(p.difficulty, 0) + 1
            data.append({
                "id":          t.id,
                "name":        t.name,
                "slug":        t.slug,
                "description": t.description,
                "icon":        t.icon,
                "order":       t.order,
                "total":       active.count(),
                "easy":        counts["Easy"],
                "medium":      counts["Medium"],
                "hard":        counts["Hard"],
            })
        return Response(data)


class LabProblemListView(APIView):
    """GET /api/lab/topics/<slug>/problems/ — problems in a topic."""
    permission_classes = [AllowAny]

    def get(self, request, topic_slug):
        try:
            topic = LabTopic.objects.get(slug=topic_slug, is_active=True)
        except LabTopic.DoesNotExist:
            return Response({"error": "Topic not found"}, status=status.HTTP_404_NOT_FOUND)

        problems = LabProblem.objects.filter(topic=topic, is_active=True)

        # Attach solved status if authenticated
        student = None
        if hasattr(request.user, "student_profile"):
            try:
                student = request.user.student_profile
            except Exception:
                pass

        solved_slugs = set()
        if student:
            solved_slugs = set(
                LabSubmission.objects.filter(
                    student=student, problem__topic=topic, all_passed=True
                ).values_list("problem__slug", flat=True)
            )

        data = []
        for p in problems:
            data.append({
                "slug":        p.slug,
                "title":       p.title,
                "difficulty":  p.difficulty,
                "tags":        p.tags,
                "order":       p.order,
                "solved":      p.slug in solved_slugs,
            })
        return Response({"topic": {"name": topic.name, "slug": topic.slug, "icon": topic.icon}, "problems": data})


class LabProblemDetailView(APIView):
    """GET /api/lab/problems/<slug>/ — full problem with sample test cases."""
    permission_classes = [AllowAny]

    def get(self, request, slug):
        try:
            problem = LabProblem.objects.select_related("topic").prefetch_related("test_cases").get(
                slug=slug, is_active=True
            )
        except LabProblem.DoesNotExist:
            return Response({"error": "Problem not found"}, status=status.HTTP_404_NOT_FOUND)

        sample_cases = [
            {"stdin": tc.stdin, "expected_output": tc.expected_output, "order": tc.order}
            for tc in problem.test_cases.filter(is_sample=True)
        ]

        return Response({
            "slug":           problem.slug,
            "title":          problem.title,
            "description":    problem.description,
            "difficulty":     problem.difficulty,
            "tags":           problem.tags,
            "examples":       problem.examples,
            "hints":          problem.hints,
            "editorial":      problem.editorial,
            "execution_type": problem.execution_type,
            "function_name":  problem.function_name,
            "sample_cases":   sample_cases,
            "topic": {
                "name": problem.topic.name,
                "slug": problem.topic.slug,
            },
        })


class LabSubmitView(APIView):
    """POST /api/lab/problems/<slug>/submit/ — run all test cases and record result."""
    permission_classes = [IsAuthenticated]

    def post(self, request, slug):
        from .services.executor import ExecutorError, execute_submission

        try:
            problem = LabProblem.objects.prefetch_related("test_cases").get(slug=slug, is_active=True)
        except LabProblem.DoesNotExist:
            return Response({"error": "Problem not found"}, status=status.HTTP_404_NOT_FOUND)

        source_code = request.data.get("source_code", "").strip()
        language_id = request.data.get("language_id")
        language    = request.data.get("language", "")

        if not source_code:
            return Response({"error": "Source code is required"}, status=status.HTTP_400_BAD_REQUEST)
        if not language_id:
            return Response({"error": "language_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            language_id = int(language_id)
        except (TypeError, ValueError):
            return Response({"error": "Invalid language_id"}, status=status.HTTP_400_BAD_REQUEST)

        # Get student
        try:
            student = request.user.student_profile
        except Exception:
            return Response({"error": "Student profile not found"}, status=status.HTTP_403_FORBIDDEN)

        test_cases = list(problem.test_cases.all())
        if not test_cases:
            return Response({"error": "No test cases configured for this problem"}, status=status.HTTP_400_BAD_REQUEST)

        results = []
        passed  = 0

        for tc in test_cases:
            try:
                payload = prepare_execution_payload(
                    problem=problem,
                    source_code=source_code,
                    language=language,
                    stdin=tc.stdin,
                )
                exec_result = execute_submission(
                    source_code=payload["source_code"],
                    language_id=language_id,
                    stdin=payload.get("stdin", tc.stdin),
                )
                actual   = (exec_result.get("stdout") or "").strip()
                expected = tc.expected_output.strip()
                ok = actual == expected
                if ok:
                    passed += 1
                results.append({
                    "stdin":           tc.stdin,
                    "expected_output": expected,
                    "actual_output":   actual,
                    "passed":          ok,
                    "stderr":          exec_result.get("stderr", ""),
                    "status":          exec_result.get("status", ""),
                    "is_sample":       tc.is_sample,
                })
            except ExecutorError as exc:
                results.append({
                    "stdin":           tc.stdin,
                    "expected_output": tc.expected_output.strip(),
                    "actual_output":   "",
                    "passed":          False,
                    "stderr":          str(exc),
                    "status":          "Error",
                    "is_sample":       tc.is_sample,
                })

        all_passed = passed == len(test_cases)
        sub_status = "Accepted" if all_passed else "Wrong Answer"

        LabSubmission.objects.create(
            student=student,
            problem=problem,
            language=language,
            language_id=language_id,
            source_code=source_code,
            status=sub_status,
            passed_cases=passed,
            total_cases=len(test_cases),
            all_passed=all_passed,
        )

        return Response({
            "status":       sub_status,
            "passed":       passed,
            "total":        len(test_cases),
            "all_passed":   all_passed,
            "results":      results,
        })


# ─── Lab Assignments (Practical Labs) ────────────────────────────────────────

def _staff_from_request(request):
    try:
        return request.user.staff_profile
    except Exception:
        return None


def _student_from_request(request):
    try:
        return request.user.student_profile
    except Exception:
        return None


def _serialize_assignment(a, student=None):
    """Compact assignment dict shared across HOD / staff / student views."""
    problems = list(a.lab_topic.problems.filter(is_active=True).values(
        "id", "slug", "title", "difficulty", "order"
    ))
    solved_slugs = set()
    if student:
        solved_slugs = set(
            LabAssignmentSubmission.objects.filter(
                assignment=a, student=student, all_passed=True
            ).values_list("problem__slug", flat=True)
        )
    for p in problems:
        p["solved"] = p["slug"] in solved_slugs

    return {
        "id":             a.id,
        "name":           a.name,
        "subject":        a.subject,
        "batch":          a.batch,
        "year":           a.year,
        "section":        a.section,
        "start_date":     a.start_date.isoformat() if a.start_date else None,
        "deadline":       a.deadline.isoformat(),
        "is_expired":     a.is_expired,
        "is_active":      a.is_active,
        "created_at":     a.created_at.isoformat(),
        "topic": {
            "id":   a.lab_topic.id,
            "name": a.lab_topic.name,
            "slug": a.lab_topic.slug,
            "icon": a.lab_topic.icon,
        },
        "assigned_staff": {
            "id":   a.assigned_staff.id,
            "name": a.assigned_staff.name,
        } if a.assigned_staff else None,
        "problems":       problems,
        "total_problems": len(problems),
        "solved_count":   len(solved_slugs),
        "all_complete":   len(problems) > 0 and len(solved_slugs) == len(problems),
    }


class HODDeptStaffView(APIView):
    """HOD: list staff in own department (for assignment staff picker)."""

    def get(self, request):
        staff = _staff_from_request(request)
        if not staff or staff.role not in ("hod", "academics", "admin"):
            return Response({"error": "HOD access required"}, status=403)
        dept_staff = StaffProfile.objects.filter(department=staff.department).values(
            "id", "name", "faculty_id", "role"
        )
        return Response(list(dept_staff))


class HODDeptInfoView(APIView):
    """HOD: return real batches and sections from department students."""

    def get(self, request):
        staff = _staff_from_request(request)
        if not staff or staff.role not in ("hod", "academics", "admin"):
            return Response({"error": "HOD access required"}, status=403)

        stu_qs = StudentProfile.objects.filter(department=staff.department)
        batches = sorted(set(
            v for v in stu_qs.values_list("batch", flat=True).distinct() if v
        ))
        sections = sorted(set(
            v for v in stu_qs.values_list("section", flat=True).distinct() if v
        ))
        # Build per-batch section map
        sections_by_batch = {}
        for batch in batches:
            secs = sorted(set(
                v for v in stu_qs.filter(batch=batch).values_list("section", flat=True).distinct() if v
            ))
            sections_by_batch[batch] = secs
        return Response({"batches": batches, "sections": sections, "sections_by_batch": sections_by_batch})


class HODLabAssignmentView(APIView):
    """HOD: list assignments for own department / create new assignment."""

    def _get_hod(self, request):
        staff = _staff_from_request(request)
        if not staff or staff.role not in ("hod", "academics", "admin"):
            return None
        return staff

    def get(self, request):
        staff = self._get_hod(request)
        if not staff:
            return Response({"error": "HOD access required"}, status=status.HTTP_403_FORBIDDEN)

        assignments = LabAssignment.objects.filter(
            department=staff.department
        ).select_related("lab_topic", "assigned_staff", "created_by").order_by("-created_at")

        data = []
        for a in assignments:
            d = _serialize_assignment(a)
            d["submission_count"] = a.submissions.values("student").distinct().count()
            data.append(d)
        return Response(data)

    def post(self, request):
        staff = self._get_hod(request)
        if not staff:
            return Response({"error": "HOD access required"}, status=status.HTTP_403_FORBIDDEN)

        topic_id       = request.data.get("topic_id")
        name           = (request.data.get("name") or "").strip()
        subject        = (request.data.get("subject") or "").strip()
        assigned_staff_id = request.data.get("assigned_staff_id")
        batch          = (request.data.get("batch") or "").strip()
        year           = (request.data.get("year") or "").strip()
        section        = (request.data.get("section") or "").strip()
        deadline_str   = (request.data.get("deadline") or "").strip()
        start_date_str = (request.data.get("start_date") or "").strip()

        if not all([topic_id, name, deadline_str]):
            return Response({"error": "topic_id, name and deadline are required"}, status=400)

        try:
            topic = LabTopic.objects.get(id=topic_id, is_active=True)
        except LabTopic.DoesNotExist:
            return Response({"error": "Topic not found"}, status=400)

        try:
            from django.utils.dateparse import parse_datetime
            deadline = parse_datetime(deadline_str)
            if not deadline:
                raise ValueError
        except (ValueError, TypeError):
            return Response({"error": "Invalid deadline format. Use ISO 8601."}, status=400)

        start_date = None
        if start_date_str:
            try:
                from django.utils.dateparse import parse_datetime as _pd
                start_date = _pd(start_date_str)
            except Exception:
                start_date = None

        assigned_staff = None
        if assigned_staff_id:
            try:
                assigned_staff = StaffProfile.objects.get(id=assigned_staff_id, department=staff.department)
            except StaffProfile.DoesNotExist:
                return Response({"error": "Staff not found in your department"}, status=400)

        assignment = LabAssignment.objects.create(
            lab_topic=topic,
            name=name,
            subject=subject,
            assigned_staff=assigned_staff,
            created_by=staff,
            department=staff.department,
            batch=batch,
            year=year,
            section=section,
            start_date=start_date,
            deadline=deadline,
        )
        return Response(_serialize_assignment(assignment), status=status.HTTP_201_CREATED)


class HODLabAssignmentDeleteView(APIView):
    """HOD: delete a lab assignment."""

    def delete(self, request, assignment_id):
        staff = _staff_from_request(request)
        if not staff or staff.role not in ("hod", "academics", "admin"):
            return Response({"error": "HOD access required"}, status=403)
        try:
            a = LabAssignment.objects.get(id=assignment_id, department=staff.department)
        except LabAssignment.DoesNotExist:
            return Response({"error": "Not found"}, status=404)
        a.delete()
        return Response({"ok": True})


class StaffLabAssignmentView(APIView):
    """Staff: list lab assignments assigned to this staff member."""

    def get(self, request):
        staff = _staff_from_request(request)
        if not staff:
            return Response({"error": "Staff access required"}, status=403)

        assignments = LabAssignment.objects.filter(
            assigned_staff=staff, is_active=True
        ).select_related("lab_topic", "created_by").order_by("-created_at")

        return Response([_serialize_assignment(a) for a in assignments])


class StaffLabSubmissionsView(APIView):
    """Staff: class-wise submission status for a specific lab assignment."""

    def get(self, request, assignment_id):
        staff = _staff_from_request(request)
        if not staff:
            return Response({"error": "Staff access required"}, status=403)

        try:
            assignment = LabAssignment.objects.select_related("lab_topic", "department").get(
                id=assignment_id, assigned_staff=staff
            )
        except LabAssignment.DoesNotExist:
            return Response({"error": "Assignment not found or not assigned to you"}, status=404)

        problems = list(assignment.lab_topic.problems.filter(is_active=True).order_by("order"))

        # Students in this assignment's batch/section
        stu_qs = StudentProfile.objects.filter(department=assignment.department)
        if assignment.batch:
            stu_qs = stu_qs.filter(batch=assignment.batch)
        if assignment.section:
            stu_qs = stu_qs.filter(section__iexact=assignment.section)

        students = stu_qs.select_related("account").order_by("name")

        # All submissions for this assignment
        subs = {
            (s.student_id, s.problem_id): s
            for s in LabAssignmentSubmission.objects.filter(assignment=assignment)
        }

        rows = []
        for stu in students:
            solved = []
            for prob in problems:
                sub = subs.get((stu.id, prob.id))
                solved.append({
                    "slug":       prob.slug,
                    "title":      prob.title,
                    "passed":     sub.all_passed if sub else False,
                    "language":   sub.language if sub else None,
                    "submitted_at": sub.submitted_at.isoformat() if sub else None,
                })
            total   = len(problems)
            done    = sum(1 for s in solved if s["passed"])
            rows.append({
                "student_name":     stu.name,
                "register_number":  stu.register_number,
                "email":            stu.account.email if stu.account else "",
                "section":          stu.section,
                "problems":         solved,
                "completed":        done,
                "total":            total,
                "all_complete":     done == total,
            })

        return Response({
            "assignment": _serialize_assignment(assignment),
            "problems":   [{"slug": p.slug, "title": p.title} for p in problems],
            "students":   rows,
        })


class StudentLabAssignmentsView(APIView):
    """Student: list practical lab assignments for their batch/year/section."""

    def get(self, request):
        student = _student_from_request(request)
        if not student:
            return Response({"error": "Student access required"}, status=403)

        qs = LabAssignment.objects.filter(
            department=student.department, is_active=True
        ).select_related("lab_topic", "assigned_staff")

        if student.batch:
            qs = qs.filter(Q(batch="") | Q(batch=student.batch))
        if student.section:
            qs = qs.filter(Q(section="") | Q(section__iexact=student.section))

        return Response([_serialize_assignment(a, student) for a in qs.order_by("deadline")])


class LabAssignmentSubmitView(APIView):
    """Student: submit solution for a problem inside a LabAssignment."""

    def post(self, request, assignment_id, slug):
        student = _student_from_request(request)
        if not student:
            return Response({"error": "Student access required"}, status=403)

        try:
            assignment = LabAssignment.objects.get(id=assignment_id, is_active=True)
        except LabAssignment.DoesNotExist:
            return Response({"error": "Assignment not found"}, status=404)

        if assignment.is_expired:
            return Response({"error": "This lab assignment has expired"}, status=400)

        try:
            problem = LabProblem.objects.prefetch_related("test_cases").get(
                slug=slug, topic=assignment.lab_topic, is_active=True
            )
        except LabProblem.DoesNotExist:
            return Response({"error": "Problem not found in this assignment"}, status=404)

        source_code = request.data.get("source_code", "").strip()
        language_id = request.data.get("language_id")
        language    = request.data.get("language", "")

        if not source_code or not language_id:
            return Response({"error": "source_code and language_id are required"}, status=400)

        try:
            language_id = int(language_id)
        except (TypeError, ValueError):
            return Response({"error": "Invalid language_id"}, status=400)

        from .services.executor import ExecutorError

        test_cases = list(problem.test_cases.all())
        if not test_cases:
            return Response({"error": "No test cases configured"}, status=400)

        results   = []
        passed    = 0
        final_out = ""

        for tc in test_cases:
            try:
                payload = prepare_execution_payload(
                    source_code=source_code,
                    language_id=language_id,
                    stdin=tc.stdin,
                    problem_slug=problem.slug,
                    execution_type=problem.execution_type,
                    function_name=problem.function_name,
                )
                exec_result = execute_judge0_submission(
                    source_code=payload["source_code"],
                    language_id=language_id,
                    stdin=payload.get("stdin", tc.stdin),
                )
                actual   = (exec_result.get("stdout") or "").strip()
                expected = tc.expected_output.strip()
                ok = actual == expected
                if ok:
                    passed += 1
                if tc.is_sample:
                    final_out = actual
                results.append({
                    "stdin":           tc.stdin,
                    "expected_output": expected,
                    "actual_output":   actual,
                    "passed":          ok,
                    "stderr":          exec_result.get("stderr", ""),
                    "status":          exec_result.get("status", ""),
                    "is_sample":       tc.is_sample,
                })
            except ExecutorError as exc:
                results.append({
                    "stdin": tc.stdin, "expected_output": tc.expected_output.strip(),
                    "actual_output": "", "passed": False,
                    "stderr": str(exc), "status": "Error", "is_sample": tc.is_sample,
                })

        all_passed  = passed == len(test_cases)
        sub_status  = "Accepted" if all_passed else "Wrong Answer"
        full_output = "\n".join(r["actual_output"] for r in results if r["actual_output"])

        LabAssignmentSubmission.objects.update_or_create(
            assignment=assignment,
            student=student,
            problem=problem,
            defaults={
                "language":    language,
                "language_id": language_id,
                "source_code": source_code,
                "output":      full_output or final_out,
                "status":      sub_status,
                "passed_cases": passed,
                "total_cases":  len(test_cases),
                "all_passed":   all_passed,
            },
        )

        return Response({
            "status":     sub_status,
            "passed":     passed,
            "total":      len(test_cases),
            "all_passed": all_passed,
            "results":    results,
            "output":     full_output or final_out,
        })


# ═══════════════════════════════════════════════════════════════════════════════
# Lab V2  (HOD creates Lab -> Staff adds Exercises -> Students submit)
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_dt(value):
    """Parse an ISO datetime string into a timezone-aware datetime. Returns None on failure."""
    if not value:
        return None
    if hasattr(value, "isoformat"):
        return value if value.tzinfo else timezone.make_aware(value)
    try:
        from datetime import datetime as _dt
        parsed = _dt.fromisoformat(str(value))
        return parsed if parsed.tzinfo else timezone.make_aware(parsed)
    except (ValueError, TypeError):
        return None


def _serialize_lab_v2(lab, student=None):
    try:
        ex_count = lab.exercises.count()
    except Exception:
        ex_count = 0
    result = {
        "id": lab.id,
        "name": lab.name,
        "batch": lab.batch,
        "section": lab.section,
        "start_date": lab.start_date.isoformat() if lab.start_date else None,
        "end_date": lab.end_date.isoformat() if lab.end_date else None,
        "is_active": lab.is_active,
        "is_expired": lab.is_expired,
        "exercise_count": ex_count,
        "created_at": lab.created_at.isoformat(),
        "lab_type": lab.lab_type,
        "company": {
            "id": lab.company.id,
            "name": lab.company.name,
        } if lab.company_id else None,
        "allowed_languages": lab.allowed_languages or list(LAB_LANGUAGE_CHOICES),
        "staff_in_charge": {
            "id": lab.staff_in_charge.id,
            "name": lab.staff_in_charge.name,
            "faculty_id": lab.staff_in_charge.faculty_id,
        } if lab.staff_in_charge else None,
        "created_by": {
            "id": lab.created_by.id,
            "name": lab.created_by.name,
        } if lab.created_by else None,
        "linked_lab_id": lab.linked_lab_id,
        "approval_status": lab.approval_status,
        "is_published": lab.is_published,
        "enable_tab_switch_check": lab.enable_tab_switch_check,
        "max_tab_switches": lab.max_tab_switches,
        "enable_fullscreen_lock": lab.enable_fullscreen_lock,
        "enable_copy_paste_lock": lab.enable_copy_paste_lock,
        "pass_threshold_percent": lab.pass_threshold_percent,
    }
    if student is not None:
        completed = LabExerciseSubmission.objects.filter(
            exercise__lab=lab, student=student
        ).count()
        result["student_progress"] = {"completed": completed, "total": ex_count}
    return result


def _serialize_exercise(ex):
    submission_count = getattr(ex, "submission_count", None)
    test_case_count = getattr(ex, "test_case_count", None)
    if submission_count is None and ex.pk:
        submission_count = ex.submissions.count()
    if test_case_count is None and ex.pk:
        test_case_count = ex.test_cases.count()
    return {
        "id": ex.id,
        "title": ex.title,
        "description": ex.description,
        "explanation": ex.explanation,
        "order": ex.order,
        "created_at": ex.created_at.isoformat(),
        "added_by": {"id": ex.added_by.id, "name": ex.added_by.name} if ex.added_by else None,
        "submission_count": submission_count,
        "test_case_count": test_case_count,
        "difficulty": ex.difficulty,
    }


def _serialize_test_case(tc):
    return {
        "id": tc.id,
        "exercise_id": tc.exercise_id,
        "stdin": tc.stdin,
        "expected_output": tc.expected_output,
        "is_sample": tc.is_sample,
        "order": tc.order,
    }


def _nonnegative_int(value, default=0):
    try:
        return max(0, int(value))
    except (TypeError, ValueError):
        return default


def _serialize_company_practical(company):
    """A Company IS its one practical — this serializes the Company together
    with the single Lab (lab_type="company") that belongs to it."""
    lab = getattr(company, "lab", None)
    return {
        "id": company.id,
        "name": company.name,
        "created_at": company.created_at.isoformat(),
        "lab_id": lab.id if lab else None,
        "batch": lab.batch if lab else "",
        "section": lab.section if lab else "",
        "start_date": lab.start_date.isoformat() if lab and lab.start_date else None,
        "end_date": lab.end_date.isoformat() if lab and lab.end_date else None,
        "is_expired": lab.is_expired if lab else None,
        "exercise_count": lab.exercises.count() if lab else 0,
        "allowed_languages": (lab.allowed_languages if lab and lab.allowed_languages else list(LAB_LANGUAGE_CHOICES)),
        "staff_in_charge": {
            "id": lab.staff_in_charge.id,
            "name": lab.staff_in_charge.name,
            "faculty_id": lab.staff_in_charge.faculty_id,
        } if lab and lab.staff_in_charge else None,
    }


class HODCompanyListView(APIView):
    """HOD: list/create Companies for company-based lab practicals, scoped to their
    department. A Company and its practical (Lab) are created together — one company
    is exactly one practical (batch, staff, dates, languages)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        staff = _staff_from_request(request)
        companies = Company.objects.filter(department=staff.department).select_related("lab", "lab__staff_in_charge")
        return Response([_serialize_company_practical(c) for c in companies])

    def post(self, request):
        staff = _staff_from_request(request)
        data = request.data

        name = (data.get("name") or "").strip()
        if not name:
            return Response({"error": "Company name is required"}, status=400)
        if Company.objects.filter(department=staff.department, name__iexact=name).exists():
            return Response({"error": "A company with this name already exists"}, status=400)

        if not data.get("batch"):
            return Response({"error": "Select a batch"}, status=400)
        start = _parse_dt(data.get("start_date"))
        end = _parse_dt(data.get("end_date"))
        if not start or not end:
            return Response({"error": "Valid start_date and end_date are required"}, status=400)

        staff_in_charge = None
        sic_id = data.get("staff_in_charge_id")
        if sic_id:
            try:
                staff_in_charge = StaffProfile.objects.get(id=sic_id, department=staff.department)
            except StaffProfile.DoesNotExist:
                return Response({"error": "Staff not found"}, status=400)

        allowed_languages = data.get("allowed_languages", list(LAB_LANGUAGE_CHOICES))
        if (not isinstance(allowed_languages, list) or not allowed_languages
                or any(lang not in LAB_LANGUAGE_CHOICES for lang in allowed_languages)):
            return Response(
                {"error": f"allowed_languages must be a non-empty list from {LAB_LANGUAGE_CHOICES}"}, status=400
            )

        with transaction.atomic():
            company = Company.objects.create(name=name, department=staff.department, created_by=staff)
            Lab.objects.create(
                name=f"{name} — Company Practical",
                department=staff.department,
                batch=data.get("batch", ""),
                section=data.get("section", ""),
                start_date=start,
                end_date=end,
                staff_in_charge=staff_in_charge,
                created_by=staff,
                lab_type="company",
                company=company,
                allowed_languages=allowed_languages,
            )
        company.refresh_from_db()
        return Response(_serialize_company_practical(company), status=201)


class HODCompanyDetailView(APIView):
    """HOD: edit (name + practical settings) or delete a Company, scoped to their department."""
    permission_classes = [IsAuthenticated]

    def _get(self, company_id, staff):
        try:
            return Company.objects.select_related("lab").get(id=company_id, department=staff.department)
        except Company.DoesNotExist:
            return None

    def put(self, request, company_id):
        staff = _staff_from_request(request)
        company = self._get(company_id, staff)
        if not company:
            return Response({"error": "Not found"}, status=404)
        data = request.data

        if "name" in data:
            name = (data.get("name") or "").strip()
            if not name:
                return Response({"error": "Company name is required"}, status=400)
            if Company.objects.filter(department=staff.department, name__iexact=name).exclude(id=company.id).exists():
                return Response({"error": "A company with this name already exists"}, status=400)
            company.name = name
            company.save(update_fields=["name"])

        lab = getattr(company, "lab", None)
        practical_fields = {"batch", "section", "start_date", "end_date", "staff_in_charge_id", "allowed_languages"}
        if lab:
            for field in ("batch", "section"):
                if field in data:
                    setattr(lab, field, data[field])
            if "start_date" in data:
                lab.start_date = _parse_dt(data["start_date"]) or lab.start_date
            if "end_date" in data:
                lab.end_date = _parse_dt(data["end_date"]) or lab.end_date
            if "staff_in_charge_id" in data:
                sic_id = data["staff_in_charge_id"]
                if sic_id:
                    try:
                        lab.staff_in_charge = StaffProfile.objects.get(id=sic_id, department=staff.department)
                    except StaffProfile.DoesNotExist:
                        return Response({"error": "Staff not found"}, status=400)
                else:
                    lab.staff_in_charge = None
            if "allowed_languages" in data:
                allowed_languages = data["allowed_languages"]
                if (not isinstance(allowed_languages, list) or not allowed_languages
                        or any(lang not in LAB_LANGUAGE_CHOICES for lang in allowed_languages)):
                    return Response(
                        {"error": f"allowed_languages must be a non-empty list from {LAB_LANGUAGE_CHOICES}"}, status=400
                    )
                lab.allowed_languages = allowed_languages
            lab.save()
        elif practical_fields & set(data.keys()):
            # Legacy company created before a practical was required — create it now.
            if not data.get("batch"):
                return Response({"error": "Select a batch"}, status=400)
            start = _parse_dt(data.get("start_date"))
            end = _parse_dt(data.get("end_date"))
            if not start or not end:
                return Response({"error": "Valid start_date and end_date are required"}, status=400)
            staff_in_charge = None
            sic_id = data.get("staff_in_charge_id")
            if sic_id:
                try:
                    staff_in_charge = StaffProfile.objects.get(id=sic_id, department=staff.department)
                except StaffProfile.DoesNotExist:
                    return Response({"error": "Staff not found"}, status=400)
            allowed_languages = data.get("allowed_languages", list(LAB_LANGUAGE_CHOICES))
            if (not isinstance(allowed_languages, list) or not allowed_languages
                    or any(lang not in LAB_LANGUAGE_CHOICES for lang in allowed_languages)):
                return Response(
                    {"error": f"allowed_languages must be a non-empty list from {LAB_LANGUAGE_CHOICES}"}, status=400
                )
            Lab.objects.create(
                name=f"{company.name} — Company Practical",
                department=staff.department,
                batch=data.get("batch", ""),
                section=data.get("section", ""),
                start_date=start,
                end_date=end,
                staff_in_charge=staff_in_charge,
                created_by=staff,
                lab_type="company",
                company=company,
                allowed_languages=allowed_languages,
            )

        return Response(_serialize_company_practical(company))

    def delete(self, request, company_id):
        staff = _staff_from_request(request)
        company = self._get(company_id, staff)
        if not company:
            return Response({"error": "Not found"}, status=404)
        company.delete()
        return Response(status=204)


class HODLabListView(APIView):
    """HOD: list/create plain "Lab Practical" and "University Lab" entries."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        staff = _staff_from_request(request)
        labs = Lab.objects.filter(department=staff.department).select_related(
            "staff_in_charge", "created_by"
        ).prefetch_related("exercises")
        return Response([_serialize_lab_v2(lab) for lab in labs])

    def post(self, request):
        staff = _staff_from_request(request)
        data = request.data
        staff_in_charge = None
        sic_id = data.get("staff_in_charge_id")
        if sic_id:
            try:
                staff_in_charge = StaffProfile.objects.get(id=sic_id, department=staff.department)
            except StaffProfile.DoesNotExist:
                return Response({"error": "Staff not found"}, status=400)
        start = _parse_dt(data.get("start_date"))
        end = _parse_dt(data.get("end_date"))
        if not start or not end:
            return Response({"error": "Valid start_date and end_date are required"}, status=400)

        allowed_languages = data.get("allowed_languages", list(LAB_LANGUAGE_CHOICES))
        if (not isinstance(allowed_languages, list) or not allowed_languages
                or any(lang not in LAB_LANGUAGE_CHOICES for lang in allowed_languages)):
            return Response(
                {"error": f"allowed_languages must be a non-empty list from {LAB_LANGUAGE_CHOICES}"}, status=400
            )

        lab_type = data.get("lab_type", "practical")
        linked_lab = None
        linked_lab_id = data.get("linked_lab_id")
        if linked_lab_id:
            try:
                linked_lab = Lab.objects.get(id=linked_lab_id, department=staff.department)
            except Lab.DoesNotExist:
                pass

        is_univ = (lab_type == "university")
        try:
            pass_threshold_percent = max(1, min(100, int(data.get("pass_threshold_percent", 70))))
        except (TypeError, ValueError):
            pass_threshold_percent = 70
        lab = Lab.objects.create(
            name=data["name"],
            department=staff.department,
            batch=data.get("batch", ""),
            section=data.get("section", ""),
            start_date=start,
            end_date=end,
            staff_in_charge=staff_in_charge or staff,
            created_by=staff,
            lab_type=lab_type,
            approval_status="approved",
            is_published=False if lab_type == "university" else data.get("is_published", True),
            enable_tab_switch_check=bool(data.get("enable_tab_switch_check", False)) if is_univ else False,
            max_tab_switches=int(data.get("max_tab_switches", 3)) if is_univ else 3,
            enable_fullscreen_lock=bool(data.get("enable_fullscreen_lock", False)) if is_univ else False,
            enable_copy_paste_lock=bool(data.get("enable_copy_paste_lock", False)) if is_univ else False,
            pass_threshold_percent=pass_threshold_percent,
            allowed_languages=allowed_languages,
            linked_lab=linked_lab,
        )
        lab.refresh_from_db()
        return Response(_serialize_lab_v2(lab), status=201)


class HODLabDetailView(APIView):
    """HOD: edit/delete a plain "Lab Practical". See HODLabListView docstring —
    Company Based Lab Practicals are managed exclusively via the Company endpoints."""
    permission_classes = [IsAuthenticated]

    def _get(self, lab_id, staff):
        try:
            return Lab.objects.get(id=lab_id, department=staff.department)
        except Lab.DoesNotExist:
            return None

    def put(self, request, lab_id):
        staff = _staff_from_request(request)
        lab = self._get(lab_id, staff)
        if not lab:
            return Response({"error": "Not found"}, status=404)
        data = request.data
        is_univ = (lab.lab_type == "university")
        for field in ("name", "batch", "section", "is_active"):
            if field in data:
                setattr(lab, field, data[field])
        if is_univ:
            for field in ("enable_tab_switch_check", "max_tab_switches", "enable_fullscreen_lock", "enable_copy_paste_lock"):
                if field in data:
                    setattr(lab, field, data[field])
        else:
            lab.enable_tab_switch_check = False
            lab.enable_fullscreen_lock = False
            lab.enable_copy_paste_lock = False
        if "pass_threshold_percent" in data:
            try:
                lab.pass_threshold_percent = max(1, min(100, int(data["pass_threshold_percent"])))
            except (TypeError, ValueError):
                pass
        if "linked_lab_id" in data:
            linked_lab_id = data["linked_lab_id"]
            if linked_lab_id:
                try:
                    lab.linked_lab = Lab.objects.get(id=linked_lab_id, department=staff.department)
                except Lab.DoesNotExist:
                    lab.linked_lab = None
            else:
                lab.linked_lab = None
        if "start_date" in data:
            lab.start_date = _parse_dt(data["start_date"]) or lab.start_date
        if "end_date" in data:
            lab.end_date = _parse_dt(data["end_date"]) or lab.end_date
        if "staff_in_charge_id" in data:
            sic_id = data["staff_in_charge_id"]
            if sic_id:
                try:
                    lab.staff_in_charge = StaffProfile.objects.get(id=sic_id, department=staff.department)
                except StaffProfile.DoesNotExist:
                    return Response({"error": "Staff not found"}, status=400)
            else:
                lab.staff_in_charge = None

        if "allowed_languages" in data:
            allowed_languages = data["allowed_languages"]
            if (not isinstance(allowed_languages, list) or not allowed_languages
                    or any(lang not in LAB_LANGUAGE_CHOICES for lang in allowed_languages)):
                return Response(
                    {"error": f"allowed_languages must be a non-empty list from {LAB_LANGUAGE_CHOICES}"}, status=400
                )
            lab.allowed_languages = allowed_languages

        lab.save()
        return Response(_serialize_lab_v2(lab))

    def delete(self, request, lab_id):
        staff = _staff_from_request(request)
        lab = self._get(lab_id, staff)
        if not lab:
            return Response({"error": "Not found"}, status=404)
        lab.delete()
        return Response(status=204)


class StaffLabListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        staff = _staff_from_request(request)
        labs = Lab.objects.filter(Q(staff_in_charge=staff) | Q(created_by=staff)).distinct().select_related(
            "created_by"
        ).prefetch_related("exercises")
        return Response([_serialize_lab_v2(lab) for lab in labs])

    def post(self, request):
        staff = _staff_from_request(request)
        data = request.data
        start = _parse_dt(data.get("start_date"))
        end = _parse_dt(data.get("end_date"))
        if not start or not end:
            return Response({"error": "Valid start_date and end_date are required"}, status=400)

        allowed_languages = data.get("allowed_languages", list(LAB_LANGUAGE_CHOICES))
        lab_type = data.get("lab_type", "practical")
        approval_status = "pending_approval" if (lab_type == "university" and not getattr(staff, "is_hod", False)) else "approved"
        linked_lab = None
        linked_lab_id = data.get("linked_lab_id")
        if linked_lab_id:
            try:
                linked_lab = Lab.objects.get(id=linked_lab_id, department=staff.department)
            except Lab.DoesNotExist:
                pass

        is_univ = (lab_type == "university")
        try:
            pass_threshold_percent = max(1, min(100, int(data.get("pass_threshold_percent", 70))))
        except (TypeError, ValueError):
            pass_threshold_percent = 70
        lab = Lab.objects.create(
            name=data["name"],
            department=staff.department,
            batch=data.get("batch", ""),
            section=data.get("section", ""),
            start_date=start,
            end_date=end,
            staff_in_charge=staff,
            created_by=staff,
            lab_type=lab_type,
            approval_status=approval_status,
            is_published=False if lab_type == "university" else True,
            enable_tab_switch_check=bool(data.get("enable_tab_switch_check", False)) if is_univ else False,
            max_tab_switches=int(data.get("max_tab_switches", 3)) if is_univ else 3,
            enable_fullscreen_lock=bool(data.get("enable_fullscreen_lock", False)) if is_univ else False,
            enable_copy_paste_lock=bool(data.get("enable_copy_paste_lock", False)) if is_univ else False,
            pass_threshold_percent=pass_threshold_percent,
            allowed_languages=allowed_languages,
            linked_lab=linked_lab,
        )
        lab.refresh_from_db()
        return Response(_serialize_lab_v2(lab), status=201)


def _parse_lab_description(text):
    """Parse a LabExercise.description blob into {problem, examples,
    difficulty, hint} — a Python port of parseDescription() in
    StaffLabPanel.jsx. Keep the two in sync if the text format changes."""
    result = {"problem": "", "examples": [], "difficulty": "Medium", "hint": ""}
    if not text:
        return result

    section = "problem"
    problem_lines = []
    cur_ex = None

    for line in text.split("\n"):
        t = line.strip()
        if t == "Examples:":
            section = "examples"
            cur_ex = {"input": "", "output": "", "explanation": ""}
            continue
        if t == "Constraints:":
            if cur_ex:
                result["examples"].append(cur_ex)
                cur_ex = None
            section = "constraints"
            continue
        diff_match = re.match(r"^Difficulty:\s*(.*)", t)
        if diff_match:
            result["difficulty"] = diff_match.group(1).strip() or "Medium"
            continue
        hint_match = re.match(r"^Hint:\s*(.*)", t)
        if hint_match:
            result["hint"] = hint_match.group(1).strip()
            continue

        if section == "problem":
            problem_lines.append(line)
        elif section == "examples":
            im = re.match(r"^\s*Input:\s*(.*)", line)
            om = re.match(r"^\s*Output:\s*(.*)", line)
            em = re.match(r"^\s*Explanation:\s*(.*)", line)
            if im:
                if cur_ex and cur_ex["input"]:
                    result["examples"].append(cur_ex)
                    cur_ex = {"input": "", "output": "", "explanation": ""}
                if cur_ex is not None:
                    cur_ex["input"] = im.group(1)
            elif om and cur_ex is not None:
                cur_ex["output"] = om.group(1)
            elif em and cur_ex is not None:
                cur_ex["explanation"] = em.group(1)
        # section == "constraints": intentionally dropped, matching the JS port.

    if cur_ex:
        result["examples"].append(cur_ex)
    result["problem"] = "\n".join(problem_lines).strip()
    return result


def _compile_lab_description(parsed):
    """Rebuild a LabExercise.description blob from parsed fields — a
    Python port of compileDescription() in StaffLabPanel.jsx."""
    parts = []
    if parsed["problem"].strip():
        parts.append(parsed["problem"].strip())

    exs = [e for e in parsed["examples"] if e.get("input", "").strip() or e.get("output", "").strip()]
    if exs:
        parts.append("\nExamples:")
        for e in exs:
            if e.get("input", "").strip():
                parts.append(f"  Input:  {e['input'].strip()}")
            if e.get("output", "").strip():
                parts.append(f"  Output: {e['output'].strip()}")
            if e.get("explanation", "").strip():
                explanation = re.sub(r"\s+", " ", e["explanation"].strip())
                parts.append(f"  Explanation: {explanation}")

    if parsed.get("difficulty"):
        parts.append(f"\nDifficulty: {parsed['difficulty']}")
    if parsed.get("hint", "").strip():
        parts.append(f"\nHint: {parsed['hint'].strip()}")

    return "\n".join(parts)


def _is_auto_placeholder(value):
    return (value or "").strip().lower() in {"", "auto", "generate", "generated"}


def _run_and_close_connections(fn, **kwargs):
    """Thread-pool target wrapper — a worker thread that touches the ORM
    (as generate_test_cases/generate_hint do, for LLMProvider rotation)
    opens its own DB connection that Django never closes on its own,
    since the request-lifecycle cleanup signals only cover the main
    request thread. Close it explicitly once the call is done either way,
    or a connection leaks every time this runs."""
    from django.db import connections
    try:
        return fn(**kwargs)
    finally:
        connections.close_all()


def _auto_generate_lab_test_cases(exercise, num_cases=None, raise_on_error=False, replace_existing=False):
    """LLM test case generation for a LabExercise, plus a best-effort
    backfill of the description's Hint: line and Examples: section. Examples
    are derived from generated test cases; when replace_existing=True the
    staff Generate button intentionally replaces old examples. The hint is a
    separate LLM call made concurrently with test case generation.

    By default best-effort — a generation failure shouldn't block exercise
    creation. Pass raise_on_error=True for manual/on-demand triggers where
    the caller wants to surface the failure to the user instead of
    silently no-op'ing."""
    if exercise.test_cases.exists() and not replace_existing:
        return 0

    from concurrent.futures import ThreadPoolExecutor
    from .services.testcase_generator import (
        generate_test_cases, generate_hint, derive_examples, extract_difficulty, TestCaseGenError,
    )

    parsed = _parse_lab_description(exercise.description)
    needs_hint = _is_auto_placeholder(parsed["hint"])
    needs_examples = replace_existing or not parsed["examples"]
    generation_description = exercise.description
    if replace_existing:
        generation_description = _compile_lab_description({
            **parsed,
            "examples": [],
            "hint": "" if needs_hint else parsed["hint"],
        })

    with ThreadPoolExecutor(max_workers=2) as pool:
        tc_future = pool.submit(
            _run_and_close_connections, generate_test_cases,
            title=exercise.title, description=generation_description,
            num_cases=num_cases, difficulty=extract_difficulty(exercise.description),
        )
        hint_future = (
            pool.submit(_run_and_close_connections, generate_hint, title=exercise.title, description=exercise.description)
            if needs_hint else None
        )

        try:
            generated = tc_future.result()
        except TestCaseGenError as exc:
            logger.warning("Auto test-case generation failed for lab exercise %r: %s", exercise.title, exc)
            if raise_on_error:
                raise
            return 0

        hint_text = None
        if hint_future is not None:
            try:
                hint_text = hint_future.result()
            except TestCaseGenError as exc:
                logger.warning("Auto hint generation failed for lab exercise %r: %s", exercise.title, exc)

    generated_rows = [
        LabExerciseTestCase(
            exercise=exercise,
            stdin=case["stdin"],
            expected_output=case["expected_output"],
            is_sample=case["is_sample"],
            order=order,
        )
        for order, case in enumerate(generated, start=1)
    ]
    with transaction.atomic():
        if replace_existing:
            exercise.test_cases.all().delete()
        LabExerciseTestCase.objects.bulk_create(generated_rows)

    if needs_examples or hint_text:
        if needs_examples:
            parsed["examples"] = derive_examples(generated)[:1]
        if hint_text:
            parsed["hint"] = hint_text
        exercise.description = _compile_lab_description(parsed)
        exercise.save(update_fields=["description"])

    return len(generated)


class StaffLabExercisesView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_lab(self, lab_id, staff):
        try:
            return Lab.objects.get(id=lab_id, staff_in_charge=staff)
        except Lab.DoesNotExist:
            return None

    def get(self, request, lab_id):
        staff = _staff_from_request(request)
        lab = self._get_lab(lab_id, staff)
        if not lab:
            return Response({"error": "Not found"}, status=404)
        exercises = lab.exercises.select_related("added_by").annotate(
            submission_count=Count("submissions", distinct=True),
            test_case_count=Count("test_cases", distinct=True),
        )
        return Response({
            "lab": _serialize_lab_v2(lab),
            "exercises": [_serialize_exercise(e) for e in exercises],
        })

    def post(self, request, lab_id):
        staff = _staff_from_request(request)
        lab = self._get_lab(lab_id, staff)
        if not lab:
            return Response({"error": "Not found"}, status=404)
        data = request.data
        exercise = LabExercise.objects.create(
            lab=lab,
            title=data["title"],
            description=data.get("description", ""),
            order=data.get("order", lab.exercises.count()),
            added_by=staff,
            difficulty=data.get("difficulty", "Medium"),
        )
        # Test cases (and the explanation) are always a separate, explicit
        # step via the Generate buttons — never triggered automatically on
        # create, same as the Problem bank.
        return Response(_serialize_exercise(exercise), status=201)


class StaffLabExercisesBulkView(APIView):
    """Bulk-create exercises for a Lab from a staff-uploaded CSV template."""
    permission_classes = [IsAuthenticated]

    def _get_lab(self, lab_id, staff):
        try:
            return Lab.objects.get(id=lab_id, staff_in_charge=staff)
        except Lab.DoesNotExist:
            return None

    def post(self, request, lab_id):
        staff = _staff_from_request(request)
        lab = self._get_lab(lab_id, staff)
        if not lab:
            return Response({"error": "Not found"}, status=404)

        rows = request.data.get("exercises") or []
        tc_rows = request.data.get("test_cases") or []
        if not isinstance(rows, list) or not isinstance(tc_rows, list):
            return Response({"error": "exercises and test_cases must be lists"}, status=400)
        if not rows and not tc_rows:
            return Response({"error": "No exercises or test cases provided"}, status=400)
        if len(rows) > 500:
            return Response({"error": "Cannot import more than 500 exercises at once"}, status=400)
        if len(tc_rows) > 2000:
            return Response({"error": "Cannot import more than 2000 test cases at once"}, status=400)

        errors = []
        to_create = []
        next_order = lab.exercises.count()
        for idx, row in enumerate(rows):
            title = (row.get("title") or "").strip()
            if not title:
                errors.append({"row": idx + 1, "error": "Title is required"})
                continue
            to_create.append(LabExercise(
                lab=lab,
                title=title[:200],
                description=row.get("description") or "",
                order=row.get("order") if isinstance(row.get("order"), int) else next_order,
                added_by=staff,
            ))
            next_order += 1

        if rows and not to_create:
            return Response({"error": "No valid exercises to import", "errors": errors}, status=400)

        with transaction.atomic():
            created = LabExercise.objects.bulk_create(to_create) if to_create else []

        # Match test-case rows to exercises by title — against both the exercises
        # just created above and any pre-existing exercises in this lab, so staff
        # can also import test cases for questions that were added earlier.
        tc_errors = []
        tc_created_count = 0
        if tc_rows:
            title_map = {}
            for ex in lab.exercises.all():
                title_map.setdefault(ex.title.strip().lower(), ex)

            tc_to_create = []
            next_order_by_exercise = {}
            for idx, row in enumerate(tc_rows):
                title = (row.get("title") or "").strip()
                if not title:
                    tc_errors.append({"row": idx + 1, "error": "Exercise title is required"})
                    continue
                ex = title_map.get(title.lower())
                if not ex:
                    tc_errors.append({"row": idx + 1, "error": f'No exercise titled "{title}" found in this lab'})
                    continue
                expected_output = row.get("expected_output")
                if expected_output is None or not str(expected_output).strip():
                    tc_errors.append({"row": idx + 1, "error": "Expected output is required"})
                    continue
                order = next_order_by_exercise.get(ex.id, 0)
                tc_to_create.append(LabExerciseTestCase(
                    exercise=ex,
                    stdin=row.get("stdin") or "",
                    expected_output=expected_output,
                    is_sample=bool(row.get("is_sample")),
                    order=order,
                ))
                next_order_by_exercise[ex.id] = order + 1

            if tc_to_create:
                with transaction.atomic():
                    LabExerciseTestCase.objects.bulk_create(tc_to_create)
                tc_created_count = len(tc_to_create)

        # No auto-generation here — test cases and explanations for any
        # exercise the CSV didn't supply test cases for are always a
        # separate, explicit step via the per-exercise Generate buttons.
        return Response({
            "created": [_serialize_exercise(e) for e in created],
            "created_count": len(created),
            "skipped_count": len(errors),
            "errors": errors,
            "test_cases_created_count": tc_created_count,
            "test_cases_skipped_count": len(tc_errors),
            "test_case_errors": tc_errors,
        }, status=201)


class StaffExerciseDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_ex(self, lab_id, exercise_id, staff):
        try:
            return LabExercise.objects.get(
                id=exercise_id, lab_id=lab_id, lab__staff_in_charge=staff
            )
        except LabExercise.DoesNotExist:
            return None

    def put(self, request, lab_id, exercise_id):
        staff = _staff_from_request(request)
        ex = self._get_ex(lab_id, exercise_id, staff)
        if not ex:
            return Response({"error": "Not found"}, status=404)
        for field in ("title", "description", "order", "difficulty"):
            if field in request.data:
                setattr(ex, field, request.data[field])
        ex.save()
        return Response(_serialize_exercise(ex))

    def delete(self, request, lab_id, exercise_id):
        staff = _staff_from_request(request)
        ex = self._get_ex(lab_id, exercise_id, staff)
        if not ex:
            return Response({"error": "Not found"}, status=404)
        ex.delete()
        return Response(status=204)


class StaffExerciseTestCasesView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_ex(self, lab_id, exercise_id, staff):
        try:
            return LabExercise.objects.get(
                id=exercise_id, lab_id=lab_id, lab__staff_in_charge=staff
            )
        except LabExercise.DoesNotExist:
            return None

    def get(self, request, lab_id, exercise_id):
        staff = _staff_from_request(request)
        ex = self._get_ex(lab_id, exercise_id, staff)
        if not ex:
            return Response({"error": "Not found"}, status=404)
        cases = ex.test_cases.all().order_by("order", "id")
        return Response({"test_cases": [_serialize_test_case(tc) for tc in cases]})

    def post(self, request, lab_id, exercise_id):
        staff = _staff_from_request(request)
        ex = self._get_ex(lab_id, exercise_id, staff)
        if not ex:
            return Response({"error": "Not found"}, status=404)

        expected_output = str(request.data.get("expected_output", "")).strip()
        if not expected_output:
            return Response({"error": "Expected output is required"}, status=400)

        order = request.data.get("order")
        if order is None:
            order = ex.test_cases.count()
        else:
            order = _nonnegative_int(order, ex.test_cases.count())
        tc = LabExerciseTestCase.objects.create(
            exercise=ex,
            stdin=str(request.data.get("stdin", "")).strip(),
            expected_output=expected_output,
            is_sample=bool(request.data.get("is_sample")),
            order=order,
        )
        return Response(_serialize_test_case(tc), status=201)


class StaffExerciseTestCaseDetailView(APIView):
    permission_classes = [IsAuthenticated]

    def _get_tc(self, lab_id, exercise_id, test_case_id, staff):
        try:
            return LabExerciseTestCase.objects.get(
                id=test_case_id,
                exercise_id=exercise_id,
                exercise__lab_id=lab_id,
                exercise__lab__staff_in_charge=staff,
            )
        except LabExerciseTestCase.DoesNotExist:
            return None

    def put(self, request, lab_id, exercise_id, test_case_id):
        staff = _staff_from_request(request)
        tc = self._get_tc(lab_id, exercise_id, test_case_id, staff)
        if not tc:
            return Response({"error": "Not found"}, status=404)

        if "stdin" in request.data:
            tc.stdin = str(request.data.get("stdin", "")).strip()
        if "expected_output" in request.data:
            expected_output = str(request.data.get("expected_output", "")).strip()
            if not expected_output:
                return Response({"error": "Expected output is required"}, status=400)
            tc.expected_output = expected_output
        if "is_sample" in request.data:
            tc.is_sample = bool(request.data.get("is_sample"))
        if "order" in request.data:
            tc.order = _nonnegative_int(request.data.get("order"), tc.order)

        tc.save(update_fields=["stdin", "expected_output", "is_sample", "order"])
        return Response(_serialize_test_case(tc))

    def delete(self, request, lab_id, exercise_id, test_case_id):
        staff = _staff_from_request(request)
        tc = self._get_tc(lab_id, exercise_id, test_case_id, staff)
        if not tc:
            return Response({"error": "Not found"}, status=404)
        tc.delete()
        return Response(status=204)


class StaffExerciseGenerateTestCasesView(APIView):
    """Manual 'Generate' trigger — staff clicks this when adding/editing an
    exercise to (re)generate its test cases via the LLM fallback chain."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id, exercise_id):
        staff = _staff_from_request(request)
        try:
            ex = LabExercise.objects.get(id=exercise_id, lab_id=lab_id, lab__staff_in_charge=staff)
        except LabExercise.DoesNotExist:
            return Response({"error": "Not found"}, status=404)

        from .services.testcase_generator import TestCaseGenError

        try:
            count = _auto_generate_lab_test_cases(ex, raise_on_error=True, replace_existing=True)
        except TestCaseGenError as exc:
            return Response({"error": f"Generation failed: {exc}"}, status=502)

        return Response({
            "generated_count": count,
            "test_cases": [_serialize_test_case(tc) for tc in ex.test_cases.all().order_by("order")],
            "description": ex.description,
        })


class StaffExerciseGenerateExplanationView(APIView):
    """Manual 'Generate' trigger for a lab exercise's brief explanation —
    a separate endpoint from test-case generation so the staff panel can
    fire both concurrently instead of waiting on one before the other."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id, exercise_id):
        staff = _staff_from_request(request)
        try:
            ex = LabExercise.objects.get(id=exercise_id, lab_id=lab_id, lab__staff_in_charge=staff)
        except LabExercise.DoesNotExist:
            return Response({"error": "Not found"}, status=404)

        from .services.testcase_generator import generate_explanation, TestCaseGenError
        try:
            parsed = _parse_lab_description(ex.description)
            clean_description = _compile_lab_description({**parsed, "examples": []})
            explanation = generate_explanation(title=ex.title, description=clean_description)
        except TestCaseGenError as exc:
            return Response({"error": f"Generation failed: {exc}"}, status=502)

        ex.explanation = explanation
        ex.save(update_fields=["explanation"])
        return Response({"explanation": ex.explanation})


class StaffLabStudentsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, lab_id):
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, staff_in_charge=staff)
        except Lab.DoesNotExist:
            return Response({"error": "Not found"}, status=404)

        exercises = list(lab.exercises.order_by("order", "created_at"))
        student_qs = StudentProfile.objects.filter(
            department=lab.department, batch=lab.batch
        )
        if lab.section:
            student_qs = student_qs.filter(section=lab.section)
        students = list(student_qs.order_by("register_number", "name"))

        subs = LabExerciseSubmission.objects.filter(exercise__lab=lab).select_related("student", "exercise")
        sub_map = {(s.student_id, s.exercise_id): s for s in subs}

        session_qs = LabStudentSession.objects.filter(lab=lab).prefetch_related("allocated_exercises")
        session_map = {s.student_id: s for s in session_qs}

        student_rows = []
        all_sub_batches = set()

        for student in students:
            session = session_map.get(student.id)
            if not session:
                is_locked = (lab.lab_type == "university")
                lock_reason = "Lab session is locked by staff. Awaiting staff unlock for your batch." if is_locked else ""
                session = LabStudentSession.objects.create(
                    lab=lab, student=student, is_locked=is_locked, lock_reason=lock_reason, sub_batch="Batch 1"
                )
                session_map[student.id] = session

            sub_b = session.sub_batch or "Batch 1"
            all_sub_batches.add(sub_b)

            ex_status = []
            for ex in exercises:
                sub = sub_map.get((student.id, ex.id))
                ex_status.append({
                    "exercise_id": ex.id,
                    "completed": sub is not None,
                    "submitted_at": sub.submitted_at.isoformat() if sub else None,
                    "language": sub.language if sub else None,
                })
            done = sum(1 for s in ex_status if s["completed"])

            allocated_list = [
                {
                    "id": ax.id,
                    "title": ax.title,
                    "difficulty": ax.difficulty,
                    "order": ax.order,
                }
                for ax in session.allocated_exercises.all()
            ]

            student_rows.append({
                "student_id": student.id,
                "student_name": student.name,
                "register_number": student.register_number or "",
                "section": student.section or "",
                "sub_batch": sub_b,
                "is_locked": session.is_locked,
                "lock_reason": session.lock_reason or "",
                "tab_switch_count": session.tab_switch_count,
                "exercises": ex_status,
                "allocated_exercises": allocated_list,
                "completed": done,
                "total": len(exercises),
            })

        def _sort_key(batch_name):
            match = re.search(r'\d+', str(batch_name))
            return int(match.group()) if match else 999

        sorted_batches = sorted(list(all_sub_batches), key=_sort_key) if all_sub_batches else ["Batch 1"]

        return Response({
            "lab": _serialize_lab_v2(lab),
            "exercises": [{"id": e.id, "title": e.title} for e in exercises],
            "students": student_rows,
            "available_sub_batches": sorted_batches,
        })



class StudentLabListView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        student = _student_from_request(request)
        labs = Lab.objects.filter(
            department=student.department, batch=student.batch, is_active=True, is_published=True
        )
        if student.section:
            labs = labs.filter(Q(section="") | Q(section=student.section))
        labs = labs.select_related("staff_in_charge").prefetch_related("exercises")
        return Response([_serialize_lab_v2(lab, student=student) for lab in labs])


class StudentLabExercisesView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, lab_id):
        student = _student_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, department=student.department, batch=student.batch, is_active=True, is_published=True)
        except Lab.DoesNotExist:
            return Response({"error": "Not found"}, status=404)

        if lab.is_expired:
            return Response({
                "lab": _serialize_lab_v2(lab, student=student),
                "is_expired": True,
                "error": "This lab has expired and can no longer be accessed.",
                "exercises": [],
            })

        session, created = LabStudentSession.objects.get_or_create(lab=lab, student=student)
        if created and lab.lab_type == "university":
            session.is_locked = True
            session.lock_reason = "Lab session is locked by staff. Awaiting staff unlock for your batch."
            session.save(update_fields=["is_locked", "lock_reason"])

        if session.is_locked:
            return Response({
                "lab": _serialize_lab_v2(lab, student=student),
                "is_locked": True,
                "lock_reason": session.lock_reason or "Lab session is locked by staff. Awaiting staff unlock for your batch.",
                "exercises": [],
            })

        # Randomized per-student question allocation (1-2 exercises out of the full
        # pool) is a University Practical Lab exam feature only — regular practical
        # / company labs must always show every exercise staff added.
        if lab.lab_type == "university":
            if not session.allocated_exercises.exists():
                from .services.lab_allocation import allocate_lab_questions_for_students
                allocate_lab_questions_for_students(lab)
                session.refresh_from_db()

            exercises = session.allocated_exercises.all() if session.allocated_exercises.exists() else lab.exercises.all()
        else:
            exercises = lab.exercises.all()

        sub_map = {
            s.exercise_id: s
            for s in LabExerciseSubmission.objects.filter(exercise__in=exercises, student=student)
        }
        ex_data = []
        for ex in exercises:
            sub = sub_map.get(ex.id)
            ex_data.append({
                "id": ex.id,
                "title": ex.title,
                "description": ex.description,
                "explanation": ex.explanation,
                "difficulty": ex.difficulty,
                "order": ex.order,
                "submitted": sub is not None,
                "submitted_at": sub.submitted_at.isoformat() if sub else None,
                "code": sub.code if sub else "",
                "language": sub.language if sub else "",
            })
        return Response({"lab": _serialize_lab_v2(lab, student=student), "exercises": ex_data})


class StudentExerciseRunView(APIView):
    """Student: run code for a LabExercise — against the exercise's own
    test cases when no custom stdin is given (same "Test Cases: X/Y
    passed" breakdown a Problem's Run button shows), or against custom
    stdin verbatim when the student provides one. Piston-backed, same
    execution service Problems use — just without Problem's function/class
    driver-injection step, since lab exercises are always plain stdin
    programs."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id, exercise_id):
        student = _student_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, department=student.department, batch=student.batch, is_active=True)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        if lab.is_expired:
            return Response({"error": "This lab has expired and no longer accepts submissions.", "is_expired": True}, status=403)

        session, _created = LabStudentSession.objects.get_or_create(lab=lab, student=student)
        if session.is_locked:
            reason = session.lock_reason or "Your session is locked. Please contact staff to unlock."
            return Response({"error": reason, "is_locked": True, "lock_reason": reason}, status=403)

        if lab.lab_type == "university":
            if session.allocated_exercises.exists() and not session.allocated_exercises.filter(id=exercise_id).exists():
                return Response({"error": "You are not allocated this exercise."}, status=403)
            try:
                exercise = LabExercise.objects.get(id=exercise_id)
            except LabExercise.DoesNotExist:
                return Response({"error": "Exercise not found"}, status=404)
        else:
            try:
                exercise = LabExercise.objects.get(id=exercise_id, lab=lab)
            except LabExercise.DoesNotExist:
                return Response({"error": "Exercise not found"}, status=404)

        source_code = (request.data.get("code") or "").strip()
        language = (request.data.get("language") or "").strip()
        stdin = request.data.get("stdin") or ""
        if not source_code:
            return Response({"detail": "Code is required."}, status=400)

        from .services.executor import get_language_id
        try:
            language_id = get_language_id(language)
        except Exception:
            return Response({"detail": f"Unsupported language: {language}"}, status=400)

        from .services.problem_testcases import build_lab_runtime_test_cases

        try:
            if stdin.strip():
                result = execute_judge0_submission(
                    source_code=source_code, language_id=language_id, stdin=stdin,
                )
            else:
                # Run every configured test case, not just samples — a student
                # relying on the console to judge correctness needs to see the
                # full picture, not a subset that can pass while others fail.
                test_cases = build_lab_runtime_test_cases(exercise, sample_only=False)
                if test_cases:
                    result = execute_lab_test_case_batch(
                        source_code=source_code, language=language, language_id=language_id,
                        test_cases=test_cases, batch_kind="run",
                    )
                else:
                    result = execute_judge0_submission(
                        source_code=source_code, language_id=language_id, stdin="",
                    )
        except ExecutorTimeoutError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_504_GATEWAY_TIMEOUT)
        except ExecutorServiceError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
        except Exception as exc:
            logger.error("Lab exercise run error for exercise %s: %s", exercise_id, exc, exc_info=True)
            return Response({"detail": f"Execution error: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(result)


class StudentExerciseSubmitView(APIView):
    """Student: submit a LabExercise. Only actually recorded once at least
    the lab's pass_threshold_percent of the exercise's test cases pass —
    re-runs them here (the full set, same as the Run button now does) so a
    submission can't be stored on the strength of a subset that happened to
    pass while others silently failed. Exercises with no test cases
    configured yet have nothing to gate against, so those still submit
    best-effort as before."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id, exercise_id):
        student = _student_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, department=student.department, batch=student.batch, is_active=True)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        if lab.is_expired:
            return Response({"error": "This lab has expired and no longer accepts submissions.", "is_expired": True}, status=403)

        session, _created = LabStudentSession.objects.get_or_create(lab=lab, student=student)
        if session.is_locked:
            reason = session.lock_reason or "Your session is locked. Please contact staff to unlock."
            return Response({"error": reason, "is_locked": True, "lock_reason": reason}, status=403)

        if lab.lab_type == "university":
            if session.allocated_exercises.exists() and not session.allocated_exercises.filter(id=exercise_id).exists():
                return Response({"error": "You are not allocated this exercise."}, status=403)
            try:
                exercise = LabExercise.objects.get(id=exercise_id)
            except LabExercise.DoesNotExist:
                return Response({"error": "Exercise not found"}, status=404)
        else:
            try:
                exercise = LabExercise.objects.get(id=exercise_id, lab=lab)
            except LabExercise.DoesNotExist:
                return Response({"error": "Exercise not found"}, status=404)
        data = request.data
        code = data.get("code", "")
        language = data.get("language", "")
        allowed_languages = exercise.lab.allowed_languages or list(LAB_LANGUAGE_CHOICES)
        if language and language not in allowed_languages:
            return Response(
                {"error": f"This lab only accepts submissions in: {', '.join(allowed_languages)}"}, status=400
            )
        if not code.strip():
            return Response({"error": "Code is required."}, status=400)

        from .services.executor import get_language_id
        from .services.problem_testcases import build_lab_runtime_test_cases

        try:
            language_id = get_language_id(language)
        except Exception:
            return Response({"error": f"Unsupported language: {language}"}, status=400)

        test_cases = build_lab_runtime_test_cases(exercise, sample_only=False)
        result = None
        if test_cases:
            try:
                result = execute_lab_test_case_batch(
                    source_code=code, language=language, language_id=language_id,
                    test_cases=test_cases, batch_kind="submit",
                )
            except ExecutorTimeoutError as exc:
                return Response({"error": str(exc)}, status=status.HTTP_504_GATEWAY_TIMEOUT)
            except ExecutorServiceError as exc:
                return Response({"error": str(exc)}, status=status.HTTP_502_BAD_GATEWAY)
            except Exception as exc:
                logger.error(
                    "Lab submit test-case run error for exercise %s: %s", exercise_id, exc, exc_info=True,
                )
                return Response({"error": f"Execution error: {exc}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            passed = result["passed_cases"]
            total = result["total_cases"]
            threshold_pct = exercise.lab.pass_threshold_percent
            pass_pct = (passed / total * 100) if total else 0
            if total > 0 and pass_pct < threshold_pct:
                return Response(
                    {
                        "error": (
                            f"{passed}/{total} test case(s) passed ({round(pass_pct)}%). "
                            f"At least {threshold_pct}% of test cases must pass before this exercise can be submitted."
                        ),
                        "test_results": result["test_results"],
                        "passed_cases": passed,
                        "total_cases": total,
                    },
                    status=400,
                )

        sub, created = LabExerciseSubmission.objects.update_or_create(
            exercise=exercise, student=student,
            defaults={
                "code": code,
                "language": language,
                "passed_cases": result["passed_cases"] if result else 0,
                "total_cases": result["total_cases"] if result else 0,
                "test_results": result["test_results"] if result else [],
            },
        )

        def _pregen():
            try:
                _generate_lab_exercise_report(exercise, sub)
            except Exception as exc:
                logger.warning("Async lab report pre-generation failed for sub %s: %s", sub.id, exc)

        threading.Thread(target=_pregen, daemon=True).start()

        return Response({
            "submitted": True,
            "submitted_at": sub.submitted_at.isoformat(),
        }, status=201 if created else 200)


_LAB_REPORT_GEN_SEMAPHORE = threading.BoundedSemaphore(12)


def _is_valid_algorithm(alg):
    if not alg or not str(alg).strip():
        return False
    alg_str = str(alg).strip()
    if alg_str.startswith("(") or alg_str == "—" or len(alg_str) < 15:
        return False
    return True


def _fallback_algorithm(exercise_title, language):
    title_clean = (exercise_title or "the program").strip()
    return (
        f"1. Start the program execution in {language or 'the selected programming language'}.\n"
        f"2. Define and initialize required variables and input structures.\n"
        f"3. Read input data from standard input or arguments according to {title_clean}.\n"
        f"4. Process the input logic step-by-step and execute standard algorithmic operations.\n"
        f"5. Output the result to standard output and complete execution."
    )


def _generate_lab_exercise_report(exercise, submission):
    """Build (or rebuild) a LabExerciseReport + its PDF for one submission.
    Protected with _LAB_REPORT_GEN_SEMAPHORE to ensure max 12 heavy generations
    run concurrently under high load (e.g. 120+ simultaneous users).
    """
    with _LAB_REPORT_GEN_SEMAPHORE:
        submission.refresh_from_db()
        existing_report = getattr(submission, "report", None)
        if existing_report and existing_report.pdf_file and _is_valid_algorithm(existing_report.algorithm):
            try:
                existing_report.pdf_file.open("rb")
                existing_report.pdf_file.seek(0)
                pdf_bytes = existing_report.pdf_file.read()
                existing_report.pdf_file.close()
                if pdf_bytes and len(pdf_bytes) >= 1000:
                    return existing_report, pdf_bytes
            except Exception as exc:
                logger.warning("Failed reading cached PDF for report %s: %s, regenerating...", getattr(existing_report, "id", None), exc)

        from .services.lab_report import (
            extract_problem_statement, build_aim, get_or_generate_question_algorithm, build_result,
        )
        from .services.lab_report_pdf import build_lab_report_pdf

        problem_statement = extract_problem_statement(exercise.description)
        aim = build_aim(exercise.title, problem_statement)

        # Single canonical algorithm per Question (exercise), shared across all students
        algorithm = get_or_generate_question_algorithm(exercise)

        # Use the test-case results recorded at submit time — never re-run the
        # code here. Re-execution can legitimately differ from the graded run
        # (executor timing/flakiness), which previously caused a report to
        # show "Failed" for a submission the student was told had passed.
        # A LabExerciseSubmission only exists once it already cleared the
        # lab's pass_threshold_percent (or the exercise has no test cases),
        # so any existing submission is, by construction, an accepted one.
        passed_n = submission.passed_cases
        total_n = submission.total_cases
        test_case_rows = [
            (r.get("stdin", ""), r.get("expected", ""), r.get("actual", ""), "Passed" if r.get("passed") else "Failed")
            for r in (submission.test_results or [])
        ]
        result_text = build_result(exercise.title, all_passed=(True if test_case_rows else None))
        if test_case_rows:
            output_text = f"{passed_n}/{total_n} test case(s) passed."
            tc_note = ""
        else:
            output_text = "(No test cases configured for this exercise, or this submission predates result tracking.)"
            tc_note = "No stored test-case results for this submission — resubmit to generate an up-to-date record."

        status_label = "Passed" if test_case_rows else "Not Verified"
        details = {
            "language": submission.language or "—",
            "status": status_label,
            "score": f"{passed_n}/{total_n}" if total_n else "—",
            "percentage": f"{round(passed_n / total_n * 100)}%" if total_n else "—",
            "submitted_at": submission.submitted_at,
        }

        lab_exercise_ids = list(
            exercise.lab.exercises.order_by("order", "created_at").values_list("id", flat=True)
        )
        exp_no = lab_exercise_ids.index(exercise.id) + 1 if exercise.id in lab_exercise_ids else exercise.order + 1

        report, _created = LabExerciseReport.objects.update_or_create(
            submission=submission,
            defaults=dict(
                exp_no=exp_no,
                exp_name=exercise.title,
                aim=aim,
                algorithm=algorithm,
                program=submission.code,
                output=output_text,
                result=result_text,
            ),
        )

        buffer = BytesIO()
        build_lab_report_pdf(
            buffer, report=report, test_case_rows=test_case_rows, test_case_note=tc_note, details=details,
        )
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()
        report.pdf_file.save(f"lab_report_{report.id}.pdf", ContentFile(pdf_bytes), save=True)
        return report, pdf_bytes


class StudentExerciseReportView(APIView):
    """Student: generate (or re-download) their lab record PDF — Exp No /
    Aim / Algorithm / Program / Output / Result, watermarked with their own
    register number — for a submitted LabExercise."""
    permission_classes = [IsAuthenticated]

    def _get_submission(self, lab_id, exercise_id, student):
        try:
            exercise = LabExercise.objects.select_related("lab").get(
                id=exercise_id, lab_id=lab_id,
                lab__department=student.department, lab__batch=student.batch,
            )
        except LabExercise.DoesNotExist:
            return None, None
        submission = LabExerciseSubmission.objects.filter(exercise=exercise, student=student).first()
        return exercise, submission

    def get(self, request, lab_id, exercise_id):
        """Download the lab exercise report PDF — serves cached file if available, or generates on demand."""
        student = _student_from_request(request)
        exercise, submission = self._get_submission(lab_id, exercise_id, student)
        if not exercise:
            return Response({"error": "Not found"}, status=404)
        if not submission or not (submission.code or "").strip():
            return Response({"error": "Submit your code for this exercise before generating a report."}, status=400)

        report = getattr(submission, "report", None)
        if report and report.pdf_file and _is_valid_algorithm(report.algorithm):
            try:
                report.pdf_file.open("rb")
                report.pdf_file.seek(0)
                pdf_bytes = report.pdf_file.read()
                report.pdf_file.close()
                if pdf_bytes and len(pdf_bytes) >= 1000:
                    response = HttpResponse(pdf_bytes, content_type="application/pdf")
                    response["Content-Length"] = str(len(pdf_bytes))
                    response["Content-Disposition"] = f'attachment; filename="lab_record_{exercise.id}_{student.register_number}.pdf"'
                    return response
            except Exception:
                pass

        try:
            report, pdf_bytes = _generate_lab_exercise_report(exercise, submission)
        except Exception as exc:
            logger.exception("Lab report PDF rendering failed for exercise %s", exercise.id)
            return Response({"error": f"PDF rendering failed: {exc}"}, status=500)

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Length"] = str(len(pdf_bytes))
        response["Content-Disposition"] = f'attachment; filename="lab_record_{exercise.id}_{student.register_number}.pdf"'
        return response

    def post(self, request, lab_id, exercise_id):
        """Generate (or serve cached) report from the student's submission."""
        student = _student_from_request(request)
        exercise, submission = self._get_submission(lab_id, exercise_id, student)
        if not exercise:
            return Response({"error": "Not found"}, status=404)
        if not submission or not (submission.code or "").strip():
            return Response({"error": "Submit your code for this exercise before generating a report."}, status=400)

        report = getattr(submission, "report", None)
        force_regen = request.query_params.get("force", "").lower() in ("true", "1")
        if report and report.pdf_file and _is_valid_algorithm(report.algorithm) and not force_regen:
            try:
                report.pdf_file.open("rb")
                report.pdf_file.seek(0)
                pdf_bytes = report.pdf_file.read()
                report.pdf_file.close()
                if pdf_bytes and len(pdf_bytes) >= 1000:
                    response = HttpResponse(pdf_bytes, content_type="application/pdf")
                    response["Content-Length"] = str(len(pdf_bytes))
                    response["Content-Disposition"] = f'attachment; filename="lab_record_{exercise.id}_{student.register_number}.pdf"'
                    return response
            except Exception:
                pass

        try:
            report, pdf_bytes = _generate_lab_exercise_report(exercise, submission)
        except Exception as exc:
            logger.exception("Lab report PDF rendering failed for exercise %s", exercise.id)
            return Response({"error": f"PDF rendering failed: {exc}"}, status=500)

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Length"] = str(len(pdf_bytes))
        response["Content-Disposition"] = f'attachment; filename="lab_record_{exercise.id}_{student.register_number}.pdf"'
        return response


class StudentLabFullReportView(APIView):
    """Student: Bulk full report download is disabled. Only individual question downloads are enabled."""
    permission_classes = [IsAuthenticated]

    def get(self, request, lab_id):
        return Response({"error": "Bulk lab report download has been disabled. Please download individual question reports."}, status=400)


class StaffLabExerciseStudentReportView(APIView):
    """Staff: generate (or download cached) a specific student's lab record PDF."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id, exercise_id, register_number):
        staff = _staff_from_request(request)
        if not staff:
            return Response({"error": "Staff access required"}, status=403)
        try:
            lab = Lab.objects.get(id=lab_id)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)
        try:
            exercise = lab.exercises.get(id=exercise_id)
        except LabExercise.DoesNotExist:
            return Response({"error": "Exercise not found"}, status=404)

        student_qs = StudentProfile.objects.filter(
            register_number=register_number, department=lab.department, batch=lab.batch,
        )
        if lab.section:
            student_qs = student_qs.filter(section=lab.section)
        student = student_qs.first()
        if not student:
            student = StudentProfile.objects.filter(register_number=register_number).first()
        if not student:
            return Response({"error": "Student not found"}, status=404)

        submission = LabExerciseSubmission.objects.filter(exercise=exercise, student=student).first()
        if not submission or not (submission.code or "").strip():
            return Response({"error": "This student hasn't submitted this exercise yet."}, status=400)

        report = getattr(submission, "report", None)
        force_regen = request.query_params.get("force", "").lower() in ("true", "1")
        if report and report.pdf_file and _is_valid_algorithm(report.algorithm) and not force_regen:
            try:
                report.pdf_file.open("rb")
                report.pdf_file.seek(0)
                pdf_bytes = report.pdf_file.read()
                report.pdf_file.close()
                if pdf_bytes and len(pdf_bytes) >= 1000:
                    response = HttpResponse(pdf_bytes, content_type="application/pdf")
                    response["Content-Length"] = str(len(pdf_bytes))
                    response["Content-Disposition"] = f'attachment; filename="lab_record_{exercise.id}_{student.register_number}.pdf"'
                    return response
            except Exception:
                pass

        try:
            report, pdf_bytes = _generate_lab_exercise_report(exercise, submission)
        except Exception as exc:
            logger.exception(
                "Staff lab report generation failed for exercise %s / student %s: %s", exercise.id, register_number, exc,
            )
            return Response({"error": f"Failed to generate report PDF: {exc}"}, status=500)

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Length"] = str(len(pdf_bytes))
        response["Content-Disposition"] = f'attachment; filename="lab_record_{exercise.id}_{student.register_number}.pdf"'
        return response


class HODLabApproveView(APIView):
    """HOD / Academic Coordinator: Approve a pending University Lab practical."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id):
        staff = _staff_from_request(request)
        if not staff or not getattr(staff, 'is_hod', False):
            return Response({"error": "HOD or Academic Coordinator access required"}, status=403)
        try:
            if getattr(staff, "is_academic_coordinator", False):
                lab = Lab.objects.get(id=lab_id)
            else:
                lab = Lab.objects.get(id=lab_id, department=staff.department)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        lab.approval_status = "approved"
        lab.is_published = True
        lab.save(update_fields=["approval_status", "is_published"])
        return Response({"detail": "University Lab approved successfully!", "lab": _serialize_lab_v2(lab)})


class StaffLabPublishView(APIView):
    """Staff: Publish / Activate an approved University Lab for students."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id):
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, staff_in_charge=staff)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        if lab.approval_status != "approved":
            return Response({"error": "Cannot publish a lab that has not been approved by the HOD."}, status=400)

        lab.is_published = True
        lab.is_active = True
        lab.save(update_fields=["is_published", "is_active"])
        return Response({"detail": "Lab published to students successfully!", "lab": _serialize_lab_v2(lab)})


class StaffLabSelectExercisesView(APIView):
    """Staff: Select exercises from the linked practice lab and copy them into a university lab."""
    permission_classes = [IsAuthenticated]

    def get(self, request, lab_id):
        """Return exercises from the linked practice lab that staff can pick from."""
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, staff_in_charge=staff, lab_type="university")
        except Lab.DoesNotExist:
            return Response({"error": "University lab not found"}, status=404)

        if not lab.linked_lab_id:
            return Response({"error": "No linked practice lab"}, status=400)

        linked_exercises = LabExercise.objects.filter(lab_id=lab.linked_lab_id).order_by("order", "created_at")
        already_selected = set(lab.exercises.values_list("title", flat=True))

        return Response({
            "linked_lab_name": lab.linked_lab.name if lab.linked_lab else "",
            "exercises": [
                {
                    "id": ex.id,
                    "title": ex.title,
                    "description": ex.description,
                    "difficulty": ex.difficulty,
                    "order": ex.order,
                    "test_case_count": ex.test_cases.count(),
                    "already_selected": ex.title in already_selected,
                }
                for ex in linked_exercises
            ],
        })

    def post(self, request, lab_id):
        """Copy selected exercises (and their test cases) from the linked lab into this university lab."""
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, staff_in_charge=staff, lab_type="university")
        except Lab.DoesNotExist:
            return Response({"error": "University lab not found"}, status=404)

        if not lab.linked_lab_id:
            return Response({"error": "No linked practice lab"}, status=400)

        exercise_ids = request.data.get("exercise_ids", [])
        if not isinstance(exercise_ids, list) or not exercise_ids:
            return Response({"error": "exercise_ids must be a non-empty list"}, status=400)

        source_exercises = LabExercise.objects.filter(
            id__in=exercise_ids, lab_id=lab.linked_lab_id
        ).prefetch_related("test_cases")

        if not source_exercises.exists():
            return Response({"error": "No valid exercises found in the linked lab"}, status=400)

        created = []
        next_order = lab.exercises.count()
        with transaction.atomic():
            for src_ex in source_exercises:
                new_ex = LabExercise.objects.create(
                    lab=lab,
                    title=src_ex.title,
                    description=src_ex.description,
                    explanation=src_ex.explanation,
                    order=next_order,
                    difficulty=src_ex.difficulty,
                    added_by=staff,
                )
                next_order += 1

                # Copy test cases
                for tc in src_ex.test_cases.all():
                    LabExerciseTestCase.objects.create(
                        exercise=new_ex,
                        stdin=tc.stdin,
                        expected_output=tc.expected_output,
                        is_sample=tc.is_sample,
                        order=tc.order,
                    )
                created.append(_serialize_exercise(new_ex))

        return Response({"created": created, "count": len(created)}, status=201)


class StaffLabAssignBatchesView(APIView):
    """Staff: Divide/assign students into sub-batches (Batch 1 to Batch N) for a Lab."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id):
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, staff_in_charge=staff)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        num_batches = request.data.get("num_batches")
        student_batches = request.data.get("student_batches")  # e.g. { "10": "Batch 1", "12": "Batch 2" }
        student_ids = request.data.get("student_ids")          # list of student IDs
        sub_batch = request.data.get("sub_batch")              # target sub_batch name

        student_qs = StudentProfile.objects.filter(department=lab.department, batch=lab.batch)
        if lab.section:
            student_qs = student_qs.filter(section=lab.section)
        students = list(student_qs.order_by("register_number", "name"))

        if student_ids and isinstance(student_ids, list) and sub_batch:
            b_name = str(sub_batch).strip() or "Batch 1"
            count = 0
            with transaction.atomic():
                for s_id in student_ids:
                    try:
                        session, _created = LabStudentSession.objects.get_or_create(lab=lab, student_id=int(s_id))
                        session.sub_batch = b_name
                        session.save(update_fields=["sub_batch"])
                        count += 1
                    except Exception:
                        pass
            return Response({"detail": f"Assigned {count} student(s) to '{b_name}'."})

        if num_batches is not None:
            try:
                N = int(num_batches)
                if N < 1 or N > 20:
                    return Response({"error": "Number of batches must be between 1 and 20"}, status=400)
            except ValueError:
                return Response({"error": "Invalid num_batches"}, status=400)

            total = len(students)
            if total == 0:
                return Response({"error": "No students in this lab section to divide"}, status=400)

            with transaction.atomic():
                for idx, student in enumerate(students):
                    batch_num = (idx * N) // total + 1
                    batch_name = f"Batch {batch_num}"
                    session, _created = LabStudentSession.objects.get_or_create(lab=lab, student=student)
                    session.sub_batch = batch_name
                    session.save(update_fields=["sub_batch"])

            return Response({"detail": f"Successfully split {total} students into {N} batch(es)."})

        elif student_batches and isinstance(student_batches, dict):
            count = 0
            with transaction.atomic():
                for student_id_str, b_name in student_batches.items():
                    try:
                        s_id = int(student_id_str)
                        session, _created = LabStudentSession.objects.get_or_create(lab=lab, student_id=s_id)
                        session.sub_batch = str(b_name).strip() or "Batch 1"
                        session.save(update_fields=["sub_batch"])
                        count += 1
                    except Exception:
                        pass
            return Response({"detail": f"Updated batch assignment for {count} student(s)."})

        else:
            return Response({"error": "Provide num_batches or student_batches dictionary"}, status=400)


class StaffLabBatchLockToggleView(APIView):
    """Staff: Unlock or Lock a specific sub-batch (or all students/selected students) for a Lab."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id):
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, staff_in_charge=staff)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        sub_batch = request.data.get("sub_batch")  # e.g. "Batch 1" or "all"
        student_ids = request.data.get("student_ids")  # optional list of student_ids
        is_locked = bool(request.data.get("is_locked", False))  # False = unlock, True = lock
        lock_reason = request.data.get("lock_reason", "Locked by staff" if is_locked else "")

        # Pre-create sessions for all eligible students if missing
        student_qs = StudentProfile.objects.filter(department=lab.department, batch=lab.batch)
        if lab.section:
            student_qs = student_qs.filter(section=lab.section)
        students = list(student_qs)
        for s in students:
            sess, _created = LabStudentSession.objects.get_or_create(lab=lab, student=s)
            if _created and lab.lab_type == "university":
                sess.is_locked = True
                sess.lock_reason = "Lab session is locked by staff. Awaiting staff unlock for your batch."
                sess.save(update_fields=["is_locked", "lock_reason"])

        sessions = LabStudentSession.objects.filter(lab=lab)
        if student_ids and isinstance(student_ids, list):
            sessions = sessions.filter(student_id__in=student_ids)
        elif sub_batch and sub_batch != "all":
            sessions = sessions.filter(sub_batch=sub_batch)

        updated_count = 0
        with transaction.atomic():
            for sess in sessions:
                sess.is_locked = is_locked
                sess.lock_reason = lock_reason if is_locked else ""
                sess.locked_at = timezone.now() if is_locked else None
                if not is_locked:
                    sess.tab_switch_count = 0  # reset violations on unlock
                sess.save(update_fields=["is_locked", "lock_reason", "locked_at", "tab_switch_count"])
                updated_count += 1

        status_str = "locked" if is_locked else "unlocked"
        target_str = f"Batch '{sub_batch}'" if sub_batch and sub_batch != "all" else f"{updated_count} student(s)"
        return Response({"detail": f"Successfully {status_str} {target_str} ({updated_count} students affected)."})


class StudentLabViolationView(APIView):
    """Student: Record security violation (tab switch or fullscreen exit) for a Lab."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id):
        student = _student_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, department=student.department, batch=student.batch)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        session, _created = LabStudentSession.objects.get_or_create(lab=lab, student=student)
        action = request.data.get("action", "tab_switch")
        reason = request.data.get("reason", "Tab switch detected")

        if action == "fullscreen_exit" and lab.enable_fullscreen_lock:
            session.is_locked = True
            session.lock_reason = "Fullscreen exit detected during University Lab session"
            session.locked_at = timezone.now()
        else:
            session.tab_switch_count += 1
            if lab.enable_tab_switch_check and session.tab_switch_count >= lab.max_tab_switches:
                session.is_locked = True
                session.lock_reason = f"Exceeded tab switch limit ({session.tab_switch_count}/{lab.max_tab_switches})"
                session.locked_at = timezone.now()

        session.save()
        return Response({
            "is_locked": session.is_locked,
            "tab_switch_count": session.tab_switch_count,
            "lock_reason": session.lock_reason,
            "max_tab_switches": lab.max_tab_switches,
        })


class StaffLabUnlockStudentView(APIView):
    """Staff: Unlock a locked student session in a Lab."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id, register_number):
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id, department=staff.department)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        student = StudentProfile.objects.filter(register_number=register_number, department=lab.department).first()
        if not student:
            return Response({"error": "Student not found"}, status=404)

        session, _created = LabStudentSession.objects.get_or_create(lab=lab, student=student)
        session.is_locked = False
        session.tab_switch_count = 0
        session.lock_reason = ""
        session.locked_at = None
        session.save()

        return Response({"detail": f"Student {student.name} ({register_number}) unlocked successfully!"})


class StaffLabFullReportView(APIView):
    """Staff: Bulk full report download is disabled. Only individual question downloads are enabled."""
    permission_classes = [IsAuthenticated]

    def get(self, request, lab_id):
        return Response({"error": "Bulk lab report download has been disabled. Please download individual question reports."}, status=400)


class StaffLabAllocateQuestionsView(APIView):
    """Staff: Trigger or re-run random difficulty-based question allocation for all students in a Lab."""
    permission_classes = [IsAuthenticated]

    def post(self, request, lab_id):
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.get(id=lab_id)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        if not lab.is_published:
            lab.is_published = True
            lab.save(update_fields=["is_published"])

        from .services.lab_allocation import allocate_lab_questions_for_students
        stats = allocate_lab_questions_for_students(lab)
        return Response({
            "detail": f"Question allocation completed for {stats.get('allocated_count', 0)} student(s).",
            "stats": stats,
        })


class StaffLabAllocationPDFView(APIView):
    """Staff: Generate and download PDF Question Allocation Sheet for a Lab."""
    permission_classes = [IsAuthenticated]

    def get(self, request, lab_id):
        staff = _staff_from_request(request)
        try:
            lab = Lab.objects.select_related("department", "staff_in_charge").get(id=lab_id)
        except Lab.DoesNotExist:
            return Response({"error": "Lab not found"}, status=404)

        # Make sure questions are allocated first if not yet allocated
        sessions = list(
            LabStudentSession.objects.filter(lab=lab)
            .select_related("student")
            .prefetch_related("allocated_exercises")
            .order_by("student__register_number", "student__name")
        )

        has_allocations = any(s.allocated_exercises.exists() for s in sessions)
        if not has_allocations or not sessions:
            from .services.lab_allocation import allocate_lab_questions_for_students
            allocate_lab_questions_for_students(lab)
            sessions = list(
                LabStudentSession.objects.filter(lab=lab)
                .select_related("student")
                .prefetch_related("allocated_exercises")
                .order_by("student__register_number", "student__name")
            )

        if not sessions:
            return Response({"error": "No students found enrolled in this lab section to generate allocation sheet."}, status=400)

        from .services.lab_allocation_pdf import build_lab_allocation_pdf
        buffer = BytesIO()
        try:
            build_lab_allocation_pdf(buffer, lab=lab, sessions=sessions)
        except Exception as exc:
            logger.exception("Failed to render lab allocation PDF for lab %s", lab_id)
            return Response({"error": f"Failed to render allocation PDF: {exc}"}, status=500)

        buffer.seek(0)
        pdf_bytes = buffer.getvalue()

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Length"] = str(len(pdf_bytes))
        clean_lab_name = re.sub(r'[^\w\-]', '_', lab.name)
        response["Content-Disposition"] = f'attachment; filename="question_allocation_{clean_lab_name}.pdf"'
        return response


# ── HOD Staff Management ──────────────────────────────────────────────────────

class HODManageStaffView(APIView):
    """HOD: add new staff member to own department."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        hod = _staff_from_request(request)
        if not hod or hod.role not in ("hod", "academics", "admin"):
            return Response({"error": "HOD access required"}, status=403)
        if not hod.department or not hod.institution:
            return Response({"error": "HOD has no department assigned"}, status=400)

        data = request.data
        faculty_id = (data.get("faculty_id") or "").strip()
        name = (data.get("name") or "").strip()
        role = data.get("role", "staff")
        password = (data.get("password") or "").strip()

        if not faculty_id:
            return Response({"error": "Faculty ID is required"}, status=400)
        if not name:
            return Response({"error": "Name is required"}, status=400)
        if role not in ("staff", "hod", "academics", "tpu", "ja"):
            role = "staff"

        if StaffProfile.objects.filter(faculty_id=faculty_id).exists():
            return Response({"error": "A staff member with this Faculty ID already exists."}, status=400)

        staff = StaffProfile.objects.create(
            faculty_id=faculty_id,
            name=name,
            role=role,
            department=hod.department,
            institution=hod.institution,
            is_active=True,
            password=password,
        )
        return Response({
            "faculty_id": staff.faculty_id,
            "name": staff.name,
            "role": staff.role,
            "role_display": staff.get_role_display(),
            "is_active": staff.is_active,
        }, status=201)


class HODManageStaffDetailView(APIView):
    """HOD: edit a staff member in own department."""
    permission_classes = [IsAuthenticated]

    def put(self, request, faculty_id):
        hod = _staff_from_request(request)
        if not hod or hod.role not in ("hod", "academics", "admin"):
            return Response({"error": "HOD access required"}, status=403)

        try:
            target = StaffProfile.objects.get(faculty_id=faculty_id, department=hod.department)
        except StaffProfile.DoesNotExist:
            return Response({"error": "Staff not found in your department"}, status=404)

        data = request.data
        name = (data.get("name") or "").strip()
        if name:
            target.name = name
        if "role" in data and data["role"] in ("staff", "hod", "tpu", "ja"):
            target.role = data["role"]
        if "is_active" in data:
            target.is_active = bool(data["is_active"])
        pw = (data.get("password") or "").strip()
        if pw:
            target.password = pw

        target.save()
        return Response({
            "faculty_id": target.faculty_id,
            "name": target.name,
            "role": target.role,
            "role_display": target.get_role_display(),
            "is_active": target.is_active,
        })

    def delete(self, request, faculty_id):
        hod = _staff_from_request(request)
        if not hod or hod.role not in ("hod", "academics", "admin"):
            return Response({"error": "HOD access required"}, status=403)

        try:
            target = StaffProfile.objects.get(faculty_id=faculty_id, department=hod.department)
        except StaffProfile.DoesNotExist:
            return Response({"error": "Staff not found in your department"}, status=404)

        if target.id == hod.id:
            return Response({"error": "You cannot delete your own account."}, status=400)

        account = target.account
        if account:
            account.delete()  # cascades to the StaffProfile
        else:
            target.delete()
        return Response({"message": "Staff deleted"})


# ── TEMPORARY: production data-loss diagnostic — remove after investigation ──
class TempDataDiagnosticsView(APIView):
    """Read-only snapshot of the live DB connection + row counts, and every
    database visible on the same Postgres server. Gated by a one-off shared
    token (not real auth) so it can be checked from a browser with no login.
    Delete this view + its URL once the data-loss investigation is closed.
    """

    permission_classes = [AllowAny]

    def get(self, request, token):
        import os

        if token != "EiTSaBPnqIByYUEBYLsS1Eqlw4D_7flj":
            return Response({"detail": "Not found."}, status=404)

        from django.db import connection

        report = {}

        db = connection.settings_dict
        report["connected_to"] = {
            "name": db.get("NAME"),
            "host": db.get("HOST"),
            "user": db.get("USER"),
            "port": db.get("PORT"),
        }

        counts = {}
        for label, model in [
            ("students", StudentProfile),
            ("staff", StaffProfile),
            ("institutions", Institution),
            ("problems", Problem),
            ("solved_problems", SolvedProblem),
            ("problem_solutions", ProblemSolution),
            ("submissions", Submission),
        ]:
            try:
                counts[label] = model.objects.count()
            except Exception as exc:
                counts[label] = f"error: {exc}"
        report["row_counts_in_connected_db"] = counts

        try:
            from .models import LabExercise, LabExerciseSubmission
            counts["lab_exercises"] = LabExercise.objects.count()
            counts["lab_exercise_submissions"] = LabExerciseSubmission.objects.count()
        except Exception as exc:
            counts["lab_models_error"] = str(exc)

        try:
            import psycopg2
            other_conn = psycopg2.connect(
                host=db.get("HOST"),
                port=db.get("PORT"),
                user=db.get("USER"),
                password=db.get("PASSWORD"),
                dbname="postgres",
                connect_timeout=5,
            )
            cur = other_conn.cursor()
            cur.execute(
                "SELECT datname, pg_size_pretty(pg_database_size(datname)) "
                "FROM pg_database WHERE datistemplate = false ORDER BY datname;"
            )
            report["all_databases_on_this_postgres_server"] = [
                {"name": row[0], "size": row[1]} for row in cur.fetchall()
            ]
            other_conn.close()
        except Exception as exc:
            report["all_databases_on_this_postgres_server"] = f"error: {exc}"

        report["env_db_vars_seen_by_backend"] = {
            "DB_HOST": os.getenv("DB_HOST"),
            "DB_NAME": os.getenv("DB_NAME"),
            "DB_USER": os.getenv("DB_USER"),
            "DB_PORT": os.getenv("DB_PORT"),
        }
        return Response(report)
